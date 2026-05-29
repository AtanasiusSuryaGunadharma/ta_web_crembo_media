from crembo_app.services import core as _core

# Memuat seluruh helper, service, dan objek Flask dari core agar potongan kode route
# tetap kompatibel setelah dipisah dari app.py monolitik.
globals().update({
    name: getattr(_core, name)
    for name in dir(_core)
    if not (name.startswith("__") and name.endswith("__"))
})

# Controller: Damage Controller

# Source legacy app.py lines 2316-2347 | routes: /api/damage/upload
@app.route("/api/damage/upload", methods=["POST"])
def upload_damage_photo():
    """Upload damage report photos"""
    ensure_auth_schema()
    incoming_files = request.files.getlist("files")
    if not incoming_files:
        single_file = request.files.get("file")
        incoming_files = [single_file] if single_file and single_file.filename else []

    if not incoming_files:
        return jsonify({"success": False, "error": "Tidak ada file yang dipilih"}), 400

    saved_files: list[dict[str, object]] = []
    try:
        for file in incoming_files:
            if not file or not file.filename:
                continue
            # Validate file extension for images only
            file_ext = Path(file.filename).suffix.lower()
            if file_ext not in {".jpg", ".jpeg", ".png", ".gif", ".webp", ".bmp"}:
                return jsonify({"success": False, "error": f"Tipe file {file_ext} tidak diizinkan. Gunakan JPG, PNG, GIF, WebP, atau BMP"}), 400
            saved_files.append(save_uploaded_attachment(file))
    except ValueError as exc:
        return jsonify({"success": False, "error": str(exc)}), 400

    if not saved_files:
        return jsonify({"success": False, "error": "Gagal menyimpan file"}), 400

    return jsonify({
        "success": True,
        "files": saved_files,
    })


# Source legacy app.py lines 2349-2443 | routes: /api/kerusakan
@app.route("/api/kerusakan", methods=["POST"])
def submit_damage_report():
    """Submit damage report form"""
    ensure_auth_schema()
    user_context = current_user_context()
    
    if not user_context.get("logged_in"):
        return jsonify({"success": False, "error": "Anda harus login terlebih dahulu"}), 401
    
    member_id = user_context.get("user_id")
    
    # Menangani form data biasa karena request dari FE dikirim via FormData
    barang_id = request.form.get("barangId")
    barang_name = normalize_text(request.form.get("itemName"), "")
    barang_code = normalize_text(request.form.get("itemCode"), "-")
    tingkat_kerusakan = normalize_text(request.form.get("severity"), "Sedang")
    deskripsi = normalize_text(request.form.get("chronology"), "")
    waktu_kejadian = request.form.get("incidentDate")
    incident_time = request.form.get("incidentTime")
    
    if not barang_name:
        return jsonify({"success": False, "error": "Nama barang harus diisi"}), 400
    
    if not deskripsi:
        return jsonify({"success": False, "error": "Deskripsi kerusakan harus diisi"}), 400
    
    if tingkat_kerusakan not in ["Ringan", "Sedang", "Berat", "Hilang"]:
        tingkat_kerusakan = "Sedang"
    
    # Process the photo if it exists in the form data
    foto_url = None
    photo_file = request.files.get("photo")
    if photo_file and photo_file.filename:
        try:
            saved = save_uploaded_attachment(photo_file)
            foto_url = saved.get("url")
        except ValueError as exc:
            return jsonify({"success": False, "error": str(exc)}), 400
    
    # Create damage report ID
    damage_id = f"krk-{int(time.time() * 1000)}-{uuid.uuid4().hex[:6]}"
    
    conn = mysql_connection()
    cursor = conn.cursor()
    try:
        ensure_damage_schema(cursor)
        
        # Simpan member details untuk referensi
        member_name = str(session.get("nama") or session.get("username") or "Anggota")
        member_identifier = str(session.get("username") or session.get("email"))
        
        cursor.execute("""
            INSERT INTO `form_kerusakan_barang` 
            (`id`, `member_id`, `barang_id`, `barang_name`, `barang_code`, `tingkat_kerusakan`, `status`, `deskripsi_kerusakan`, `waktu_kejadian`, `foto_kerusakan`)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            damage_id,
            str(member_id),
            barang_id if barang_id else None,
            barang_name,
            barang_code,
            tingkat_kerusakan,
            "Pending Review",
            deskripsi,
            f"{waktu_kejadian} {incident_time}" if waktu_kejadian and incident_time else datetime.now(),
            json.dumps([{"url": foto_url, "name": photo_file.filename}] if foto_url else [], ensure_ascii=False)
        ))
        
        # Create admin notification
        try:
            ensure_notifications_schema()
            create_notification(
                cursor, 
                "kerusakan",
                "Laporan Kerusakan Barang Baru",
                f"Laporan kerusakan barang dari {member_name}: {barang_name}",
                f"/hasil-form-kerusakan-barang.html",
                {"report_id": damage_id, "member_id": member_id, "barang_name": barang_name},
                target_role="admin"
            )
        except Exception:
            pass
            
        conn.commit()
        return jsonify({
            "success": True,
            "damageId": damage_id,
            "message": "Laporan kerusakan barang berhasil dikirim"
        })
    except Exception as e:
        conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 400
    finally:
        cursor.close()
        conn.close()


# Source legacy app.py lines 2445-2513 | routes: /api/form-kerusakan/history
@app.route("/api/form-kerusakan/history", methods=["GET"])
def get_damage_history():
    """Get damage report history"""
    ensure_auth_schema()
    user_context = current_user_context()
    
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        ensure_damage_schema(cursor)
        
        role = user_context.get("role", "")
        member_id = user_context.get("user_id")
        
        query = "SELECT d.*, a.nama as member_name FROM `form_kerusakan_barang` d LEFT JOIN `anggota` a ON d.member_id = a.id WHERE 1=1"
        params = []
        
        # Non-admin users only see their own damage reports
        if role not in ["admin", "super_admin"]:
            if not member_id:
                return jsonify({"success": True, "items": []})
            query += " AND d.`member_id` = %s"
            params.append(str(member_id))
        
        # Apply filters
        status_filter = request.args.get('status')
        severity_filter = request.args.get('severity')
        search_query = request.args.get('search', '').strip()
        
        if status_filter and status_filter != 'all':
            query += " AND d.`status` = %s"
            params.append(status_filter)
        
        if severity_filter and severity_filter != 'all':
            query += " AND d.`tingkat_kerusakan` = %s"
            params.append(severity_filter)
        
        if search_query:
            query += " AND (d.`barang_name` LIKE %s OR d.`deskripsi_kerusakan` LIKE %s OR a.`nama` LIKE %s)"
            search_param = f"%{search_query}%"
            params.extend([search_param, search_param, search_param])
        
        query += " ORDER BY d.`created_at` DESC"
        
        cursor.execute(query, tuple(params))
        rows = cursor.fetchall() or []
        
        items = []
        for r in rows:
            items.append({
                "id": r.get("id"),
                "memberId": r.get("member_id"),
                "memberName": r.get("member_name"),
                "barangId": r.get("barang_id"),
                "barangName": r.get("barang_name"),
                "barangCode": r.get("barang_code"),
                "tingkatKerusakan": r.get("tingkat_kerusakan"),
                "status": r.get("status"),
                "deskripsiKerusakan": r.get("deskripsi_kerusakan"),
                "waktuKejadian": r.get("waktu_kejadian").isoformat() if r.get("waktu_kejadian") else None,
                "fotoKerusakan": json.loads(r.get("foto_kerusakan") or "[]"),
                "createdAt": r.get("created_at").isoformat() if r.get("created_at") else None,
                "updatedAt": r.get("updated_at").isoformat() if r.get("updated_at") else None,
            })
        
        return jsonify({"success": True, "items": items})
    finally:
        cursor.close()
        conn.close()


# Source legacy app.py lines 2515-2563 | routes: /api/form-kerusakan/<damage_id>
@app.route("/api/form-kerusakan/<damage_id>", methods=["GET"])
def get_damage_detail(damage_id):
    """Get single damage report detail"""
    ensure_auth_schema()
    user_context = current_user_context()
    
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        ensure_damage_schema(cursor)
        
        cursor.execute("""
            SELECT d.*, a.nama as member_name FROM `form_kerusakan_barang` d
            LEFT JOIN `anggota` a ON d.member_id = a.id
            WHERE d.`id` = %s LIMIT 1
        """, (damage_id,))
        
        row = cursor.fetchone()
        if not row:
            return jsonify({"success": False, "error": "Laporan tidak ditemukan"}), 404
        
        # Check access permission
        role = user_context.get("role", "")
        member_id = str(user_context.get("user_id"))
        
        if role not in ["admin", "super_admin"] and str(row.get("member_id")) != member_id:
            return jsonify({"success": False, "error": "Akses ditolak"}), 403
        
        return jsonify({
            "success": True,
            "item": {
                "id": row.get("id"),
                "memberId": row.get("member_id"),
                "memberName": row.get("member_name"),
                "barangId": row.get("barang_id"),
                "barangName": row.get("barang_name"),
                "barangCode": row.get("barang_code"),
                "tingkatKerusakan": row.get("tingkat_kerusakan"),
                "status": row.get("status"),
                "deskripsiKerusakan": row.get("deskripsi_kerusakan"),
                "waktuKejadian": row.get("waktu_kejadian").isoformat() if row.get("waktu_kejadian") else None,
                "fotoKerusakan": json.loads(row.get("foto_kerusakan") or "[]"),
                "createdAt": row.get("created_at").isoformat() if row.get("created_at") else None,
                "updatedAt": row.get("updated_at").isoformat() if row.get("updated_at") else None,
            }
        })
    finally:
        cursor.close()
        conn.close()


# Source legacy app.py lines 2565-2628 | routes: /api/form-kerusakan/<damage_id>/status
@app.route("/api/form-kerusakan/<damage_id>/status", methods=["PUT"])
def update_damage_status(damage_id):
    """Update damage report status (admin only)"""
    ensure_auth_schema()
    user_context = current_user_context()
    role = user_context.get("role", "")
    
    if role not in ["admin", "super_admin"]:
        return jsonify({"success": False, "error": "Anda tidak memiliki izin untuk operasi ini"}), 403
    
    data = request.get_json(silent=True) or {}
    new_status = data.get("status", "").strip()
    
    # HANYA 3 STATUS: Pending Review, Diproses, Selesai (Ditolak dihapus)
    if new_status not in ["Pending Review", "Diproses", "Selesai"]:
        return jsonify({"success": False, "error": "Status tidak valid"}), 400
    
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        ensure_damage_schema(cursor)
        
        cursor.execute("SELECT `member_id` FROM `form_kerusakan_barang` WHERE `id` = %s LIMIT 1", (damage_id,))
        row = cursor.fetchone()
        if not row:
            return jsonify({"success": False, "error": "Laporan tidak ditemukan"}), 404
        
        cursor.execute("""
            UPDATE `form_kerusakan_barang` 
            SET `status` = %s, `updated_at` = NOW()
            WHERE `id` = %s
        """, (new_status, damage_id))
        
        try:
            ensure_notifications_schema()
            cursor.execute("SELECT `nama` FROM `anggota` WHERE `id` = %s LIMIT 1", (row.get("member_id"),))
            member_row = cursor.fetchone()
            
            status_indo = {
                "Pending Review": "Dalam Review",
                "Diproses": "Sedang Diproses",
                "Selesai": "Selesai"
            }.get(new_status, new_status)
            
            create_notification(
                cursor,
                "kerusakan",
                "Status Laporan Kerusakan Diperbarui",
                f"Status laporan kerusakan Anda telah diubah menjadi: <b>{status_indo}</b>",
                f"/riwayat-form-kerusakan-barang-anggota.html",
                {"report_id": damage_id, "target_user_id": row.get("member_id")},
                target_role="user"
            )
        except Exception as e:
            print(f"[WARN] Gagal mengirim notifikasi update status kerusakan: {e}")
        
        conn.commit()
        return jsonify({"success": True})
    except Exception as e:
        conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 400
    finally:
        cursor.close()
        conn.close()


# Source legacy app.py lines 9728-9854 | routes: /api/form-kerusakan/export.xlsx
@app.route("/api/form-kerusakan/export.xlsx", methods=["GET"])
def export_kerusakan_excel():
    ensure_auth_schema()
    user_context = current_user_context()
    role = user_context.get("role", "")
    
    if role not in ["admin", "super_admin"]:
        abort(403)
        
    search_query = request.args.get('search', '').strip()
    status_filter = request.args.get('status')
    severity_filter = request.args.get('severity')
    
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        ensure_damage_schema(cursor)
        
        query = "SELECT d.*, a.nama as member_name FROM `form_kerusakan_barang` d LEFT JOIN `anggota` a ON d.member_id = a.id WHERE 1=1"
        params = []
        
        if status_filter and status_filter != 'all':
            query += " AND d.`status` = %s"
            params.append(status_filter)
        
        if severity_filter and severity_filter != 'all':
            query += " AND d.`tingkat_kerusakan` = %s"
            params.append(severity_filter)
        
        if search_query:
            query += " AND (d.`barang_name` LIKE %s OR d.`deskripsi_kerusakan` LIKE %s OR a.`nama` LIKE %s)"
            search_param = f"%{search_query}%"
            params.extend([search_param, search_param, search_param])
        
        query += " ORDER BY d.`created_at` DESC"
        
        cursor.execute(query, tuple(params))
        items = cursor.fetchall() or []
        
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        
        # Susunan Header baru
        headers = [
            "ID Laporan", "Barang", "Kode Barang", "Pelapor", 
            "Tingkat Kerusakan", "Deskripsi Kerusakan", "Waktu Kejadian", "Diinput", "Foto Bukti", "Status"
        ]
        
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Laporan Kerusakan"
        worksheet.append(headers)
        
        base_url = PUBLIC_BASE_URL
        
        for item in items:
            waktu_kej = item.get("waktu_kejadian").strftime("%Y-%m-%d %H:%M:%S") if item.get("waktu_kejadian") else "-"
            diinput = item.get("created_at").strftime("%Y-%m-%d %H:%M:%S") if item.get("created_at") else "-"
            
            # Format foto ke full URL
            foto_links = []
            foto_raw = safe_json_loads(item.get("foto_kerusakan"), [])
            if foto_raw:
                for f in foto_raw:
                    if isinstance(f, dict) and f.get("url"):
                        url = str(f.get("url"))
                        if url.startswith("/"): url = f"{base_url}{url}"
                        foto_links.append(url)
            foto_label = "\n".join(foto_links) if foto_links else "-"
            
            row = [
                item.get("id"),
                item.get("barang_name") or "-",
                item.get("barang_code") or "-",
                item.get("member_name") or "-",
                item.get("tingkat_kerusakan") or "-",
                item.get("deskripsi_kerusakan") or "-",
                waktu_kej,
                diinput,
                foto_label,
                item.get("status") or "-",
            ]
            worksheet.append(row)
            
        header_fill = PatternFill("solid", fgColor="7F0000")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2), 2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrapText=True)

        for column_cells in worksheet.columns:
            max_length = 0
            column_letter = column_cells[0].column_letter
            for cell in column_cells:
                try:
                    cell_value = str(cell.value or "")
                    if "\n" in cell_value:
                        lines = cell_value.split("\n")
                        line_lengths = [len(l) for l in lines]
                        if line_lengths and max(line_lengths) > max_length:
                            max_length = max(line_lengths)
                    elif len(cell_value) > max_length:
                        max_length = len(cell_value)
                except Exception:
                    continue
            worksheet.column_dimensions[column_letter].width = min(max_length + 4, 60)

        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        
        filename = f"laporan-kerusakan-{datetime.now().strftime('%Y%m%d%H%M')}.xlsx"
        
        return send_file(
            buffer,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    finally:
        cursor.close()
        conn.close()


# Source legacy app.py lines 9856-10005 | routes: /api/form-kerusakan/export.pdf
@app.route("/api/form-kerusakan/export.pdf", methods=["GET"])
def export_kerusakan_pdf():
    ensure_auth_schema()
    user_context = current_user_context()
    role = user_context.get("role", "")
    
    if role not in ["admin", "super_admin"]:
        abort(403)
        
    search_query = request.args.get('search', '').strip()
    status_filter = request.args.get('status')
    severity_filter = request.args.get('severity')
    
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        ensure_damage_schema(cursor)
        
        query = "SELECT d.*, a.nama as member_name FROM `form_kerusakan_barang` d LEFT JOIN `anggota` a ON d.member_id = a.id WHERE 1=1"
        params = []
        
        if status_filter and status_filter != 'all':
            query += " AND d.`status` = %s"
            params.append(status_filter)
        
        if severity_filter and severity_filter != 'all':
            query += " AND d.`tingkat_kerusakan` = %s"
            params.append(severity_filter)
        
        if search_query:
            query += " AND (d.`barang_name` LIKE %s OR d.`deskripsi_kerusakan` LIKE %s OR a.`nama` LIKE %s)"
            search_param = f"%{search_query}%"
            params.extend([search_param, search_param, search_param])
        
        query += " ORDER BY d.`created_at` DESC"
        
        cursor.execute(query, tuple(params))
        items = cursor.fetchall() or []
        
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Spacer, Table, TableStyle, Paragraph

        # Susunan header baru di PDF
        headers = ["Barang", "Pelapor", "Tingkat", "Deskripsi", "Waktu", "Diinput", "Foto Bukti", "Status"]
        rows = []
        base_url = PUBLIC_BASE_URL
        
        for item in items:
            waktu_kej = item.get("waktu_kejadian").strftime("%Y-%m-%d %H:%M") if item.get("waktu_kejadian") else "-"
            diinput = item.get("created_at").strftime("%Y-%m-%d %H:%M") if item.get("created_at") else "-"
            
            foto_links = []
            foto_raw = safe_json_loads(item.get("foto_kerusakan"), [])
            if foto_raw:
                for f in foto_raw:
                    if isinstance(f, dict) and f.get("url"):
                        url = str(f.get("url"))
                        if url.startswith("/"): url = f"{base_url}{url}"
                        # Kita bungkus url foto menggunakan tag link agar bisa di klik di PDF
                        foto_links.append(f"<a href='{url}' color='blue'>{url}</a>")
            foto_label = "<br/>".join(foto_links) if foto_links else "-"

            rows.append([
                item.get("barang_name") or "-",
                item.get("member_name") or "-",
                item.get("tingkat_kerusakan") or "-",
                item.get("deskripsi_kerusakan") or "-",
                waktu_kej,
                diinput,
                foto_label,
                item.get("status") or "-"
            ])
            
        buffer = BytesIO()
        document = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            rightMargin=8 * mm,
            leftMargin=8 * mm,
            topMargin=10 * mm,
            bottomMargin=10 * mm,
        )
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle("DamageTitle", parent=styles["Title"], fontName="Helvetica-Bold", fontSize=14, leading=16, textColor=colors.HexColor("#7F0000"))
        normal_style = ParagraphStyle("DamageNormal", parent=styles["BodyText"], fontName="Helvetica", fontSize=8, leading=10)
        cell_style = ParagraphStyle("CellNormal", parent=styles["BodyText"], fontName="Helvetica", fontSize=7, leading=9, wordWrap='CJK')

        elements = [
            Paragraph("Rekap Data Laporan Kerusakan Barang", title_style),
            Spacer(1, 4 * mm),
            Paragraph(f"Dicetak pada: {datetime.now().strftime('%d %B %Y %H:%M')} WIB", normal_style),
            Spacer(1, 4 * mm),
        ]

        wrapped_rows = []
        for row in rows:
            wrapped_row = []
            for cell_data in row:
                text = str(cell_data).replace('\n', '<br/>')
                wrapped_row.append(Paragraph(text, cell_style))
            wrapped_rows.append(wrapped_row)

        table_data = [headers] + wrapped_rows if rows else [headers, ["-" for _ in headers]]
        
        page_width = landscape(A4)[0] - (16 * mm) 
        # Sesuaikan proporsi persentase ke-8 kolom 
        col_widths = [
            0.15 * page_width,
            0.12 * page_width,
            0.08 * page_width,
            0.20 * page_width,
            0.10 * page_width,
            0.10 * page_width,
            0.15 * page_width,
            0.10 * page_width
        ]
        
        table = Table(table_data, colWidths=col_widths, repeatRows=1)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#7F0000")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 7),
                    ("ALIGN", (0, 0), (0, -1), "CENTER"),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D1D5DB")),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.HexColor("#FFFFFF")]),
                ]
            )
        )
        elements.append(table)
        document.build(elements)
        buffer.seek(0)
        
        filename = f"laporan-kerusakan-{datetime.now().strftime('%Y%m%d%H%M')}.pdf"
        
        return send_file(
            buffer,
            as_attachment=False,  # FALSE agar terbuka di tab browser dulu (preview) sebelum didownload
            download_name=filename,
            mimetype="application/pdf",
        )
    finally:
        cursor.close()
        conn.close()


