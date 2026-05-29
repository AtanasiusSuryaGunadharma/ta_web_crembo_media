from crembo_app.services import core as _core

# Memuat seluruh helper, service, dan objek Flask dari core agar potongan kode route
# tetap kompatibel setelah dipisah dari app.py monolitik.
globals().update({
    name: getattr(_core, name)
    for name in dir(_core)
    if not (name.startswith("__") and name.endswith("__"))
})

# Controller: Evaluation Controller

# Source legacy app.py lines 14986-15003 | routes: /api/evaluasi-streaming/settings
@app.route("/api/evaluasi-streaming/settings", methods=["GET", "POST"])
def api_streaming_evaluation_settings():
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True, buffered=True)
    try:
        ensure_streaming_evaluation_schema(cursor)
        if request.method == "POST":
            if normalize_role_value(session.get("role") or "") not in {"admin", "super_admin"}:
                return jsonify({"success": False, "error": "Akses ditolak."}), 403
            data = request.get_json(silent=True) or {}
            month = max(1, min(12, parse_required_int(data.get("startMonth"), 5)))
            year = max(2000, parse_required_int(data.get("startYear"), 2026))
            cursor.execute("UPDATE streaming_evaluation_settings SET start_month = %s, start_year = %s WHERE id = 1", (month, year))
            conn.commit()
        return jsonify({"success": True, "settings": eval_get_settings(cursor)})
    finally:
        cursor.close()
        conn.close()


# Source legacy app.py lines 15006-15032 | routes: /api/evaluasi-streaming/questions
@app.route("/api/evaluasi-streaming/questions", methods=["GET", "POST"])
def api_streaming_evaluation_questions():
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True, buffered=True)
    try:
        ensure_streaming_evaluation_schema(cursor)
        if request.method == "POST":
            if normalize_role_value(session.get("role") or "") not in {"admin", "super_admin"}:
                return jsonify({"success": False, "error": "Akses ditolak."}), 403
            data = request.get_json(silent=True) or {}
            questions = eval_normalize_question_payload(data.get("questions"))
            # Boleh kosong: pertanyaan tambahan evaluasi tidak wajib ada.
            cursor.execute("DELETE FROM streaming_evaluation_questions")
            for idx, q in enumerate(questions, start=1):
                cursor.execute(
                    """
                    INSERT INTO streaming_evaluation_questions
                    (id, label, question_type, required, help_text, options_json, order_index)
                    VALUES (%s, %s, %s, %s, %s, %s, %s)
                    """,
                    (q["id"], q["label"], q["type"], 1 if q.get("required") else 0, q.get("helpText") or "", json.dumps(q.get("options") or [], ensure_ascii=False), idx),
                )
            conn.commit()
        return jsonify({"success": True, "questions": eval_fetch_questions(cursor)})
    finally:
        cursor.close()
        conn.close()


# Source legacy app.py lines 15035-15057 | routes: /api/evaluasi-streaming/questions/reset
@app.route("/api/evaluasi-streaming/questions/reset", methods=["POST"])
def api_streaming_evaluation_questions_reset():
    if normalize_role_value(session.get("role") or "") not in {"admin", "super_admin"}:
        return jsonify({"success": False, "error": "Akses ditolak."}), 403
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True, buffered=True)
    try:
        ensure_streaming_evaluation_schema(cursor)
        cursor.execute("DELETE FROM streaming_evaluation_questions")
        for idx, q in enumerate(EVAL_DEFAULT_QUESTIONS, start=1):
            cursor.execute(
                """
                INSERT INTO streaming_evaluation_questions
                (id, label, question_type, required, help_text, options_json, order_index)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
                """,
                (q["id"], q["label"], q["type"], 1 if q.get("required") else 0, q.get("helpText") or "", json.dumps(q.get("options") or [], ensure_ascii=False), idx),
            )
        conn.commit()
        return jsonify({"success": True, "questions": eval_fetch_questions(cursor)})
    finally:
        cursor.close()
        conn.close()


# Source legacy app.py lines 15060-15069 | routes: /api/evaluasi-streaming/members
@app.route("/api/evaluasi-streaming/members", methods=["GET"])
def api_streaming_evaluation_members():
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True, buffered=True)
    try:
        ensure_auth_schema()
        return jsonify({"success": True, "members": eval_fetch_active_members(cursor), "currentUser": current_user_context()})
    finally:
        cursor.close()
        conn.close()


# Source legacy app.py lines 15072-15118 | routes: /api/evaluasi-streaming/schedules
@app.route("/api/evaluasi-streaming/schedules", methods=["GET"])
def api_streaming_evaluation_schedules():
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True, buffered=True)
    try:
        ensure_streaming_evaluation_schema(cursor)
        now = datetime.now()
        settings = eval_get_settings(cursor)
        start_setting = eval_start_date_from_settings(settings)
        kind = normalize_text(request.args.get("kind") or "all")
        mode = normalize_text(request.args.get("mode") or "form")
        include_evaluated = normalize_text(request.args.get("includeEvaluated") or "0") in {"1", "true", "yes"}
        # Search a useful window: from configured start until now for public form; current year for member/admin if requested.
        start_date = parse_optional_date(request.args.get("startDate"))
        end_date = parse_optional_date(request.args.get("endDate"))
        if start_date:
            start = datetime.strptime(start_date, "%Y-%m-%d").date()
        else:
            start = start_setting
        if end_date:
            end = datetime.strptime(end_date, "%Y-%m-%d").date()
        else:
            end = now.date()
        schedules = eval_all_schedules(cursor, start, end, kind)
        evals = eval_evaluation_map(cursor, start.isoformat(), end.isoformat())
        sort_mode = normalize_text(request.args.get("sort") or "date_asc")
        rows = []
        for schedule in schedules:
            dt = eval_datetime_from_parts(schedule.get("date"), schedule.get("time"))
            is_due = bool(dt and dt <= now)
            has_staff = eval_schedule_has_staff(schedule)
            ev = evals.get((schedule.get("kind"), schedule.get("scheduleKey")))
            schedule["evaluated"] = bool(ev)
            schedule["evaluationId"] = ev.get("id") if ev else None
            schedule["due"] = is_due
            schedule["hasStaff"] = has_staff
            if mode == "form" and (not is_due or not has_staff or (ev and not include_evaluated)):
                continue
            rows.append(schedule)
        if sort_mode in {"date_desc", "newest", "terbaru"}:
            rows.sort(key=lambda item: (item.get("date") or "", item.get("time") or "", item.get("misaName") or ""), reverse=True)
        else:
            rows.sort(key=lambda item: (item.get("date") or "", item.get("time") or "", item.get("misaName") or ""))
        return jsonify({"success": True, "schedules": rows, "questions": eval_fetch_questions(cursor), "settings": settings, "currentUser": current_user_context()})
    finally:
        cursor.close()
        conn.close()


# Source legacy app.py lines 15121-15136 | routes: /api/evaluasi-streaming/schedule-detail
@app.route("/api/evaluasi-streaming/schedule-detail", methods=["GET"])
def api_streaming_evaluation_schedule_detail():
    schedule_id = normalize_text(request.args.get("scheduleId"))
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True, buffered=True)
    try:
        ensure_streaming_evaluation_schema(cursor)
        schedule = eval_find_schedule(cursor, schedule_id)
        if not schedule:
            return jsonify({"success": False, "error": "Jadwal tidak ditemukan."}), 404
        ev_map = eval_evaluation_map(cursor, schedule.get("date"), schedule.get("date"))
        ev = ev_map.get((schedule.get("kind"), schedule.get("scheduleKey")))
        return jsonify({"success": True, "schedule": schedule, "evaluation": eval_row_to_dict(ev) if ev else None, "questions": eval_fetch_questions(cursor), "members": eval_fetch_active_members(cursor), "currentUser": current_user_context()})
    finally:
        cursor.close()
        conn.close()


# Source legacy app.py lines 15139-15234 | routes: /api/evaluasi-streaming/submit
@app.route("/api/evaluasi-streaming/submit", methods=["POST"])
def api_streaming_evaluation_submit():
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True, buffered=True)
    try:
        ensure_streaming_evaluation_schema(cursor)
        payload = request.get_json(silent=True) or {}
        valid, message = eval_validate_submit_payload(payload)
        if not valid:
            return jsonify({"success": False, "error": message}), 400
        schedule = eval_find_schedule(cursor, normalize_text(payload.get("scheduleId")))
        if not schedule:
            return jsonify({"success": False, "error": "Jadwal tidak ditemukan."}), 404
        dt = eval_datetime_from_parts(schedule.get("date"), schedule.get("time"))
        if not dt or dt > datetime.now():
            return jsonify({"success": False, "error": "Evaluasi baru bisa diisi setelah jadwal misa dimulai."}), 400
        if not eval_schedule_has_staff(schedule):
            return jsonify({"success": False, "error": "Jadwal ini belum memiliki petugas."}), 400
        cursor.execute("SELECT id FROM streaming_evaluations WHERE schedule_kind = %s AND schedule_key = %s LIMIT 1", (schedule.get("kind"), schedule.get("scheduleKey")))
        if cursor.fetchone():
            return jsonify({"success": False, "error": "Evaluasi untuk jadwal ini sudah pernah diisi."}), 409

        viewer = current_user_context()
        evaluator_id = viewer.get("user_id") if viewer.get("logged_in") else payload.get("evaluatorId")
        evaluator_name = normalize_text(viewer.get("nama") if viewer.get("logged_in") else payload.get("evaluatorName"))
        evaluator_role = normalize_text(payload.get("evaluatorRole"))
        if viewer.get("logged_in") and not evaluator_role:
            # Find role for this user in selected schedule.
            for staff in schedule.get("staff") or []:
                if normalize_text(staff.get("memberId")) == normalize_text(viewer.get("user_id")):
                    evaluator_role = normalize_text(staff.get("role"))
                    break
        if not evaluator_name:
            return jsonify({"success": False, "error": "Nama pengisi evaluasi wajib diisi."}), 400

        submitted_staff = payload.get("staff") if isinstance(payload.get("staff"), list) else schedule.get("staff") or []
        final_staff = []
        active_members_map = {str(m.get("id")): m for m in eval_fetch_active_members(cursor)}
        for raw_slot in submitted_staff:
            if not isinstance(raw_slot, dict):
                continue
            slot = dict(raw_slot)
            slot.setdefault("role", raw_slot.get("role"))
            slot.setdefault("roleId", raw_slot.get("roleId"))
            slot.setdefault("assignmentId", raw_slot.get("assignmentId"))
            attendance = normalize_text(raw_slot.get("attendance") or "present")
            actual_id = normalize_text(raw_slot.get("actualMemberId") or raw_slot.get("memberId"))
            if attendance == "not_attend":
                slot["actualMemberId"] = ""
                slot["actualMemberName"] = "Tidak datang"
                slot["attendance"] = "not_attend"
            else:
                member_info = active_members_map.get(actual_id)
                slot["actualMemberId"] = actual_id
                slot["actualMemberName"] = normalize_text(member_info.get("name") if member_info else raw_slot.get("actualMemberName") or raw_slot.get("memberName"))
                slot["attendance"] = "present"
            final_staff.append(slot)
            if schedule.get("kind") == "misa_biasa":
                eval_upsert_regular_assignment(cursor, schedule, slot)
            else:
                eval_upsert_misa_besar_assignment(cursor, schedule, slot)

        extra_staff = payload.get("extraStaff") if isinstance(payload.get("extraStaff"), list) else []
        staff_evals = payload.get("staffEvaluations") if isinstance(payload.get("staffEvaluations"), list) else []
        dynamic_answers = payload.get("dynamicAnswers") if isinstance(payload.get("dynamicAnswers"), list) else []
        checklist = payload.get("checklist") if isinstance(payload.get("checklist"), dict) else {}
        general = normalize_text(payload.get("generalAssessment"))

        cursor.execute(
            """
            INSERT INTO streaming_evaluations
            (schedule_kind, schedule_key, schedule_date, schedule_time, misa_name, misa_type_label,
             evaluator_id, evaluator_name, evaluator_role, staff_json, extra_staff_json, staff_evaluations_json,
             technical_issue, nontechnical_issue, checklist_json, final_note, dynamic_answers_json, general_assessment, submitted_at)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, NOW())
            """,
            (
                schedule.get("kind"), schedule.get("scheduleKey"), schedule.get("date"), schedule.get("time"),
                schedule.get("misaName"), schedule.get("kindLabel"), evaluator_id, evaluator_name, evaluator_role,
                json.dumps(final_staff, ensure_ascii=False), json.dumps(extra_staff, ensure_ascii=False), json.dumps(staff_evals, ensure_ascii=False),
                normalize_text(payload.get("technicalIssue")), normalize_text(payload.get("nontechnicalIssue")), json.dumps(checklist, ensure_ascii=False),
                normalize_text(payload.get("finalNote")), json.dumps(dynamic_answers, ensure_ascii=False), general,
            ),
        )
        evaluation_id = cursor.lastrowid
        cursor.execute("SELECT * FROM streaming_evaluations WHERE id = %s", (evaluation_id,))
        evaluation = eval_row_to_dict(cursor.fetchone())
        eval_create_urgent_notifications(cursor, evaluation)
        conn.commit()
        return jsonify({"success": True, "evaluation": evaluation})
    except Exception as exc:
        conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 400
    finally:
        cursor.close()
        conn.close()


# Source legacy app.py lines 15237-15311 | routes: /api/evaluasi-streaming/member
@app.route("/api/evaluasi-streaming/member", methods=["GET"])
def api_streaming_evaluation_member():
    if not session.get("logged_in"):
        return jsonify({"success": False, "error": "Unauthorized"}), 401
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True, buffered=True)
    try:
        ensure_streaming_evaluation_schema(cursor)
        viewer = current_user_context()
        member_id = normalize_text(viewer.get("user_id"))
        now = datetime.now()
        month = request.args.get("month")
        year = request.args.get("year")
        selected_month = parse_required_int(month, now.month) if normalize_text(month) not in {"all", ""} else now.month
        selected_year = parse_required_int(year, now.year) if normalize_text(year) not in {"all", ""} else now.year
        start, end = eval_period_bounds("month", selected_year, selected_month)
        kind = normalize_text(request.args.get("kind") or "all")
        status = normalize_text(request.args.get("status") or "all")
        schedules = eval_all_schedules(cursor, start, end, kind)
        evals = eval_evaluation_map(cursor, start.isoformat(), end.isoformat())
        items = []
        total_all = 0
        displayed = 0
        completed = 0
        not_filled = 0
        progress_by_kind: dict[str, dict[str, object]] = {}
        for s in schedules:
            if member_id not in eval_schedule_member_ids(s):
                continue
            total_all += 1
            dt = eval_datetime_from_parts(s.get("date"), s.get("time"))
            is_due = bool(dt and dt <= now)
            ev = evals.get((s.get("kind"), s.get("scheduleKey")))
            if is_due:
                displayed += 1
                if ev:
                    completed += 1
                else:
                    not_filled += 1
            p = progress_by_kind.setdefault(s.get("kind"), {"kind": s.get("kind"), "kindLabel": s.get("kindLabel"), "total": 0, "done": 0})
            p["total"] += 1
            if is_due:
                p["done"] += 1
            s["evaluated"] = bool(ev)
            s["evaluationId"] = ev.get("id") if ev else None
            s["due"] = is_due
            # Form evaluasi anggota hanya boleh tampil setelah jadwal benar-benar dimulai.
            # Total/progress tetap menghitung semua tugas bulan itu, tetapi daftar yang bisa diisi/review
            # tidak menampilkan jadwal masa depan agar user tidak bisa mengisi sebelum hari-H/jam mulai.
            if not is_due:
                continue
            if status == "filled" and not ev:
                continue
            if status == "empty" and ev:
                continue
            items.append(s)
        progress = []
        for row in progress_by_kind.values():
            total = row.get("total") or 0
            done = row.get("done") or 0
            row["percentage"] = round((done / total) * 100) if total else 0
            progress.append(row)
        progress.sort(key=lambda r: r.get("kindLabel") or "")
        items.sort(key=lambda r: (r.get("date") or "", r.get("time") or ""), reverse=normalize_text(request.args.get("sort")) == "date_desc")
        return jsonify({
            "success": True,
            "currentUser": viewer,
            "stats": {"totalSessions": total_all, "displayed": displayed, "filled": completed, "empty": not_filled},
            "progress": progress,
            "items": items,
            "questions": eval_fetch_questions(cursor),
        })
    finally:
        cursor.close()
        conn.close()


# Source legacy app.py lines 15314-15370 | routes: /api/evaluasi-streaming/admin/results
@app.route("/api/evaluasi-streaming/admin/results", methods=["GET"])
def api_streaming_evaluation_admin_results():
    if normalize_role_value(session.get("role") or "") not in {"admin", "super_admin"}:
        return jsonify({"success": False, "error": "Akses ditolak."}), 403
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True, buffered=True)
    try:
        ensure_streaming_evaluation_schema(cursor)
        now = datetime.now()
        scale = normalize_text(request.args.get("scale") or "month")
        year = parse_required_int(request.args.get("year"), now.year)
        month = parse_required_int(request.args.get("month"), now.month)
        week = parse_required_int(request.args.get("week"), int(now.strftime("%V")))
        kind = normalize_text(request.args.get("kind") or "all")
        search = normalize_text(request.args.get("search"))
        sort = normalize_text(request.args.get("sort") or "date_asc")
        start, end = eval_period_bounds(scale, year, month, week)
        schedules = eval_all_schedules(cursor, start, end, kind)
        schedule_due = []
        for s in schedules:
            dt = eval_datetime_from_parts(s.get("date"), s.get("time"))
            if dt and dt <= now and eval_schedule_has_staff(s):
                schedule_due.append(s)
        evaluations = eval_fetch_evaluations(cursor, start.isoformat(), end.isoformat(), kind, search, sort)
        eval_keys = {(e.get("kind"), e.get("scheduleKey")) for e in evaluations}
        all_eval_map = eval_evaluation_map(cursor, start.isoformat(), end.isoformat())
        viewer = current_user_context()
        viewer_id = normalize_text(viewer.get("user_id"))
        pending = []
        kw = search.lower()
        for s in schedule_due:
            if (s.get("kind"), s.get("scheduleKey")) in all_eval_map:
                continue
            if kw and kw not in " ".join([normalize_text(s.get("misaName")), normalize_text(s.get("kindLabel")), normalize_text(s.get("displayDateTime"))]).lower():
                continue
            pending_item = dict(s)
            pending_item["canFillByCurrentUser"] = bool(
                viewer_id and any(normalize_text(staff.get("memberId")) == viewer_id for staff in (s.get("staff") or []))
            )
            pending_item["currentUserRole"] = normalize_role_value(viewer.get("role") or "")
            pending.append(pending_item)
        pending.sort(key=lambda r: (r.get("date") or "", r.get("time") or ""))
        urgent_count = sum(1 for e in evaluations if ("urgent" in normalize_text(e.get("generalAssessment")).lower() or "serius" in normalize_text(e.get("generalAssessment")).lower()))
        total_misa = len([s for s in schedules if eval_datetime_from_parts(s.get("date"), s.get("time")) and eval_datetime_from_parts(s.get("date"), s.get("time")) <= now])
        summary = eval_build_admin_summary(evaluations, pending)
        return jsonify({
            "success": True,
            "period": {"start": start.isoformat(), "end": end.isoformat(), "scale": scale, "month": month, "year": year, "week": week},
            "metrics": {"evaluationsIn": len(evaluations), "missed": len(pending), "totalMisa": total_misa, "urgent": urgent_count},
            "summary": summary,
            "evaluations": evaluations,
            "pending": pending,
            "currentUser": viewer,
        })
    finally:
        cursor.close()
        conn.close()


# Source legacy app.py lines 15439-15452 | routes: /api/evaluasi-streaming/admin/results/<int:evaluation_id>
@app.route("/api/evaluasi-streaming/admin/results/<int:evaluation_id>", methods=["DELETE"])
def api_streaming_evaluation_delete(evaluation_id):
    if normalize_role_value(session.get("role") or "") not in {"admin", "super_admin"}:
        return jsonify({"success": False, "error": "Akses ditolak."}), 403
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True, buffered=True)
    try:
        ensure_streaming_evaluation_schema(cursor)
        cursor.execute("DELETE FROM streaming_evaluations WHERE id = %s", (evaluation_id,))
        conn.commit()
        return jsonify({"success": True})
    finally:
        cursor.close()
        conn.close()


# Source legacy app.py lines 15567-15617 | routes: /api/evaluasi-streaming/admin/export.xlsx
@app.route("/api/evaluasi-streaming/admin/export.xlsx", methods=["GET"])
def api_streaming_evaluation_export_xlsx():
    if normalize_role_value(session.get("role") or "") not in {"admin", "super_admin"}:
        return jsonify({"success": False, "error": "Akses ditolak."}), 403
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True, buffered=True)
    try:
        ensure_streaming_evaluation_schema(cursor)
        now = datetime.now()
        start, end = eval_period_bounds(
            normalize_text(request.args.get("scale") or "month"),
            parse_required_int(request.args.get("year"), now.year),
            parse_required_int(request.args.get("month"), now.month),
            parse_required_int(request.args.get("week"), int(now.strftime("%V"))),
        )
        evaluations = eval_fetch_evaluations(
            cursor,
            start.isoformat(),
            end.isoformat(),
            normalize_text(request.args.get("kind") or "all"),
            normalize_text(request.args.get("search")),
            normalize_text(request.args.get("sort") or "date_asc"),
        )
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        wb = Workbook()
        ws = wb.active
        ws.title = "Evaluasi Streaming"
        headers, rows = eval_export_rows(evaluations)
        ws.append(headers)
        for row in rows:
            ws.append(row)
        fill = PatternFill("solid", fgColor="7F0000")
        font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 12), 65)
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(buf, as_attachment=True, download_name="hasil-evaluasi-streaming.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    finally:
        cursor.close()
        conn.close()


# Source legacy app.py lines 15620-15818 | routes: /api/evaluasi-streaming/admin/export.pdf
@app.route("/api/evaluasi-streaming/admin/export.pdf", methods=["GET"])
def api_streaming_evaluation_export_pdf():
    if normalize_role_value(session.get("role") or "") not in {"admin", "super_admin"}:
        return jsonify({"success": False, "error": "Akses ditolak."}), 403
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True, buffered=True)
    try:
        ensure_streaming_evaluation_schema(cursor)
        now = datetime.now()
        scale = normalize_text(request.args.get("scale") or "month")
        year = parse_required_int(request.args.get("year"), now.year)
        month = parse_required_int(request.args.get("month"), now.month)
        week = parse_required_int(request.args.get("week"), int(now.strftime("%V")))
        kind = normalize_text(request.args.get("kind") or "all")
        search = normalize_text(request.args.get("search"))
        sort = normalize_text(request.args.get("sort") or "date_asc")
        start, end = eval_period_bounds(scale, year, month, week)
        evaluations = eval_fetch_evaluations(cursor, start.isoformat(), end.isoformat(), kind, search, sort)

        from reportlab.lib.pagesizes import A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, KeepTogether
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER, TA_LEFT

        def esc_pdf(value) -> str:
            return html.escape(eval_export_value(value)).replace("\n", "<br/>")

        def answer_text(answer) -> str:
            if isinstance(answer, list):
                return ", ".join(normalize_text(v) for v in answer if normalize_text(v)) or "-"
            return normalize_text(answer) or "-"

        def staff_name_from_slot(slot: dict[str, object]) -> str:
            if normalize_text(slot.get("attendance")) == "not_attend":
                return "Tidak datang"
            return normalize_text(slot.get("actualMemberName") or slot.get("actualName") or slot.get("memberName") or slot.get("name")) or "-"

        def compact_period_label() -> str:
            month_names = ["", "Januari", "Februari", "Maret", "April", "Mei", "Juni", "Juli", "Agustus", "September", "Oktober", "November", "Desember"]
            if scale == "all":
                return "Semua Periode"
            if scale == "year":
                return f"Tahun {year}"
            if scale == "week":
                return f"Minggu ke-{week}, {year}"
            return f"{month_names[month] if 1 <= month <= 12 else month} {year}"

        kind_label = {"all": "Misa Biasa & Misa Besar", "misa_biasa": "Misa Biasa", "misa_besar": "Misa Besar"}.get(kind, "Misa Biasa & Misa Besar")

        buf = BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=28, rightMargin=28, topMargin=26, bottomMargin=24)
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle("eval-title", parent=styles["Title"], fontSize=18, leading=22, alignment=TA_CENTER, spaceAfter=8)
        subtitle_style = ParagraphStyle("eval-subtitle", parent=styles["BodyText"], fontSize=9, leading=12, alignment=TA_CENTER, textColor=colors.HexColor("#4b5563"), spaceAfter=14)
        section_style = ParagraphStyle("eval-section", parent=styles["Heading2"], fontSize=11, leading=14, textColor=colors.HexColor("#800000"), spaceBefore=8, spaceAfter=6)
        normal = ParagraphStyle("eval-normal", parent=styles["BodyText"], fontSize=8, leading=10, alignment=TA_LEFT)
        small = ParagraphStyle("eval-small", parent=styles["BodyText"], fontSize=7.2, leading=9, alignment=TA_LEFT)
        tiny = ParagraphStyle("eval-tiny", parent=styles["BodyText"], fontSize=6.6, leading=8.2, alignment=TA_LEFT)

        def P(value, style=small):
            return Paragraph(esc_pdf(value), style)

        def styled_table(data, col_widths=None, header=True):
            table = Table(data, colWidths=col_widths, repeatRows=1 if header else 0, hAlign="LEFT")
            base_style = [
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#d9b8b8")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 5),
                ("RIGHTPADDING", (0, 0), (-1, -1), 5),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
            if header:
                base_style += [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#800000")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ]
            table.setStyle(TableStyle(base_style))
            return table

        elements = [
            Paragraph("Hasil Evaluasi Streaming", title_style),
            Paragraph(f"Periode: {html.escape(compact_period_label())} &nbsp;&nbsp;|&nbsp;&nbsp; Jenis: {html.escape(kind_label)} &nbsp;&nbsp;|&nbsp;&nbsp; Total evaluasi: {len(evaluations)}", subtitle_style),
        ]

        if not evaluations:
            elements.append(Paragraph("Tidak ada data evaluasi sesuai filter yang dipilih.", normal))
        else:
            for idx, e in enumerate(evaluations, start=1):
                heading = f"{idx}. {eval_export_value(e.get('misaName'))} — {eval_export_value(e.get('kindLabel'))}"
                header_tbl = Table(
                    [[Paragraph(html.escape(heading), section_style), Paragraph(html.escape(eval_export_value(e.get('generalAssessment'))), section_style)]],
                    colWidths=[380, 130],
                    hAlign="LEFT",
                )
                header_tbl.setStyle(TableStyle([
                    ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#fff1f2")),
                    ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#800000")),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 7),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 7),
                ]))

                info_rows = [
                    [P("Tanggal"), P(f"{e.get('dayName') or '-'}, {e.get('dateText') or e.get('date') or '-'}")],
                    [P("Jam"), P(f"{e.get('time') or '-'} WIB")],
                    [P("Jenis Misa"), P(e.get("kindLabel"))],
                    [P("Judul / Nama Misa"), P(e.get("misaName"))],
                    [P("Pengisi Evaluasi"), P(f"{e.get('evaluatorName') or '-'} ({e.get('evaluatorRole') or '-'})")],
                    [P("Waktu Submit"), P((e.get("submittedAt") or "").replace("T", " ")[:16])],
                    [P("Penilaian Umum Streaming Keseluruhan"), P(e.get("generalAssessment"))],
                    [P("Kendala Teknis Selama Streaming"), P(e.get("technicalIssue"))],
                    [P("Kendala Misa Non-Teknis"), P(e.get("nontechnicalIssue"))],
                    [P("Checklist Kondisi Pelayanan"), P(eval_checklist_export_text(e.get("checklist") or {}))],
                    [P("Catatan Penutup / Rekomendasi"), P(e.get("finalNote"))],
                ]
                info_tbl = styled_table(info_rows, col_widths=[150, 360], header=False)
                info_tbl.setStyle(TableStyle([("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#fff7f7")), ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold")]))

                staff_header = [[P("Role", tiny), P("Petugas Jadwal", tiny), P("Petugas Aktual", tiny), P("Status Kehadiran", tiny)]]
                staff_rows = []
                for slot in e.get("staff") or []:
                    if not isinstance(slot, dict):
                        continue
                    staff_rows.append([
                        P(slot.get("role"), tiny),
                        P(slot.get("memberName") or slot.get("name"), tiny),
                        P(staff_name_from_slot(slot), tiny),
                        P("Tidak datang" if normalize_text(slot.get("attendance")) == "not_attend" else "Hadir/Digantikan", tiny),
                    ])
                if not staff_rows:
                    staff_rows = [[P("-", tiny), P("-", tiny), P("-", tiny), P("-", tiny)]]
                staff_tbl = styled_table(staff_header + staff_rows, col_widths=[95, 135, 135, 145])

                eval_header = [[P("Nama Petugas", tiny), P("Role", tiny), P("Penilaian Singkat", tiny), P("Catatan Performa", tiny)]]
                eval_rows = []
                for item in e.get("staffEvaluations") or []:
                    if not isinstance(item, dict):
                        continue
                    eval_rows.append([
                        P(item.get("memberName") or item.get("name"), tiny),
                        P(item.get("role"), tiny),
                        P(item.get("rating") or item.get("assessment"), tiny),
                        P(item.get("note") or item.get("catatan") or item.get("comment"), tiny),
                    ])
                if not eval_rows:
                    eval_rows = [[P("-", tiny), P("-", tiny), P("-", tiny), P("-", tiny)]]
                staff_eval_tbl = styled_table(eval_header + eval_rows, col_widths=[125, 95, 115, 175])

                extra_header = [[P("Nama Petugas Tambahan", tiny), P("Role Bantuan", tiny)]]
                extra_rows = []
                for item in e.get("extraStaff") or []:
                    if not isinstance(item, dict):
                        continue
                    extra_rows.append([P(item.get("name") or item.get("memberName"), tiny), P(item.get("role") or item.get("helpRole"), tiny)])
                if not extra_rows:
                    extra_rows = [[P("-", tiny), P("-", tiny)]]
                extra_tbl = styled_table(extra_header + extra_rows, col_widths=[255, 255])

                dynamic_header = [[P("Pertanyaan Tambahan", tiny), P("Jawaban", tiny)]]
                dynamic_rows = []
                for item in e.get("dynamicAnswers") or []:
                    if not isinstance(item, dict):
                        continue
                    dynamic_rows.append([P(item.get("question") or item.get("label") or item.get("text"), tiny), P(answer_text(item.get("answer")), tiny)])
                if not dynamic_rows:
                    dynamic_rows = [[P("-", tiny), P("-", tiny)]]
                dynamic_tbl = styled_table(dynamic_header + dynamic_rows, col_widths=[255, 255])

                block = [
                    header_tbl,
                    Spacer(1, 5),
                    info_tbl,
                    Spacer(1, 6),
                    Paragraph("Petugas Bertugas: Jadwal dan Aktual", section_style),
                    staff_tbl,
                    Spacer(1, 6),
                    Paragraph("Evaluasi Per Petugas", section_style),
                    staff_eval_tbl,
                    Spacer(1, 6),
                    Paragraph("Petugas Tambahan", section_style),
                    extra_tbl,
                    Spacer(1, 6),
                    Paragraph("Jawaban Pertanyaan Tambahan Dinamis", section_style),
                    dynamic_tbl,
                    Spacer(1, 12),
                ]
                elements.extend(block)
                if idx < len(evaluations):
                    elements.append(PageBreak())

        doc.build(elements)
        buf.seek(0)
        return send_file(buf, as_attachment=False, download_name="hasil-evaluasi-streaming.pdf", mimetype="application/pdf")
    finally:
        cursor.close()
        conn.close()


# Source legacy app.py lines 15959-15968 | routes: /api/evaluasi-streaming/reminders/run
@app.route("/api/evaluasi-streaming/reminders/run", methods=["POST"])
def api_streaming_evaluation_reminders_run():
    """Endpoint manual untuk admin/super_admin menjalankan generator reminder evaluasi."""
    if normalize_role_value(session.get("role") or "") not in {"admin", "super_admin"}:
        return jsonify({"success": False, "error": "Akses ditolak."}), 403
    try:
        created = create_streaming_evaluation_reminder_notifications()
        return jsonify({"success": True, "created": created})
    except Exception as exc:
        return jsonify({"success": False, "error": str(exc)}), 400


