"""Activity Log Controller.

File ini berisi route/controller yang dipisahkan dari app.py server lama.
Logika helper tetap dipanggil dari crembo_app.services.core agar perilaku produksi tetap sama.
"""

from crembo_app.services import core as _core

globals().update({
    name: getattr(_core, name)
    for name in dir(_core)
    if not (name.startswith("__") and name.endswith("__"))
})


# Route dari app.py server: /api/activity-logs/context
@app.route("/api/activity-logs/context", methods=["GET"])
def api_activity_logs_context():
    denied = require_activity_log_auth()
    if denied:
        return denied
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        ensure_activity_log_schema(cursor)
        current = current_user_context()
        users = activity_visible_users(cursor)
        scope_label = "Semua pengguna" if activity_current_role() == "super_admin" else ("Admin sendiri + anggota" if activity_current_role() == "admin" else "Diri sendiri")
        role_values = activity_visible_role_values()
        module_scope, module_params = activity_scope_where("l")
        cursor.execute(f"SELECT DISTINCT `module` FROM `activity_logs` l WHERE {module_scope} ORDER BY `module` ASC", tuple(module_params))
        modules = [row.get("module") for row in cursor.fetchall() or [] if row.get("module")]
        cursor.execute(f"SELECT DISTINCT `action` FROM `activity_logs` l WHERE {module_scope} ORDER BY `action` ASC", tuple(module_params))
        actions = [row.get("action") for row in cursor.fetchall() or [] if row.get("action")]
        default_month, default_year = activity_month_year_defaults()
        return jsonify({
            "ok": True,
            "currentUser": current,
            "scopeLabel": scope_label,
            "users": users,
            "roles": [{"value": value, "label": activity_role_label(value)} for value in role_values],
            "modules": modules,
            "actions": actions,
            "defaultMonth": default_month,
            "defaultYear": default_year,
        })
    finally:
        cursor.close()
        conn.close()


# Route dari app.py server: /api/activity-logs
@app.route("/api/activity-logs", methods=["GET"])
def api_activity_logs():
    denied = require_activity_log_auth()
    if denied:
        return denied
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        ensure_activity_log_schema(cursor)
        base_sql, params, meta = activity_filter_query(request.args)
        cursor.execute(f"SELECT COUNT(*) AS total {base_sql}", tuple(params))
        total = fetch_scalar_value(cursor.fetchone(), 0) or 0
        offset = (meta["page"] - 1) * meta["perPage"]
        cursor.execute(
            f"""
            SELECT l.*
            {base_sql}
            ORDER BY {meta["orderSql"]}
            LIMIT %s OFFSET %s
            """,
            tuple(params + [meta["perPage"], offset]),
        )
        logs = [activity_log_row_to_dict(row) for row in cursor.fetchall() or []]
        return jsonify({
            "ok": True,
            "logs": logs,
            "total": total,
            "page": meta["page"],
            "perPage": meta["perPage"],
            "totalPages": max(1, (int(total) + meta["perPage"] - 1) // meta["perPage"]),
            "period": {
                "month": meta["month"],
                "year": meta["year"],
                "start": meta["start"].isoformat(),
                "end": meta["end"].isoformat(),
            },
        })
    finally:
        cursor.close()
        conn.close()


# Route dari app.py server: /api/activity-logs/export.xlsx
@app.route("/api/activity-logs/export.xlsx", methods=["GET"])
def api_activity_logs_export_xlsx():
    denied = require_activity_log_auth()
    if denied:
        return denied
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        ensure_activity_log_schema(cursor)
        logs = fetch_activity_logs_for_export(cursor, request.args)
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = "Log Aktivitas"
        headers, rows = activity_export_rows(logs)
        ws.append(headers)
        for row in rows:
            ws.append(row)
        fill = PatternFill("solid", fgColor="800000")
        font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for row_cells in ws.iter_rows(min_row=2):
            for cell in row_cells:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for column_cells in ws.columns:
            max_length = 0
            column_letter = column_cells[0].column_letter
            for cell in column_cells:
                max_length = max(max_length, len(str(cell.value or "")))
            ws.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 45)

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        record_activity("EXPORT", "Log Aktivitas", "Mengekspor log aktivitas ke Excel", "activity_logs.export_xlsx", method="GET")
        return send_file(
            buffer,
            as_attachment=True,
            download_name=f"log-aktivitas-{datetime.now().strftime('%Y%m%d%H%M')}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    finally:
        cursor.close()
        conn.close()


# Route dari app.py server: /api/activity-logs/export.pdf
@app.route("/api/activity-logs/export.pdf", methods=["GET"])
def api_activity_logs_export_pdf():
    denied = require_activity_log_auth()
    if denied:
        return denied
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        ensure_activity_log_schema(cursor)
        logs = fetch_activity_logs_for_export(cursor, request.args)
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Spacer, Table, TableStyle, Paragraph

        def p(value, style):
            return Paragraph(html.escape(normalize_text(value) or "-").replace("\n", "<br/>"), style)

        headers, rows = activity_export_rows(logs)
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=8 * mm, leftMargin=8 * mm, topMargin=10 * mm, bottomMargin=10 * mm)
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle("activity-title", parent=styles["Title"], fontName="Helvetica-Bold", fontSize=16, textColor=colors.HexColor("#800000"), alignment=1)
        cell_style = ParagraphStyle("activity-cell", parent=styles["BodyText"], fontSize=7, leading=9)
        header_style = ParagraphStyle("activity-header", parent=cell_style, fontName="Helvetica-Bold", textColor=colors.white)

        data = [[p(h, header_style) for h in headers]]
        for row in rows:
            data.append([p(cell, cell_style) for cell in row])

        table = Table(data, repeatRows=1, colWidths=[10*mm, 34*mm, 30*mm, 28*mm, 20*mm, 20*mm, 26*mm, 72*mm, 42*mm, 16*mm])
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#800000")),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#e9d5d5")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#fff7f7")]),
            ("LEFTPADDING", (0, 0), (-1, -1), 3),
            ("RIGHTPADDING", (0, 0), (-1, -1), 3),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]))
        story = [
            Paragraph("Laporan Log Aktivitas", title_style),
            Spacer(1, 6),
            Paragraph(f"Total log: {len(rows)}", styles["Normal"]),
            Spacer(1, 8),
            table,
        ]
        doc.build(story)
        buffer.seek(0)
        record_activity("EXPORT", "Log Aktivitas", "Mengekspor log aktivitas ke PDF", "activity_logs.export_pdf", method="GET")
        return send_file(
            buffer,
            as_attachment=False,
            download_name=f"log-aktivitas-{datetime.now().strftime('%Y%m%d%H%M')}.pdf",
            mimetype="application/pdf",
        )
    finally:
        cursor.close()
        conn.close()

