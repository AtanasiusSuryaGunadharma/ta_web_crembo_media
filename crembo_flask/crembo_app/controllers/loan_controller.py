from crembo_app.services import core as _core

# Memuat seluruh helper, service, dan objek Flask dari core agar potongan kode route
# tetap kompatibel setelah dipisah dari app.py monolitik.
globals().update({
    name: getattr(_core, name)
    for name in dir(_core)
    if not (name.startswith("__") and name.endswith("__"))
})

# Controller: Loan Controller

# Source legacy app.py lines 973-1057 | routes: /api/pengajuan/<pengajuan_id>/ambil
@app.route("/api/pengajuan/<pengajuan_id>/ambil", methods=["POST"])
def input_pengambilan(pengajuan_id):
    ensure_auth_schema()
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        ensure_loan_schema(cursor)
        cursor.execute("SELECT * FROM `loan_requests` WHERE `id` = %s LIMIT 1", (pengajuan_id,))
        req = cursor.fetchone()
        if not req:
            return jsonify({"success": False, "error": "Pengajuan tidak ditemukan"}), 404

        role = session.get("role") or ""
        user_id = session.get("user_id")
        if role not in ["admin", "super_admin"] and str(req.get("member_id")) != str(user_id):
            return jsonify({"success": False, "error": "Akses ditolak"}), 403

        if req.get("status") != "approved":
            return jsonify({"success": False, "error": "Hanya pengajuan dengan status 'approved' dapat diambil"}), 400

        tanggal = request.form.get("tanggal") or request.form.get("date") or None
        waktu = request.form.get("waktu") or request.form.get("time") or None
        lokasi = request.form.get("lokasi") or request.form.get("location") or ""
        units_raw = request.form.get("unitDetails") or request.form.get("units") or None
        try:
            unit_details = json.loads(units_raw) if units_raw else []
        except Exception:
            unit_details = []

        photo_record = None
        photo_file = request.files.get("photo") or request.files.get("bukti")
        if photo_file:
            try:
                saved = save_uploaded_attachment(photo_file)
                photo_record = saved.get("url") or saved.get("uri") or saved.get("name")
            except Exception:
                photo_record = None

        pickup_info = {
            "date": tanggal,
            "time": waktu,
            "location": lokasi,
            "photo": photo_record,
            "units": unit_details,
        }

        ensure_column(cursor, "loan_requests", "pickup_info", "`pickup_info` longtext DEFAULT NULL")
        ensure_column(cursor, "loan_requests", "pickup_at", "`pickup_at` datetime DEFAULT NULL")

        cursor.execute(
            "UPDATE `loan_requests` SET `status` = %s, `pickup_info` = %s, `pickup_at` = %s WHERE `id` = %s",
            (
                "taken",
                json.dumps(pickup_info, ensure_ascii=False),
                datetime.utcnow(),
                pengajuan_id,
            ),
        )
        conn.commit()

        # Update Notifikasi Pengambilan Barang
        try:
            ensure_notifications_schema()
            user_name = session.get("nama") or session.get("username") or f"User ID {user_id}"
            safe_user = html.escape(str(user_name))
            safe_lokasi = html.escape(str(lokasi))
            
            condition_texts = [f"&bull; Unit {i+1}: {html.escape(str(u.get('status')))} - {html.escape(str(u.get('reason','-')))}" for i, u in enumerate(unit_details)]
            condition_str = "<br>".join(condition_texts)
            photo_link = f"<br><br><a href='{photo_record}' target='_blank' style='display:inline-block; padding:4px 8px; background:#7f1d1d; color:#fff; border-radius:4px; text-decoration:none; font-size:11px; font-weight:bold;'>Lihat Foto Bukti</a>" if photo_record else ""
            
            body_text = f"Pengambilan dicatat oleh <b>{safe_user}</b>.<br><b>Lokasi:</b> {safe_lokasi}<br><b>Kondisi:</b><br>{condition_str}{photo_link}"

            create_notification(cursor, "peminjaman", f"Barang Diambil: {req.get('barang_name')}", body_text, "/riwayat-peminjaman-pengembalian.html", {"pengajuan_id": pengajuan_id, "target_user_id": req.get("member_id")}, target_role="admin")
            
            create_notification(cursor, "peminjaman", f"Pengambilan Tersimpan: {req.get('barang_name')}", "Data pengambilan Anda telah tersimpan.", "/riwayat-peminjaman-barang-anggota.html", {"pengajuan_id": pengajuan_id, "target_user_id": req.get("member_id")}, target_role="user")
            
            conn.commit()
        except Exception as e:
            print("Error notif:", e)

        return jsonify({"success": True})
    finally:
        cursor.close()
        conn.close()


# Source legacy app.py lines 1059-1184 | routes: /api/pengajuan/<pengajuan_id>/kembali
@app.route("/api/pengajuan/<pengajuan_id>/kembali", methods=["POST"])
def input_pengembalian(pengajuan_id):
    ensure_auth_schema()
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        ensure_loan_schema(cursor)
        cursor.execute("SELECT * FROM `loan_requests` WHERE `id` = %s LIMIT 1", (pengajuan_id,))
        req = cursor.fetchone()
        if not req:
            return jsonify({"success": False, "error": "Pengajuan tidak ditemukan"}), 404

        role = session.get("role") or ""
        user_id = session.get("user_id")
        if role not in ["admin", "super_admin"] and str(req.get("member_id")) != str(user_id):
            return jsonify({"success": False, "error": "Akses ditolak"}), 403

        if req.get("status") not in ["taken", "approved"]:
            return jsonify({"success": False, "error": "Hanya pengajuan yang sedang dipinjam dapat dikembalikan"}), 400

        tanggal = request.form.get("tanggal") or request.form.get("date") or None
        waktu = request.form.get("waktu") or request.form.get("time") or None
        lokasi = request.form.get("lokasi") or request.form.get("location") or ""
        units_raw = request.form.get("unitDetails") or request.form.get("units") or None
        try:
            return_unit_details = json.loads(units_raw) if units_raw else []
        except Exception:
            return_unit_details = []

        photo_record = None
        photo_file = request.files.get("photo") or request.files.get("bukti")
        if photo_file:
            try:
                saved = save_uploaded_attachment(photo_file)
                photo_record = saved.get("url") or saved.get("uri") or saved.get("name")
            except Exception:
                photo_record = None

        return_info = {
            "date": tanggal,
            "time": waktu,
            "location": lokasi,
            "photo": photo_record,
            "units": return_unit_details,
        }

        ensure_column(cursor, "loan_requests", "return_info", "`return_info` longtext DEFAULT NULL")
        ensure_column(cursor, "loan_requests", "return_at", "`return_at` datetime DEFAULT NULL")

        barang_id = req.get("barang_id")
        jumlah_req = int(req.get("jumlah", 1))

        cursor.execute("SELECT `total_unit`, `available_unit`, `unit_details` FROM `inventory_items` WHERE `id` = %s LIMIT 1", (barang_id,))
        inv = cursor.fetchone()
        if inv:
            total = int(inv.get("total_unit", 1))
            available = int(inv.get("available_unit", 0))
            unit_details_raw = safe_json_loads(inv.get("unit_details"), [])

            units_restored = 0
            for idx, unit in enumerate(unit_details_raw):
                if unit.get("reason") == f"Dipinjam (Req ID: {pengajuan_id})":
                    kondisi_input = "Tersedia"
                    alasan_input = ""
                    if idx < len(return_unit_details):
                        input_val = return_unit_details[idx].get("status", "Baik")
                        if input_val == "Baik":
                            kondisi_input = "Tersedia"
                        else:
                            kondisi_input = input_val
                        alasan_input = return_unit_details[idx].get("reason", "")
                    
                    unit["status"] = kondisi_input
                    unit["reason"] = alasan_input
                    if kondisi_input == "Tersedia":
                        unit["available"] = True
                        units_restored += 1
                    else:
                        unit["available"] = False

            if units_restored == 0:
                for unit in unit_details_raw:
                    if unit.get("status") == "Dipinjam" and units_restored < jumlah_req:
                        unit["status"] = "Tersedia"
                        unit["reason"] = ""
                        unit["available"] = True
                        units_restored += 1

            new_available = min(total, available + units_restored)
            new_inv_status = "Tersedia" if new_available > 0 else "Dipinjam"

            cursor.execute(
                "UPDATE `inventory_items` SET `available_unit` = %s, `status` = %s, `unit_details` = %s WHERE `id` = %s",
                (new_available, new_inv_status, json.dumps(unit_details_raw, ensure_ascii=False), barang_id),
            )

        cursor.execute(
            "UPDATE `loan_requests` SET `status` = %s, `return_info` = %s, `return_at` = %s WHERE `id` = %s",
            ("returned", json.dumps(return_info, ensure_ascii=False), datetime.utcnow(), pengajuan_id),
        )

        # Update Notifikasi Pengembalian Barang
        try:
            ensure_notifications_schema()
            user_name = session.get("nama") or session.get("username") or f"User ID {user_id}"
            safe_user = html.escape(str(user_name))
            safe_lokasi = html.escape(str(lokasi))
            
            condition_texts = [f"&bull; Unit {i+1}: {html.escape(str(u.get('status')))} - {html.escape(str(u.get('reason','-')))}" for i, u in enumerate(return_unit_details)]
            condition_str = "<br>".join(condition_texts)
            photo_link = f"<br><br><a href='{photo_record}' target='_blank' style='display:inline-block; padding:4px 8px; background:#7f1d1d; color:#fff; border-radius:4px; text-decoration:none; font-size:11px; font-weight:bold;'>Lihat Foto Bukti</a>" if photo_record else ""
            
            body_text = f"Pengembalian dicatat oleh <b>{safe_user}</b>.<br><b>Lokasi:</b> {safe_lokasi}<br><b>Kondisi:</b><br>{condition_str}{photo_link}"

            create_notification(cursor, "peminjaman", f"Barang Dikembalikan: {req.get('barang_name')}", body_text, "/riwayat-peminjaman-pengembalian.html", {"pengajuan_id": pengajuan_id, "target_user_id": req.get("member_id")}, target_role="admin")
        except Exception as e:
            print("Error notif:", e)

        conn.commit()
        return jsonify({"success": True})
    except Exception as e:
        conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 400
    finally:
        cursor.close()
        conn.close()


# Source legacy app.py lines 1910-1981 | routes: /api/pengajuan
@app.route("/api/pengajuan", methods=["GET"])
def list_pengajuan():
    ensure_auth_schema()
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        ensure_loan_schema(cursor)
        
        status_filter = request.args.get('status')
        from_date = request.args.get('from_date')
        to_date = request.args.get('to_date')

        query = """
            SELECT l.*, a.nama as member_name 
            FROM `loan_requests` l
            LEFT JOIN `anggota` a ON l.member_id = a.id
            WHERE 1=1
        """
        params = []
        
        role = session.get("role") or ""
        if role not in ["admin", "super_admin"]:
            user_id = session.get("user_id")
            if not user_id:
                return jsonify({"success": True, "items": []})
            query += " AND l.`member_id` = %s"
            params.append(str(user_id))
        
        if status_filter and status_filter != 'all':
            query += " AND l.`status` = %s"
            params.append(status_filter)
        if from_date:
            query += " AND l.`tanggal_mulai` >= %s"
            params.append(from_date)
        if to_date:
            query += " AND l.`tanggal_mulai` <= %s"
            params.append(to_date)
            
        query += " ORDER BY l.`created_at` DESC"

        cursor.execute(query, tuple(params))
        rows = cursor.fetchall() or []
        items = []
        for r in rows:
            items.append({
                "id": r.get("id"),
                "memberId": r.get("member_id"),
                "memberNama": r.get("member_name"),
                "barangId": r.get("barang_id"),
                "barangNama": r.get("barang_name"),
                "barangCode": r.get("barang_code"),
                "barangFoto": r.get("barang_photo"),
                "jumlahDiminta": r.get("jumlah"),
                "tanggalPengajuan": r.get("tanggal_pengajuan"),
                "tanggalMulai": r.get("tanggal_mulai"),
                "waktuMulai": str(r.get("waktu_mulai") or ""),
                "tanggalSelesai": r.get("tanggal_selesai"),
                "waktuSelesai": str(r.get("waktu_selesai") or ""),
                "tujuan": r.get("tujuan"),
                "status": r.get("status"),
                "adminNote": r.get("admin_note", ""),
                "approvedBy": r.get("approved_by", ""),
                "approvedAt": r.get("approved_at"),
                "createdAt": r.get("created_at"),
                "updatedAt": r.get("updated_at"),
                "pickupInfo": safe_json_loads(r.get("pickup_info"), {}),
                "returnInfo": safe_json_loads(r.get("return_info"), {}),
            })
        return jsonify({"success": True, "items": items})
    finally:
        cursor.close()
        conn.close()


# Source legacy app.py lines 1983-2075 | routes: /api/pengajuan
@app.route("/api/pengajuan", methods=["POST"])
def create_pengajuan():
    ensure_auth_schema()
    data = request.get_json(silent=True) or {}
    barang_id = data.get("barangId")
    barang_nama = data.get("barangNama")
    jumlah = int(data.get("jumlahDiminta") or data.get("jumlah") or 0)
    tanggal_mulai = data.get("tanggalMulai")
    waktu_mulai = data.get("waktuMulai")
    tanggal_selesai = data.get("tanggalSelesai")
    waktu_selesai = data.get("waktuSelesai")
    tujuan = data.get("tujuan")

    if not barang_id or not barang_nama or not jumlah or not tanggal_mulai or not tanggal_selesai or not tujuan:
        return jsonify({"success": False, "error": "Field tidak lengkap"}), 400

    try:
        if datetime.fromisoformat(tanggal_mulai) >= datetime.fromisoformat(tanggal_selesai):
            return jsonify({"success": False, "error": "Tanggal selesai harus lebih besar dari tanggal mulai"}), 400
    except Exception:
        pass

    conn = mysql_connection()
    cursor = conn.cursor()
    try:
        ensure_loan_schema(cursor)
        pengajuan_id = f"pjn-{int(datetime.utcnow().timestamp() * 1000)}"

        barang_foto = None
        barang_code = None
        try:
            c2 = conn.cursor(dictionary=True)
            c2.execute("SELECT `code`, `photos` FROM `inventory_items` WHERE `id` = %s LIMIT 1", (barang_id,))
            row = c2.fetchone()
            if row:
                barang_code = row.get("code")
                photos_raw = row.get("photos")
                try:
                    photos = json.loads(photos_raw) if photos_raw else []
                    if isinstance(photos, list) and photos:
                        first = photos[0]
                        if isinstance(first, dict):
                            barang_foto = first.get("url") or first.get("uri") or None
                        elif isinstance(first, str):
                            barang_foto = first
                except Exception:
                    pass
            c2.close()
        except Exception:
            pass

        cursor.execute(
            """
            INSERT INTO `loan_requests` (`id`, `member_id`, `barang_id`, `barang_name`, `barang_code`, `barang_photo`, `jumlah`, `tanggal_pengajuan`, `tanggal_mulai`, `waktu_mulai`, `tanggal_selesai`, `waktu_selesai`, `tujuan`, `status`)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """,
            (
                pengajuan_id,
                session.get("user_id") or None,
                barang_id,
                barang_nama,
                barang_code,
                barang_foto,
                jumlah,
                datetime.utcnow().date(),
                tanggal_mulai,
                waktu_mulai,
                tanggal_selesai,
                waktu_selesai,
                tujuan,
                "pending",
            ),
        )
        conn.commit()
        
        # Update Notifikasi Admin
        try:
            ensure_notifications_schema()
            nc = conn.cursor()
            try:
                user_name = session.get('nama') or session.get('username') or f"User ID {session.get('user_id')}"
                safe_user = html.escape(str(user_name))
                create_notification(nc, "peminjaman", f"Pengajuan Peminjaman: {barang_nama}", f"Diajukan oleh <b>{safe_user}</b>. Harap ditinjau.", url_for('dashboard') if False else "/persetujuan-peminjaman.html", {"pengajuan_id": pengajuan_id}, target_role="admin")
                conn.commit()
            finally:
                nc.close()
        except Exception:
            pass
            
        return jsonify({"success": True, "id": pengajuan_id})
    finally:
        cursor.close()
        conn.close()


# Source legacy app.py lines 2078-2136 | routes: /api/pengajuan/<pengajuan_id>/cancel
@app.route("/api/pengajuan/<pengajuan_id>/cancel", methods=["POST"])
def cancel_pengajuan(pengajuan_id):
    ensure_auth_schema()
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        ensure_loan_schema(cursor)
        cursor.execute("SELECT `status`, `barang_id`, `jumlah` FROM `loan_requests` WHERE `id` = %s LIMIT 1", (pengajuan_id,))
        row = cursor.fetchone()
        if not row:
            return jsonify({"success": False, "error": "Pengajuan tidak ditemukan"}), 404
            
        status = row.get("status")
        
        if status not in ["pending", "approved"]:
            return jsonify({"success": False, "error": "Hanya pengajuan dengan status pending atau disetujui (sebelum diambil) yang dapat dibatalkan"}), 400
            
        if status == "approved":
            barang_id = row.get("barang_id")
            jumlah_req = int(row.get("jumlah", 1))
            
            cursor.execute("SELECT `total_unit`, `available_unit`, `unit_details` FROM `inventory_items` WHERE `id` = %s LIMIT 1", (barang_id,))
            inv = cursor.fetchone()
            if inv:
                total = int(inv.get("total_unit", 1))
                available = int(inv.get("available_unit", 0))
                
                unit_details_raw = safe_json_loads(inv.get("unit_details"), [])
                units_restored = 0
                
                for unit in unit_details_raw:
                     if unit.get("reason") == f"Dipinjam (Req ID: {pengajuan_id})":
                          unit["status"] = "Tersedia"
                          unit["reason"] = ""
                          unit["available"] = True
                          units_restored += 1
                
                if units_restored == 0:
                     for unit in unit_details_raw:
                          if unit.get("status") == "Dipinjam" and units_restored < jumlah_req:
                               unit["status"] = "Tersedia"
                               unit["reason"] = ""
                               unit["available"] = True
                               units_restored += 1
                
                new_available = min(total, available + jumlah_req)
                new_inv_status = "Tersedia" if new_available > 0 else "Dipinjam"
                
                cursor.execute(
                     "UPDATE `inventory_items` SET `available_unit` = %s, `status` = %s, `unit_details` = %s WHERE `id` = %s", 
                     (new_available, new_inv_status, json.dumps(unit_details_raw, ensure_ascii=False), barang_id)
                )

        cursor.execute("UPDATE `loan_requests` SET `status` = %s WHERE `id` = %s", ("cancelled", pengajuan_id))
        conn.commit()
        return jsonify({"success": True})
    finally:
        cursor.close()
        conn.close()


# Source legacy app.py lines 2138-2279 | routes: /api/admin/pengajuan/<pengajuan_id>/status
@app.route("/api/admin/pengajuan/<pengajuan_id>/status", methods=["PUT"])
def update_pengajuan_status(pengajuan_id):
    role = session.get("role") or ""
    if not session.get("logged_in") or role not in ["admin", "super_admin"]:
        return jsonify({"success": False, "error": "Akses Ditolak. Anda bukan Admin."}), 403
        
    data = request.get_json(silent=True) or {}
    new_status = data.get("status")
    admin_note = data.get("adminNote", "")
    
    if new_status not in ["approved", "rejected", "cancelled", "taken", "returned"]:
        return jsonify({"success": False, "error": "Status tidak valid"}), 400

    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        ensure_column(cursor, "loan_requests", "admin_note", "`admin_note` text DEFAULT NULL")
        ensure_column(cursor, "loan_requests", "approved_by", "`approved_by` varchar(150) DEFAULT NULL")
        ensure_column(cursor, "loan_requests", "approved_at", "`approved_at` datetime DEFAULT NULL")

        cursor.execute("SELECT * FROM `loan_requests` WHERE `id` = %s LIMIT 1", (pengajuan_id,))
        req = cursor.fetchone()
        if not req:
            return jsonify({"success": False, "error": "Pengajuan tidak ditemukan"}), 404
            
        current_status = req.get("status")
        barang_id = req.get("barang_id")
        jumlah_req = int(req.get("jumlah", 1))
        
        if current_status != "pending" and new_status in ["approved", "rejected"]:
             return jsonify({"success": False, "error": "Hanya pengajuan berstatus Pending yang dapat disetujui atau ditolak"}), 400

        if new_status == "approved" and current_status == "pending":
            cursor.execute("SELECT `total_unit`, `available_unit`, `can_borrow`, `unit_details` FROM `inventory_items` WHERE `id` = %s LIMIT 1", (barang_id,))
            inv = cursor.fetchone()
            
            if not inv:
                 return jsonify({"success": False, "error": "Barang tidak ditemukan di inventaris"}), 404
            if not bool(inv.get("can_borrow", 1)):
                 return jsonify({"success": False, "error": "Barang tidak dapat dipinjam"}), 400
                 
            available = int(inv.get("available_unit", 0))
            if available < jumlah_req:
                 return jsonify({"success": False, "error": f"Stok tersedia ({available}) tidak mencukupi permintaan ({jumlah_req})"}), 400
                 
            unit_details_raw = safe_json_loads(inv.get("unit_details"), [])
            units_changed = 0
            
            for unit in unit_details_raw:
                 if unit.get("status") == "Tersedia" and units_changed < jumlah_req:
                      unit["status"] = "Dipinjam"
                      unit["reason"] = f"Dipinjam (Req ID: {pengajuan_id})"
                      unit["available"] = False
                      units_changed += 1
                      
            new_available = available - jumlah_req
            new_inv_status = "Dipinjam" if new_available <= 0 else "Tersedia"
            
            cursor.execute(
                 "UPDATE `inventory_items` SET `available_unit` = %s, `status` = %s, `unit_details` = %s WHERE `id` = %s", 
                 (new_available, new_inv_status, json.dumps(unit_details_raw, ensure_ascii=False), barang_id)
            )
            
        elif (new_status in ["cancelled", "returned"]) and (current_status in ["approved", "taken"]):
             cursor.execute("SELECT `total_unit`, `available_unit`, `unit_details` FROM `inventory_items` WHERE `id` = %s LIMIT 1", (barang_id,))
             inv = cursor.fetchone()
             if inv:
                 total = int(inv.get("total_unit", 1))
                 available = int(inv.get("available_unit", 0))
                 
                 unit_details_raw = safe_json_loads(inv.get("unit_details"), [])
                 units_restored = 0
                 
                 for unit in unit_details_raw:
                      if unit.get("reason") == f"Dipinjam (Req ID: {pengajuan_id})":
                           unit["status"] = "Tersedia"
                           unit["reason"] = ""
                           unit["available"] = True
                           units_restored += 1
                 
                 if units_restored == 0:
                      for unit in unit_details_raw:
                           if unit.get("status") == "Dipinjam" and units_restored < jumlah_req:
                                unit["status"] = "Tersedia"
                                unit["reason"] = ""
                                unit["available"] = True
                                units_restored += 1
                 
                 new_available = min(total, available + jumlah_req)
                 new_inv_status = "Tersedia" if new_available > 0 else "Dipinjam"
                 
                 cursor.execute(
                      "UPDATE `inventory_items` SET `available_unit` = %s, `status` = %s, `unit_details` = %s WHERE `id` = %s", 
                      (new_available, new_inv_status, json.dumps(unit_details_raw, ensure_ascii=False), barang_id)
                 )
        
        admin_name = str(session.get("nama") or session.get("username") or "Admin")
        approved_at = datetime.utcnow() if new_status != "pending" else None
        
        cursor.execute(
            """
            UPDATE `loan_requests` 
            SET `status` = %s, `admin_note` = %s, `approved_by` = %s, `approved_at` = COALESCE(`approved_at`, %s)
            WHERE `id` = %s
            """,
            (new_status, admin_note, admin_name, approved_at, pengajuan_id)
        )
        
        try:
            ensure_notifications_schema()
            barang_name = req.get("barang_name", "Barang")
            member_id = req.get("member_id")
            
            if member_id:
                status_label = "Disetujui" if new_status == "approved" else ("Ditolak" if new_status == "rejected" else ("Dibatalkan" if new_status == "cancelled" else new_status))
                
                notif_title = f"Peminjaman {status_label}: {barang_name}"
                notif_body = f"Pengajuan peminjaman Anda untuk {barang_name} telah {status_label.lower()} oleh Admin."
                if admin_note:
                    notif_body += f"<br>Catatan: {html.escape(admin_note)}"

                create_notification(
                    cursor, 
                    "peminjaman", 
                    notif_title, 
                    notif_body, 
                    url_for('dashboard_anggota') if False else "/pengajuan-peminjaman-barang-anggota.html", 
                    {"pengajuan_id": pengajuan_id, "target_user_id": member_id}, 
                    target_role="user"
                )
        except Exception as e:
            print(f"[WARN] Gagal mengirim notifikasi update status peminjaman: {e}")

        conn.commit()
        return jsonify({"success": True})
        
    except Exception as e:
        conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 400
    finally:
        cursor.close()
        conn.close()


# Source legacy app.py lines 9467-9598 | routes: /api/pengajuan/export.xlsx
@app.route("/api/pengajuan/export.xlsx", methods=["GET"])
def export_pengajuan_excel():
    ensure_auth_schema()
    role = session.get("role", "")
    if role not in ["admin", "super_admin"]:
        abort(403)
        
    status_filter = request.args.get('status')
    from_date = request.args.get('from_date')
    to_date = request.args.get('to_date')
    
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        ensure_loan_schema(cursor)
        query = """
            SELECT l.*, a.nama as member_name 
            FROM `loan_requests` l
            LEFT JOIN `anggota` a ON l.member_id = a.id
            WHERE 1=1
        """
        params = []
        if status_filter and status_filter != 'all':
            query += " AND l.`status` = %s"
            params.append(status_filter)
        if from_date:
            query += " AND l.`tanggal_mulai` >= %s"
            params.append(from_date)
        if to_date:
            query += " AND l.`tanggal_mulai` <= %s"
            params.append(to_date)
            
        query += " ORDER BY l.`created_at` DESC"
        cursor.execute(query, tuple(params))
        items = cursor.fetchall() or []
        
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        
        headers = [
            "ID Laporan", "Peminjam", "Barang", "Qty", 
            "Mulai Pinjam", "Target Kembali", "Status",
            "Waktu Ambil", "Kondisi Ambil", "Catatan Ambil", "Bukti Ambil",
            "Waktu Kembali", "Kondisi Kembali", "Catatan Kembali", "Bukti Kembali", "Catatan Admin"
        ]
        
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Riwayat Peminjaman"
        worksheet.append(headers)
        
        base_url = PUBLIC_BASE_URL
        
        for item in items:
            p_info = safe_json_loads(item.get("pickup_info"), {})
            r_info = safe_json_loads(item.get("return_info"), {})
            
            p_kondisi = ", ".join([u.get("status", "") for u in p_info.get("units", [])]) if p_info.get("units") else "-"
            r_kondisi = ", ".join([u.get("status", "") for u in r_info.get("units", [])]) if r_info.get("units") else "-"
            
            p_catatan = " | ".join([u.get("reason", "") for u in p_info.get("units", []) if u.get("reason")]) if p_info.get("units") else "-"
            r_catatan = " | ".join([u.get("reason", "") for u in r_info.get("units", []) if u.get("reason")]) if r_info.get("units") else "-"
            
            p_foto = f"{base_url}{p_info.get('photo')}" if p_info.get("photo") and str(p_info.get("photo")).startswith("/") else (p_info.get("photo") or "-")
            r_foto = f"{base_url}{r_info.get('photo')}" if r_info.get("photo") and str(r_info.get("photo")).startswith("/") else (r_info.get("photo") or "-")

            mulai_str = f"{item.get('tanggal_mulai')} {str(item.get('waktu_mulai', ''))[:5]}"
            selesai_str = f"{item.get('tanggal_selesai')} {str(item.get('waktu_selesai', ''))[:5]}"
            p_waktu = f"{p_info.get('date', '')} {p_info.get('time', '')}" if p_info.get("date") else "-"
            r_waktu = f"{r_info.get('date', '')} {r_info.get('time', '')}" if r_info.get("date") else "-"
            
            status_map = {"pending": "Pending", "approved": "Disetujui", "taken": "Dipinjam", "returned": "Selesai", "rejected": "Ditolak", "cancelled": "Dibatalkan"}
            status_indo = status_map.get(item.get("status"), item.get("status"))
            
            row = [
                item.get("id"),
                item.get("member_name") or "-",
                f"{item.get('barang_name')} ({item.get('barang_code')})",
                item.get("jumlah", 1),
                mulai_str,
                selesai_str,
                status_indo,
                p_waktu,
                p_kondisi,
                p_catatan,
                p_foto,
                r_waktu,
                r_kondisi,
                r_catatan,
                r_foto,
                item.get("admin_note") or "-"
            ]
            worksheet.append(row)
            
        header_fill = PatternFill("solid", fgColor="7F0000")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2), 2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrapText=True)

        for column_cells in worksheet.columns:
            max_length = 0
            column_letter = column_cells[0].column_letter
            for cell in column_cells:
                try:
                    cell_value = str(cell.value or "")
                    if len(cell_value) > max_length:
                        max_length = len(cell_value)
                except Exception:
                    continue
            worksheet.column_dimensions[column_letter].width = min(max_length + 2, 40)

        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        
        filename = f"riwayat-peminjaman-{datetime.now().strftime('%Y%m%d%H%M')}.xlsx"
        
        return send_file(
            buffer,
            as_attachment=False, # Karena file excel browser cenderung akan download langsung
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    finally:
        cursor.close()
        conn.close()


# Source legacy app.py lines 9600-9726 | routes: /api/pengajuan/export.pdf
@app.route("/api/pengajuan/export.pdf", methods=["GET"])
def export_pengajuan_pdf():
    ensure_auth_schema()
    role = session.get("role", "")
    if role not in ["admin", "super_admin"]:
        abort(403)
        
    status_filter = request.args.get('status')
    from_date = request.args.get('from_date')
    to_date = request.args.get('to_date')
    
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        ensure_loan_schema(cursor)
        query = """
            SELECT l.*, a.nama as member_name 
            FROM `loan_requests` l
            LEFT JOIN `anggota` a ON l.member_id = a.id
            WHERE 1=1
        """
        params = []
        if status_filter and status_filter != 'all':
            query += " AND l.`status` = %s"
            params.append(status_filter)
        if from_date:
            query += " AND l.`tanggal_mulai` >= %s"
            params.append(from_date)
        if to_date:
            query += " AND l.`tanggal_mulai` <= %s"
            params.append(to_date)
            
        query += " ORDER BY l.`created_at` DESC"
        cursor.execute(query, tuple(params))
        items = cursor.fetchall() or []
        
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Spacer, Table, TableStyle, Paragraph

        headers = ["Peminjam", "Barang", "Tgl Ambil", "Kondisi/Bukti Ambil", "Tgl Kembali", "Kondisi/Bukti Kembali", "Status"]
        rows = []
        base_url = PUBLIC_BASE_URL
        
        for item in items:
            p_info = safe_json_loads(item.get("pickup_info"), {})
            r_info = safe_json_loads(item.get("return_info"), {})
            
            p_waktu = f"{p_info.get('date', '')} {p_info.get('time', '')}" if p_info.get("date") else "-"
            r_waktu = f"{r_info.get('date', '')} {r_info.get('time', '')}" if r_info.get("date") else "-"
            
            p_kondisi = ", ".join([u.get("status", "") for u in p_info.get("units", [])]) if p_info.get("units") else "-"
            r_kondisi = ", ".join([u.get("status", "") for u in r_info.get("units", [])]) if r_info.get("units") else "-"
            
            p_foto = f"{base_url}{p_info.get('photo')}" if p_info.get("photo") and str(p_info.get("photo")).startswith("/") else (p_info.get("photo") or "")
            r_foto = f"{base_url}{r_info.get('photo')}" if r_info.get("photo") and str(r_info.get("photo")).startswith("/") else (r_info.get("photo") or "")
            
            p_cell = f"{p_kondisi}<br/><a href='{p_foto}' color='blue'>Lihat Foto Ambil</a>" if p_foto else p_kondisi
            r_cell = f"{r_kondisi}<br/><a href='{r_foto}' color='blue'>Lihat Foto Kembali</a>" if r_foto else r_kondisi
            
            status_map = {"pending": "Pending", "approved": "Disetujui", "taken": "Dipinjam", "returned": "Selesai", "rejected": "Ditolak", "cancelled": "Dibatalkan"}
            
            rows.append([
                item.get("member_name") or "-",
                item.get("barang_name") or "-",
                p_waktu,
                p_cell,
                r_waktu,
                r_cell,
                status_map.get(item.get("status"), item.get("status"))
            ])
            
        buffer = BytesIO()
        document = SimpleDocTemplate(
            buffer, pagesize=landscape(A4),
            rightMargin=8 * mm, leftMargin=8 * mm, topMargin=10 * mm, bottomMargin=10 * mm,
        )
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle("Title", parent=styles["Title"], fontName="Helvetica-Bold", fontSize=14, leading=16, textColor=colors.HexColor("#7F0000"))
        normal_style = ParagraphStyle("Normal", parent=styles["BodyText"], fontName="Helvetica", fontSize=8, leading=10)
        cell_style = ParagraphStyle("Cell", parent=styles["BodyText"], fontName="Helvetica", fontSize=7, leading=9)

        elements = [
            Paragraph("Rekap Data Riwayat Peminjaman dan Pengembalian", title_style),
            Spacer(1, 4 * mm),
            Paragraph(f"Dicetak pada: {datetime.now().strftime('%d %B %Y %H:%M')} WIB", normal_style),
            Spacer(1, 4 * mm),
        ]

        wrapped_rows = []
        for row in rows:
            wrapped_row = []
            for cell_data in row:
                wrapped_row.append(Paragraph(str(cell_data), cell_style))
            wrapped_rows.append(wrapped_row)

        table_data = [headers] + wrapped_rows if rows else [headers, ["-" for _ in headers]]
        page_width = landscape(A4)[0] - (16 * mm) 
        col_widths = [0.15*page_width, 0.20*page_width, 0.12*page_width, 0.18*page_width, 0.12*page_width, 0.18*page_width, 0.05*page_width]
        
        table = Table(table_data, colWidths=col_widths, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#7F0000")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("ALIGN", (0, 0), (0, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D1D5DB")),
        ]))
        elements.append(table)
        document.build(elements)
        buffer.seek(0)
        
        filename = f"riwayat-peminjaman-{datetime.now().strftime('%Y%m%d%H%M')}.pdf"
        
        return send_file(
            buffer,
            as_attachment=False,  # Buka di tab baru (preview)
            download_name=filename,
            mimetype="application/pdf",
        )
    finally:
        cursor.close()
        conn.close()


