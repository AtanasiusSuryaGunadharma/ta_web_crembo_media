from crembo_app.services import core as _core

# Memuat seluruh helper, service, dan objek Flask dari core agar potongan kode route
# tetap kompatibel setelah dipisah dari app.py monolitik.
globals().update({
    name: getattr(_core, name)
    for name in dir(_core)
    if not (name.startswith("__") and name.endswith("__"))
})

# Controller: Inventory Controller

# Source legacy app.py lines 882-895 | routes: /api/inventory/categories
@app.route("/api/inventory/categories", methods=["GET"])
def get_inventory_categories():
    ensure_auth_schema()
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        cursor.execute("SELECT `name` FROM `inventory_categories` ORDER BY `order_index` ASC, `name` ASC")
        categories = [row["name"] for row in cursor.fetchall() or [] if row.get("name")]
        if not categories:
            categories = INVENTORY_DEFAULT_CATEGORIES[:]
        return jsonify({"success": True, "categories": categories})
    finally:
        cursor.close()
        conn.close()


# Source legacy app.py lines 897-916 | routes: /api/inventory/categories
@app.route("/api/inventory/categories", methods=["POST"])
def create_inventory_category():
    ensure_auth_schema()
    data = request.get_json(silent=True) or {}
    category_name = normalize_text(data.get("name"))
    if not category_name:
        return jsonify({"success": False, "error": "Nama kategori wajib diisi."}), 400

    conn = mysql_connection()
    cursor = conn.cursor()
    try:
        ensure_inventory_category(cursor, category_name)
        conn.commit()
        return jsonify({"success": True, "name": category_name})
    except Exception as exc:
        conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 400
    finally:
        cursor.close()
        conn.close()


# Source legacy app.py lines 918-948 | routes: /api/inventory/categories/<category_name>
@app.route("/api/inventory/categories/<category_name>", methods=["PUT"])
def update_inventory_category(category_name):
    ensure_auth_schema()
    data = request.get_json(silent=True) or {}
    new_name = normalize_text(data.get("name"))
    old_name = normalize_text(category_name)
    if not old_name or not new_name:
        return jsonify({"success": False, "error": "Nama kategori tidak valid."}), 400

    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        cursor.execute("SELECT `id` FROM `inventory_categories` WHERE `name` = %s LIMIT 1", (old_name,))
        existing = cursor.fetchone()
        if not existing:
            return jsonify({"success": False, "error": "Kategori tidak ditemukan."}), 404
        if new_name != old_name:
            cursor.execute("SELECT COUNT(*) FROM `inventory_categories` WHERE `name` = %s", (new_name,))
            if fetch_scalar_value(cursor.fetchone(), 0) > 0:
                return jsonify({"success": False, "error": "Nama kategori sudah dipakai."}), 400

        cursor.execute("UPDATE `inventory_categories` SET `name` = %s WHERE `name` = %s", (new_name, old_name))
        cursor.execute("UPDATE `inventory_items` SET `category` = %s WHERE `category` = %s", (new_name, old_name))
        conn.commit()
        return jsonify({"success": True, "name": new_name})
    except Exception as exc:
        conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 400
    finally:
        cursor.close()
        conn.close()


# Source legacy app.py lines 950-971 | routes: /api/inventory/categories/<category_name>
@app.route("/api/inventory/categories/<category_name>", methods=["DELETE"])
def delete_inventory_category(category_name):
    ensure_auth_schema()
    old_name = normalize_text(category_name)
    if not old_name:
        return jsonify({"success": False, "error": "Kategori tidak valid."}), 400

    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        cursor.execute("SELECT COUNT(*) FROM `inventory_items` WHERE `category` = %s", (old_name,))
        if fetch_scalar_value(cursor.fetchone(), 0) > 0:
            return jsonify({"success": False, "error": "Kategori ini masih dipakai oleh data inventaris."}), 409
        cursor.execute("DELETE FROM `inventory_categories` WHERE `name` = %s", (old_name,))
        conn.commit()
        return jsonify({"success": True})
    except Exception as exc:
        conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 400
    finally:
        cursor.close()
        conn.close()


# Source legacy app.py lines 1186-1225 | routes: /api/inventory/items
@app.route("/api/inventory/items", methods=["GET"])
def get_inventory_items():
    ensure_auth_schema()
    filters = inventory_request_filters(request.args)
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        cursor.execute(
            """
             SELECT `id`, `code`, `name`, `category`, `location`, `purchase_date`, `purchase_price`,
                 `total_unit`, `available_unit`, `has_multiple`, `can_borrow`, `status`,
                 `notes`, `photos`, `unit_details`, `created_at`, `updated_at`
            FROM `inventory_items`
            ORDER BY `updated_at` DESC, `code` ASC
            """
        )
        raw_items = [inventory_item_row_to_dict(row) for row in cursor.fetchall() or []]
        categories = []
        cursor.execute("SELECT `name` FROM `inventory_categories` ORDER BY `order_index` ASC, `name` ASC")
        for row in cursor.fetchall() or []:
            if row.get("name"):
                categories.append(row["name"])

        items = inventory_filter_items(
            raw_items,
            search=filters["search"],
            category=filters["category"],
            status=filters["status"],
            sort_mode=filters["sort"],
        )
        return jsonify({
            "success": True,
            "items": items,
            "categories": categories or INVENTORY_DEFAULT_CATEGORIES[:],
            "summary": inventory_summary_from_items(items),
            "filters": filters,
        })
    finally:
        cursor.close()
        conn.close()


# Source legacy app.py lines 1227-1250 | routes: /api/inventory/items/<item_id>
@app.route("/api/inventory/items/<item_id>", methods=["GET"])
def get_inventory_item(item_id):
    ensure_auth_schema()
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        cursor.execute(
            """
             SELECT `id`, `code`, `name`, `category`, `location`, `purchase_date`, `purchase_price`,
                 `total_unit`, `available_unit`, `has_multiple`, `can_borrow`, `status`,
                 `notes`, `photos`, `unit_details`, `created_at`, `updated_at`
            FROM `inventory_items`
            WHERE `id` = %s
            LIMIT 1
            """,
            (item_id,),
        )
        row = cursor.fetchone()
        if not row:
            return jsonify({"success": False, "error": "Barang inventaris tidak ditemukan."}), 404
        return jsonify({"success": True, "item": inventory_item_row_to_dict(row)})
    finally:
        cursor.close()
        conn.close()


# Source legacy app.py lines 1252-1298 | routes: /api/inventory/items
@app.route("/api/inventory/items", methods=["POST"])
def create_inventory_item():
    ensure_auth_schema()
    data = request.get_json(silent=True) or {}

    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        payload = inventory_item_payload_from_request(data)
        cursor.execute("SELECT COUNT(*) FROM `inventory_items` WHERE `code` = %s", (payload["code"],))
        if fetch_scalar_value(cursor.fetchone(), 0) > 0:
            return jsonify({"success": False, "error": "Kode barang sudah dipakai."}), 400

        ensure_inventory_category(cursor, payload["category"])
        item_id = normalize_text(data.get("id")) or f"inv-{int(time.time() * 1000)}"
        cursor.execute(
            """
            INSERT INTO `inventory_items`
            (`id`, `code`, `name`, `category`, `location`, `purchase_date`, `purchase_price`, `total_unit`, `available_unit`, `has_multiple`, `can_borrow`, `status`, `notes`, `photos`, `unit_details`)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """,
            (
                item_id,
                payload["code"],
                payload["name"],
                payload["category"],
                payload["location"],
                payload["purchase_date"],
                payload["purchase_price"],
                payload["total_unit"],
                payload["available_unit"],
                payload["has_multiple"],
                payload["can_borrow"],
                payload["status"],
                payload["notes"],
                json.dumps(payload["photos"], ensure_ascii=False),
                json.dumps(payload["unit_details"], ensure_ascii=False),
            ),
        )
        conn.commit()
        return jsonify({"success": True, "item": get_inventory_item_response(cursor, item_id)})
    except Exception as exc:
        conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 400
    finally:
        cursor.close()
        conn.close()


# Source legacy app.py lines 1315-1390 | routes: /api/inventory/items/<item_id>
@app.route("/api/inventory/items/<item_id>", methods=["PUT"])
def update_inventory_item(item_id):
    ensure_auth_schema()
    data = request.get_json(silent=True) or {}

    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        cursor.execute("SELECT * FROM `inventory_items` WHERE `id` = %s LIMIT 1", (item_id,))
        existing = cursor.fetchone()
        if not existing:
            return jsonify({"success": False, "error": "Barang inventaris tidak ditemukan."}), 404

        payload = inventory_item_payload_from_request(data, existing)
        cursor.execute(
            "SELECT COUNT(*) FROM `inventory_items` WHERE `code` = %s AND `id` <> %s",
            (payload["code"], item_id),
        )
        if fetch_scalar_value(cursor.fetchone(), 0) > 0:
            return jsonify({"success": False, "error": "Kode barang sudah dipakai."}), 400

        ensure_inventory_category(cursor, payload["category"])
        old_photos = normalize_inventory_photos(safe_json_loads(existing.get("photos"), []))
        new_photos = payload["photos"]

        cursor.execute(
            """
            UPDATE `inventory_items`
            SET `code` = %s,
                `name` = %s,
                `category` = %s,
                `location` = %s,
                `purchase_date` = %s,
                `purchase_price` = %s,
                `total_unit` = %s,
                `available_unit` = %s,
                `has_multiple` = %s,
                `can_borrow` = %s,
                `status` = %s,
                `notes` = %s,
                `photos` = %s,
                `unit_details` = %s
            WHERE `id` = %s
            """,
            (
                payload["code"],
                payload["name"],
                payload["category"],
                payload["location"],
                payload["purchase_date"],
                payload["purchase_price"],
                payload["total_unit"],
                payload["available_unit"],
                payload["has_multiple"],
                payload["can_borrow"],
                payload["status"],
                payload["notes"],
                json.dumps(new_photos, ensure_ascii=False),
                json.dumps(payload["unit_details"], ensure_ascii=False),
                item_id,
            ),
        )
        conn.commit()

        old_urls = {photo.get("url") for photo in old_photos if photo.get("url")}
        new_urls = {photo.get("url") for photo in new_photos if photo.get("url")}
        for removed_url in old_urls - new_urls:
            remove_physical_file(removed_url)

        return jsonify({"success": True, "item": get_inventory_item_response(cursor, item_id)})
    except Exception as exc:
        conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 400
    finally:
        cursor.close()
        conn.close()


# Source legacy app.py lines 1392-1415 | routes: /api/inventory/items/<item_id>
@app.route("/api/inventory/items/<item_id>", methods=["DELETE"])
def delete_inventory_item(item_id):
    ensure_auth_schema()
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        cursor.execute("SELECT `photos`, `name` FROM `inventory_items` WHERE `id` = %s LIMIT 1", (item_id,))
        existing = cursor.fetchone()
        if not existing:
            return jsonify({"success": False, "error": "Barang inventaris tidak ditemukan."}), 404

        cursor.execute("DELETE FROM `inventory_items` WHERE `id` = %s", (item_id,))
        conn.commit()

        for photo in normalize_inventory_photos(safe_json_loads(existing.get("photos"), [])):
            remove_physical_file(photo.get("url") or "")

        return jsonify({"success": True})
    except Exception as exc:
        conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 400
    finally:
        cursor.close()
        conn.close()


# Source legacy app.py lines 1417-1477 | routes: /api/inventory/export.xlsx
@app.route("/api/inventory/export.xlsx", methods=["GET"])
def export_inventory_excel():
    ensure_auth_schema()
    filters = inventory_request_filters(request.args)
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        items = fetch_inventory_items_for_report(cursor, filters)

        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill

        headers, rows = inventory_export_rows(items)
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Inventaris"
        worksheet.append(headers)
        for row in rows:
            worksheet.append(row)

        header_fill = PatternFill("solid", fgColor="7F0000")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2), 2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrapText=True)

        for column_cells in worksheet.columns:
            max_length = 0
            column_letter = column_cells[0].column_letter
            for cell in column_cells:
                try:
                    cell_value = str(cell.value or "")
                    if "\n" in cell_value:
                        lines = cell_value.split("\n")
                        line_lengths = [len(l) for l in lines]
                        if line_lengths and max(line_lengths) > max_length:
                            max_length = max(line_lengths)
                    elif len(cell_value) > max_length:
                        max_length = len(cell_value)
                except Exception:
                    continue
            worksheet.column_dimensions[column_letter].width = min(max_length + 4, 60)

        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        
        return send_file(
            buffer,
            as_attachment=True,
            download_name=inventory_export_filename("xlsx"),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    finally:
        cursor.close()
        conn.close()


# Source legacy app.py lines 1479-1571 | routes: /api/inventory/export.pdf
@app.route("/api/inventory/export.pdf", methods=["GET"])
def export_inventory_pdf():
    ensure_auth_schema()
    filters = inventory_request_filters(request.args)
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        items = fetch_inventory_items_for_report(cursor, filters)

        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Spacer, Table, TableStyle, Paragraph

        headers, rows = inventory_export_rows(items)
        buffer = BytesIO()
        document = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            rightMargin=8 * mm,
            leftMargin=8 * mm,
            topMargin=10 * mm,
            bottomMargin=10 * mm,
        )
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle("InventoryTitle", parent=styles["Title"], fontName="Helvetica-Bold", fontSize=14, leading=16, textColor=colors.HexColor("#7F0000"))
        normal_style = ParagraphStyle("InventoryNormal", parent=styles["BodyText"], fontName="Helvetica", fontSize=8, leading=10)
        cell_style = ParagraphStyle("CellNormal", parent=styles["BodyText"], fontName="Helvetica", fontSize=7, leading=9, wordWrap='CJK')

        elements = [
            Paragraph("Laporan Data Inventaris Barang", title_style),
            Spacer(1, 4 * mm),
            Paragraph(f"Filter: {filters.get('category', 'all')} | Cari: {filters.get('search', '') or '-'} | Urutan: {filters.get('sort', 'updated-desc')}", normal_style),
            Spacer(1, 4 * mm),
        ]

        wrapped_rows = []
        for row in rows:
            wrapped_row = []
            for cell_data in row:
                text = str(cell_data).replace('\n', '<br/>')
                wrapped_row.append(Paragraph(text, cell_style))
            wrapped_rows.append(wrapped_row)

        table_data = [headers] + wrapped_rows if rows else [headers, ["-" for _ in headers]]
        
        page_width = landscape(A4)[0] - (16 * mm) 
        col_widths = [
            0.07 * page_width,
            0.10 * page_width,
            0.06 * page_width,
            0.07 * page_width,
            0.07 * page_width,
            0.07 * page_width,
            0.03 * page_width,
            0.04 * page_width,
            0.04 * page_width,
            0.04 * page_width,
            0.18 * page_width,
            0.10 * page_width,
            0.08 * page_width,
            0.05 * page_width
        ]
        
        table = Table(table_data, colWidths=col_widths, repeatRows=1)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#7F0000")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 7),
                    ("ALIGN", (0, 0), (0, -1), "CENTER"),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D1D5DB")),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.HexColor("#FFFFFF")]),
                ]
            )
        )
        elements.append(table)
        document.build(elements)
        buffer.seek(0)
        
        return send_file(
            buffer,
            as_attachment=False, 
            download_name=inventory_export_filename("pdf"),
            mimetype="application/pdf",
        )
    finally:
        cursor.close()
        conn.close()


# Source legacy app.py lines 2283-2314 | routes: /api/inventory/search
@app.route("/api/inventory/search", methods=["GET"])
def search_inventory_items():
    """Search inventory items for auto-lookup"""
    ensure_auth_schema()
    query_str = request.args.get('q', '').strip().lower()
    if not query_str or len(query_str) < 2:
        return jsonify({"success": True, "items": []})
    
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        ensure_inventory_schema(cursor)
        
        cursor.execute("""
            SELECT `id`, `code`, `name`, `photos` FROM `inventory_items`
            WHERE LOWER(`name`) LIKE %s OR LOWER(`code`) LIKE %s
            LIMIT 20
        """, (f"%{query_str}%", f"%{query_str}%"))
        
        rows = cursor.fetchall() or []
        items = []
        for r in rows:
            items.append({
                "id": r.get("id"),
                "code": r.get("code"),
                "name": r.get("name"),
                "photos": normalize_inventory_photos(r.get("photos"))
            })
        return jsonify({"success": True, "items": items})
    finally:
        cursor.close()
        conn.close()


