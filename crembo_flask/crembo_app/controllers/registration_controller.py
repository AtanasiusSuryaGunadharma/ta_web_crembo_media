from crembo_app.services import core as _core

# Memuat seluruh helper, service, dan objek Flask dari core agar potongan kode route
# tetap kompatibel setelah dipisah dari app.py monolitik.
globals().update({
    name: getattr(_core, name)
    for name in dir(_core)
    if not (name.startswith("__") and name.endswith("__"))
})

# Controller: Registration Controller

# Source legacy app.py lines 8081-8109 | routes: /api/registration/forms
@app.route("/api/registration/forms", methods=["GET"])
def get_registration_forms():
    ensure_registration_form_schema()
    scope = (request.args.get("scope") or "public").strip().lower()
    viewer = current_user_context()
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        cursor.execute("SELECT * FROM `registration_forms` ORDER BY `updated_at` DESC, `created_at` DESC")
        rows = cursor.fetchall() or []
        counts = registration_form_submission_counts(cursor)
        forms = [registration_form_row_to_dict(row, counts.get(str(row.get("id") or ""), 0)) for row in rows]

        if scope == "admin":
            if not can_manage_registration_forms():
                return jsonify({"success": False, "error": "Forbidden"}), 403
            return jsonify(forms)

        accessible_forms = []
        for form in forms:
            if form.get("visibility") == "draft":
                continue
            # if form.get("target") == "internal" and not viewer.get("logged_in"):
            #     continue
            accessible_forms.append(form)
        return jsonify(accessible_forms)
    finally:
        cursor.close()
        conn.close()


# Source legacy app.py lines 8112-8131 | routes: /api/registration/forms/<form_id>
@app.route("/api/registration/forms/<form_id>", methods=["GET"])
def get_registration_form_detail(form_id):
    ensure_registration_form_schema()
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        form = registration_form_lookup(cursor, form_id, include_counts=True)
        if not form:
            return jsonify({"success": False, "error": "Form not found"}), 404

        viewer = current_user_context()
        if form.get("visibility") == "draft" and not can_manage_registration_forms():
            return jsonify({"success": False, "error": "Forbidden"}), 403
        # if form.get("target") == "internal" and not viewer.get("logged_in") and not can_manage_registration_forms():
        #     return jsonify({"success": False, "error": "Forbidden"}), 403

        return jsonify(form)
    finally:
        cursor.close()
        conn.close()


# Source legacy app.py lines 8135-8205 | routes: /api/registration/forms
@app.route("/api/registration/forms", methods=["POST"])
def create_registration_form():
    ensure_registration_form_schema()
    if not can_manage_registration_forms():
        return jsonify({"success": False, "error": "Forbidden"}), 403

    data = request.json or {}
    values = registration_form_payload_to_db_values(data)
    if not values["title"] or not values["open_date"] or not values["close_date"]:
        return jsonify({"success": False, "error": "Judul, tanggal pembukaan, dan tanggal penutupan wajib diisi."}), 400
    if not values["fields_json"] or values["fields_json"] == "[]":
        return jsonify({"success": False, "error": "Tambahkan minimal 1 pertanyaan."}), 400

    try:
        open_date = datetime.strptime(values["open_date"], "%Y-%m-%d").date()
        close_date = datetime.strptime(values["close_date"], "%Y-%m-%d").date()
    except ValueError:
        return jsonify({"success": False, "error": "Format tanggal tidak valid."}), 400
    if close_date < open_date:
        return jsonify({"success": False, "error": "Tanggal penutupan harus sama atau setelah tanggal pembukaan."}), 400

    form_id = data.get("id") or f"registration-form-{int(time.time() * 1000)}"
    conn = mysql_connection()
    cursor = conn.cursor()
    try:
        cursor.execute(
            """
            INSERT INTO `registration_forms`
            (`id`, `title`, `description`, `target`, `visibility`, `open_date`, `close_date`, `quota`, `image_url`, `image_name`, `attachments`, `fields_json`, `created_by`, `created_by_name`, `created_by_role`)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """,
            (
                form_id,
                values["title"],
                values["description"],
                values["target"],
                values["visibility"],
                values["open_date"],
                values["close_date"],
                values["quota"],
                values["image_url"],
                values["image_name"],
                values["attachments"],
                values["fields_json"],
                str(session.get("user_id") or ""),
                str(session.get("nama") or session.get("username") or ""),
                str(session.get("role") or ""),
            ),
        )
        conn.commit()
        try:
            ensure_notifications_schema()
            nc = conn.cursor()
            try:
                # PERBAIKAN: ubah url_for agar mengarah ke halaman public FE, bukan ke API
                create_notification(nc, "form", f"Form Pendaftaran Baru: {values['title']}", values.get('description') or "Terdapat form pendaftaran baru.", url_for('public_registration_form_detail_page', form_id=form_id), {"form_id": form_id})
                conn.commit()
            finally:
                nc.close()
        except Exception:
            pass
        return jsonify({"success": True, "id": form_id})
    except mysql.connector.IntegrityError:
        conn.rollback()
        return jsonify({"success": False, "error": "ID form sudah digunakan."}), 409
    except Exception as error:
        conn.rollback()
        return jsonify({"success": False, "error": str(error)}), 400
    finally:
        cursor.close()
        conn.close()


# Source legacy app.py lines 8208-8297 | routes: /api/registration/forms/<form_id>
@app.route("/api/registration/forms/<form_id>", methods=["PUT"])
def update_registration_form(form_id):
    ensure_registration_form_schema()
    if not can_manage_registration_forms():
        return jsonify({"success": False, "error": "Forbidden"}), 403

    data = request.json or {}
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        cursor.execute("SELECT * FROM `registration_forms` WHERE `id` = %s LIMIT 1", (form_id,))
        existing = cursor.fetchone()
        if not existing:
            return jsonify({"success": False, "error": "Form not found"}), 404

        values = registration_form_payload_to_db_values(data, existing)
        if not values["title"] or not values["open_date"] or not values["close_date"]:
            return jsonify({"success": False, "error": "Judul, tanggal pembukaan, dan tanggal penutupan wajib diisi."}), 400
        if not values["fields_json"] or values["fields_json"] == "[]":
            return jsonify({"success": False, "error": "Tambahkan minimal 1 pertanyaan."}), 400

        try:
            open_date = datetime.strptime(values["open_date"], "%Y-%m-%d").date()
            close_date = datetime.strptime(values["close_date"], "%Y-%m-%d").date()
        except ValueError:
            return jsonify({"success": False, "error": "Format tanggal tidak valid."}), 400
        if close_date < open_date:
            return jsonify({"success": False, "error": "Tanggal penutupan harus sama atau setelah tanggal pembukaan."}), 400

        cursor.execute(
            """
            UPDATE `registration_forms`
            SET `title` = %s,
                `description` = %s,
                `target` = %s,
                `visibility` = %s,
                `open_date` = %s,
                `close_date` = %s,
                `quota` = %s,
                `image_url` = %s,
                `image_name` = %s,
                `attachments` = %s,
                `fields_json` = %s
            WHERE `id` = %s
            """,
            (
                values["title"],
                values["description"],
                values["target"],
                values["visibility"],
                values["open_date"],
                values["close_date"],
                values["quota"],
                values["image_url"],
                values["image_name"],
                values["attachments"],
                values["fields_json"],
                form_id,
            ),
        )
        conn.commit()

        # Cleanup file lama yang sudah dihapus/diganti saat Edit
        if existing:
            try:
                # Cleanup Image
                if existing.get("image_url") and existing.get("image_url") != values["image_url"]:
                    remove_physical_file(existing["image_url"])
                
                # Cleanup Attachments
                old_atts_raw = existing.get("attachments") or "[]"
                old_atts = json.loads(old_atts_raw) if isinstance(old_atts_raw, str) else (old_atts_raw or [])
                new_atts_raw = values.get("attachments") or "[]"
                new_atts = json.loads(new_atts_raw) if isinstance(new_atts_raw, str) else (new_atts_raw or [])
                new_att_urls = [a.get("url") for a in new_atts if isinstance(a, dict) and a.get("url")]
                
                for oa in old_atts:
                    if isinstance(oa, dict) and oa.get("url") and oa.get("url") not in new_att_urls:
                        remove_physical_file(oa["url"])
            except Exception as e:
                print(f"[WARN] Failed to cleanup replaced files for registration form {form_id}: {e}")


        return jsonify({"success": True, "id": form_id})
    except Exception as error:
        conn.rollback()
        return jsonify({"success": False, "error": str(error)}), 400
    finally:
        cursor.close()
        conn.close()


# Source legacy app.py lines 8300-8342 | routes: /api/registration/forms/<form_id>
@app.route("/api/registration/forms/<form_id>", methods=["DELETE"])
def delete_registration_form(form_id):
    ensure_registration_form_schema()
    if not can_manage_registration_forms():
        return jsonify({"success": False, "error": "Forbidden"}), 403

    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        # Fetch form data first to get the files we need to delete
        cursor.execute("SELECT `image_url`, `attachments` FROM `registration_forms` WHERE `id` = %s LIMIT 1", (form_id,))
        form = cursor.fetchone()
        
        if not form:
            return jsonify({"success": False, "error": "Form not found"}), 404

        # Delete the record from the database
        cursor.execute("DELETE FROM `registration_forms` WHERE `id` = %s", (form_id,))
        conn.commit()

        # Clean up physical files
        try:
            # Delete image
            if form.get("image_url"):
                remove_physical_file(form["image_url"])
            
            # Delete attachments
            attachments_raw = form.get("attachments") or "[]"
            attachments = json.loads(attachments_raw) if isinstance(attachments_raw, str) else (attachments_raw or [])
            for att in attachments:
                if isinstance(att, dict) and att.get("url"):
                    remove_physical_file(att["url"])
        except Exception as e:
            print(f"[WARN] Error during physical file cleanup for form {form_id}: {e}")


        return jsonify({"success": True})
    except Exception as error:
        conn.rollback()
        return jsonify({"success": False, "error": str(error)}), 400
    finally:
        cursor.close()
        conn.close()


# Source legacy app.py lines 8345-8363 | routes: /api/registration/forms/<form_id>/submissions
@app.route("/api/registration/forms/<form_id>/submissions", methods=["GET"])
def get_registration_form_submissions(form_id):
    ensure_registration_form_schema()
    if not can_manage_registration_forms():
        return jsonify({"success": False, "error": "Forbidden"}), 403

    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        form = registration_form_lookup(cursor, form_id, include_counts=True)
        if not form:
            return jsonify({"success": False, "error": "Form not found"}), 404

        cursor.execute("SELECT * FROM `registration_form_submissions` WHERE `form_id` = %s ORDER BY `submitted_at` DESC", (form_id,))
        submissions = [registration_submission_row_to_dict(row) for row in cursor.fetchall() or []]
        return jsonify({"form": form, "submissions": submissions, "submissionCount": len(submissions)})
    finally:
        cursor.close()
        conn.close()


# Source legacy app.py lines 8419-8475 | routes: /api/registration/forms/<form_id>/export.xlsx
@app.route("/api/registration/forms/<form_id>/export.xlsx", methods=["GET"])
def export_registration_form_excel(form_id):
    ensure_registration_form_schema()
    if not can_manage_registration_forms():
        return jsonify({"success": False, "error": "Forbidden"}), 403

    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        form = registration_form_lookup(cursor, form_id, include_counts=True)
        if not form:
            return jsonify({"success": False, "error": "Form not found"}), 404
        cursor.execute("SELECT * FROM `registration_form_submissions` WHERE `form_id` = %s ORDER BY `submitted_at` DESC", (form_id,))
        submissions = [registration_submission_row_to_dict(row) for row in cursor.fetchall() or []]

        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill

        headers, rows = registration_export_rows(form, submissions)
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Pendaftar"
        worksheet.append(headers)
        for row in rows:
            worksheet.append(row)

        header_fill = PatternFill("solid", fgColor="7F0000")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for column_cells in worksheet.columns:
            max_length = 0
            column_letter = column_cells[0].column_letter
            for cell in column_cells:
                try:
                    cell_value = str(cell.value or "")
                    if len(cell_value) > max_length:
                        max_length = len(cell_value)
                except Exception:
                    continue
            worksheet.column_dimensions[column_letter].width = min(max_length + 4, 40)

        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return send_file(
            buffer,
            as_attachment=True,
            download_name=registration_export_filename(form, "xlsx"),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    finally:
        cursor.close()
        conn.close()


# Source legacy app.py lines 8478-8540 | routes: /api/registration/forms/<form_id>/export.pdf
@app.route("/api/registration/forms/<form_id>/export.pdf", methods=["GET"])
def export_registration_form_pdf(form_id):
    ensure_registration_form_schema()
    if not can_manage_registration_forms():
        return jsonify({"success": False, "error": "Forbidden"}), 403

    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        form = registration_form_lookup(cursor, form_id, include_counts=True)
        if not form:
            return jsonify({"success": False, "error": "Form not found"}), 404
        cursor.execute("SELECT * FROM `registration_form_submissions` WHERE `form_id` = %s ORDER BY `submitted_at` DESC", (form_id,))
        submissions = [registration_submission_row_to_dict(row) for row in cursor.fetchall() or []]

        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Spacer, Table, TableStyle, Paragraph

        headers, rows = registration_export_rows(form, submissions)
        buffer = BytesIO()
        document = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=8 * mm, leftMargin=8 * mm, topMargin=10 * mm, bottomMargin=10 * mm)
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle("RegistrationTitle", parent=styles["Title"], fontName="Helvetica-Bold", fontSize=14, leading=16, textColor=colors.HexColor("#7F0000"))
        normal_style = ParagraphStyle("RegistrationNormal", parent=styles["BodyText"], fontName="Helvetica", fontSize=8, leading=10)

        elements = [
            Paragraph(f"Data Pendaftar: {form.get('title')}", title_style),
            Spacer(1, 4 * mm),
            Paragraph(f"Total pendaftar: {len(submissions)}", normal_style),
            Spacer(1, 4 * mm),
        ]

        table_data = [headers] + rows if rows else [headers, ["-", "-", "-", "-"] + ["" for _ in headers[4:]]]
        table = Table(table_data, repeatRows=1)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#7F0000")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 7),
                    ("ALIGN", (0, 0), (0, -1), "CENTER"),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D1D5DB")),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.HexColor("#FFFFFF")]),
                ]
            )
        )
        elements.append(table)
        document.build(elements)
        buffer.seek(0)
        return send_file(
            buffer,
            as_attachment=True,
            download_name=registration_export_filename(form, "pdf"),
            mimetype="application/pdf",
        )
    finally:
        cursor.close()
        conn.close()


# Source legacy app.py lines 8543-8600 | routes: /api/registration/forms/<form_id>/submit
@app.route("/api/registration/forms/<form_id>/submit", methods=["POST"])
def submit_registration_form(form_id):
    ensure_registration_form_schema()
    data = request.json or {}
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        form = registration_form_lookup(cursor, form_id, include_counts=True)
        if not form:
            return jsonify({"success": False, "error": "Form not found"}), 404

        submission_count = int(form.get("submissionCount") or 0)
        state = registration_form_status(form, submission_count)
        if not state["open"]:
            return jsonify({"success": False, "error": state["text"]}), 400

        submitter, error_payload = registration_submitter_context(form, data.get("submitter"))
        if error_payload:
            error_body, status_code = error_payload
            return jsonify(error_body), status_code

        answers, error_message = registration_submission_payload_to_rows(form, data.get("answers"))
        if error_message:
            return jsonify({"success": False, "error": error_message}), 400

        cursor.execute(
            "SELECT 1 FROM `registration_form_submissions` WHERE `form_id` = %s AND `submitter_key` = %s LIMIT 1",
            (form_id, submitter["key"]),
        )
        if cursor.fetchone():
            return jsonify({"success": False, "error": "Anda sudah pernah mengirim jawaban untuk form ini."}), 409

        submission_id = f"submission-{int(time.time() * 1000)}-{uuid.uuid4().hex[:6]}"
        cursor.execute(
            """
            INSERT INTO `registration_form_submissions`
            (`id`, `form_id`, `submitter_key`, `submitter_identifier`, `submitter_role`, `submitter_user_id`, `submitter_source`, `answers_json`)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            """,
            (
                submission_id,
                form_id,
                submitter["key"],
                submitter["identifier"],
                submitter["role"],
                submitter["user_id"],
                submitter["source"],
                json.dumps(answers, ensure_ascii=False),
            ),
        )
        conn.commit()
        return jsonify({"success": True, "id": submission_id})
    except Exception as error:
        conn.rollback()
        return jsonify({"success": False, "error": str(error)}), 400
    finally:
        cursor.close()
        conn.close()


# Source legacy app.py lines 8603-8627 | routes: /api/registration/submissions/me
@app.route("/api/registration/submissions/me", methods=["GET"])
def get_my_registration_submissions():
    ensure_registration_form_schema()
    viewer = current_user_context()
    client_key = str(request.args.get("clientKey") or request.headers.get("X-Registration-Client-Key") or "").strip()

    if viewer.get("logged_in"):
        submitter_key = f"member:{viewer.get('user_id') or viewer.get('username') or viewer.get('email') or 'public'}"
    else:
        if not client_key:
            return jsonify([])
        submitter_key = client_key

    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        cursor.execute(
            "SELECT * FROM `registration_form_submissions` WHERE `submitter_key` = %s ORDER BY `submitted_at` DESC",
            (submitter_key,),
        )
        submissions = [registration_submission_row_to_dict(row) for row in cursor.fetchall() or []]
        return jsonify(submissions)
    finally:
        cursor.close()
        conn.close()


