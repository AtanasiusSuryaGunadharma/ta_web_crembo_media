from pathlib import Path
import json
from datetime import datetime, timedelta
from io import BytesIO
import os
import re
import time
import uuid
import html
import calendar
import secrets
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.image import MIMEImage
from email.utils import formataddr
import mimetypes

import mysql.connector
from flask import Flask, abort, flash, jsonify, redirect, render_template, request, send_file, send_from_directory, session, url_for
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename

BASE_DIR = Path(__file__).resolve().parent
FRONTEND_DIR = BASE_DIR / "frontend"

app = Flask(
    __name__,
    template_folder=str(FRONTEND_DIR),
    static_folder=str(FRONTEND_DIR / "static"),
    static_url_path="/static",
)
app.secret_key = "dev-secret-change-me"


@app.errorhandler(404)
def handle_api_not_found(error):
    if request.path.startswith("/api/"):
        return jsonify({"ok": False, "message": "Endpoint API tidak ditemukan."}), 404
    return error

@app.errorhandler(403)
def handle_api_forbidden(error):
    if request.path.startswith("/api/"):
        return jsonify({"ok": False, "message": "Akses ditolak."}), 403
    return error

@app.errorhandler(500)
def handle_api_internal_error(error):
    if request.path.startswith("/api/"):
        return jsonify({"ok": False, "message": "Terjadi kesalahan server pada API. Cek terminal Flask untuk detailnya."}), 500
    return error


@app.after_request
def add_monitoring_api_cors_headers(response):
    if request.path.startswith("/api/monitoring-kewajiban-tugas/"):
        origin = request.headers.get("Origin")
        if origin:
            response.headers["Access-Control-Allow-Origin"] = origin
            response.headers["Access-Control-Allow-Credentials"] = "true"
            response.headers["Vary"] = "Origin"
    return response

MYSQL_CONFIG = {
    "host": "127.0.0.1",
    "port": 3306,
    "user": "root",
    "password": "",
    "database": "crembo_db_new",
    "autocommit": False,
}

# Konfigurasi SMTP untuk fitur Lupa Password.
# Untuk deploy, sebaiknya pindahkan nilai ini ke environment variable
# SMTP_HOST, SMTP_PORT, SMTP_USERNAME, SMTP_PASSWORD, SMTP_FROM_EMAIL, SMTP_FROM_NAME.
SMTP_HOST = os.getenv("SMTP_HOST", "smtp.gmail.com")
SMTP_PORT = int(os.getenv("SMTP_PORT", "587"))
SMTP_USERNAME = os.getenv("SMTP_USERNAME") or os.getenv("SMTP_USER") or "crembomedia@gmail.com"
SMTP_PASSWORD = os.getenv("SMTP_PASSWORD") or os.getenv("SMTP_PASS") or "leezqsrjqdphonev"
SMTP_FROM_EMAIL = os.getenv("SMTP_FROM_EMAIL", SMTP_USERNAME)
SMTP_FROM_NAME = os.getenv("SMTP_FROM_NAME", "Crembo Media")
# Logo untuk body email OTP. Default mengikuti file logo di folder frontend.
EMAIL_LOGO_PATH = Path(os.getenv("EMAIL_LOGO_PATH", str(FRONTEND_DIR / "LOGO CREMBO PUTIH yg bagus.png")))
PASSWORD_RESET_OTP_MINUTES = int(os.getenv("PASSWORD_RESET_OTP_MINUTES", "5"))
PASSWORD_RESET_RESEND_SECONDS = int(os.getenv("PASSWORD_RESET_RESEND_SECONDS", "120"))
PASSWORD_RESET_MAX_ATTEMPTS = int(os.getenv("PASSWORD_RESET_MAX_ATTEMPTS", "5"))

UPLOAD_FOLDER = FRONTEND_DIR / "uploads"
ALLOWED_ATTACHMENT_EXTENSIONS = {
    ".jpg", ".jpeg", ".png", ".gif", ".webp", ".bmp", ".pdf",
    ".doc", ".docx", ".xls", ".xlsx", ".ppt", ".pptx",
    ".zip", ".rar", ".7z", ".txt", ".csv",
}
PREVIEWABLE_ATTACHMENT_EXTENSIONS = {
    ".jpg", ".jpeg", ".png", ".gif", ".webp", ".bmp", ".pdf",
}
INVENTORY_DEFAULT_CATEGORIES = ["Kamera", "Audio", "Aksesori", "Switcher", "Kabel", "Lighting", "Lainnya"]
INVENTORY_DEFAULT_ITEMS = []


ADMIN_MODULE_KEYS = ["streaming", "inventaris", "publikasi", "konten", "keanggotaan"]
ADMIN_MODULE_LABELS = {
    "streaming": "Tugas Streaming",
    "inventaris": "Inventaris & Peminjaman",
    "publikasi": "Informasi & Publikasi",
    "konten": "Konten Utama Website",
    "keanggotaan": "Keanggotaan",
}
ADMIN_PAGE_MODULE_MAP = {
    "jadwal-tugas-streaming-admin.html": "streaming",
    "registrasi-tugas-misa-besar.html": "streaming",
    "penugasan-petugas-misa.html": "streaming",
    "monitoring-tugas-anggota.html": "streaming",
    "hasil-evaluasi-streaming.html": "streaming",
    "setting-pertanyaan-evaluasi-streaming.html": "streaming",
    "data-inventaris-barang.html": "inventaris",
    "persetujuan-peminjaman.html": "inventaris",
    "riwayat-peminjaman-pengembalian.html": "inventaris",
    "hasil-form-kerusakan-barang.html": "inventaris",
    "manajemen-profil.html": "publikasi",
    "kelola-berita.html": "publikasi",
    "manajemen-agenda.html": "publikasi",
    "manajemen-form-pendaftaran.html": "publikasi",
    "kelola-carousel-home.html": "konten",
    "kelola-tentang-crembo.html": "konten",
    "kelola-embed-youtube.html": "konten",
    "kelola-embed-google-maps.html": "konten",
    "kelola-embed-instagram.html": "konten",
    "manajemen-anggota.html": "keanggotaan",
    "setting-sertifikat-anggota.html": "keanggotaan",
}

def default_admin_permissions() -> dict[str, bool]:
    return {key: True for key in ADMIN_MODULE_KEYS}

def normalize_admin_permissions(value=None, *, default_all: bool = True) -> dict[str, bool]:
    base = {key: bool(default_all) for key in ADMIN_MODULE_KEYS}
    if isinstance(value, str):
        value = safe_json_loads(value, {})
    if not isinstance(value, dict):
        return base
    normalized = {}
    for key in ADMIN_MODULE_KEYS:
        raw = value.get(key, base[key])
        normalized[key] = bool(raw) and str(raw).lower() not in {"0", "false", "no", "tidak", "off"}
    return normalized

def permission_row_to_dict(row) -> dict[str, bool]:
    if not row:
        return default_admin_permissions()
    return {key: bool(row.get(key, 1)) for key in ADMIN_MODULE_KEYS}

def ensure_admin_permissions_schema(cursor) -> None:
    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS `admin_module_permissions` (
          `member_id` int(11) NOT NULL,
          `streaming` tinyint(1) NOT NULL DEFAULT 1,
          `inventaris` tinyint(1) NOT NULL DEFAULT 1,
          `publikasi` tinyint(1) NOT NULL DEFAULT 1,
          `konten` tinyint(1) NOT NULL DEFAULT 1,
          `keanggotaan` tinyint(1) NOT NULL DEFAULT 1,
          `created_at` datetime NOT NULL DEFAULT CURRENT_TIMESTAMP,
          `updated_at` datetime NOT NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
          PRIMARY KEY (`member_id`)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_general_ci
        """
    )
    for key in ADMIN_MODULE_KEYS:
        ensure_column(cursor, "admin_module_permissions", key, f"`{key}` tinyint(1) NOT NULL DEFAULT 1")
    ensure_column(cursor, "admin_module_permissions", "created_at", "`created_at` datetime NOT NULL DEFAULT CURRENT_TIMESTAMP")
    ensure_column(cursor, "admin_module_permissions", "updated_at", "`updated_at` datetime NOT NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP")
    cursor.execute(
        """
        INSERT IGNORE INTO `admin_module_permissions`
          (`member_id`, `streaming`, `inventaris`, `publikasi`, `konten`, `keanggotaan`)
        SELECT `id`, 1, 1, 1, 1, 1
        FROM `anggota`
        WHERE `role` = 'admin'
        """
    )
    cursor.execute(
        """
        DELETE p FROM `admin_module_permissions` p
        LEFT JOIN `anggota` a ON a.id = p.member_id
        WHERE a.id IS NULL OR COALESCE(a.role, '') <> 'admin'
        """
    )

def get_admin_permissions(member_id=None, role: str | None = None) -> dict[str, bool]:
    normalized_role = normalize_role_value(role or session.get("role") or "")
    if normalized_role == "super_admin":
        return default_admin_permissions()
    if normalized_role != "admin":
        return default_admin_permissions()
    try:
        target_id = int(member_id if member_id is not None else session.get("user_id"))
    except (TypeError, ValueError):
        return default_admin_permissions()

    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        ensure_admin_permissions_schema(cursor)
        cursor.execute("SELECT streaming, inventaris, publikasi, konten, keanggotaan FROM admin_module_permissions WHERE member_id = %s LIMIT 1", (target_id,))
        row = cursor.fetchone()
        if not row:
            cursor.execute(
                """
                INSERT IGNORE INTO admin_module_permissions
                  (member_id, streaming, inventaris, publikasi, konten, keanggotaan)
                VALUES (%s, 1, 1, 1, 1, 1)
                """,
                (target_id,),
            )
            conn.commit()
            return default_admin_permissions()
        return permission_row_to_dict(row)
    finally:
        cursor.close()
        conn.close()

def admin_has_module_access(module_key: str) -> bool:
    if not session.get("logged_in"):
        return False
    role = normalize_role_value(session.get("role") or "")
    if role == "super_admin":
        return True
    if role != "admin":
        return False
    permissions = get_admin_permissions(session.get("user_id"), role)
    return bool(permissions.get(module_key))

def require_super_admin_api():
    if not session.get("logged_in") or normalize_role_value(session.get("role") or "") != "super_admin":
        return jsonify({"ok": False, "message": "Hanya Super Admin yang dapat mengakses Kelola Data Admin."}), 403
    return None

def template_exists(template_name: str) -> bool:
    return (FRONTEND_DIR / template_name).is_file()

def mysql_connection():
    return mysql.connector.connect(**MYSQL_CONFIG)

def ensure_upload_folder() -> Path:
    UPLOAD_FOLDER.mkdir(parents=True, exist_ok=True)
    return UPLOAD_FOLDER

def normalize_text(value, fallback: str = "") -> str:
    return str(value if value is not None else fallback).strip()

def parse_optional_int(value) -> int | None:
    if value in (None, ""):
        return None
    try:
        return max(0, int(float(value)))
    except (TypeError, ValueError):
        return None

def parse_required_int(value, fallback: int = 0) -> int:
    parsed = parse_optional_int(value)
    return fallback if parsed is None else parsed

def fetch_scalar_value(row, default=0):
    if row is None:
        return default
    if isinstance(row, dict):
        return next(iter(row.values()), default)
    if isinstance(row, (list, tuple)):
        return row[0] if row else default
    return row

def parse_optional_date(value) -> str | None:
    raw = normalize_text(value)
    if not raw:
        return None
    try:
        datetime.strptime(raw, "%Y-%m-%d")
    except ValueError:
        return None
    return raw

def safe_json_loads(value, fallback):
    if isinstance(value, (list, dict)):
        return value
    if not value:
        return fallback
    try:
        return json.loads(value)
    except (TypeError, ValueError):
        return fallback

def attachment_record_from_url(url_value: str, *, name: str | None = None, mime_type: str | None = None, size: int | None = None) -> dict[str, object]:
    clean_url = (url_value or "").strip()
    filename = (name or Path(clean_url).name or "lampiran").strip()
    extension = Path(filename).suffix.lower() or Path(clean_url).suffix.lower()
    previewable = extension in PREVIEWABLE_ATTACHMENT_EXTENSIONS
    kind = "image" if extension in {".jpg", ".jpeg", ".png", ".gif", ".webp", ".bmp"} else ("pdf" if extension == ".pdf" else "file")
    return {
        "url": clean_url,
        "name": filename,
        "mimeType": mime_type or "",
        "size": int(size) if isinstance(size, int) else 0,
        "previewable": previewable,
        "kind": kind,
    }

def normalize_attachment_payload(value) -> list[dict[str, object]]:
    if value in (None, "", []):
        return []

    raw_items = value
    if isinstance(value, str):
        raw_text = value.strip()
        if not raw_text:
            return []
        try:
            parsed = json.loads(raw_text)
            raw_items = parsed if isinstance(parsed, list) else [raw_text]
        except (TypeError, ValueError):
            raw_items = [raw_text]

    if isinstance(raw_items, dict):
        raw_items = [raw_items]

    normalized_items: list[dict[str, object]] = []
    for item in raw_items or []:
        if isinstance(item, str):
            clean_item = item.strip()
            if clean_item:
                normalized_items.append(attachment_record_from_url(clean_item))
            continue

        if not isinstance(item, dict):
            continue

        url_value = str(item.get("url") or item.get("path") or item.get("fileUrl") or "").strip()
        if not url_value:
            continue

        normalized_items.append(
            attachment_record_from_url(
                url_value,
                name=str(item.get("name") or item.get("filename") or "").strip() or None,
                mime_type=str(item.get("mimeType") or item.get("mime_type") or "").strip() or None,
                size=item.get("size") if isinstance(item.get("size"), int) else None,
            )
        )

    return normalized_items

def save_uploaded_attachment(file_storage) -> dict[str, object]:
    filename = secure_filename(file_storage.filename or "")
    if not filename:
        raise ValueError("Nama file tidak valid.")

    extension = Path(filename).suffix.lower()
    if extension not in ALLOWED_ATTACHMENT_EXTENSIONS:
        raise ValueError("Format file tidak didukung.")

    upload_folder = ensure_upload_folder()
    unique_name = f"{Path(filename).stem}_{uuid.uuid4().hex}{extension}"
    destination = upload_folder / unique_name
    file_storage.save(destination)

    return attachment_record_from_url(
        url_for("serve_uploaded_file", filename=unique_name),
        name=filename,
        mime_type=file_storage.mimetype,
        size=destination.stat().st_size,
    )

def remove_physical_file(file_url: str):
    if not file_url:
        return
    try:
        filename = file_url.split('/')[-1]
        file_path = UPLOAD_FOLDER / filename
        if file_path.exists() and file_path.is_file():
            file_path.unlink()
    except Exception as e:
        print(f"[WARN] Failed to delete file {file_url}: {e}")

def normalize_inventory_photos(value) -> list[dict[str, object]]:
    photos = normalize_attachment_payload(value)
    normalized: list[dict[str, object]] = []
    for photo in photos:
        if not isinstance(photo, dict):
            continue
        url_value = normalize_text(photo.get("url"))
        if not url_value:
            continue
        normalized.append(
            {
                "url": url_value,
                "name": normalize_text(photo.get("name")) or Path(url_value).name,
                "mimeType": normalize_text(photo.get("mimeType")),
                "size": parse_required_int(photo.get("size"), 0),
                "previewable": True,
                "kind": "image",
            }
        )
    return normalized

INVENTORY_UNIT_STATUSES = {"Tersedia", "Dipinjam", "Perbaikan", "Rusak", "Hilang"}

def inventory_unit_reason_for_status(status: str) -> str:
    return {
        "Dipinjam": "Sedang digunakan",
        "Perbaikan": "Sedang diperbaiki",
        "Rusak": "Kondisi rusak",
        "Hilang": "Belum ditemukan",
    }.get(normalize_text(status), "")

def normalize_inventory_unit_detail(value, index: int, fallback_status: str = "Tersedia") -> dict[str, object]:
    raw = value if isinstance(value, dict) else {}
    if isinstance(value, str):
        raw_status = normalize_text(value)
        if raw_status in INVENTORY_UNIT_STATUSES:
            raw = {"status": raw_status}
    status = normalize_text(raw.get("status")) or normalize_text(fallback_status) or "Tersedia"
    if status not in INVENTORY_UNIT_STATUSES:
        status = "Tersedia"
    reason = normalize_text(raw.get("reason") or raw.get("note") or raw.get("keterangan"))
    if not reason and status != "Tersedia":
        reason = inventory_unit_reason_for_status(status)
    return {
        "index": index + 1,
        "label": f"Unit {index + 1}",
        "status": status,
        "reason": reason,
        "available": status == "Tersedia",
    }

def normalize_inventory_unit_details(value, total_unit: int, available_unit: int | None = None, fallback_status: str = "Tersedia") -> list[dict[str, object]]:
    total = max(1, parse_required_int(total_unit, 1))
    raw_items = safe_json_loads(value, [])
    if isinstance(raw_items, dict):
        raw_items = [raw_items]
    if not isinstance(raw_items, list):
        raw_items = []

    normalized: list[dict[str, object]] = []
    available_target = total if available_unit is None else max(0, min(total, parse_required_int(available_unit, total)))
    fallback = normalize_text(fallback_status) or "Tersedia"
    if fallback not in INVENTORY_UNIT_STATUSES:
        fallback = "Dipinjam"
    if fallback == "Tersedia" and available_target < total:
        fallback = "Dipinjam"

    for index in range(total):
        raw_item = raw_items[index] if index < len(raw_items) else {}
        if not isinstance(raw_item, dict):
            raw_item = {}
        detail = normalize_inventory_unit_detail(raw_item, index, fallback_status=fallback)
        if index >= len(raw_items):
            if index < available_target:
                detail["status"] = "Tersedia"
                detail["reason"] = ""
                detail["available"] = True
            else:
                detail["status"] = fallback
                detail["reason"] = inventory_unit_reason_for_status(fallback) or detail["reason"]
                detail["available"] = False
        normalized.append(detail)

    return normalized

def inventory_status_from_unit_details(unit_details: list[dict[str, object]], fallback: str = "Tersedia") -> str:
    statuses = [normalize_text(unit.get("status")) for unit in unit_details if isinstance(unit, dict)]
    if not statuses or all(status == "Tersedia" for status in statuses):
        return "Tersedia"
    if any(status == "Dipinjam" for status in statuses):
        return "Dipinjam"
    if any(status == "Rusak" for status in statuses):
        return "Rusak"
    if any(status == "Perbaikan" for status in statuses):
        return "Perbaikan"
    if any(status == "Hilang" for status in statuses):
        return "Hilang"
    return normalize_text(fallback) or "Tersedia"

def inventory_unit_details_text(unit_details: list[dict[str, object]]) -> str:
    parts: list[str] = []
    for unit in unit_details:
        if not isinstance(unit, dict):
            continue
        label = normalize_text(unit.get("label")) or f"Unit {unit.get('index') or len(parts) + 1}"
        status = normalize_text(unit.get("status")) or "Tersedia"
        reason = normalize_text(unit.get("reason"))
        if reason:
            parts.append(f"{label}: {status} - {reason}")
        else:
            parts.append(f"{label}: {status}")
    return "; \n".join(parts)

def inventory_item_row_to_dict(row: dict[str, object]) -> dict[str, object]:
    photos = normalize_inventory_photos(safe_json_loads(row.get("photos"), []))
    total_unit = parse_required_int(row.get("total_unit"), 1)
    fallback_status = normalize_text(row.get("status")) or "Tersedia"
    unit_details = normalize_inventory_unit_details(
        safe_json_loads(row.get("unit_details"), []),
        total_unit,
        parse_required_int(row.get("available_unit"), total_unit),
        fallback_status=fallback_status,
    )
    available_unit = sum(1 for unit in unit_details if unit.get("status") == "Tersedia")
    purchase_date = row.get("purchase_date")
    purchase_price = row.get("purchase_price")
    updated_at = row.get("updated_at")
    created_at = row.get("created_at")
    photo_url = photos[0]["url"] if photos else ""
    return {
        "id": row.get("id") or "",
        "code": normalize_text(row.get("code")).upper(),
        "name": normalize_text(row.get("name")),
        "category": normalize_text(row.get("category")) or "Lainnya",
        "location": normalize_text(row.get("location")),
        "purchaseDate": purchase_date.isoformat() if hasattr(purchase_date, "isoformat") else normalize_text(purchase_date),
        "purchasePrice": int(purchase_price) if purchase_price not in (None, "") else None,
        "totalUnit": total_unit,
        "availableUnit": available_unit,
        "borrowedUnit": max(total_unit - available_unit, 0),
        "hasMultiple": bool(row.get("has_multiple")),
        "canBorrow": bool(row.get("can_borrow", True)),
        "status": inventory_status_from_unit_details(unit_details, fallback=fallback_status),
        "notes": normalize_text(row.get("notes")),
        "photos": photos,
        "unitDetails": unit_details,
        "photo": photo_url,
        "createdAt": created_at.isoformat() if hasattr(created_at, "isoformat") else normalize_text(created_at),
        "updatedAt": updated_at.isoformat() if hasattr(updated_at, "isoformat") else normalize_text(updated_at),
    }

def inventory_summary_from_items(items: list[dict[str, object]]) -> dict[str, int]:
    total_unit = 0
    total_available = 0
    total_borrowed = 0
    for item in items:
        current_total = parse_required_int(item.get("totalUnit"), 0)
        current_available = parse_required_int(item.get("availableUnit"), 0)
        total_unit += current_total
        total_available += current_available
        total_borrowed += max(current_total - current_available, 0)
    return {
        "totalJenisBarang": len(items),
        "totalUnit": total_unit,
        "totalAvailable": total_available,
        "totalBorrowed": total_borrowed,
    }

def inventory_sort_items(items: list[dict[str, object]], sort_mode: str) -> list[dict[str, object]]:
    mode = normalize_text(sort_mode) or "updated-desc"
    def item_time(item: dict[str, object]) -> datetime:
        raw = normalize_text(item.get("updatedAt")) or "1970-01-01T00:00:00"
        try:
            return datetime.fromisoformat(raw.replace("Z", "+00:00"))
        except ValueError:
            return datetime(1970, 1, 1)

    if mode == "name-asc":
        return sorted(items, key=lambda item: normalize_text(item.get("name")).lower())
    if mode == "name-desc":
        return sorted(items, key=lambda item: normalize_text(item.get("name")).lower(), reverse=True)
    if mode == "updated-asc":
        return sorted(items, key=item_time)
    return sorted(items, key=item_time, reverse=True)

def inventory_filter_items(items: list[dict[str, object]], search: str = "", category: str = "all", status: str = "all", sort_mode: str = "updated-desc") -> list[dict[str, object]]:
    keyword = normalize_text(search).lower()
    category_value = normalize_text(category) or "all"
    status_value = normalize_text(status) or "all"

    filtered: list[dict[str, object]] = []
    for item in items:
        haystack = " ".join([
            normalize_text(item.get("code")),
            normalize_text(item.get("name")),
            normalize_text(item.get("category")),
            normalize_text(item.get("location")),
            normalize_text(item.get("notes")),
        ]).lower()
        matches_keyword = not keyword or keyword in haystack
        matches_category = category_value == "all" or normalize_text(item.get("category")) == category_value
        matches_status = status_value == "all" or normalize_text(item.get("status")) == status_value
        if matches_keyword and matches_category and matches_status:
            filtered.append(item)

    return inventory_sort_items(filtered, sort_mode)

def inventory_request_filters(args) -> dict[str, str]:
    return {
        "search": normalize_text(args.get("search")),
        "category": normalize_text(args.get("category")) or "all",
        "status": normalize_text(args.get("status")) or "all",
        "sort": normalize_text(args.get("sort")) or "updated-desc",
    }

def ensure_column(cursor, table_name: str, column_name: str, definition: str) -> None:
    cursor.execute(
        """
        SELECT COUNT(*)
        FROM information_schema.COLUMNS
        WHERE TABLE_SCHEMA = %s AND TABLE_NAME = %s AND COLUMN_NAME = %s
        """,
        (MYSQL_CONFIG["database"], table_name, column_name),
    )
    row = cursor.fetchone()
    if isinstance(row, dict):
        exists = list(row.values())[0] > 0
    else:
        exists = row[0] > 0 if row else False
        
    if not exists:
        try:
            cursor.execute(f"ALTER TABLE `{table_name}` ADD COLUMN {definition}")
        except mysql.connector.Error as exc:
            # Intermittent race antar request: dua request bisa ALTER kolom yang sama.
            # 1060 = Duplicate column name.
            if getattr(exc, "errno", None) != 1060:
                raise

def ensure_inventory_schema(cursor) -> None:
    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS `inventory_categories` (
          `id` varchar(100) NOT NULL,
          `name` varchar(255) NOT NULL,
          `order_index` int(11) NOT NULL DEFAULT 0,
          `created_at` datetime NOT NULL DEFAULT CURRENT_TIMESTAMP,
          `updated_at` datetime NOT NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
          PRIMARY KEY (`id`),
          UNIQUE KEY `uniq_inventory_category_name` (`name`)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_general_ci
        """
    )
    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS `inventory_items` (
          `id` varchar(100) NOT NULL,
          `code` varchar(100) NOT NULL,
          `name` varchar(255) NOT NULL,
          `category` varchar(255) NOT NULL,
          `location` varchar(255) NOT NULL,
          `purchase_date` date DEFAULT NULL,
          `purchase_price` bigint(20) DEFAULT NULL,
          `total_unit` int(11) NOT NULL DEFAULT 1,
          `available_unit` int(11) NOT NULL DEFAULT 1,
          `has_multiple` tinyint(1) NOT NULL DEFAULT 0,
          `can_borrow` tinyint(1) NOT NULL DEFAULT 1,
          `status` varchar(50) NOT NULL DEFAULT 'Tersedia',
          `notes` text DEFAULT NULL,
          `photos` longtext DEFAULT NULL,
          `unit_details` longtext DEFAULT NULL,
          `created_at` datetime NOT NULL DEFAULT CURRENT_TIMESTAMP,
          `updated_at` datetime NOT NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
          PRIMARY KEY (`id`),
          UNIQUE KEY `uniq_inventory_item_code` (`code`),
          KEY `idx_inventory_item_category` (`category`),
          KEY `idx_inventory_item_updated_at` (`updated_at`)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_general_ci
        """
    )
    try:
        cursor.execute("ALTER TABLE `inventory_items` DROP COLUMN `condition`")
    except Exception:
        pass

    ensure_column(cursor, "inventory_items", "unit_details", "`unit_details` longtext DEFAULT NULL")

    cursor.execute("SELECT COUNT(*) FROM `inventory_categories`")
    if cursor.fetchone()[0] == 0:
        for order_index, category_name in enumerate(INVENTORY_DEFAULT_CATEGORIES, start=1):
            cursor.execute(
                """
                INSERT INTO `inventory_categories` (`id`, `name`, `order_index`)
                VALUES (%s, %s, %s)
                """,
                (f"inv-cat-{order_index:02d}", category_name, order_index),
            )

def inventory_category_exists(cursor, category_name: str) -> bool:
    cursor.execute("SELECT COUNT(*) FROM `inventory_categories` WHERE `name` = %s", (category_name,))
    row = cursor.fetchone()
    if isinstance(row, dict):
        return next(iter(row.values())) > 0
    return row[0] > 0 if row else False

def ensure_inventory_category(cursor, category_name: str) -> None:
    clean_name = normalize_text(category_name)
    if not clean_name:
        raise ValueError("Kategori wajib diisi.")
    if inventory_category_exists(cursor, clean_name):
        return
    cursor.execute("SELECT COALESCE(MAX(`order_index`), 0) + 1 FROM `inventory_categories`")
    row = cursor.fetchone()
    if isinstance(row, dict):
        next_order = next(iter(row.values())) or 1
    else:
        next_order = row[0] or 1 if row else 1
    cursor.execute(
        """
        INSERT INTO `inventory_categories` (`id`, `name`, `order_index`)
        VALUES (%s, %s, %s)
        """,
        (f"inv-cat-{uuid.uuid4().hex[:12]}", clean_name, next_order),
    )

def inventory_item_payload_from_request(data: dict[str, object], existing: dict[str, object] | None = None) -> dict[str, object]:
    code = normalize_text(data.get("code")).upper()
    name = normalize_text(data.get("name"))
    category = normalize_text(data.get("category"))
    location = normalize_text(data.get("location"))
    if not code:
        raise ValueError("Kode barang wajib diisi.")
    if not name:
        raise ValueError("Nama barang wajib diisi.")
    if not category:
        raise ValueError("Kategori wajib dipilih.")
    if not location:
        raise ValueError("Lokasi simpan wajib diisi.")

    has_multiple = bool(data.get("hasMultiple"))
    total_unit = parse_required_int(data.get("totalUnit"), 1) if has_multiple else 1
    total_unit = max(1, total_unit)
    available_unit = parse_required_int(data.get("availableUnit"), total_unit) if has_multiple else 1
    available_unit = min(total_unit, max(0, available_unit))
    fallback_status = normalize_text(data.get("status")) or (normalize_text(existing.get("status")) if existing else "Tersedia")
    purchase_date = parse_optional_date(data.get("purchaseDate"))
    purchase_price = parse_optional_int(data.get("purchasePrice"))
    can_borrow = bool(data.get("canBorrow", True))
    notes = normalize_text(data.get("notes"))

    if data.get("photos") is not None:
        photos = normalize_inventory_photos(data.get("photos"))
    elif data.get("photo"):
        photos = normalize_inventory_photos([data.get("photo")])
    elif existing is not None:
        photos = normalize_inventory_photos(safe_json_loads(existing.get("photos"), []))
    else:
        photos = []

    unit_details_source = data.get("unitDetails")
    if unit_details_source is None and existing is not None:
        unit_details_source = safe_json_loads(existing.get("unit_details"), [])
    unit_details = normalize_inventory_unit_details(unit_details_source, total_unit, available_unit, fallback_status=fallback_status)
    available_unit = sum(1 for unit in unit_details if unit.get("status") == "Tersedia")
    status = inventory_status_from_unit_details(unit_details, fallback=fallback_status)
    if status not in INVENTORY_UNIT_STATUSES:
        status = "Tersedia"
    if not can_borrow and status == "Dipinjam":
        status = "Tersedia"

    return {
        "code": code,
        "name": name,
        "category": category,
        "location": location,
        "purchase_date": purchase_date,
        "purchase_price": purchase_price,
        "total_unit": total_unit,
        "available_unit": available_unit,
        "has_multiple": 1 if has_multiple or total_unit > 1 else 0,
        "can_borrow": 1 if can_borrow else 0,
        "status": status,
        "notes": notes,
        "photos": photos,
        "unit_details": unit_details,
    }

def inventory_export_rows(items: list[dict[str, object]]):
    headers = [
        "Kode Barang",
        "Nama Barang",
        "Kategori",
        "Lokasi Simpan",
        "Tgl Pembelian",
        "Harga",
        "Total",
        "Tersedia",
        "Dipinjam",
        "Bisa Pinjam",
        "Rincian Unit",
        "Foto",
        "Keterangan",
        "Update",
    ]
    rows = []
    base_url = "http://127.0.0.1:5000" 

    for item in items:
        photos = item.get("photos") or []
        photo_links = []
        for p in photos:
            if isinstance(p, dict) and p.get("url"):
                url = str(p.get("url"))
                if url.startswith("/"):
                    url = f"{base_url}{url}"
                photo_links.append(url)
                
        photo_label = "\n".join(photo_links) if photo_links else "-"
        rincian = inventory_unit_details_text(item.get("unitDetails") or [])
        keterangan = normalize_text(item.get("notes")) or "-"

        rows.append([
            normalize_text(item.get("code")),
            normalize_text(item.get("name")),
            normalize_text(item.get("category")),
            normalize_text(item.get("location")),
            normalize_text(item.get("purchaseDate")) or "-",
            format_currency(item.get("purchasePrice")),
            str(parse_required_int(item.get("totalUnit"), 0)),
            str(parse_required_int(item.get("availableUnit"), 0)),
            str(parse_required_int(item.get("borrowedUnit"), max(parse_required_int(item.get("totalUnit"), 0) - parse_required_int(item.get("availableUnit"), 0), 0))),
            "Ya" if item.get("canBorrow") else "Tidak",
            rincian,
            photo_label,
            keterangan,
            normalize_text(item.get("updatedAt"))[:10] if item.get("updatedAt") else "-",
        ])
    return headers, rows

def inventory_export_filename(extension: str) -> str:
    return f"laporan-data-inventaris-barang.{extension}"

def fetch_inventory_items_for_report(cursor, filters: dict[str, str]) -> list[dict[str, object]]:
    cursor.execute(
        """
        SELECT `id`, `code`, `name`, `category`, `location`, `purchase_date`, `purchase_price`,
                   `total_unit`, `available_unit`, `has_multiple`, `can_borrow`, `status`,
                   `notes`, `photos`, `unit_details`, `created_at`, `updated_at`
        FROM `inventory_items`
        ORDER BY `updated_at` DESC, `code` ASC
        """
    )
    items = [inventory_item_row_to_dict(row) for row in cursor.fetchall() or []]
    return inventory_filter_items(
        items,
        search=filters.get("search", ""),
        category=filters.get("category", "all"),
        status=filters.get("status", "all"),
        sort_mode=filters.get("sort", "updated-desc"),
    )

def format_currency(value):
    if value is None or value == "" or str(value).lower() == "none":
        return "-"
    try:
        val = int(value)
        return f"Rp {val:,}".replace(",", ".")
    except (ValueError, TypeError):
        return "-"

@app.route("/api/inventory/categories", methods=["GET"])
def get_inventory_categories():
    ensure_auth_schema()
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        cursor.execute("SELECT `name` FROM `inventory_categories` ORDER BY `order_index` ASC, `name` ASC")
        categories = [row["name"] for row in cursor.fetchall() or [] if row.get("name")]
        if not categories:
            categories = INVENTORY_DEFAULT_CATEGORIES[:]
        return jsonify({"success": True, "categories": categories})
    finally:
        cursor.close()
        conn.close()

@app.route("/api/inventory/categories", methods=["POST"])
def create_inventory_category():
    ensure_auth_schema()
    data = request.get_json(silent=True) or {}
    category_name = normalize_text(data.get("name"))
    if not category_name:
        return jsonify({"success": False, "error": "Nama kategori wajib diisi."}), 400

    conn = mysql_connection()
    cursor = conn.cursor()
    try:
        ensure_inventory_category(cursor, category_name)
        conn.commit()
        return jsonify({"success": True, "name": category_name})
    except Exception as exc:
        conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 400
    finally:
        cursor.close()
        conn.close()

@app.route("/api/inventory/categories/<category_name>", methods=["PUT"])
def update_inventory_category(category_name):
    ensure_auth_schema()
    data = request.get_json(silent=True) or {}
    new_name = normalize_text(data.get("name"))
    old_name = normalize_text(category_name)
    if not old_name or not new_name:
        return jsonify({"success": False, "error": "Nama kategori tidak valid."}), 400

    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        cursor.execute("SELECT `id` FROM `inventory_categories` WHERE `name` = %s LIMIT 1", (old_name,))
        existing = cursor.fetchone()
        if not existing:
            return jsonify({"success": False, "error": "Kategori tidak ditemukan."}), 404
        if new_name != old_name:
            cursor.execute("SELECT COUNT(*) FROM `inventory_categories` WHERE `name` = %s", (new_name,))
            if fetch_scalar_value(cursor.fetchone(), 0) > 0:
                return jsonify({"success": False, "error": "Nama kategori sudah dipakai."}), 400

        cursor.execute("UPDATE `inventory_categories` SET `name` = %s WHERE `name` = %s", (new_name, old_name))
        cursor.execute("UPDATE `inventory_items` SET `category` = %s WHERE `category` = %s", (new_name, old_name))
        conn.commit()
        return jsonify({"success": True, "name": new_name})
    except Exception as exc:
        conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 400
    finally:
        cursor.close()
        conn.close()

@app.route("/api/inventory/categories/<category_name>", methods=["DELETE"])
def delete_inventory_category(category_name):
    ensure_auth_schema()
    old_name = normalize_text(category_name)
    if not old_name:
        return jsonify({"success": False, "error": "Kategori tidak valid."}), 400

    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        cursor.execute("SELECT COUNT(*) FROM `inventory_items` WHERE `category` = %s", (old_name,))
        if fetch_scalar_value(cursor.fetchone(), 0) > 0:
            return jsonify({"success": False, "error": "Kategori ini masih dipakai oleh data inventaris."}), 409
        cursor.execute("DELETE FROM `inventory_categories` WHERE `name` = %s", (old_name,))
        conn.commit()
        return jsonify({"success": True})
    except Exception as exc:
        conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 400
    finally:
        cursor.close()
        conn.close()

@app.route("/api/pengajuan/<pengajuan_id>/ambil", methods=["POST"])
def input_pengambilan(pengajuan_id):
    ensure_auth_schema()
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        ensure_loan_schema(cursor)
        cursor.execute("SELECT * FROM `loan_requests` WHERE `id` = %s LIMIT 1", (pengajuan_id,))
        req = cursor.fetchone()
        if not req:
            return jsonify({"success": False, "error": "Pengajuan tidak ditemukan"}), 404

        role = session.get("role") or ""
        user_id = session.get("user_id")
        if role not in ["admin", "super_admin"] and str(req.get("member_id")) != str(user_id):
            return jsonify({"success": False, "error": "Akses ditolak"}), 403

        if req.get("status") != "approved":
            return jsonify({"success": False, "error": "Hanya pengajuan dengan status 'approved' dapat diambil"}), 400

        tanggal = request.form.get("tanggal") or request.form.get("date") or None
        waktu = request.form.get("waktu") or request.form.get("time") or None
        lokasi = request.form.get("lokasi") or request.form.get("location") or ""
        units_raw = request.form.get("unitDetails") or request.form.get("units") or None
        try:
            unit_details = json.loads(units_raw) if units_raw else []
        except Exception:
            unit_details = []

        photo_record = None
        photo_file = request.files.get("photo") or request.files.get("bukti")
        if photo_file:
            try:
                saved = save_uploaded_attachment(photo_file)
                photo_record = saved.get("url") or saved.get("uri") or saved.get("name")
            except Exception:
                photo_record = None

        pickup_info = {
            "date": tanggal,
            "time": waktu,
            "location": lokasi,
            "photo": photo_record,
            "units": unit_details,
        }

        ensure_column(cursor, "loan_requests", "pickup_info", "`pickup_info` longtext DEFAULT NULL")
        ensure_column(cursor, "loan_requests", "pickup_at", "`pickup_at` datetime DEFAULT NULL")

        cursor.execute(
            "UPDATE `loan_requests` SET `status` = %s, `pickup_info` = %s, `pickup_at` = %s WHERE `id` = %s",
            (
                "taken",
                json.dumps(pickup_info, ensure_ascii=False),
                datetime.utcnow(),
                pengajuan_id,
            ),
        )
        conn.commit()

        # Update Notifikasi Pengambilan Barang
        try:
            ensure_notifications_schema()
            user_name = session.get("nama") or session.get("username") or f"User ID {user_id}"
            safe_user = html.escape(str(user_name))
            safe_lokasi = html.escape(str(lokasi))
            
            condition_texts = [f"&bull; Unit {i+1}: {html.escape(str(u.get('status')))} - {html.escape(str(u.get('reason','-')))}" for i, u in enumerate(unit_details)]
            condition_str = "<br>".join(condition_texts)
            photo_link = f"<br><br><a href='{photo_record}' target='_blank' style='display:inline-block; padding:4px 8px; background:#7f1d1d; color:#fff; border-radius:4px; text-decoration:none; font-size:11px; font-weight:bold;'>Lihat Foto Bukti</a>" if photo_record else ""
            
            body_text = f"Pengambilan dicatat oleh <b>{safe_user}</b>.<br><b>Lokasi:</b> {safe_lokasi}<br><b>Kondisi:</b><br>{condition_str}{photo_link}"

            create_notification(cursor, "peminjaman", f"Barang Diambil: {req.get('barang_name')}", body_text, "/riwayat-peminjaman-pengembalian.html", {"pengajuan_id": pengajuan_id, "target_user_id": req.get("member_id")}, target_role="admin")
            
            create_notification(cursor, "peminjaman", f"Pengambilan Tersimpan: {req.get('barang_name')}", "Data pengambilan Anda telah tersimpan.", "/riwayat-peminjaman-barang-anggota.html", {"pengajuan_id": pengajuan_id, "target_user_id": req.get("member_id")}, target_role="user")
            
            conn.commit()
        except Exception as e:
            print("Error notif:", e)

        return jsonify({"success": True})
    finally:
        cursor.close()
        conn.close()

@app.route("/api/pengajuan/<pengajuan_id>/kembali", methods=["POST"])
def input_pengembalian(pengajuan_id):
    ensure_auth_schema()
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        ensure_loan_schema(cursor)
        cursor.execute("SELECT * FROM `loan_requests` WHERE `id` = %s LIMIT 1", (pengajuan_id,))
        req = cursor.fetchone()
        if not req:
            return jsonify({"success": False, "error": "Pengajuan tidak ditemukan"}), 404

        role = session.get("role") or ""
        user_id = session.get("user_id")
        if role not in ["admin", "super_admin"] and str(req.get("member_id")) != str(user_id):
            return jsonify({"success": False, "error": "Akses ditolak"}), 403

        if req.get("status") not in ["taken", "approved"]:
            return jsonify({"success": False, "error": "Hanya pengajuan yang sedang dipinjam dapat dikembalikan"}), 400

        tanggal = request.form.get("tanggal") or request.form.get("date") or None
        waktu = request.form.get("waktu") or request.form.get("time") or None
        lokasi = request.form.get("lokasi") or request.form.get("location") or ""
        units_raw = request.form.get("unitDetails") or request.form.get("units") or None
        try:
            return_unit_details = json.loads(units_raw) if units_raw else []
        except Exception:
            return_unit_details = []

        photo_record = None
        photo_file = request.files.get("photo") or request.files.get("bukti")
        if photo_file:
            try:
                saved = save_uploaded_attachment(photo_file)
                photo_record = saved.get("url") or saved.get("uri") or saved.get("name")
            except Exception:
                photo_record = None

        return_info = {
            "date": tanggal,
            "time": waktu,
            "location": lokasi,
            "photo": photo_record,
            "units": return_unit_details,
        }

        ensure_column(cursor, "loan_requests", "return_info", "`return_info` longtext DEFAULT NULL")
        ensure_column(cursor, "loan_requests", "return_at", "`return_at` datetime DEFAULT NULL")

        barang_id = req.get("barang_id")
        jumlah_req = int(req.get("jumlah", 1))

        cursor.execute("SELECT `total_unit`, `available_unit`, `unit_details` FROM `inventory_items` WHERE `id` = %s LIMIT 1", (barang_id,))
        inv = cursor.fetchone()
        if inv:
            total = int(inv.get("total_unit", 1))
            available = int(inv.get("available_unit", 0))
            unit_details_raw = safe_json_loads(inv.get("unit_details"), [])

            units_restored = 0
            for idx, unit in enumerate(unit_details_raw):
                if unit.get("reason") == f"Dipinjam (Req ID: {pengajuan_id})":
                    kondisi_input = "Tersedia"
                    alasan_input = ""
                    if idx < len(return_unit_details):
                        input_val = return_unit_details[idx].get("status", "Baik")
                        if input_val == "Baik":
                            kondisi_input = "Tersedia"
                        else:
                            kondisi_input = input_val
                        alasan_input = return_unit_details[idx].get("reason", "")
                    
                    unit["status"] = kondisi_input
                    unit["reason"] = alasan_input
                    if kondisi_input == "Tersedia":
                        unit["available"] = True
                        units_restored += 1
                    else:
                        unit["available"] = False

            if units_restored == 0:
                for unit in unit_details_raw:
                    if unit.get("status") == "Dipinjam" and units_restored < jumlah_req:
                        unit["status"] = "Tersedia"
                        unit["reason"] = ""
                        unit["available"] = True
                        units_restored += 1

            new_available = min(total, available + units_restored)
            new_inv_status = "Tersedia" if new_available > 0 else "Dipinjam"

            cursor.execute(
                "UPDATE `inventory_items` SET `available_unit` = %s, `status` = %s, `unit_details` = %s WHERE `id` = %s",
                (new_available, new_inv_status, json.dumps(unit_details_raw, ensure_ascii=False), barang_id),
            )

        cursor.execute(
            "UPDATE `loan_requests` SET `status` = %s, `return_info` = %s, `return_at` = %s WHERE `id` = %s",
            ("returned", json.dumps(return_info, ensure_ascii=False), datetime.utcnow(), pengajuan_id),
        )

        # Update Notifikasi Pengembalian Barang
        try:
            ensure_notifications_schema()
            user_name = session.get("nama") or session.get("username") or f"User ID {user_id}"
            safe_user = html.escape(str(user_name))
            safe_lokasi = html.escape(str(lokasi))
            
            condition_texts = [f"&bull; Unit {i+1}: {html.escape(str(u.get('status')))} - {html.escape(str(u.get('reason','-')))}" for i, u in enumerate(return_unit_details)]
            condition_str = "<br>".join(condition_texts)
            photo_link = f"<br><br><a href='{photo_record}' target='_blank' style='display:inline-block; padding:4px 8px; background:#7f1d1d; color:#fff; border-radius:4px; text-decoration:none; font-size:11px; font-weight:bold;'>Lihat Foto Bukti</a>" if photo_record else ""
            
            body_text = f"Pengembalian dicatat oleh <b>{safe_user}</b>.<br><b>Lokasi:</b> {safe_lokasi}<br><b>Kondisi:</b><br>{condition_str}{photo_link}"

            create_notification(cursor, "peminjaman", f"Barang Dikembalikan: {req.get('barang_name')}", body_text, "/riwayat-peminjaman-pengembalian.html", {"pengajuan_id": pengajuan_id, "target_user_id": req.get("member_id")}, target_role="admin")
        except Exception as e:
            print("Error notif:", e)

        conn.commit()
        return jsonify({"success": True})
    except Exception as e:
        conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 400
    finally:
        cursor.close()
        conn.close()

@app.route("/api/inventory/items", methods=["GET"])
def get_inventory_items():
    ensure_auth_schema()
    filters = inventory_request_filters(request.args)
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        cursor.execute(
            """
             SELECT `id`, `code`, `name`, `category`, `location`, `purchase_date`, `purchase_price`,
                 `total_unit`, `available_unit`, `has_multiple`, `can_borrow`, `status`,
                 `notes`, `photos`, `unit_details`, `created_at`, `updated_at`
            FROM `inventory_items`
            ORDER BY `updated_at` DESC, `code` ASC
            """
        )
        raw_items = [inventory_item_row_to_dict(row) for row in cursor.fetchall() or []]
        categories = []
        cursor.execute("SELECT `name` FROM `inventory_categories` ORDER BY `order_index` ASC, `name` ASC")
        for row in cursor.fetchall() or []:
            if row.get("name"):
                categories.append(row["name"])

        items = inventory_filter_items(
            raw_items,
            search=filters["search"],
            category=filters["category"],
            status=filters["status"],
            sort_mode=filters["sort"],
        )
        return jsonify({
            "success": True,
            "items": items,
            "categories": categories or INVENTORY_DEFAULT_CATEGORIES[:],
            "summary": inventory_summary_from_items(items),
            "filters": filters,
        })
    finally:
        cursor.close()
        conn.close()

@app.route("/api/inventory/items/<item_id>", methods=["GET"])
def get_inventory_item(item_id):
    ensure_auth_schema()
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        cursor.execute(
            """
             SELECT `id`, `code`, `name`, `category`, `location`, `purchase_date`, `purchase_price`,
                 `total_unit`, `available_unit`, `has_multiple`, `can_borrow`, `status`,
                 `notes`, `photos`, `unit_details`, `created_at`, `updated_at`
            FROM `inventory_items`
            WHERE `id` = %s
            LIMIT 1
            """,
            (item_id,),
        )
        row = cursor.fetchone()
        if not row:
            return jsonify({"success": False, "error": "Barang inventaris tidak ditemukan."}), 404
        return jsonify({"success": True, "item": inventory_item_row_to_dict(row)})
    finally:
        cursor.close()
        conn.close()

@app.route("/api/inventory/items", methods=["POST"])
def create_inventory_item():
    ensure_auth_schema()
    data = request.get_json(silent=True) or {}

    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        payload = inventory_item_payload_from_request(data)
        cursor.execute("SELECT COUNT(*) FROM `inventory_items` WHERE `code` = %s", (payload["code"],))
        if fetch_scalar_value(cursor.fetchone(), 0) > 0:
            return jsonify({"success": False, "error": "Kode barang sudah dipakai."}), 400

        ensure_inventory_category(cursor, payload["category"])
        item_id = normalize_text(data.get("id")) or f"inv-{int(time.time() * 1000)}"
        cursor.execute(
            """
            INSERT INTO `inventory_items`
            (`id`, `code`, `name`, `category`, `location`, `purchase_date`, `purchase_price`, `total_unit`, `available_unit`, `has_multiple`, `can_borrow`, `status`, `notes`, `photos`, `unit_details`)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """,
            (
                item_id,
                payload["code"],
                payload["name"],
                payload["category"],
                payload["location"],
                payload["purchase_date"],
                payload["purchase_price"],
                payload["total_unit"],
                payload["available_unit"],
                payload["has_multiple"],
                payload["can_borrow"],
                payload["status"],
                payload["notes"],
                json.dumps(payload["photos"], ensure_ascii=False),
                json.dumps(payload["unit_details"], ensure_ascii=False),
            ),
        )
        conn.commit()
        return jsonify({"success": True, "item": get_inventory_item_response(cursor, item_id)})
    except Exception as exc:
        conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 400
    finally:
        cursor.close()
        conn.close()

def get_inventory_item_response(cursor, item_id: str) -> dict[str, object]:
    cursor.execute(
        """
        SELECT `id`, `code`, `name`, `category`, `location`, `purchase_date`, `purchase_price`,
         `total_unit`, `available_unit`, `has_multiple`, `can_borrow`, `status`,
         `notes`, `photos`, `unit_details`, `created_at`, `updated_at`
        FROM `inventory_items`
        WHERE `id` = %s
        LIMIT 1
        """,
        (item_id,),
    )
    row = cursor.fetchone() or {}
    return inventory_item_row_to_dict(row)

@app.route("/api/inventory/items/<item_id>", methods=["PUT"])
def update_inventory_item(item_id):
    ensure_auth_schema()
    data = request.get_json(silent=True) or {}

    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        cursor.execute("SELECT * FROM `inventory_items` WHERE `id` = %s LIMIT 1", (item_id,))
        existing = cursor.fetchone()
        if not existing:
            return jsonify({"success": False, "error": "Barang inventaris tidak ditemukan."}), 404

        payload = inventory_item_payload_from_request(data, existing)
        cursor.execute(
            "SELECT COUNT(*) FROM `inventory_items` WHERE `code` = %s AND `id` <> %s",
            (payload["code"], item_id),
        )
        if fetch_scalar_value(cursor.fetchone(), 0) > 0:
            return jsonify({"success": False, "error": "Kode barang sudah dipakai."}), 400

        ensure_inventory_category(cursor, payload["category"])
        old_photos = normalize_inventory_photos(safe_json_loads(existing.get("photos"), []))
        new_photos = payload["photos"]

        cursor.execute(
            """
            UPDATE `inventory_items`
            SET `code` = %s,
                `name` = %s,
                `category` = %s,
                `location` = %s,
                `purchase_date` = %s,
                `purchase_price` = %s,
                `total_unit` = %s,
                `available_unit` = %s,
                `has_multiple` = %s,
                `can_borrow` = %s,
                `status` = %s,
                `notes` = %s,
                `photos` = %s,
                `unit_details` = %s
            WHERE `id` = %s
            """,
            (
                payload["code"],
                payload["name"],
                payload["category"],
                payload["location"],
                payload["purchase_date"],
                payload["purchase_price"],
                payload["total_unit"],
                payload["available_unit"],
                payload["has_multiple"],
                payload["can_borrow"],
                payload["status"],
                payload["notes"],
                json.dumps(new_photos, ensure_ascii=False),
                json.dumps(payload["unit_details"], ensure_ascii=False),
                item_id,
            ),
        )
        conn.commit()

        old_urls = {photo.get("url") for photo in old_photos if photo.get("url")}
        new_urls = {photo.get("url") for photo in new_photos if photo.get("url")}
        for removed_url in old_urls - new_urls:
            remove_physical_file(removed_url)

        return jsonify({"success": True, "item": get_inventory_item_response(cursor, item_id)})
    except Exception as exc:
        conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 400
    finally:
        cursor.close()
        conn.close()

@app.route("/api/inventory/items/<item_id>", methods=["DELETE"])
def delete_inventory_item(item_id):
    ensure_auth_schema()
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        cursor.execute("SELECT `photos`, `name` FROM `inventory_items` WHERE `id` = %s LIMIT 1", (item_id,))
        existing = cursor.fetchone()
        if not existing:
            return jsonify({"success": False, "error": "Barang inventaris tidak ditemukan."}), 404

        cursor.execute("DELETE FROM `inventory_items` WHERE `id` = %s", (item_id,))
        conn.commit()

        for photo in normalize_inventory_photos(safe_json_loads(existing.get("photos"), [])):
            remove_physical_file(photo.get("url") or "")

        return jsonify({"success": True})
    except Exception as exc:
        conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 400
    finally:
        cursor.close()
        conn.close()

@app.route("/api/inventory/export.xlsx", methods=["GET"])
def export_inventory_excel():
    ensure_auth_schema()
    filters = inventory_request_filters(request.args)
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        items = fetch_inventory_items_for_report(cursor, filters)

        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill

        headers, rows = inventory_export_rows(items)
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Inventaris"
        worksheet.append(headers)
        for row in rows:
            worksheet.append(row)

        header_fill = PatternFill("solid", fgColor="7F0000")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2), 2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrapText=True)

        for column_cells in worksheet.columns:
            max_length = 0
            column_letter = column_cells[0].column_letter
            for cell in column_cells:
                try:
                    cell_value = str(cell.value or "")
                    if "\n" in cell_value:
                        lines = cell_value.split("\n")
                        line_lengths = [len(l) for l in lines]
                        if line_lengths and max(line_lengths) > max_length:
                            max_length = max(line_lengths)
                    elif len(cell_value) > max_length:
                        max_length = len(cell_value)
                except Exception:
                    continue
            worksheet.column_dimensions[column_letter].width = min(max_length + 4, 60)

        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        
        return send_file(
            buffer,
            as_attachment=True,
            download_name=inventory_export_filename("xlsx"),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    finally:
        cursor.close()
        conn.close()

@app.route("/api/inventory/export.pdf", methods=["GET"])
def export_inventory_pdf():
    ensure_auth_schema()
    filters = inventory_request_filters(request.args)
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        items = fetch_inventory_items_for_report(cursor, filters)

        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Spacer, Table, TableStyle, Paragraph

        headers, rows = inventory_export_rows(items)
        buffer = BytesIO()
        document = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            rightMargin=8 * mm,
            leftMargin=8 * mm,
            topMargin=10 * mm,
            bottomMargin=10 * mm,
        )
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle("InventoryTitle", parent=styles["Title"], fontName="Helvetica-Bold", fontSize=14, leading=16, textColor=colors.HexColor("#7F0000"))
        normal_style = ParagraphStyle("InventoryNormal", parent=styles["BodyText"], fontName="Helvetica", fontSize=8, leading=10)
        cell_style = ParagraphStyle("CellNormal", parent=styles["BodyText"], fontName="Helvetica", fontSize=7, leading=9, wordWrap='CJK')

        elements = [
            Paragraph("Laporan Data Inventaris Barang", title_style),
            Spacer(1, 4 * mm),
            Paragraph(f"Filter: {filters.get('category', 'all')} | Cari: {filters.get('search', '') or '-'} | Urutan: {filters.get('sort', 'updated-desc')}", normal_style),
            Spacer(1, 4 * mm),
        ]

        wrapped_rows = []
        for row in rows:
            wrapped_row = []
            for cell_data in row:
                text = str(cell_data).replace('\n', '<br/>')
                wrapped_row.append(Paragraph(text, cell_style))
            wrapped_rows.append(wrapped_row)

        table_data = [headers] + wrapped_rows if rows else [headers, ["-" for _ in headers]]
        
        page_width = landscape(A4)[0] - (16 * mm) 
        col_widths = [
            0.07 * page_width,
            0.10 * page_width,
            0.06 * page_width,
            0.07 * page_width,
            0.07 * page_width,
            0.07 * page_width,
            0.03 * page_width,
            0.04 * page_width,
            0.04 * page_width,
            0.04 * page_width,
            0.18 * page_width,
            0.10 * page_width,
            0.08 * page_width,
            0.05 * page_width
        ]
        
        table = Table(table_data, colWidths=col_widths, repeatRows=1)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#7F0000")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 7),
                    ("ALIGN", (0, 0), (0, -1), "CENTER"),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D1D5DB")),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.HexColor("#FFFFFF")]),
                ]
            )
        )
        elements.append(table)
        document.build(elements)
        buffer.seek(0)
        
        return send_file(
            buffer,
            as_attachment=False, 
            download_name=inventory_export_filename("pdf"),
            mimetype="application/pdf",
        )
    finally:
        cursor.close()
        conn.close()

def _date_to_iso(value) -> str:
    if not value:
        return ""
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return str(value)

def sync_session_from_member_row(row: dict[str, object] | None) -> None:
    """Sinkronkan session dari 1 row anggota tanpa menjalankan proses tulis DB."""
    if not row:
        return
    session["user_id"] = row.get("id")
    session["username"] = row.get("username") or ""
    session["nama"] = row.get("nama") or ""
    session["role"] = normalize_role_value(row.get("role") or "user")
    session["telp"] = row.get("telp") or ""
    session["email"] = row.get("email") or ""
    session["alamat"] = row.get("alamat") or ""
    session["tgl_lahir"] = _date_to_iso(row.get("tgl_lahir"))
    session["status_akun"] = row.get("status_akun") or "aktif"


def current_user_context_from_row(row: dict[str, object] | None, *, include_admin_permissions: bool = False) -> dict[str, object]:
    """Bangun context user dari row yang sudah dibaca.

    Fungsi ini sengaja tidak memanggil apply_membership_status_transitions(),
    ensure_auth_schema(), atau query tambahan lain agar endpoint read-only seperti
    profil anggota tidak memicu deadlock ketika beberapa request berjalan bersamaan.
    """
    if not row:
        return {
            "logged_in": bool(session.get("logged_in")),
            "user_id": session.get("user_id"),
            "username": session.get("username") or "",
            "nama": session.get("nama") or "",
            "role": session.get("role") or "",
            "telp": session.get("telp") or "",
            "email": session.get("email") or "",
            "alamat": session.get("alamat") or "",
            "tgl_lahir": _date_to_iso(session.get("tgl_lahir")),
            "status_akun": session.get("status_akun") or "",
            "inactive_until": "",
            "inactive_from": "",
            "inactive_type": "",
            "inactive_reason": "",
            "admin_permissions": {},
            "created_at": "",
            "updated_at": "",
        }

    role_value = normalize_role_value(row.get("role") or "user")
    admin_permissions = {}
    if include_admin_permissions and session.get("logged_in") and role_value in {"admin", "super_admin"}:
        admin_permissions = get_admin_permissions(row.get("id"), role_value)

    return {
        "logged_in": bool(session.get("logged_in")),
        "user_id": row.get("id"),
        "username": row.get("username") or "",
        "nama": row.get("nama") or "",
        "role": role_value,
        "telp": row.get("telp") or "",
        "email": row.get("email") or "",
        "alamat": row.get("alamat") or "",
        "tgl_lahir": _date_to_iso(row.get("tgl_lahir")),
        "status_akun": row.get("status_akun") or "aktif",
        "inactive_until": _date_to_iso(row.get("inactive_until")),
        "inactive_from": _date_to_iso(row.get("inactive_from")),
        "inactive_type": row.get("inactive_type") or "",
        "inactive_reason": row.get("inactive_reason") or "",
        "admin_permissions": admin_permissions,
        "created_at": row.get("created_at").isoformat() if row.get("created_at") else "",
        "updated_at": row.get("updated_at").isoformat() if row.get("updated_at") else "",
    }


def refresh_session_user_from_db() -> dict[str, object] | None:
    """Ambil ulang user login dari database dengan proses read-only.

    Sebelumnya fungsi ini menjalankan apply_membership_status_transitions() setiap
    kali halaman profil/dashboard dirender. Karena proses tersebut melakukan
    UPDATE ke tabel anggota dan membership_status_requests, request paralel dari
    halaman profil anggota bisa saling mengunci dan memunculkan error MySQL 1213
    deadlock. Refresh session cukup membaca data user login saja.
    """
    if not session.get("logged_in") or not session.get("user_id"):
        return None

    conn = None
    cursor = None
    try:
        conn = mysql_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute(
            """
            SELECT id, nama, username, telp, password, role, tgl_lahir, email, alamat,
                   status_akun, inactive_until, inactive_from, inactive_type,
                   inactive_reason, inactive_note, created_at, updated_at
            FROM anggota
            WHERE id = %s
            LIMIT 1
            """,
            (session.get("user_id"),),
        )
        row = cursor.fetchone()
        if not row:
            return None

        sync_session_from_member_row(row)
        return row
    except Exception:
        return None
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()


def current_user_context() -> dict[str, object]:
    row = refresh_session_user_from_db()
    return current_user_context_from_row(row, include_admin_permissions=True)

def normalize_phone(value: str) -> str:
    cleaned = re.sub(r"[\s\-()]+", "", (value or "").strip())
    if cleaned.startswith("+62"):
        return "62" + cleaned[3:]
    if cleaned.startswith("08"):
        return "62" + cleaned[1:]
    return cleaned

def normalize_status(value: str) -> str:
    raw = (value or "").strip().lower()
    return "aktif" if raw in {"", "aktif", "active"} else "nonaktif"

def normalize_role_value(value: str) -> str:
    raw = (value or "").strip().lower().replace(" ", "_")
    if raw in {"super_admin", "superadmin"}:
        return "super_admin"
    if raw == "admin":
        return "admin"
    return "user"

def member_row_to_dict(row: dict[str, object]) -> dict[str, object]:
    birth_date = row.get("tgl_lahir") or ""
    status_value = row.get("status_akun") or "aktif"
    inactive_until_value = row.get("inactive_until") or ""
    if hasattr(inactive_until_value, "isoformat"):
        inactive_until_value = inactive_until_value.isoformat()
    return {
        "id": row.get("id"),
        "name": row.get("nama") or "Anggota",
        "fullName": row.get("nama") or "Anggota",
        "username": row.get("username") or "",
        "email": row.get("email") or "",
        "phone": row.get("telp") or "",
        "address": row.get("alamat") or "",
        "birthDate": birth_date,
        "tanggalLahir": birth_date,
        "role": normalize_role_value(row.get("role") or "user"),
        "status": "Aktif" if normalize_status(status_value) == "aktif" else "Nonaktif",
        "password": row.get("password") or "",
        "note": "",
        "registeredAt": row.get("created_at") or "",
        "createdAt": row.get("created_at") or "",
        "updatedAt": row.get("updated_at") or "",
        "inactiveUntil": inactive_until_value or "",
        "inactiveFrom": _date_to_iso(row.get("inactive_from")),
        "inactiveType": row.get("inactive_type") or "",
        "inactiveReason": row.get("inactive_reason") or "",
        "inactiveNote": row.get("inactive_note") or "",
    }

def read_member_rows() -> list[dict[str, object]]:
    ensure_auth_schema()
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute(
        """
        SELECT id, nama, username, telp, password, role, tgl_lahir, email, alamat, status_akun, inactive_until, inactive_from, inactive_type, inactive_reason, inactive_note
        , created_at, updated_at
        FROM anggota
        ORDER BY id ASC
        """
    )
    rows = [member_row_to_dict(row) for row in cursor.fetchall()]
    cursor.close()
    conn.close()
    return rows

def hash_member_password(password_value: str, birth_date: str) -> str:
    candidate = (password_value or "").strip()
    if not candidate:
        candidate = default_password_from_birth_date(birth_date)
    if candidate.startswith(("scrypt:", "pbkdf2:", "argon2:", "bcrypt:")):
        return candidate
    return generate_password_hash(candidate)

def default_password_from_birth_date(birth_date: str) -> str:
    raw = (birth_date or "").strip()
    if not raw:
        return ""
    parts = raw.split("-")
    if len(parts) == 3:
        return f"{parts[2]}/{parts[1]}/{parts[0]}"
    return {
        "logged_in": bool(session.get("logged_in")),
        "user_id": session.get("user_id"),
        "username": session.get("username") or "",
        "nama": session.get("nama") or "",
        "role": session.get("role") or "",
        "telp": session.get("telp") or "",
        "email": session.get("email") or "",
        "alamat": session.get("alamat") or "",
    }

def ensure_loan_schema(cursor) -> None:
    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS `loan_requests` (
          `id` varchar(100) NOT NULL,
          `member_id` varchar(100) DEFAULT NULL,
          `barang_id` varchar(100) NOT NULL,
          `barang_name` varchar(255) NOT NULL,
          `barang_code` varchar(100) DEFAULT NULL,
          `barang_photo` text DEFAULT NULL,
          `jumlah` int(11) NOT NULL DEFAULT 1,
          `tanggal_pengajuan` date DEFAULT NULL,
          `tanggal_mulai` date DEFAULT NULL,
          `waktu_mulai` time DEFAULT NULL,
          `tanggal_selesai` date DEFAULT NULL,
          `waktu_selesai` time DEFAULT NULL,
          `tujuan` text DEFAULT NULL,
          `status` varchar(50) NOT NULL DEFAULT 'pending',
          `pickup_info` longtext DEFAULT NULL,
          `pickup_at` datetime DEFAULT NULL,
          `return_info` longtext DEFAULT NULL,
          `return_at` datetime DEFAULT NULL,
          `created_at` datetime NOT NULL DEFAULT CURRENT_TIMESTAMP,
          `updated_at` datetime NOT NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
          PRIMARY KEY (`id`),
          KEY `idx_loan_status` (`status`),
          KEY `idx_loan_member` (`member_id`)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_general_ci
        """
    )
    ensure_column(cursor, "loan_requests", "waktu_mulai", "`waktu_mulai` time DEFAULT NULL")
    ensure_column(cursor, "loan_requests", "waktu_selesai", "`waktu_selesai` time DEFAULT NULL")

def ensure_misa_besar_schema():
    conn = mysql_connection()
    cursor = conn.cursor(buffered=True)
    try:
        # Pastikan tabel utama ada
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS `misa_besar` (
              `id` int AUTO_INCREMENT PRIMARY KEY,
              `misa_name` varchar(255) NOT NULL,
              `misa_date` date NOT NULL,
              `misa_time` time NOT NULL
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_general_ci;
        """)
        conn.commit()

        # Tambahkan kolom baru (Safe Alter)
        for col, col_def in [
            ("misa_note", "text"),
            ("allow_member_request", "tinyint(1) DEFAULT 0"),
            ("status", "enum('draft','published') DEFAULT 'draft'"),
            ("created_at", "timestamp DEFAULT CURRENT_TIMESTAMP")
        ]:
            try:
                cursor.execute(f"ALTER TABLE misa_besar ADD COLUMN {col} {col_def}")
            except:
                pass 
        conn.commit()
                
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS `misa_besar_names` (
              `id` int AUTO_INCREMENT PRIMARY KEY,
              `misa_id` int NOT NULL,
              `role_name` varchar(100) NOT NULL,
              FOREIGN KEY (misa_id) REFERENCES misa_besar(id) ON DELETE CASCADE
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_general_ci;
        """)
        conn.commit()

        try:
            cursor.execute("ALTER TABLE misa_besar_names ADD COLUMN required_count int NOT NULL DEFAULT 1")
        except:
            pass
        conn.commit()
            
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS `misa_besar_assignments` (
              `id` int AUTO_INCREMENT PRIMARY KEY,
              `role_id` int NOT NULL,
              `member_id` int NOT NULL,
              FOREIGN KEY (role_id) REFERENCES misa_besar_names(id) ON DELETE CASCADE,
              FOREIGN KEY (member_id) REFERENCES anggota(id) ON DELETE CASCADE,
              UNIQUE KEY `uniq_role_member` (`role_id`, `member_id`)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_general_ci;
        """)
        try:
            ensure_column(cursor, "misa_besar_assignments", "created_at", "`created_at` datetime NOT NULL DEFAULT CURRENT_TIMESTAMP")
            ensure_column(cursor, "misa_besar_assignments", "request_source", "`request_source` varchar(30) NOT NULL DEFAULT 'admin'")
            conn.commit()
        except Exception as exc:
            print(f"[WARN] Failed to ensure misa besar assignment metadata: {exc}")
    finally:
        cursor.close()
        conn.close()

def ensure_damage_schema(cursor) -> None:
    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS `form_kerusakan_barang` (
          `id` varchar(100) NOT NULL,
          `member_id` varchar(100) DEFAULT NULL,
          `barang_id` varchar(100) DEFAULT NULL,
          `barang_name` varchar(255) NOT NULL,
          `barang_code` varchar(100) DEFAULT NULL,
          `tingkat_kerusakan` varchar(50) NOT NULL DEFAULT 'Sedang',
          `status` varchar(50) NOT NULL DEFAULT 'Pending Review',
          `deskripsi_kerusakan` longtext DEFAULT NULL,
          `waktu_kejadian` datetime DEFAULT NULL,
          `foto_kerusakan` longtext DEFAULT NULL,
          `created_at` datetime NOT NULL DEFAULT CURRENT_TIMESTAMP,
          `updated_at` datetime NOT NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
          PRIMARY KEY (`id`),
          KEY `idx_damage_status` (`status`),
          KEY `idx_damage_member` (`member_id`),
          KEY `idx_damage_created` (`created_at`)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_general_ci
        """
    )

@app.route("/api/pengajuan", methods=["GET"])
def list_pengajuan():
    ensure_auth_schema()
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        ensure_loan_schema(cursor)
        
        status_filter = request.args.get('status')
        from_date = request.args.get('from_date')
        to_date = request.args.get('to_date')

        query = """
            SELECT l.*, a.nama as member_name 
            FROM `loan_requests` l
            LEFT JOIN `anggota` a ON l.member_id = a.id
            WHERE 1=1
        """
        params = []
        
        role = session.get("role") or ""
        if role not in ["admin", "super_admin"]:
            user_id = session.get("user_id")
            if not user_id:
                return jsonify({"success": True, "items": []})
            query += " AND l.`member_id` = %s"
            params.append(str(user_id))
        
        if status_filter and status_filter != 'all':
            query += " AND l.`status` = %s"
            params.append(status_filter)
        if from_date:
            query += " AND l.`tanggal_mulai` >= %s"
            params.append(from_date)
        if to_date:
            query += " AND l.`tanggal_mulai` <= %s"
            params.append(to_date)
            
        query += " ORDER BY l.`created_at` DESC"

        cursor.execute(query, tuple(params))
        rows = cursor.fetchall() or []
        items = []
        for r in rows:
            items.append({
                "id": r.get("id"),
                "memberId": r.get("member_id"),
                "memberNama": r.get("member_name"),
                "barangId": r.get("barang_id"),
                "barangNama": r.get("barang_name"),
                "barangCode": r.get("barang_code"),
                "barangFoto": r.get("barang_photo"),
                "jumlahDiminta": r.get("jumlah"),
                "tanggalPengajuan": r.get("tanggal_pengajuan"),
                "tanggalMulai": r.get("tanggal_mulai"),
                "waktuMulai": str(r.get("waktu_mulai") or ""),
                "tanggalSelesai": r.get("tanggal_selesai"),
                "waktuSelesai": str(r.get("waktu_selesai") or ""),
                "tujuan": r.get("tujuan"),
                "status": r.get("status"),
                "adminNote": r.get("admin_note", ""),
                "approvedBy": r.get("approved_by", ""),
                "approvedAt": r.get("approved_at"),
                "createdAt": r.get("created_at"),
                "updatedAt": r.get("updated_at"),
                "pickupInfo": safe_json_loads(r.get("pickup_info"), {}),
                "returnInfo": safe_json_loads(r.get("return_info"), {}),
            })
        return jsonify({"success": True, "items": items})
    finally:
        cursor.close()
        conn.close()

@app.route("/api/pengajuan", methods=["POST"])
def create_pengajuan():
    ensure_auth_schema()
    data = request.get_json(silent=True) or {}
    barang_id = data.get("barangId")
    barang_nama = data.get("barangNama")
    jumlah = int(data.get("jumlahDiminta") or data.get("jumlah") or 0)
    tanggal_mulai = data.get("tanggalMulai")
    waktu_mulai = data.get("waktuMulai")
    tanggal_selesai = data.get("tanggalSelesai")
    waktu_selesai = data.get("waktuSelesai")
    tujuan = data.get("tujuan")

    if not barang_id or not barang_nama or not jumlah or not tanggal_mulai or not tanggal_selesai or not tujuan:
        return jsonify({"success": False, "error": "Field tidak lengkap"}), 400

    try:
        if datetime.fromisoformat(tanggal_mulai) >= datetime.fromisoformat(tanggal_selesai):
            return jsonify({"success": False, "error": "Tanggal selesai harus lebih besar dari tanggal mulai"}), 400
    except Exception:
        pass

    conn = mysql_connection()
    cursor = conn.cursor()
    try:
        ensure_loan_schema(cursor)
        pengajuan_id = f"pjn-{int(datetime.utcnow().timestamp() * 1000)}"

        barang_foto = None
        barang_code = None
        try:
            c2 = conn.cursor(dictionary=True)
            c2.execute("SELECT `code`, `photos` FROM `inventory_items` WHERE `id` = %s LIMIT 1", (barang_id,))
            row = c2.fetchone()
            if row:
                barang_code = row.get("code")
                photos_raw = row.get("photos")
                try:
                    photos = json.loads(photos_raw) if photos_raw else []
                    if isinstance(photos, list) and photos:
                        first = photos[0]
                        if isinstance(first, dict):
                            barang_foto = first.get("url") or first.get("uri") or None
                        elif isinstance(first, str):
                            barang_foto = first
                except Exception:
                    pass
            c2.close()
        except Exception:
            pass

        cursor.execute(
            """
            INSERT INTO `loan_requests` (`id`, `member_id`, `barang_id`, `barang_name`, `barang_code`, `barang_photo`, `jumlah`, `tanggal_pengajuan`, `tanggal_mulai`, `waktu_mulai`, `tanggal_selesai`, `waktu_selesai`, `tujuan`, `status`)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """,
            (
                pengajuan_id,
                session.get("user_id") or None,
                barang_id,
                barang_nama,
                barang_code,
                barang_foto,
                jumlah,
                datetime.utcnow().date(),
                tanggal_mulai,
                waktu_mulai,
                tanggal_selesai,
                waktu_selesai,
                tujuan,
                "pending",
            ),
        )
        conn.commit()
        
        # Update Notifikasi Admin
        try:
            ensure_notifications_schema()
            nc = conn.cursor()
            try:
                user_name = session.get('nama') or session.get('username') or f"User ID {session.get('user_id')}"
                safe_user = html.escape(str(user_name))
                create_notification(nc, "peminjaman", f"Pengajuan Peminjaman: {barang_nama}", f"Diajukan oleh <b>{safe_user}</b>. Harap ditinjau.", url_for('dashboard') if False else "/persetujuan-peminjaman.html", {"pengajuan_id": pengajuan_id}, target_role="admin")
                conn.commit()
            finally:
                nc.close()
        except Exception:
            pass
            
        return jsonify({"success": True, "id": pengajuan_id})
    finally:
        cursor.close()
        conn.close()


@app.route("/api/pengajuan/<pengajuan_id>/cancel", methods=["POST"])
def cancel_pengajuan(pengajuan_id):
    ensure_auth_schema()
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        ensure_loan_schema(cursor)
        cursor.execute("SELECT `status`, `barang_id`, `jumlah` FROM `loan_requests` WHERE `id` = %s LIMIT 1", (pengajuan_id,))
        row = cursor.fetchone()
        if not row:
            return jsonify({"success": False, "error": "Pengajuan tidak ditemukan"}), 404
            
        status = row.get("status")
        
        if status not in ["pending", "approved"]:
            return jsonify({"success": False, "error": "Hanya pengajuan dengan status pending atau disetujui (sebelum diambil) yang dapat dibatalkan"}), 400
            
        if status == "approved":
            barang_id = row.get("barang_id")
            jumlah_req = int(row.get("jumlah", 1))
            
            cursor.execute("SELECT `total_unit`, `available_unit`, `unit_details` FROM `inventory_items` WHERE `id` = %s LIMIT 1", (barang_id,))
            inv = cursor.fetchone()
            if inv:
                total = int(inv.get("total_unit", 1))
                available = int(inv.get("available_unit", 0))
                
                unit_details_raw = safe_json_loads(inv.get("unit_details"), [])
                units_restored = 0
                
                for unit in unit_details_raw:
                     if unit.get("reason") == f"Dipinjam (Req ID: {pengajuan_id})":
                          unit["status"] = "Tersedia"
                          unit["reason"] = ""
                          unit["available"] = True
                          units_restored += 1
                
                if units_restored == 0:
                     for unit in unit_details_raw:
                          if unit.get("status") == "Dipinjam" and units_restored < jumlah_req:
                               unit["status"] = "Tersedia"
                               unit["reason"] = ""
                               unit["available"] = True
                               units_restored += 1
                
                new_available = min(total, available + jumlah_req)
                new_inv_status = "Tersedia" if new_available > 0 else "Dipinjam"
                
                cursor.execute(
                     "UPDATE `inventory_items` SET `available_unit` = %s, `status` = %s, `unit_details` = %s WHERE `id` = %s", 
                     (new_available, new_inv_status, json.dumps(unit_details_raw, ensure_ascii=False), barang_id)
                )

        cursor.execute("UPDATE `loan_requests` SET `status` = %s WHERE `id` = %s", ("cancelled", pengajuan_id))
        conn.commit()
        return jsonify({"success": True})
    finally:
        cursor.close()
        conn.close()

@app.route("/api/admin/pengajuan/<pengajuan_id>/status", methods=["PUT"])
def update_pengajuan_status(pengajuan_id):
    role = session.get("role") or ""
    if not session.get("logged_in") or role not in ["admin", "super_admin"]:
        return jsonify({"success": False, "error": "Akses Ditolak. Anda bukan Admin."}), 403
        
    data = request.get_json(silent=True) or {}
    new_status = data.get("status")
    admin_note = data.get("adminNote", "")
    
    if new_status not in ["approved", "rejected", "cancelled", "taken", "returned"]:
        return jsonify({"success": False, "error": "Status tidak valid"}), 400

    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        ensure_column(cursor, "loan_requests", "admin_note", "`admin_note` text DEFAULT NULL")
        ensure_column(cursor, "loan_requests", "approved_by", "`approved_by` varchar(150) DEFAULT NULL")
        ensure_column(cursor, "loan_requests", "approved_at", "`approved_at` datetime DEFAULT NULL")

        cursor.execute("SELECT * FROM `loan_requests` WHERE `id` = %s LIMIT 1", (pengajuan_id,))
        req = cursor.fetchone()
        if not req:
            return jsonify({"success": False, "error": "Pengajuan tidak ditemukan"}), 404
            
        current_status = req.get("status")
        barang_id = req.get("barang_id")
        jumlah_req = int(req.get("jumlah", 1))
        
        if current_status != "pending" and new_status in ["approved", "rejected"]:
             return jsonify({"success": False, "error": "Hanya pengajuan berstatus Pending yang dapat disetujui atau ditolak"}), 400

        if new_status == "approved" and current_status == "pending":
            cursor.execute("SELECT `total_unit`, `available_unit`, `can_borrow`, `unit_details` FROM `inventory_items` WHERE `id` = %s LIMIT 1", (barang_id,))
            inv = cursor.fetchone()
            
            if not inv:
                 return jsonify({"success": False, "error": "Barang tidak ditemukan di inventaris"}), 404
            if not bool(inv.get("can_borrow", 1)):
                 return jsonify({"success": False, "error": "Barang tidak dapat dipinjam"}), 400
                 
            available = int(inv.get("available_unit", 0))
            if available < jumlah_req:
                 return jsonify({"success": False, "error": f"Stok tersedia ({available}) tidak mencukupi permintaan ({jumlah_req})"}), 400
                 
            unit_details_raw = safe_json_loads(inv.get("unit_details"), [])
            units_changed = 0
            
            for unit in unit_details_raw:
                 if unit.get("status") == "Tersedia" and units_changed < jumlah_req:
                      unit["status"] = "Dipinjam"
                      unit["reason"] = f"Dipinjam (Req ID: {pengajuan_id})"
                      unit["available"] = False
                      units_changed += 1
                      
            new_available = available - jumlah_req
            new_inv_status = "Dipinjam" if new_available <= 0 else "Tersedia"
            
            cursor.execute(
                 "UPDATE `inventory_items` SET `available_unit` = %s, `status` = %s, `unit_details` = %s WHERE `id` = %s", 
                 (new_available, new_inv_status, json.dumps(unit_details_raw, ensure_ascii=False), barang_id)
            )
            
        elif (new_status in ["cancelled", "returned"]) and (current_status in ["approved", "taken"]):
             cursor.execute("SELECT `total_unit`, `available_unit`, `unit_details` FROM `inventory_items` WHERE `id` = %s LIMIT 1", (barang_id,))
             inv = cursor.fetchone()
             if inv:
                 total = int(inv.get("total_unit", 1))
                 available = int(inv.get("available_unit", 0))
                 
                 unit_details_raw = safe_json_loads(inv.get("unit_details"), [])
                 units_restored = 0
                 
                 for unit in unit_details_raw:
                      if unit.get("reason") == f"Dipinjam (Req ID: {pengajuan_id})":
                           unit["status"] = "Tersedia"
                           unit["reason"] = ""
                           unit["available"] = True
                           units_restored += 1
                 
                 if units_restored == 0:
                      for unit in unit_details_raw:
                           if unit.get("status") == "Dipinjam" and units_restored < jumlah_req:
                                unit["status"] = "Tersedia"
                                unit["reason"] = ""
                                unit["available"] = True
                                units_restored += 1
                 
                 new_available = min(total, available + jumlah_req)
                 new_inv_status = "Tersedia" if new_available > 0 else "Dipinjam"
                 
                 cursor.execute(
                      "UPDATE `inventory_items` SET `available_unit` = %s, `status` = %s, `unit_details` = %s WHERE `id` = %s", 
                      (new_available, new_inv_status, json.dumps(unit_details_raw, ensure_ascii=False), barang_id)
                 )
        
        admin_name = str(session.get("nama") or session.get("username") or "Admin")
        approved_at = datetime.utcnow() if new_status != "pending" else None
        
        cursor.execute(
            """
            UPDATE `loan_requests` 
            SET `status` = %s, `admin_note` = %s, `approved_by` = %s, `approved_at` = COALESCE(`approved_at`, %s)
            WHERE `id` = %s
            """,
            (new_status, admin_note, admin_name, approved_at, pengajuan_id)
        )
        
        try:
            ensure_notifications_schema()
            barang_name = req.get("barang_name", "Barang")
            member_id = req.get("member_id")
            
            if member_id:
                status_label = "Disetujui" if new_status == "approved" else ("Ditolak" if new_status == "rejected" else ("Dibatalkan" if new_status == "cancelled" else new_status))
                
                notif_title = f"Peminjaman {status_label}: {barang_name}"
                notif_body = f"Pengajuan peminjaman Anda untuk {barang_name} telah {status_label.lower()} oleh Admin."
                if admin_note:
                    notif_body += f"<br>Catatan: {html.escape(admin_note)}"

                create_notification(
                    cursor, 
                    "peminjaman", 
                    notif_title, 
                    notif_body, 
                    url_for('dashboard_anggota') if False else "/pengajuan-peminjaman-barang-anggota.html", 
                    {"pengajuan_id": pengajuan_id, "target_user_id": member_id}, 
                    target_role="user"
                )
        except Exception as e:
            print(f"[WARN] Gagal mengirim notifikasi update status peminjaman: {e}")

        conn.commit()
        return jsonify({"success": True})
        
    except Exception as e:
        conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 400
    finally:
        cursor.close()
        conn.close()

# --- DAMAGE REPORT ENDPOINTS ---

@app.route("/api/inventory/search", methods=["GET"])
def search_inventory_items():
    """Search inventory items for auto-lookup"""
    ensure_auth_schema()
    query_str = request.args.get('q', '').strip().lower()
    if not query_str or len(query_str) < 2:
        return jsonify({"success": True, "items": []})
    
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        ensure_inventory_schema(cursor)
        
        cursor.execute("""
            SELECT `id`, `code`, `name`, `photos` FROM `inventory_items`
            WHERE LOWER(`name`) LIKE %s OR LOWER(`code`) LIKE %s
            LIMIT 20
        """, (f"%{query_str}%", f"%{query_str}%"))
        
        rows = cursor.fetchall() or []
        items = []
        for r in rows:
            items.append({
                "id": r.get("id"),
                "code": r.get("code"),
                "name": r.get("name"),
                "photos": normalize_inventory_photos(r.get("photos"))
            })
        return jsonify({"success": True, "items": items})
    finally:
        cursor.close()
        conn.close()

@app.route("/api/damage/upload", methods=["POST"])
def upload_damage_photo():
    """Upload damage report photos"""
    ensure_auth_schema()
    incoming_files = request.files.getlist("files")
    if not incoming_files:
        single_file = request.files.get("file")
        incoming_files = [single_file] if single_file and single_file.filename else []

    if not incoming_files:
        return jsonify({"success": False, "error": "Tidak ada file yang dipilih"}), 400

    saved_files: list[dict[str, object]] = []
    try:
        for file in incoming_files:
            if not file or not file.filename:
                continue
            # Validate file extension for images only
            file_ext = Path(file.filename).suffix.lower()
            if file_ext not in {".jpg", ".jpeg", ".png", ".gif", ".webp", ".bmp"}:
                return jsonify({"success": False, "error": f"Tipe file {file_ext} tidak diizinkan. Gunakan JPG, PNG, GIF, WebP, atau BMP"}), 400
            saved_files.append(save_uploaded_attachment(file))
    except ValueError as exc:
        return jsonify({"success": False, "error": str(exc)}), 400

    if not saved_files:
        return jsonify({"success": False, "error": "Gagal menyimpan file"}), 400

    return jsonify({
        "success": True,
        "files": saved_files,
    })

@app.route("/api/kerusakan", methods=["POST"])
def submit_damage_report():
    """Submit damage report form"""
    ensure_auth_schema()
    user_context = current_user_context()
    
    if not user_context.get("logged_in"):
        return jsonify({"success": False, "error": "Anda harus login terlebih dahulu"}), 401
    
    member_id = user_context.get("user_id")
    
    # Menangani form data biasa karena request dari FE dikirim via FormData
    barang_id = request.form.get("barangId")
    barang_name = normalize_text(request.form.get("itemName"), "")
    barang_code = normalize_text(request.form.get("itemCode"), "-")
    tingkat_kerusakan = normalize_text(request.form.get("severity"), "Sedang")
    deskripsi = normalize_text(request.form.get("chronology"), "")
    waktu_kejadian = request.form.get("incidentDate")
    incident_time = request.form.get("incidentTime")
    
    if not barang_name:
        return jsonify({"success": False, "error": "Nama barang harus diisi"}), 400
    
    if not deskripsi:
        return jsonify({"success": False, "error": "Deskripsi kerusakan harus diisi"}), 400
    
    if tingkat_kerusakan not in ["Ringan", "Sedang", "Berat", "Hilang"]:
        tingkat_kerusakan = "Sedang"
    
    # Process the photo if it exists in the form data
    foto_url = None
    photo_file = request.files.get("photo")
    if photo_file and photo_file.filename:
        try:
            saved = save_uploaded_attachment(photo_file)
            foto_url = saved.get("url")
        except ValueError as exc:
            return jsonify({"success": False, "error": str(exc)}), 400
    
    # Create damage report ID
    damage_id = f"krk-{int(time.time() * 1000)}-{uuid.uuid4().hex[:6]}"
    
    conn = mysql_connection()
    cursor = conn.cursor()
    try:
        ensure_damage_schema(cursor)
        
        # Simpan member details untuk referensi
        member_name = str(session.get("nama") or session.get("username") or "Anggota")
        member_identifier = str(session.get("username") or session.get("email"))
        
        cursor.execute("""
            INSERT INTO `form_kerusakan_barang` 
            (`id`, `member_id`, `barang_id`, `barang_name`, `barang_code`, `tingkat_kerusakan`, `status`, `deskripsi_kerusakan`, `waktu_kejadian`, `foto_kerusakan`)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            damage_id,
            str(member_id),
            barang_id if barang_id else None,
            barang_name,
            barang_code,
            tingkat_kerusakan,
            "Pending Review",
            deskripsi,
            f"{waktu_kejadian} {incident_time}" if waktu_kejadian and incident_time else datetime.now(),
            json.dumps([{"url": foto_url, "name": photo_file.filename}] if foto_url else [], ensure_ascii=False)
        ))
        
        # Create admin notification
        try:
            ensure_notifications_schema()
            create_notification(
                cursor, 
                "kerusakan",
                "Laporan Kerusakan Barang Baru",
                f"Laporan kerusakan barang dari {member_name}: {barang_name}",
                f"/hasil-form-kerusakan-barang.html",
                {"report_id": damage_id, "member_id": member_id, "barang_name": barang_name},
                target_role="admin"
            )
        except Exception:
            pass
            
        conn.commit()
        return jsonify({
            "success": True,
            "damageId": damage_id,
            "message": "Laporan kerusakan barang berhasil dikirim"
        })
    except Exception as e:
        conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 400
    finally:
        cursor.close()
        conn.close()

@app.route("/api/form-kerusakan/history", methods=["GET"])
def get_damage_history():
    """Get damage report history"""
    ensure_auth_schema()
    user_context = current_user_context()
    
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        ensure_damage_schema(cursor)
        
        role = user_context.get("role", "")
        member_id = user_context.get("user_id")
        
        query = "SELECT d.*, a.nama as member_name FROM `form_kerusakan_barang` d LEFT JOIN `anggota` a ON d.member_id = a.id WHERE 1=1"
        params = []
        
        # Non-admin users only see their own damage reports
        if role not in ["admin", "super_admin"]:
            if not member_id:
                return jsonify({"success": True, "items": []})
            query += " AND d.`member_id` = %s"
            params.append(str(member_id))
        
        # Apply filters
        status_filter = request.args.get('status')
        severity_filter = request.args.get('severity')
        search_query = request.args.get('search', '').strip()
        
        if status_filter and status_filter != 'all':
            query += " AND d.`status` = %s"
            params.append(status_filter)
        
        if severity_filter and severity_filter != 'all':
            query += " AND d.`tingkat_kerusakan` = %s"
            params.append(severity_filter)
        
        if search_query:
            query += " AND (d.`barang_name` LIKE %s OR d.`deskripsi_kerusakan` LIKE %s OR a.`nama` LIKE %s)"
            search_param = f"%{search_query}%"
            params.extend([search_param, search_param, search_param])
        
        query += " ORDER BY d.`created_at` DESC"
        
        cursor.execute(query, tuple(params))
        rows = cursor.fetchall() or []
        
        items = []
        for r in rows:
            items.append({
                "id": r.get("id"),
                "memberId": r.get("member_id"),
                "memberName": r.get("member_name"),
                "barangId": r.get("barang_id"),
                "barangName": r.get("barang_name"),
                "barangCode": r.get("barang_code"),
                "tingkatKerusakan": r.get("tingkat_kerusakan"),
                "status": r.get("status"),
                "deskripsiKerusakan": r.get("deskripsi_kerusakan"),
                "waktuKejadian": r.get("waktu_kejadian").isoformat() if r.get("waktu_kejadian") else None,
                "fotoKerusakan": json.loads(r.get("foto_kerusakan") or "[]"),
                "createdAt": r.get("created_at").isoformat() if r.get("created_at") else None,
                "updatedAt": r.get("updated_at").isoformat() if r.get("updated_at") else None,
            })
        
        return jsonify({"success": True, "items": items})
    finally:
        cursor.close()
        conn.close()

@app.route("/api/form-kerusakan/<damage_id>", methods=["GET"])
def get_damage_detail(damage_id):
    """Get single damage report detail"""
    ensure_auth_schema()
    user_context = current_user_context()
    
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        ensure_damage_schema(cursor)
        
        cursor.execute("""
            SELECT d.*, a.nama as member_name FROM `form_kerusakan_barang` d
            LEFT JOIN `anggota` a ON d.member_id = a.id
            WHERE d.`id` = %s LIMIT 1
        """, (damage_id,))
        
        row = cursor.fetchone()
        if not row:
            return jsonify({"success": False, "error": "Laporan tidak ditemukan"}), 404
        
        # Check access permission
        role = user_context.get("role", "")
        member_id = str(user_context.get("user_id"))
        
        if role not in ["admin", "super_admin"] and str(row.get("member_id")) != member_id:
            return jsonify({"success": False, "error": "Akses ditolak"}), 403
        
        return jsonify({
            "success": True,
            "item": {
                "id": row.get("id"),
                "memberId": row.get("member_id"),
                "memberName": row.get("member_name"),
                "barangId": row.get("barang_id"),
                "barangName": row.get("barang_name"),
                "barangCode": row.get("barang_code"),
                "tingkatKerusakan": row.get("tingkat_kerusakan"),
                "status": row.get("status"),
                "deskripsiKerusakan": row.get("deskripsi_kerusakan"),
                "waktuKejadian": row.get("waktu_kejadian").isoformat() if row.get("waktu_kejadian") else None,
                "fotoKerusakan": json.loads(row.get("foto_kerusakan") or "[]"),
                "createdAt": row.get("created_at").isoformat() if row.get("created_at") else None,
                "updatedAt": row.get("updated_at").isoformat() if row.get("updated_at") else None,
            }
        })
    finally:
        cursor.close()
        conn.close()

@app.route("/api/form-kerusakan/<damage_id>/status", methods=["PUT"])
def update_damage_status(damage_id):
    """Update damage report status (admin only)"""
    ensure_auth_schema()
    user_context = current_user_context()
    role = user_context.get("role", "")
    
    if role not in ["admin", "super_admin"]:
        return jsonify({"success": False, "error": "Anda tidak memiliki izin untuk operasi ini"}), 403
    
    data = request.get_json(silent=True) or {}
    new_status = data.get("status", "").strip()
    
    # HANYA 3 STATUS: Pending Review, Diproses, Selesai (Ditolak dihapus)
    if new_status not in ["Pending Review", "Diproses", "Selesai"]:
        return jsonify({"success": False, "error": "Status tidak valid"}), 400
    
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        ensure_damage_schema(cursor)
        
        cursor.execute("SELECT `member_id` FROM `form_kerusakan_barang` WHERE `id` = %s LIMIT 1", (damage_id,))
        row = cursor.fetchone()
        if not row:
            return jsonify({"success": False, "error": "Laporan tidak ditemukan"}), 404
        
        cursor.execute("""
            UPDATE `form_kerusakan_barang` 
            SET `status` = %s, `updated_at` = NOW()
            WHERE `id` = %s
        """, (new_status, damage_id))
        
        try:
            ensure_notifications_schema()
            cursor.execute("SELECT `nama` FROM `anggota` WHERE `id` = %s LIMIT 1", (row.get("member_id"),))
            member_row = cursor.fetchone()
            
            status_indo = {
                "Pending Review": "Dalam Review",
                "Diproses": "Sedang Diproses",
                "Selesai": "Selesai"
            }.get(new_status, new_status)
            
            create_notification(
                cursor,
                "kerusakan",
                "Status Laporan Kerusakan Diperbarui",
                f"Status laporan kerusakan Anda telah diubah menjadi: <b>{status_indo}</b>",
                f"/riwayat-form-kerusakan-barang-anggota.html",
                {"report_id": damage_id, "target_user_id": row.get("member_id")},
                target_role="user"
            )
        except Exception as e:
            print(f"[WARN] Gagal mengirim notifikasi update status kerusakan: {e}")
        
        conn.commit()
        return jsonify({"success": True})
    except Exception as e:
        conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 400
    finally:
        cursor.close()
        conn.close()

def ensure_news_schema() -> None:
    conn = mysql_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS `news_categories` (
                `id` VARCHAR(100) PRIMARY KEY,
                `name` VARCHAR(255) NOT NULL UNIQUE,
                `description` TEXT,
                `order_index` INT DEFAULT 0,
                `created_at` DATETIME DEFAULT CURRENT_TIMESTAMP,
                `updated_at` DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_general_ci
        """)

        cursor.execute("""
            CREATE TABLE IF NOT EXISTS `news` (
                `id` VARCHAR(100) PRIMARY KEY,
                `title` VARCHAR(500) NOT NULL,
                `slug` VARCHAR(500) NOT NULL UNIQUE,
                `content` LONGTEXT NOT NULL,
                `summary` TEXT,
                `thumbnails` JSON,
                `attachments` JSON,
                `status` VARCHAR(20) DEFAULT 'draft',
                `published_at` DATETIME,
                `created_at` DATETIME DEFAULT CURRENT_TIMESTAMP,
                `updated_at` DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_general_ci
        """)

        cursor.execute("""
            CREATE TABLE IF NOT EXISTS `news_category_mapping` (
                `id` INT AUTO_INCREMENT PRIMARY KEY,
                `news_id` VARCHAR(100) NOT NULL,
                `category_id` VARCHAR(100) NOT NULL,
                `created_at` DATETIME DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (`news_id`) REFERENCES `news`(`id`) ON DELETE CASCADE,
                FOREIGN KEY (`category_id`) REFERENCES `news_categories`(`id`) ON DELETE CASCADE,
                UNIQUE KEY `unique_news_category` (`news_id`, `category_id`)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_general_ci
        """)

        conn.commit()
    except Exception as e:
        print(f"[INFO] News schema ensure: {e}")
    finally:
        cursor.close()
        conn.close()

def ensure_agenda_schema() -> None:
    conn = mysql_connection()
    cursor = conn.cursor()
    try:
        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS `agendas` (
              `id` varchar(100) NOT NULL,
              `title` varchar(255) NOT NULL,
              `description` longtext DEFAULT NULL,
              `start_date` date NOT NULL,
              `start_time` time NOT NULL,
              `end_date` date DEFAULT NULL,
              `end_time` time DEFAULT NULL,
              `location` varchar(255) DEFAULT NULL,
              `registration_link` varchar(500) DEFAULT NULL,
              `image_url` text DEFAULT NULL,
              `image_name` varchar(255) DEFAULT NULL,
              `attachments` longtext DEFAULT NULL,
              `status` varchar(20) NOT NULL DEFAULT 'active',
              `order_index` int(11) NOT NULL DEFAULT 0,
              `created_at` datetime NOT NULL DEFAULT CURRENT_TIMESTAMP,
              `updated_at` datetime NOT NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
              PRIMARY KEY (`id`)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_general_ci
            """
        )
        conn.commit()
    finally:
        cursor.close()
        conn.close()

def ensure_registration_form_schema() -> None:
    conn = mysql_connection()
    cursor = conn.cursor()
    try:
        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS `registration_forms` (
              `id` varchar(100) NOT NULL,
              `title` varchar(255) NOT NULL,
              `description` longtext DEFAULT NULL,
              `target` varchar(20) NOT NULL DEFAULT 'public',
              `visibility` varchar(20) NOT NULL DEFAULT 'visible',
              `open_date` date DEFAULT NULL,
              `close_date` date DEFAULT NULL,
              `quota` int(11) NOT NULL DEFAULT 0,
              `image_url` text DEFAULT NULL,
              `image_name` varchar(255) DEFAULT NULL,
              `attachments` longtext DEFAULT NULL,
              `fields_json` longtext DEFAULT NULL,
              `created_by` varchar(100) DEFAULT NULL,
              `created_by_name` varchar(255) DEFAULT NULL,
              `created_by_role` varchar(50) DEFAULT NULL,
              `created_at` datetime NOT NULL DEFAULT CURRENT_TIMESTAMP,
              `updated_at` datetime NOT NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
              PRIMARY KEY (`id`),
              KEY `idx_registration_forms_target` (`target`),
              KEY `idx_registration_forms_visibility` (`visibility`),
              KEY `idx_registration_forms_dates` (`open_date`, `close_date`)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_general_ci
            """
        )
        ensure_column(cursor, "registration_forms", "image_url", "`image_url` text DEFAULT NULL")
        ensure_column(cursor, "registration_forms", "image_name", "`image_name` varchar(255) DEFAULT NULL")
        ensure_column(cursor, "registration_forms", "attachments", "`attachments` longtext DEFAULT NULL")

        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS `registration_form_submissions` (
              `id` varchar(100) NOT NULL,
              `form_id` varchar(100) NOT NULL,
              `submitter_key` varchar(255) NOT NULL,
              `submitter_identifier` varchar(255) NOT NULL DEFAULT '',
              `submitter_role` varchar(50) NOT NULL DEFAULT 'public',
              `submitter_user_id` varchar(100) DEFAULT NULL,
              `submitter_source` varchar(20) NOT NULL DEFAULT 'public',
              `answers_json` longtext NOT NULL,
              `submitted_at` datetime NOT NULL DEFAULT CURRENT_TIMESTAMP,
              PRIMARY KEY (`id`),
              UNIQUE KEY `uniq_registration_submission` (`form_id`, `submitter_key`),
              KEY `idx_registration_submissions_form` (`form_id`),
              KEY `idx_registration_submissions_submitter` (`submitter_key`),
              CONSTRAINT `fk_registration_submissions_form`
                FOREIGN KEY (`form_id`) REFERENCES `registration_forms` (`id`) ON DELETE CASCADE
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_general_ci
            """
        )
        conn.commit()
    finally:
        cursor.close()
        conn.close()

def ensure_notifications_schema() -> None:
    conn = mysql_connection()
    cursor = conn.cursor()
    try:
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS `notifications` (
              `id` varchar(100) NOT NULL,
              `type` varchar(50) DEFAULT NULL,
              `title` varchar(255) DEFAULT NULL,
              `body` text DEFAULT NULL,
              `url` varchar(255) DEFAULT NULL,
              `data` longtext DEFAULT NULL,
              `created_at` datetime NOT NULL DEFAULT CURRENT_TIMESTAMP,
              PRIMARY KEY (`id`)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_general_ci
        ''')

        cursor.execute('''
            CREATE TABLE IF NOT EXISTS `notification_reads` (
              `notification_id` varchar(100) NOT NULL,
              `user_key` varchar(255) NOT NULL,
              `read_at` datetime NOT NULL DEFAULT CURRENT_TIMESTAMP,
              PRIMARY KEY (`notification_id`,`user_key`)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_general_ci
        ''')
        ensure_column(cursor, "notifications", "target_role", "`target_role` varchar(50) DEFAULT NULL")
        conn.commit()
    finally:
        cursor.close()
        conn.close()

def create_notification(cursor, type_value: str, title: str, body: str, url: str | None = None, data: dict | None = None, target_role: str | None = None):
    nid = f"notif-{int(time.time() * 1000)}-{uuid.uuid4().hex[:6]}"
    cursor.execute(
        """
        INSERT INTO `notifications` (`id`, `type`, `title`, `body`, `url`, `data`, `target_role`)
        VALUES (%s, %s, %s, %s, %s, %s, %s)
        """,
        (
            nid,
            str(type_value or ""),
            str(title or ""),
            str(body or ""),
            str(url or "") if url else None,
            json.dumps(data or {}, ensure_ascii=False),
            target_role 
        ),
    )
    return nid

def create_notification_once(
    cursor,
    type_value: str,
    title: str,
    body: str,
    url: str | None = None,
    data: dict | None = None,
    target_role: str | None = None,
    *,
    dedupe_key: str | None = None,
):
    """Buat notifikasi sekali saja berdasarkan dedupe_key di payload data.

    Dedupe dibuat tahan terhadap variasi JSON yang memakai spasi ataupun tidak,
    karena data notifikasi lama di database bisa berasal dari beberapa versi kode.
    """
    payload = dict(data or {})
    clean_key = normalize_text(dedupe_key or payload.get("dedupe_key"))
    if clean_key:
        payload["dedupe_key"] = clean_key
        cursor.execute(
            """
            SELECT `id` FROM `notifications`
            WHERE `type` = %s
              AND (`data` LIKE %s OR `data` LIKE %s)
            LIMIT 1
            """,
            (
                str(type_value or ""),
                f'%"dedupe_key": "{clean_key}"%',
                f'%"dedupe_key":"{clean_key}"%',
            ),
        )
        existing = cursor.fetchone()
        if existing:
            return existing.get("id") if isinstance(existing, dict) else existing[0]
    return create_notification(cursor, type_value, title, body, url, payload, target_role)


def notification_target_url_for_member_role(role_value: str | None, *, default_user_url: str = "/jadwal-tugas-misa-anggota.html") -> str:
    normalized = normalize_role_value(role_value or "user")
    if normalized in {"admin", "super_admin"}:
        return "/jadwal-tugas-streaming-admin.html"
    return default_user_url


def create_task_success_notification(
    cursor,
    *,
    member_id: object,
    member_role: str | None = "user",
    misa_type: str,
    misa_name: str,
    role_name: str,
    date_text: str,
    time_text: str,
    source: str = "member_request",
    misa_besar_id: object | None = None,
):
    """Notif langsung setelah anggota berhasil mengambil/request jadwal mandiri."""
    safe_role = normalize_text(role_name) or "Role"
    safe_misa = normalize_text(misa_name) or ("Misa Besar" if misa_type == "misa_besar" else "Misa Biasa")
    time_label = format_time_hhmm(time_text)
    day_label = request_task_day_name(date_text) if 'request_task_day_name' in globals() else "-"
    date_label = request_task_format_date(date_text) if 'request_task_format_date' in globals() else normalize_text(date_text)
    title = f"Request Tugas Berhasil: {safe_role}"
    body = (
        f"Anda berhasil terdaftar sebagai <b>{html.escape(safe_role)}</b> untuk "
        f"<b>{html.escape(safe_misa)}</b> pada hari {html.escape(day_label)}, "
        f"{html.escape(date_label)} jam {html.escape(time_label)} WIB."
    )
    return create_notification(
        cursor,
        "tugas",
        title,
        body,
        "/request-tugas-anggota.html",
        {
            "target_user_id": str(member_id),
            "notification_kind": "member_request_success",
            "misa_type": misa_type,
            "misa_besar_id": misa_besar_id,
            "misa_name": safe_misa,
            "role": safe_role,
            "misa_date": date_text,
            "misa_time": time_label,
            "source": source,
        },
        target_role=None,
    )


def create_due_task_reminder_notifications() -> int:
    """Generate pengingat H-1 dan hari-H untuk semua petugas yang sudah terdaftar.

    Fungsi ini dipanggil saat dashboard memuat notifikasi. Dedupe key mencegah
    notifikasi H-1 / hari-H dibuat berulang untuk jadwal dan petugas yang sama.
    """
    ensure_task_request_schema()
    ensure_notifications_schema()
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True, buffered=True)
    sent = 0
    try:
        today = datetime.now().date()
        tomorrow = today + timedelta(days=1)
        today_text = today.strftime("%Y-%m-%d")
        tomorrow_text = tomorrow.strftime("%Y-%m-%d")

        reminder_labels = {
            today_text: ("hari_h", "Hari Ini", "Pengingat Tugas Hari Ini"),
            tomorrow_text: ("h_1", "H-1", "Pengingat Tugas H-1"),
        }

        # Misa Biasa
        cursor.execute(
            """
            SELECT DATE_FORMAT(sa.schedule_date, '%Y-%m-%d') AS schedule_date,
                   DATE_FORMAT(sa.schedule_time, '%H:%i') AS schedule_time,
                   sa.role_name, sa.member_id,
                   COALESCE(a.nama, CONCAT('ID ', sa.member_id)) AS member_name,
                   COALESCE(a.role, 'user') AS member_role,
                   COALESCE(cfg.mass_name, 'Misa Biasa') AS mass_name
            FROM streaming_assignments sa
            JOIN anggota a ON a.id = sa.member_id
            LEFT JOIN streaming_weekly_config cfg
              ON cfg.day_name = CASE WEEKDAY(sa.schedule_date)
                WHEN 0 THEN 'Senin' WHEN 1 THEN 'Selasa' WHEN 2 THEN 'Rabu'
                WHEN 3 THEN 'Kamis' WHEN 4 THEN 'Jumat' WHEN 5 THEN 'Sabtu'
                ELSE 'Minggu' END
              AND DATE_FORMAT(cfg.start_time, '%H:%i') = DATE_FORMAT(sa.schedule_time, '%H:%i')
            WHERE sa.schedule_date IN (%s, %s)
              AND LOWER(COALESCE(a.status_akun, 'aktif')) IN ('', 'aktif', 'active')
            """,
            (today_text, tomorrow_text),
        )
        for row in cursor.fetchall() or []:
            date_text = normalize_text(row.get("schedule_date"))
            reminder_code, reminder_label, title_prefix = reminder_labels.get(date_text, ("", "", ""))
            if not reminder_code:
                continue
            time_text = format_time_hhmm(row.get("schedule_time"))
            role_name = normalize_text(row.get("role_name")) or "Role"
            misa_name = normalize_text(row.get("mass_name")) or "Misa Biasa"
            member_id = normalize_text(row.get("member_id"))
            member_role = normalize_text(row.get("member_role")) or "user"
            day_label = request_task_day_name(date_text)
            date_label = request_task_format_date(date_text)
            title = f"{title_prefix}: {role_name}"
            body = (
                f"{html.escape(reminder_label)}: Anda bertugas sebagai <b>{html.escape(role_name)}</b> "
                f"untuk <b>{html.escape(misa_name)}</b> pada hari {html.escape(day_label)}, "
                f"{html.escape(date_label)} jam {html.escape(time_text)} WIB."
            )
            dedupe_key = f"task-reminder:biasa:{date_text}:{time_text}:{role_name}:{member_id}:{reminder_code}"
            create_notification_once(
                cursor, "tugas", title, body,
                notification_target_url_for_member_role(member_role),
                {
                    "target_user_id": member_id,
                    "notification_kind": "task_reminder",
                    "reminder": reminder_code,
                    "reminder_label": reminder_label,
                    "misa_type": "misa_biasa",
                    "misa_name": misa_name,
                    "role": role_name,
                    "misa_date": date_text,
                    "misa_time": time_text,
                },
                target_role=None,
                dedupe_key=dedupe_key,
            )
            sent += 1

        # Misa Besar
        cursor.execute(
            """
            SELECT mb.id AS misa_id, mb.misa_name, DATE_FORMAT(mb.misa_date, '%Y-%m-%d') AS misa_date,
                   DATE_FORMAT(mb.misa_time, '%H:%i') AS misa_time,
                   n.id AS role_id, n.role_name, a.member_id,
                   COALESCE(ag.nama, CONCAT('ID ', a.member_id)) AS member_name,
                   COALESCE(ag.role, 'user') AS member_role
            FROM misa_besar_assignments a
            JOIN misa_besar_names n ON n.id = a.role_id
            JOIN misa_besar mb ON mb.id = n.misa_id
            JOIN anggota ag ON ag.id = a.member_id
            WHERE mb.status = 'published'
              AND mb.misa_date IN (%s, %s)
              AND LOWER(COALESCE(ag.status_akun, 'aktif')) IN ('', 'aktif', 'active')
            """,
            (today_text, tomorrow_text),
        )
        for row in cursor.fetchall() or []:
            date_text = normalize_text(row.get("misa_date"))
            reminder_code, reminder_label, title_prefix = reminder_labels.get(date_text, ("", "", ""))
            if not reminder_code:
                continue
            time_text = format_time_hhmm(row.get("misa_time"))
            role_name = normalize_text(row.get("role_name")) or "Role"
            misa_name = normalize_text(row.get("misa_name")) or "Misa Besar"
            misa_id = row.get("misa_id")
            member_id = normalize_text(row.get("member_id"))
            member_role = normalize_text(row.get("member_role")) or "user"
            day_label = request_task_day_name(date_text)
            date_label = request_task_format_date(date_text)
            title = f"{title_prefix}: {role_name}"
            body = (
                f"{html.escape(reminder_label)}: Anda bertugas sebagai <b>{html.escape(role_name)}</b> "
                f"untuk <b>{html.escape(misa_name)}</b> pada hari {html.escape(day_label)}, "
                f"{html.escape(date_label)} jam {html.escape(time_text)} WIB."
            )
            dedupe_key = f"task-reminder:besar:{misa_id}:{row.get('role_id')}:{member_id}:{reminder_code}"
            create_notification_once(
                cursor, "tugas", title, body,
                notification_target_url_for_member_role(member_role),
                {
                    "target_user_id": member_id,
                    "notification_kind": "task_reminder",
                    "reminder": reminder_code,
                    "reminder_label": reminder_label,
                    "misa_type": "misa_besar",
                    "misa_besar_id": misa_id,
                    "misa_name": misa_name,
                    "role": role_name,
                    "misa_date": date_text,
                    "misa_time": time_text,
                },
                target_role=None,
                dedupe_key=dedupe_key,
            )
            sent += 1

        conn.commit()
        return sent
    except Exception:
        conn.rollback()
        raise
    finally:
        cursor.close()
        conn.close()


@app.route("/api/notifications", methods=["GET"])
def get_notifications():
    ensure_notifications_schema()
    try:
        create_due_task_reminder_notifications()
    except Exception as exc:
        print(f"[WARN] Gagal membuat notifikasi pengingat tugas: {exc}")
    try:
        create_monthly_requirement_notifications()
    except Exception as exc:
        print(f"[WARN] Gagal membuat notifikasi target bulanan: {exc}")
    try:
        create_streaming_evaluation_reminder_notifications()
    except Exception as exc:
        print(f"[WARN] Gagal membuat notifikasi pengingat evaluasi streaming: {exc}")
    viewer = current_user_context()
    client_key = str(request.args.get("clientKey") or request.headers.get("X-Registration-Client-Key") or "").strip()
    user_key = None
    role = "user" 
    
    if viewer.get("logged_in"):
        user_key = f"member:{viewer.get('user_id') or viewer.get('username') or viewer.get('email') or 'member'}"
        role = viewer.get("role") or "user"
    else:
        if client_key:
            user_key = client_key

    limit = int(request.args.get("limit") or 50)
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        target_roles = ["admin", "super_admin"] if role in ["admin", "super_admin"] else ["user"]
        format_strings = ','.join(['%s'] * len(target_roles))

        # Filter target_user_id langsung di SQL, bukan hanya di frontend.
        # Sebelumnya API mengambil 50 notifikasi global terlebih dahulu, lalu dashboard
        # memfilter target_user_id di browser. Jika ada banyak notifikasi milik user lain,
        # notifikasi evaluasi milik user yang sedang login bisa tidak ikut terbawa limit.
        user_id = normalize_text(viewer.get("user_id")) if viewer.get("logged_in") else ""
        target_user_clause = ""
        target_user_params = []
        if user_id:
            target_user_clause = """
              AND (
                `data` LIKE %s
                OR `data` LIKE %s
                OR `data` LIKE %s
                OR `data` LIKE %s
                OR `target_role` IS NOT NULL
                OR `type` NOT IN ('tugas','evaluasi','tukar','keanggotaan','peminjaman','kerusakan')
              )
            """
            target_user_params = [
                f'%"target_user_id": "{user_id}"%',
                f'%"target_user_id":"{user_id}"%',
                f'%"target_user_id": {user_id}%',
                f'%"target_user_id":{user_id}%',
            ]

        inactive_type_clause = ""
        if viewer.get("logged_in") and normalize_role_value(role) == "user" and normalize_status(viewer.get("status_akun") or "aktif") == "nonaktif":
            inactive_type_clause = " AND `type` = 'keanggotaan'"

        query = f"""
            SELECT * FROM `notifications`
            WHERE (`target_role` IS NULL OR `target_role` IN ({format_strings}))
            {target_user_clause}
            {inactive_type_clause}
            ORDER BY `created_at` DESC, `id` DESC LIMIT %s
        """
        params = tuple(target_roles) + tuple(target_user_params) + (limit,)

        cursor.execute(query, params)
        rows = cursor.fetchall() or []
        results = []
        viewer_user_id = normalize_text(viewer.get("user_id")) if viewer.get("logged_in") else ""
        viewer_created_at = None
        if viewer_user_id:
            try:
                cursor.execute("SELECT created_at FROM anggota WHERE id = %s LIMIT 1", (viewer_user_id,))
                member_created_row = cursor.fetchone()
                viewer_created_at = member_created_row.get("created_at") if member_created_row else None
            except Exception:
                viewer_created_at = None

        user_scoped_types = {"tugas", "evaluasi", "tukar", "keanggotaan", "peminjaman", "kerusakan"}
        for r in rows:
            try:
                payload = json.loads(r.get("data") or "{}")
            except Exception:
                payload = {}

            # Filter ulang di Python supaya notifikasi lama/orphan tidak bocor ke user baru.
            payload_target_user = normalize_text(payload.get("target_user_id"))
            notif_type = normalize_text(r.get("type"))
            notif_created = r.get("created_at")

            if viewer_user_id:
                if payload_target_user and payload_target_user != viewer_user_id:
                    continue
                # Notifikasi personal seperti Tugas Baru lama yang belum punya target_user_id
                # tidak boleh tampil ke semua anggota. Broadcast tetap boleh jika target_role diisi.
                if notif_type in user_scoped_types and not payload_target_user and not normalize_text(r.get("target_role")):
                    continue
                # Jika id user pernah dipakai/akun baru dibuat setelah notifikasi lama, jangan tampilkan.
                if viewer_created_at and notif_created and notif_created < viewer_created_at and notif_type in user_scoped_types:
                    continue

            is_read = False
            if user_key:
                cursor.execute("SELECT 1 FROM `notification_reads` WHERE `notification_id` = %s AND `user_key` = %s LIMIT 1", (r["id"], user_key))
                is_read = cursor.fetchone() is not None
            results.append({
                "id": r.get("id"),
                "type": r.get("type"),
                "title": r.get("title"),
                "body": r.get("body"),
                "url": r.get("url"),
                "data": payload,
                "createdAt": r.get("created_at").isoformat() if r.get("created_at") else None,
                "read": bool(is_read),
            })
        return jsonify(results)
    finally:
        cursor.close()
        conn.close()


def dashboard_period_bounds(range_value: str) -> tuple[object, object]:
    """Return inclusive start/end dates for dashboard range."""
    now = datetime.now()
    mode = normalize_text(range_value).lower() or "month"
    if mode == "year":
        return datetime(now.year, 1, 1).date(), datetime(now.year, 12, 31).date()
    if mode == "week":
        start_dt = now - timedelta(days=now.weekday())
        start = datetime(start_dt.year, start_dt.month, start_dt.day).date()
        return start, start + timedelta(days=6)
    start = datetime(now.year, now.month, 1).date()
    end = datetime(now.year, now.month, calendar.monthrange(now.year, now.month)[1]).date()
    return start, end


def dashboard_series_template(range_value: str) -> tuple[list[str], list[object]]:
    now = datetime.now()
    mode = normalize_text(range_value).lower() or "month"
    month_names = ["Jan", "Feb", "Mar", "Apr", "Mei", "Jun", "Jul", "Agu", "Sep", "Okt", "Nov", "Des"]
    if mode == "year":
        labels = month_names[:]
        keys = [(now.year, month) for month in range(1, 13)]
        return labels, keys
    if mode == "week":
        start, _ = dashboard_period_bounds("week")
        labels = ["Sen", "Sel", "Rab", "Kam", "Jum", "Sab", "Min"]
        keys = [(start + timedelta(days=idx)).isoformat() for idx in range(7)]
        return labels, keys
    start, end = dashboard_period_bounds("month")
    labels = []
    keys = []
    total_days = end.day
    for day in range(1, total_days + 1):
        labels.append(str(day) if day == 1 or day == total_days or day % 3 == 0 else "")
        keys.append(datetime(now.year, now.month, day).date().isoformat())
    return labels, keys


def dashboard_count_schedules_by_range(cursor, range_value: str) -> int:
    start, end = dashboard_period_bounds(range_value)
    schedules = eval_all_schedules(cursor, start, end, "all")
    return len(schedules)


def dashboard_schedule_series(cursor, range_value: str) -> dict[str, object]:
    mode = normalize_text(range_value).lower() or "month"
    start, end = dashboard_period_bounds(mode)
    labels, keys = dashboard_series_template(mode)
    counts = {key: 0 for key in keys}
    schedules = eval_all_schedules(cursor, start, end, "all")
    for item in schedules:
        date_text = normalize_text(item.get("date"))
        if not date_text:
            continue
        if mode == "year":
            try:
                parsed = datetime.strptime(date_text, "%Y-%m-%d")
                key = (parsed.year, parsed.month)
            except Exception:
                continue
        else:
            key = date_text
        if key in counts:
            counts[key] += 1
    return {"range": mode, "labels": labels, "values": [counts.get(key, 0) for key in keys]}


def dashboard_loan_series(cursor, range_value: str) -> dict[str, object]:
    mode = normalize_text(range_value).lower() or "year"
    start, end = dashboard_period_bounds(mode)
    labels, keys = dashboard_series_template(mode)
    counts = {key: 0 for key in keys}
    cursor.execute(
        """
        SELECT DATE(COALESCE(`tanggal_pengajuan`, `created_at`)) AS item_date, COUNT(*) AS total
        FROM `loan_requests`
        WHERE DATE(COALESCE(`tanggal_pengajuan`, `created_at`)) BETWEEN %s AND %s
        GROUP BY DATE(COALESCE(`tanggal_pengajuan`, `created_at`))
        """,
        (start.isoformat(), end.isoformat()),
    )
    for row in cursor.fetchall() or []:
        raw_date = row.get("item_date") if isinstance(row, dict) else None
        date_text = raw_date.isoformat() if hasattr(raw_date, "isoformat") else normalize_text(raw_date)
        if not date_text:
            continue
        if mode == "year":
            try:
                parsed = datetime.strptime(date_text, "%Y-%m-%d")
                key = (parsed.year, parsed.month)
            except Exception:
                continue
        else:
            key = date_text
        if key in counts:
            counts[key] = int(row.get("total") or 0)
    return {"range": mode, "labels": labels, "values": [counts.get(key, 0) for key in keys]}


def dashboard_pending_evaluation_count(cursor) -> int:
    now = datetime.now()
    start, end = dashboard_period_bounds("month")
    schedules = eval_all_schedules(cursor, start, end, "all")
    evaluation_map = eval_evaluation_map(cursor, start.isoformat(), end.isoformat())
    pending = 0
    for schedule in schedules:
        schedule_dt = eval_datetime_from_parts(schedule.get("date"), schedule.get("time"))
        if not schedule_dt or schedule_dt > now:
            continue
        if not eval_schedule_has_staff(schedule):
            continue
        key = (normalize_text(schedule.get("kind")), normalize_text(schedule.get("scheduleKey")))
        if key not in evaluation_map:
            pending += 1
    return pending




# -----------------------------------------------------------------------------
# Dashboard Anggota: ringkasan backend real-time untuk panel anggota
# -----------------------------------------------------------------------------

def member_dashboard_schedule_dt(date_text: str, time_text: str = "00:00") -> datetime:
    try:
        return datetime.strptime(f"{normalize_text(date_text)} {format_time_hhmm(time_text)}", "%Y-%m-%d %H:%M")
    except Exception:
        try:
            return datetime.strptime(normalize_text(date_text), "%Y-%m-%d")
        except Exception:
            return datetime(1970, 1, 1)


def member_dashboard_type_filter(value: str) -> str:
    raw = normalize_text(value).lower()
    if raw in {"biasa", "misa_biasa"}:
        return "biasa"
    if raw in {"besar", "misa_besar"}:
        return "besar"
    return "all"


def member_dashboard_kind_label(kind: str) -> str:
    return "Misa Besar" if kind == "besar" else "Misa Biasa"


def member_dashboard_registered_tasks(cursor, member_id: object, *, month: int, year: int, kind_filter: str = "all") -> list[dict[str, object]]:
    """Ambil tugas terdaftar user untuk tabel dashboard anggota.

    Data dibangun dari assignment aktif. Tugas yang sudah dibatalkan tidak muncul
    lagi di tabel ini karena histori pembatalan sudah tersedia di Riwayat Tugas.
    """
    kind_filter = member_dashboard_type_filter(kind_filter)
    start_date = f"{year:04d}-{month:02d}-01"
    end_date = f"{year:04d}-{month:02d}-{calendar.monthrange(year, month)[1]:02d}"
    evaluation_map = {}
    try:
        evaluation_map = eval_evaluation_map(cursor, start_date, end_date)
    except Exception as exc:
        print(f"[WARN] Dashboard anggota gagal membaca evaluasi: {exc}")

    now = datetime.now()
    rows: list[dict[str, object]] = []

    if kind_filter in {"all", "biasa"}:
        cursor.execute(
            """
            SELECT sa.id AS assignment_id,
                   DATE_FORMAT(sa.schedule_date, '%Y-%m-%d') AS date,
                   DATE_FORMAT(sa.schedule_time, '%H:%i') AS time,
                   sa.role_name AS role,
                   COALESCE(sa.request_source, 'admin') AS request_source,
                   COALESCE(cfg.mass_name, 'Misa Biasa') AS mass_name
            FROM streaming_assignments sa
            LEFT JOIN streaming_weekly_config cfg
              ON cfg.day_name = CASE WEEKDAY(sa.schedule_date)
                WHEN 0 THEN 'Senin' WHEN 1 THEN 'Selasa' WHEN 2 THEN 'Rabu'
                WHEN 3 THEN 'Kamis' WHEN 4 THEN 'Jumat' WHEN 5 THEN 'Sabtu'
                ELSE 'Minggu' END
              AND DATE_FORMAT(cfg.start_time, '%H:%i') = DATE_FORMAT(sa.schedule_time, '%H:%i')
            LEFT JOIN streaming_cancelled sc
              ON sc.mass_date = sa.schedule_date
             AND DATE_FORMAT(sc.mass_time, '%H:%i') = DATE_FORMAT(sa.schedule_time, '%H:%i')
            LEFT JOIN misa_besar mb
              ON mb.status = 'published'
             AND mb.misa_date = sa.schedule_date
             AND DATE_FORMAT(mb.misa_time, '%H:%i') = DATE_FORMAT(sa.schedule_time, '%H:%i')
            WHERE sa.member_id = %s
              AND MONTH(sa.schedule_date) = %s
              AND YEAR(sa.schedule_date) = %s
              AND sc.id IS NULL
              AND mb.id IS NULL
            ORDER BY sa.schedule_date DESC, sa.schedule_time DESC, sa.id DESC
            """,
            (member_id, month, year),
        )
        for row in cursor.fetchall() or []:
            date_text = normalize_text(row.get("date"))
            time_text = format_time_hhmm(row.get("time"))
            schedule_dt = member_dashboard_schedule_dt(date_text, time_text)
            schedule_key = f"biasa:{date_text}:{time_text}"
            evaluated = ("misa_biasa", schedule_key) in evaluation_map
            can_cancel = schedule_dt > now and cancel_task_can_cancel(date_text)
            rows.append({
                "id": f"biasa:{row.get('assignment_id')}",
                "type": "biasa",
                "typeLabel": "Misa Biasa",
                "assignmentId": row.get("assignment_id"),
                "misaId": None,
                "roleId": None,
                "misaName": normalize_text(row.get("mass_name")) or "Misa Biasa",
                "date": date_text,
                "dateLabel": request_task_format_date(date_text),
                "dayName": request_task_day_name(date_text),
                "time": time_text,
                "role": normalize_text(row.get("role")) or "Role",
                "status": "Selesai" if schedule_dt < now else "Terdaftar",
                "evaluation": "Sudah Diisi" if evaluated else "Belum Diisi",
                "evaluated": evaluated,
                "source": normalize_text(row.get("request_source")) or "admin",
                "sourceLabel": "Request Mandiri" if normalize_text(row.get("request_source")) == "member_request" else "Ditugaskan Admin",
                "canCancel": bool(can_cancel),
                "cancelDeadlineLabel": cancel_task_rule_label(date_text),
            })

    if kind_filter in {"all", "besar"}:
        cursor.execute(
            """
            SELECT a.id AS assignment_id, mb.id AS misa_id, n.id AS role_id,
                   mb.misa_name,
                   DATE_FORMAT(mb.misa_date, '%Y-%m-%d') AS date,
                   DATE_FORMAT(mb.misa_time, '%H:%i') AS time,
                   n.role_name AS role,
                   COALESCE(a.request_source, 'admin') AS request_source
            FROM misa_besar_assignments a
            JOIN misa_besar_names n ON n.id = a.role_id
            JOIN misa_besar mb ON mb.id = n.misa_id
            WHERE a.member_id = %s
              AND mb.status = 'published'
              AND MONTH(mb.misa_date) = %s
              AND YEAR(mb.misa_date) = %s
            ORDER BY mb.misa_date DESC, mb.misa_time DESC, a.id DESC
            """,
            (member_id, month, year),
        )
        for row in cursor.fetchall() or []:
            date_text = normalize_text(row.get("date"))
            time_text = format_time_hhmm(row.get("time"))
            schedule_dt = member_dashboard_schedule_dt(date_text, time_text)
            schedule_key = f"besar:{row.get('misa_id')}"
            evaluated = ("misa_besar", schedule_key) in evaluation_map
            can_cancel = schedule_dt > now and cancel_task_can_cancel(date_text)
            rows.append({
                "id": f"besar:{row.get('assignment_id')}",
                "type": "besar",
                "typeLabel": "Misa Besar",
                "assignmentId": row.get("assignment_id"),
                "misaId": row.get("misa_id"),
                "roleId": row.get("role_id"),
                "misaName": normalize_text(row.get("misa_name")) or "Misa Besar",
                "date": date_text,
                "dateLabel": request_task_format_date(date_text),
                "dayName": request_task_day_name(date_text),
                "time": time_text,
                "role": normalize_text(row.get("role")) or "Role",
                "status": "Selesai" if schedule_dt < now else "Terdaftar",
                "evaluation": "Sudah Diisi" if evaluated else "Belum Diisi",
                "evaluated": evaluated,
                "source": normalize_text(row.get("request_source")) or "admin",
                "sourceLabel": "Request Mandiri" if normalize_text(row.get("request_source")) == "member_request" else "Ditugaskan Admin",
                "canCancel": bool(can_cancel),
                "cancelDeadlineLabel": cancel_task_rule_label(date_text),
            })

    rows.sort(key=lambda item: member_dashboard_schedule_dt(item.get("date"), item.get("time")), reverse=True)
    return rows


def member_dashboard_request_chart(cursor, member_id: object, range_value: str) -> dict[str, object]:
    mode = normalize_text(range_value).lower() or "month"
    if mode not in {"week", "month", "year"}:
        mode = "month"
    start, end = dashboard_period_bounds(mode)
    labels, keys = dashboard_series_template(mode)
    counts = {key: 0 for key in keys}

    def add_rows(table_sql: str, params: tuple[object, ...]):
        cursor.execute(table_sql, params)
        for row in cursor.fetchall() or []:
            raw_date = row.get("item_date") if isinstance(row, dict) else None
            date_text = raw_date.isoformat() if hasattr(raw_date, "isoformat") else normalize_text(raw_date)
            if not date_text:
                continue
            if mode == "year":
                try:
                    parsed = datetime.strptime(date_text, "%Y-%m-%d")
                    key = (parsed.year, parsed.month)
                except Exception:
                    continue
            else:
                key = date_text
            if key in counts:
                counts[key] += parse_required_int(row.get("total") if isinstance(row, dict) else 0, 0)

    add_rows(
        """
        SELECT DATE(schedule_date) AS item_date, COUNT(*) AS total
        FROM streaming_assignments
        WHERE member_id = %s
          AND DATE(schedule_date) BETWEEN %s AND %s
        GROUP BY DATE(schedule_date)
        """,
        (member_id, start.isoformat(), end.isoformat()),
    )
    add_rows(
        """
        SELECT DATE(mb.misa_date) AS item_date, COUNT(*) AS total
        FROM misa_besar_assignments a
        JOIN misa_besar_names n ON n.id = a.role_id
        JOIN misa_besar mb ON mb.id = n.misa_id
        WHERE a.member_id = %s
          AND mb.status = 'published'
          AND DATE(mb.misa_date) BETWEEN %s AND %s
        GROUP BY DATE(mb.misa_date)
        """,
        (member_id, start.isoformat(), end.isoformat()),
    )
    return {"range": mode, "labels": labels, "values": [counts.get(key, 0) for key in keys]}


def member_dashboard_recommended_tasks(cursor, member_id: object, *, target_minimum: int, regular_count: int) -> list[dict[str, object]]:
    if regular_count >= target_minimum:
        return []
    now = datetime.now()
    items = request_task_regular_items(cursor, now.month, now.year, str(member_id))
    result: list[dict[str, object]] = []
    for item in items:
        if not item.get("canRequest"):
            continue
        schedule_dt = member_dashboard_schedule_dt(item.get("date"), item.get("time"))
        if schedule_dt < now:
            continue
        roles = item.get("roles") or []
        filled_count = sum(1 for role in roles if isinstance(role, dict) and role.get("filled"))
        total_roles = len(roles)
        open_roles = [normalize_text(role.get("role")) for role in (item.get("openRoles") or []) if isinstance(role, dict) and normalize_text(role.get("role"))]
        first_role = open_roles[0] if open_roles else "Role kosong"
        result.append({
            "id": item.get("id"),
            "type": "biasa",
            "typeLabel": "Misa Biasa",
            "misaName": item.get("misaName"),
            "date": item.get("date"),
            "dateLabel": item.get("dateLabel"),
            "dayName": item.get("dayName"),
            "time": item.get("time"),
            "role": first_role,
            "openRoles": open_roles,
            "filledCount": filled_count,
            "totalRoles": total_roles,
            "href": f"request-tugas-anggota.html?jenis=biasa&month={now.month}&year={now.year}",
        })
        if len(result) >= 3:
            break
    return result


@app.route("/api/dashboard/anggota-overview", methods=["GET"])
def api_dashboard_anggota_overview():
    ensure_auth_schema()
    ensure_task_request_schema()
    ensure_task_cancellation_schema()
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True, buffered=True)
    try:
        ensure_monthly_monitoring_schema(cursor)
        ensure_streaming_evaluation_schema(cursor)
        member, error = member_monitoring_current_user(cursor)
        if error:
            return error
        member_id = member.get("id")
        now = datetime.now()

        schedule_range = normalize_text(request.args.get("scheduleRange") or "week").lower()
        done_range = normalize_text(request.args.get("doneRange") or "month").lower()
        chart_range = normalize_text(request.args.get("chartRange") or "month").lower()
        if schedule_range not in {"week", "month", "year"}:
            schedule_range = "week"
        if done_range not in {"week", "month", "year"}:
            done_range = "month"
        if chart_range not in {"week", "month", "year"}:
            chart_range = "month"

        task_month = min(12, max(1, parse_required_int(request.args.get("taskMonth"), now.month)))
        task_year = parse_required_int(request.args.get("taskYear"), now.year)
        task_type = member_dashboard_type_filter(request.args.get("taskType") or "all")

        # Metric 1: semua jadwal tersedia pada periode terpilih.
        schedule_start, schedule_end = dashboard_period_bounds(schedule_range)
        available_schedule_count = len(eval_all_schedules(cursor, schedule_start, schedule_end, "all"))

        # Target minimum berlaku untuk Misa Biasa bulan berjalan.
        target_minimum = monitoring_get_target_minimum(cursor)
        current_regular_count = member_monitoring_regular_count_for_period(cursor, member_id, now.year, now.month)
        pending_minimum = max(0, target_minimum - current_regular_count)

        # Metric tugas akan datang selalu bulan berjalan, semua jenis tugas.
        current_month_tasks = member_monitoring_fetch_all_tasks(cursor, member_id, month=str(now.month), year=str(now.year), include_big=True)
        upcoming_count = sum(1 for task in current_month_tasks if member_dashboard_schedule_dt(task.get("date"), task.get("time")) > now)

        # Metric selesai mengikuti select Minggu/Bulan/Tahun, semua jenis tugas.
        done_start, done_end = dashboard_period_bounds(done_range)
        done_tasks = member_monitoring_fetch_all_tasks(cursor, member_id, month="all", year=str(now.year), include_big=True)
        done_count = 0
        for task in done_tasks:
            task_dt = member_dashboard_schedule_dt(task.get("date"), task.get("time"))
            if done_start <= task_dt.date() <= done_end and task_dt <= now:
                done_count += 1

        request_chart = member_dashboard_request_chart(cursor, member_id, chart_range)
        recommendations = member_dashboard_recommended_tasks(cursor, member_id, target_minimum=target_minimum, regular_count=current_regular_count)
        registered_tasks = member_dashboard_registered_tasks(cursor, member_id, month=task_month, year=task_year, kind_filter=task_type)

        conn.commit()
        return jsonify({
            "success": True,
            "currentUser": current_user_context_from_row({
                "id": member.get("id"),
                "nama": member.get("name"),
                "username": member.get("username") or "",
                "role": member.get("role") or "user",
                "status_akun": member.get("status_akun") or "aktif",
            }),
            "metrics": {
                "availableSchedules": {"range": schedule_range, "total": available_schedule_count},
                "targetMinimum": target_minimum,
                "regularTasksThisMonth": current_regular_count,
                "pendingMinimum": pending_minimum,
                "upcomingThisMonth": upcoming_count,
                "doneTasks": {"range": done_range, "total": done_count},
            },
            "requestChart": request_chart,
            "tasksToTake": recommendations,
            "registeredTasks": registered_tasks,
            "filters": {"taskMonth": task_month, "taskYear": task_year, "taskType": task_type},
        })
    except Exception as exc:
        conn.rollback()
        print(f"[ERROR] Dashboard anggota gagal memuat data: {exc}")
        return jsonify({"success": False, "error": f"Gagal memuat dashboard anggota: {exc}"}), 500
    finally:
        cursor.close()
        conn.close()


@app.route("/api/dashboard/admin-overview", methods=["GET"])
def api_dashboard_admin_overview():
    role = normalize_role_value(session.get("role") or "")
    if not session.get("logged_in") or role not in {"admin", "super_admin"}:
        return jsonify({"success": False, "error": "Akses ditolak."}), 403

    metric_range = normalize_text(request.args.get("scheduleRange") or "week")
    schedule_chart_range = normalize_text(request.args.get("scheduleChartRange") or "week")
    loan_chart_range = normalize_text(request.args.get("loanChartRange") or "year")
    if metric_range not in {"week", "month", "year"}:
        metric_range = "week"
    if schedule_chart_range not in {"week", "month", "year"}:
        schedule_chart_range = "week"
    if loan_chart_range not in {"week", "month", "year"}:
        loan_chart_range = "year"

    # Beberapa ensure_* memakai koneksi sendiri, jadi jalankan sebelum query utama.
    ensure_agenda_schema()

    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True, buffered=True)
    try:
        ensure_loan_schema(cursor)
        ensure_activity_log_schema(cursor)
        ensure_streaming_evaluation_schema(cursor)
        conn.commit()

        schedule_total = dashboard_count_schedules_by_range(cursor, metric_range)
        schedule_chart = dashboard_schedule_series(cursor, schedule_chart_range)
        loan_chart = dashboard_loan_series(cursor, loan_chart_range)

        cursor.execute("SELECT COUNT(*) AS total FROM `loan_requests` WHERE `status` = 'pending'")
        loan_pending = int(fetch_scalar_value(cursor.fetchone(), 0) or 0)

        evaluation_pending = dashboard_pending_evaluation_count(cursor)

        cursor.execute(
            """
            SELECT COUNT(*) AS total
            FROM `agendas`
            WHERE LOWER(COALESCE(`status`, 'active')) = 'active'
              AND COALESCE(`end_date`, `start_date`) >= CURDATE()
            """
        )
        agenda_active = int(fetch_scalar_value(cursor.fetchone(), 0) or 0)

        scope_sql, scope_params = activity_scope_where("l")
        cursor.execute(
            f"""
            SELECT l.*
            FROM `activity_logs` l
            WHERE {scope_sql}
            ORDER BY l.`created_at` DESC, l.`id` DESC
            LIMIT 5
            """,
            tuple(scope_params),
        )
        logs = [activity_log_row_to_dict(row) for row in cursor.fetchall() or []]

        return jsonify({
            "success": True,
            "metrics": {
                "schedule": {"range": metric_range, "total": schedule_total},
                "loanPending": loan_pending,
                "evaluationPending": evaluation_pending,
                "agendaActive": agenda_active,
            },
            "scheduleChart": schedule_chart,
            "loanChart": loan_chart,
            "logs": logs,
            "currentUser": current_user_context(),
        })
    finally:
        cursor.close()
        conn.close()


@app.route("/api/notifications/<notif_id>/mark_read", methods=["POST"])
def mark_notification_read(notif_id):
    ensure_notifications_schema()
    viewer = current_user_context()
    client_key = str(request.args.get("clientKey") or request.headers.get("X-Registration-Client-Key") or "").strip()
    if viewer.get("logged_in"):
        user_key = f"member:{viewer.get('user_id') or viewer.get('username') or viewer.get('email') or 'member'}"
    else:
        user_key = client_key or None
    if not user_key:
        return jsonify({"success": False, "error": "Missing client key"}), 400

    conn = mysql_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("INSERT IGNORE INTO `notification_reads` (`notification_id`, `user_key`) VALUES (%s, %s)", (notif_id, user_key))
        conn.commit()
        return jsonify({"success": True})
    except Exception as e:
        conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 400
    finally:
        cursor.close()
        conn.close()

@app.route("/api/notifications/<notif_id>/toggle_read", methods=["POST"])
def toggle_notification_read(notif_id):
    ensure_notifications_schema()
    viewer = current_user_context()
    client_key = str(request.args.get("clientKey") or request.headers.get("X-Registration-Client-Key") or "").strip()
    if viewer.get("logged_in"):
        user_key = f"member:{viewer.get('user_id') or viewer.get('username') or viewer.get('email') or 'member'}"
    else:
        user_key = client_key or None
    if not user_key:
        return jsonify({"success": False, "error": "Missing client key"}), 400

    conn = mysql_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT 1 FROM `notification_reads` WHERE `notification_id` = %s AND `user_key` = %s LIMIT 1", (notif_id, user_key))
        if cursor.fetchone():
            cursor.execute("DELETE FROM `notification_reads` WHERE `notification_id` = %s AND `user_key` = %s", (notif_id, user_key))
            is_read = False
        else:
            cursor.execute("INSERT IGNORE INTO `notification_reads` (`notification_id`, `user_key`) VALUES (%s, %s)", (notif_id, user_key))
            is_read = True
        conn.commit()
        return jsonify({"success": True, "read": is_read})
    except Exception as e:
        conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 400
    finally:
        cursor.close()
        conn.close()

@app.route("/api/notifications/mark_all_read", methods=["POST"])
def mark_all_notifications_read():
    ensure_notifications_schema()
    viewer = current_user_context()
    client_key = str(request.args.get("clientKey") or request.headers.get("X-Registration-Client-Key") or "").strip()
    if viewer.get("logged_in"):
        user_key = f"member:{viewer.get('user_id') or viewer.get('username') or viewer.get('email') or 'member'}"
    else:
        user_key = client_key or None
    if not user_key:
        return jsonify({"success": False, "error": "Missing client key"}), 400

    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        cursor.execute("SELECT `id` FROM `notifications`")
        rows = cursor.fetchall() or []
        for r in rows:
            try:
                cursor.execute("INSERT IGNORE INTO `notification_reads` (`notification_id`, `user_key`) VALUES (%s, %s)", (r["id"], user_key))
            except Exception:
                pass
        conn.commit()
        return jsonify({"success": True})
    finally:
        cursor.close()
        conn.close()


def ensure_streaming_schema() -> None:
    conn = mysql_connection()
    # PERBAIKAN: Gunakan buffered=True agar tidak ada error unread result
    cursor = conn.cursor(buffered=True)
    try:
        cursor.execute("CREATE TABLE IF NOT EXISTS `streaming_weekly_config` (`id` int AUTO_INCREMENT PRIMARY KEY, `day_name` varchar(20), `start_time` time, `mass_name` varchar(255))")
        cursor.execute("CREATE TABLE IF NOT EXISTS `streaming_cancelled` (`id` int AUTO_INCREMENT PRIMARY KEY, `mass_date` date, `mass_time` time)")
        cursor.execute("CREATE TABLE IF NOT EXISTS `streaming_roles` (`id` int AUTO_INCREMENT PRIMARY KEY, `role_name` varchar(100) UNIQUE, `order_index` int DEFAULT 0)")
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS `streaming_assignments` (
              `id` int AUTO_INCREMENT PRIMARY KEY,
              `schedule_date` date NOT NULL,
              `schedule_time` time NOT NULL,
              `role_name` varchar(100) NOT NULL,
              `member_id` int NOT NULL,
              UNIQUE KEY `uniq_assignment` (`schedule_date`, `schedule_time`, `role_name`)
            )
        """)
        try:
            ensure_column(cursor, "streaming_assignments", "created_at", "`created_at` datetime NOT NULL DEFAULT CURRENT_TIMESTAMP")
            ensure_column(cursor, "streaming_assignments", "request_source", "`request_source` varchar(30) NOT NULL DEFAULT 'admin'")
            conn.commit()
        except Exception as exc:
            print(f"[WARN] Failed to ensure streaming assignment metadata: {exc}")
        
        cursor.execute("SELECT COUNT(*) FROM `streaming_roles`")
        res = cursor.fetchone()
        if res and res[0] == 0:
            cursor.execute("INSERT IGNORE INTO `streaming_roles` (role_name, order_index) VALUES ('Produser', 1), ('Operator Streaming', 2), ('Kameramen 1', 3), ('Kameramen 2', 4), ('Kameramen 3', 5)")
            conn.commit()
    except Exception as e:
        print(f"[WARN] Error in schema: {e}")
    finally:
        cursor.close()
        conn.close()

def format_time_hhmm(t_obj):
    """Mencegah bug timedelta slicing '7:30:' menjadi '07:30'"""
    if t_obj is None:
        return "00:00"
    if isinstance(t_obj, str):
        return t_obj[:5].zfill(5)
    if hasattr(t_obj, 'total_seconds'):
        hours, remainder = divmod(int(t_obj.total_seconds()), 3600)
        minutes, _ = divmod(remainder, 60)
        return f"{hours:02d}:{minutes:02d}"
    return str(t_obj)[:5].zfill(5)

DAYS_INDO = ["Senin", "Selasa", "Rabu", "Kamis", "Jumat", "Sabtu", "Minggu"]


def clear_streaming_assignments_for_published_misa_besar(cursor, misa_date, misa_time, status) -> int:
    """Hapus penugasan Misa Biasa bila Misa Besar published memakai tanggal+jam yang sama.

    Ini menjaga aturan operasional: dalam tanggal dan jam mulai yang sama hanya ada satu misa.
    Misa Besar yang masih draft tidak mengubah penugasan Misa Biasa.
    """
    if normalize_misa_besar_status(status) != "published":
        return 0

    if hasattr(misa_date, "isoformat"):
        date_value = misa_date.isoformat()
    else:
        date_value = normalize_text(misa_date)
    time_value = format_time_hhmm(misa_time)
    if not date_value or not time_value:
        return 0

    cursor.execute(
        """
        DELETE FROM `streaming_assignments`
        WHERE `schedule_date` = %s AND `schedule_time` = %s
        """,
        (date_value, time_value),
    )
    return cursor.rowcount or 0

def normalize_misa_besar_status(value: str | None) -> str:
    raw = normalize_text(value).lower()
    return "published" if raw in {"published", "publish", "publikasi"} else "draft"

def fetch_misa_besar_notification_snapshot(cursor, misa_id: int) -> dict[str, object] | None:
    """Ambil snapshot Misa Besar + petugas untuk kebutuhan notifikasi.

    Snapshot ini dipakai untuk dua jenis notifikasi:
    1. notifikasi petugas yang memang ditugaskan;
    2. notifikasi ke anggota aktif jika Misa Besar published masih memiliki role/slot kosong.
    """
    cursor.execute(
        """
        SELECT id, misa_name, DATE_FORMAT(misa_date, '%Y-%m-%d') AS misa_date,
               DATE_FORMAT(misa_time, '%H:%i') AS misa_time, misa_note,
               allow_member_request, status
        FROM misa_besar
        WHERE id = %s
        LIMIT 1
        """,
        (misa_id,),
    )
    event = cursor.fetchone()
    if not event:
        return None

    cursor.execute(
        """
        SELECT n.id AS role_id, n.role_name, n.required_count, a.member_id
        FROM misa_besar_names n
        LEFT JOIN misa_besar_assignments a ON a.role_id = n.id
        WHERE n.misa_id = %s
        ORDER BY n.id ASC, a.id ASC
        """,
        (misa_id,),
    )

    assignments: list[dict[str, str]] = []
    seen_assignments: set[tuple[str, str]] = set()
    roles_by_id: dict[str, dict[str, object]] = {}

    for row in cursor.fetchall() or []:
        if not isinstance(row, dict):
            continue

        role_id = normalize_text(row.get("role_id"))
        role_name = normalize_text(row.get("role_name")) or "Role"
        required_count = parse_required_int(row.get("required_count"), 1) or 1
        required_count = max(1, required_count)

        if role_id and role_id not in roles_by_id:
            roles_by_id[role_id] = {
                "roleId": role_id,
                "role": role_name,
                "requiredCount": required_count,
                "memberIds": [],
            }

        member_id = row.get("member_id")
        if member_id in (None, ""):
            continue

        member_id_text = str(member_id)
        if role_id and role_id in roles_by_id and member_id_text not in roles_by_id[role_id]["memberIds"]:
            roles_by_id[role_id]["memberIds"].append(member_id_text)

        key = (role_name, member_id_text)
        if key in seen_assignments:
            continue
        seen_assignments.add(key)
        assignments.append({"role": role_name, "memberId": member_id_text})

    roles: list[dict[str, object]] = []
    for role in roles_by_id.values():
        filled_count = len(role.get("memberIds") or [])
        required_count = parse_required_int(role.get("requiredCount"), 1) or 1
        role["filledCount"] = filled_count
        role["emptyCount"] = max(required_count - filled_count, 0)
        roles.append(role)

    return {
        "id": int(event.get("id") or misa_id),
        "misaName": normalize_text(event.get("misa_name")) or "Misa Besar",
        "misaDate": normalize_text(event.get("misa_date")),
        "misaTime": format_time_hhmm(event.get("misa_time")),
        "misaNote": normalize_text(event.get("misa_note")),
        "allowMemberRequest": bool(event.get("allow_member_request")),
        "status": normalize_misa_besar_status(event.get("status")),
        "assignments": assignments,
        "roles": roles,
    }

def misa_besar_assignment_keys(snapshot: dict[str, object] | None) -> set[tuple[str, str]]:
    if not snapshot:
        return set()
    keys: set[tuple[str, str]] = set()
    for item in snapshot.get("assignments") or []:
        if not isinstance(item, dict):
            continue
        role_name = normalize_text(item.get("role")) or "Role"
        member_id = normalize_text(item.get("memberId"))
        if member_id:
            keys.add((role_name, member_id))
    return keys

def create_misa_besar_assignment_notifications(
    cursor,
    snapshot: dict[str, object] | None,
    *,
    only_keys: set[tuple[str, str]] | None = None,
) -> int:
    """Kirim notif tugas untuk petugas Misa Besar yang statusnya sudah published."""
    if not snapshot or normalize_misa_besar_status(snapshot.get("status")) != "published":
        return 0

    try:
        date_obj = datetime.strptime(normalize_text(snapshot.get("misaDate")), "%Y-%m-%d")
        day_name = DAYS_INDO[date_obj.weekday()]
        date_formatted = date_obj.strftime("%d/%m/%Y")
    except Exception:
        day_name = "-"
        date_formatted = normalize_text(snapshot.get("misaDate")) or "-"

    misa_id = snapshot.get("id")
    misa_name = normalize_text(snapshot.get("misaName")) or "Misa Besar"
    misa_time = format_time_hhmm(snapshot.get("misaTime"))
    sent = 0
    sent_keys: set[tuple[str, str]] = set()

    for item in snapshot.get("assignments") or []:
        if not isinstance(item, dict):
            continue
        role_name = normalize_text(item.get("role")) or "Role"
        member_id = normalize_text(item.get("memberId"))
        key = (role_name, member_id)
        if not member_id or key in sent_keys:
            continue
        if only_keys is not None and key not in only_keys:
            continue

        title = f"Tugas Baru: {role_name}"
        body = (
            f"Anda ditugaskan sebagai <b>{html.escape(role_name)}</b> untuk "
            f"<b>{html.escape(misa_name)}</b> pada hari {html.escape(day_name)}, "
            f"{html.escape(date_formatted)} jam {html.escape(misa_time)} WIB."
        )
        create_notification(
            cursor,
            "tugas",
            title,
            body,
            "/jadwal-tugas-misa-anggota.html",
            {
                "target_user_id": member_id,
                "misa_besar_id": misa_id,
                "misa_type": "misa_besar",
                "role": role_name,
                "misa_name": misa_name,
                "misa_date": snapshot.get("misaDate"),
                "misa_time": misa_time,
            },
            target_role=None,
        )
        sent_keys.add(key)
        sent += 1
    return sent

def notify_misa_besar_assignment_changes(cursor, old_snapshot: dict[str, object] | None, new_snapshot: dict[str, object] | None) -> int:
    """
    Aturan notifikasi Misa Besar:
    - draft tidak mengirim notifikasi;
    - saat berubah ke published, semua petugas terisi dapat notifikasi;
    - saat sudah published, hanya petugas baru/peran baru yang dapat notifikasi.
    """
    if not new_snapshot or normalize_misa_besar_status(new_snapshot.get("status")) != "published":
        return 0

    ensure_notifications_schema()

    old_status = normalize_misa_besar_status(old_snapshot.get("status")) if old_snapshot else "draft"
    if old_status != "published":
        return create_misa_besar_assignment_notifications(cursor, new_snapshot)

    new_keys = misa_besar_assignment_keys(new_snapshot)
    old_keys = misa_besar_assignment_keys(old_snapshot)
    changed_keys = new_keys - old_keys
    if not changed_keys:
        return 0
    return create_misa_besar_assignment_notifications(cursor, new_snapshot, only_keys=changed_keys)


def misa_besar_open_role_labels(snapshot: dict[str, object] | None) -> list[str]:
    """Kembalikan daftar role/slot Misa Besar yang masih kosong."""
    if not snapshot:
        return []

    labels: list[str] = []
    for role in snapshot.get("roles") or []:
        if not isinstance(role, dict):
            continue
        role_name = normalize_text(role.get("role")) or "Role"
        empty_count = parse_required_int(role.get("emptyCount"), 0)
        if empty_count <= 0:
            continue
        if empty_count > 1:
            labels.append(f"{role_name} ({empty_count} slot)")
        else:
            labels.append(role_name)
    return labels

def create_misa_besar_open_role_notifications(cursor, snapshot: dict[str, object] | None) -> int:
    """Kirim notifikasi kategori tugas ke anggota aktif biasa bila published Misa Besar masih punya slot kosong.

    Notifikasi dibuat per anggota agar dashboard anggota bisa tetap memfilter menggunakan target_user_id,
    sama seperti notifikasi tugas lain.
    """
    if not snapshot or normalize_misa_besar_status(snapshot.get("status")) != "published":
        return 0

    empty_roles = misa_besar_open_role_labels(snapshot)
    if not empty_roles:
        return 0

    ensure_notifications_schema()

    try:
        date_obj = datetime.strptime(normalize_text(snapshot.get("misaDate")), "%Y-%m-%d")
        day_name = DAYS_INDO[date_obj.weekday()]
        date_formatted = date_obj.strftime("%d/%m/%Y")
    except Exception:
        day_name = "-"
        date_formatted = normalize_text(snapshot.get("misaDate")) or "-"

    misa_id = snapshot.get("id")
    misa_name = normalize_text(snapshot.get("misaName")) or "Misa Besar"
    misa_time = format_time_hhmm(snapshot.get("misaTime"))
    request_label = "dibuka" if snapshot.get("allowMemberRequest") else "ditutup"
    empty_role_text = ", ".join(empty_roles)

    cursor.execute(
        """
        SELECT id, nama
        FROM anggota
        WHERE LOWER(COALESCE(role, '')) = 'user'
          AND LOWER(COALESCE(status_akun, '')) = 'aktif'
        ORDER BY nama ASC, id ASC
        """
    )
    members = cursor.fetchall() or []

    sent = 0
    for member in members:
        if not isinstance(member, dict):
            continue
        member_id = normalize_text(member.get("id"))
        if not member_id:
            continue

        title = "Misa Besar Butuh Petugas"
        body = (
            f"Ada Misa Besar baru <b>{html.escape(misa_name)}</b> pada hari {html.escape(day_name)}, "
            f"{html.escape(date_formatted)} jam {html.escape(misa_time)} WIB yang masih memiliki slot kosong: "
            f"<b>{html.escape(empty_role_text)}</b>. "
            f"Status request anggota: <b>{html.escape(request_label)}</b>. "
            "Silakan buka halaman request tugas untuk melihat detail dan mengajukan tugas."
        )

        create_notification(
            cursor,
            "tugas",
            title,
            body,
            f"/request-tugas-anggota.html?misaBesarId={misa_id}",
            {
                "target_user_id": member_id,
                "notification_kind": "misa_besar_open_roles",
                "misa_besar_id": misa_id,
                "misa_type": "misa_besar",
                "misa_name": misa_name,
                "misa_date": snapshot.get("misaDate"),
                "misa_time": misa_time,
                "empty_roles": empty_roles,
                "allow_member_request": bool(snapshot.get("allowMemberRequest")),
            },
            target_role="user",
        )
        sent += 1

    return sent

def notify_misa_besar_open_roles_if_newly_published(cursor, old_snapshot: dict[str, object] | None, new_snapshot: dict[str, object] | None) -> int:
    """Kirim notifikasi slot kosong hanya saat Misa Besar baru masuk status published."""
    if not new_snapshot or normalize_misa_besar_status(new_snapshot.get("status")) != "published":
        return 0

    old_status = normalize_misa_besar_status(old_snapshot.get("status")) if old_snapshot else "draft"
    if old_status == "published":
        return 0

    return create_misa_besar_open_role_notifications(cursor, new_snapshot)

def ensure_auth_schema() -> None:
    conn = mysql_connection()
    cursor = conn.cursor()

    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS `anggota` (
          `id` int(11) NOT NULL,
          `nama` varchar(255) DEFAULT NULL,
          `username` varchar(150) DEFAULT NULL,
          `telp` varchar(50) DEFAULT NULL,
          `password` varchar(255) DEFAULT NULL,
          `role` varchar(50) DEFAULT NULL,
          `tgl_lahir` varchar(50) DEFAULT NULL,
          `email` varchar(255) DEFAULT NULL,
          `alamat` text DEFAULT NULL,
          `status_akun` varchar(20) NOT NULL DEFAULT 'aktif',
          `created_at` datetime NOT NULL DEFAULT CURRENT_TIMESTAMP,
          `updated_at` datetime NOT NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
          PRIMARY KEY (`id`),
          UNIQUE KEY `uniq_anggota_username` (`username`),
          UNIQUE KEY `uniq_anggota_email` (`email`),
          UNIQUE KEY `uniq_anggota_telp` (`telp`)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_general_ci
        """
    )

    ensure_column(cursor, "anggota", "alamat", "`alamat` text DEFAULT NULL")
    ensure_column(cursor, "anggota", "status_akun", "`status_akun` varchar(20) NOT NULL DEFAULT 'aktif'")
    ensure_column(cursor, "anggota", "inactive_until", "`inactive_until` date DEFAULT NULL")
    ensure_column(cursor, "anggota", "inactive_from", "`inactive_from` date DEFAULT NULL")
    ensure_column(cursor, "anggota", "inactive_type", "`inactive_type` varchar(30) DEFAULT NULL")
    ensure_column(cursor, "anggota", "inactive_reason", "`inactive_reason` varchar(255) DEFAULT NULL")
    ensure_column(cursor, "anggota", "inactive_note", "`inactive_note` text DEFAULT NULL")
    ensure_membership_request_schema(cursor)
    ensure_admin_permissions_schema(cursor)
    ensure_column(cursor, "anggota", "created_at", "`created_at` datetime NOT NULL DEFAULT CURRENT_TIMESTAMP")
    ensure_column(cursor, "anggota", "updated_at", "`updated_at` datetime NOT NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP")

    cursor.execute("UPDATE `anggota` SET `status_akun` = 'aktif' WHERE `status_akun` IS NULL OR `status_akun` = ''")
    cursor.execute(
        """
        UPDATE `anggota`
        SET `created_at` = COALESCE(`created_at`, CURRENT_TIMESTAMP),
            `updated_at` = COALESCE(`updated_at`, CURRENT_TIMESTAMP)
        WHERE `created_at` IS NULL OR `updated_at` IS NULL
        """
    )

    cursor.execute(
        """
        SELECT COUNT(*)
        FROM information_schema.TABLES
        WHERE TABLE_SCHEMA = %s AND TABLE_NAME = 'db_anggota'
        """,
        (MYSQL_CONFIG["database"],),
    )
    legacy_exists = cursor.fetchone()[0] > 0
    if legacy_exists:
        legacy_cursor = conn.cursor(dictionary=True)
        legacy_cursor.execute(
            "SELECT id, nama, username, telp, password, role, tgl_lahir, email, alamat, status_akun FROM db_anggota"
        )
        for legacy_row in legacy_cursor.fetchall():
            cursor.execute(
                """
                INSERT IGNORE INTO anggota
                (id, nama, username, telp, password, role, tgl_lahir, email, alamat, status_akun)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """,
                (
                    legacy_row.get("id"),
                    legacy_row.get("nama"),
                    legacy_row.get("username"),
                    legacy_row.get("telp"),
                    legacy_row.get("password"),
                    legacy_row.get("role"),
                    legacy_row.get("tgl_lahir"),
                    legacy_row.get("email"),
                    legacy_row.get("alamat"),
                    legacy_row.get("status_akun") or "aktif",
                ),
            )
        legacy_cursor.close()
        cursor.execute("DROP TABLE `db_anggota`")

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS `carousel_slides` (
          `id` varchar(100) NOT NULL,
          `title` varchar(255) DEFAULT NULL,
          `slug` varchar(255) DEFAULT NULL,
          `description` text DEFAULT NULL,
          `button_text` varchar(100) DEFAULT NULL,
          `button_link` varchar(255) DEFAULT NULL,
          `background_image` text DEFAULT NULL,
          `order_index` int(11) DEFAULT 0,
          `is_visible` tinyint(1) DEFAULT 1,
          PRIMARY KEY (`id`)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_general_ci
    ''')

    cursor.execute(
        '''
        CREATE TABLE IF NOT EXISTS `tentang_crembo_config` (
          `id` int(11) NOT NULL AUTO_INCREMENT,
          `description` text DEFAULT NULL,
          `button_text` varchar(255) DEFAULT NULL,
          `button_link` varchar(255) DEFAULT NULL,
          `auto_seconds` int(11) DEFAULT 5,
          PRIMARY KEY (`id`)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_general_ci
        '''
    )
    cursor.execute("SELECT COUNT(*) FROM `tentang_crembo_config`")
    if cursor.fetchone()[0] == 0:
        cursor.execute("INSERT INTO `tentang_crembo_config` (`id`, `description`, `button_text`, `button_link`, `auto_seconds`) VALUES (1, 'Ringkasan profil organisasi, visi pelayanan multimedia, serta peran Crembo dalam mendukung kegiatan liturgi dan agenda komunitas.', 'Pelajari Lebih Lanjut', 'profil.html', 5)")

    cursor.execute(
        '''
        CREATE TABLE IF NOT EXISTS `tentang_crembo_media` (
          `id` varchar(100) NOT NULL,
          `type` varchar(50) DEFAULT 'image',
          `url` text DEFAULT NULL,
          `order_index` int(11) DEFAULT 0,
          `is_visible` tinyint(1) DEFAULT 1,
          PRIMARY KEY (`id`)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_general_ci
        '''
    )

    cursor.execute(
        '''
        CREATE TABLE IF NOT EXISTS `instagram_posts` (
          `id_instagram` varchar(100) NOT NULL,
          `judul_instagram` varchar(200) NOT NULL,
          `url_instagram` varchar(255) NOT NULL,
          `urutan` int(11) NOT NULL DEFAULT 0,
          `tgl_instagram` datetime DEFAULT CURRENT_TIMESTAMP,
          `ip` varchar(25) DEFAULT NULL,
          `status` tinyint(1) NOT NULL DEFAULT 1,
          PRIMARY KEY (`id_instagram`)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_general_ci
        '''
    )

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS `youtube_embeds` (
          `id` varchar(100) NOT NULL,
          `url` text NOT NULL,
          `embed_type` varchar(30) NOT NULL DEFAULT 'video',
          `title` varchar(255) DEFAULT NULL,
          `order_index` int(11) DEFAULT 0,
          `is_visible` tinyint(1) DEFAULT 1,
          PRIMARY KEY (`id`)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_general_ci
    ''')
    ensure_column(cursor, "youtube_embeds", "embed_type", "`embed_type` varchar(30) NOT NULL DEFAULT 'video'")
    ensure_column(cursor, "youtube_embeds", "title", "`title` varchar(255) DEFAULT NULL")
    try:
        cursor.execute("ALTER TABLE `youtube_embeds` MODIFY COLUMN `url` text NOT NULL")
    except Exception:
        pass

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS `google_maps_embed` (
          `id` int(11) NOT NULL,
          `url` text DEFAULT NULL,
          PRIMARY KEY (`id`)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_general_ci
    ''')
    cursor.execute("SELECT COUNT(*) FROM `google_maps_embed`")
    if cursor.fetchone()[0] == 0:
        cursor.execute("INSERT INTO `google_maps_embed` (`id`, `url`) VALUES (1, '')")

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS `sertifikat_config` (
          `id` int(11) NOT NULL AUTO_INCREMENT,
          `ketua_name` varchar(255) DEFAULT 'Ketua Crembo Media',
          `pembina_name` varchar(255) DEFAULT 'Pembina Crembo Media',
          `ketua_sign_url` text DEFAULT NULL,
          `pembina_sign_url` text DEFAULT NULL,
          `updated_at` datetime NOT NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
          PRIMARY KEY (`id`)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_general_ci
    ''')
    cursor.execute("SELECT COUNT(*) FROM `sertifikat_config`")
    if cursor.fetchone()[0] == 0:
        cursor.execute("INSERT INTO `sertifikat_config` (`id`, `ketua_name`, `pembina_name`, `ketua_sign_url`, `pembina_sign_url`) VALUES (1, 'Ketua Crembo Media', 'Pembina Crembo Media', '', '')")

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS `organization_profiles` (
          `id` varchar(100) NOT NULL,
          `title` varchar(255) NOT NULL,
          `description` text DEFAULT NULL,
          `attachment_url` text DEFAULT NULL,
          `order_index` int(11) DEFAULT 0,
          `is_visible` tinyint(1) DEFAULT 1,
          `created_at` datetime NOT NULL DEFAULT CURRENT_TIMESTAMP,
          `updated_at` datetime NOT NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
          PRIMARY KEY (`id`)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_general_ci
    ''')

    ensure_inventory_schema(cursor)

    conn.commit()
    cursor.close()
    conn.close()

def fetch_member(identifier: str):
    ensure_auth_schema()
    needle = (identifier or "").strip()
    needle_lower = needle.lower()
    needle_phone = normalize_phone(needle)

    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute(
        """
        SELECT id, nama, username, telp, password, role, tgl_lahir, email, alamat, status_akun, inactive_until, inactive_from, inactive_type, inactive_reason, inactive_note
        , created_at, updated_at
        FROM anggota
        ORDER BY id ASC
        """
    )
    rows = cursor.fetchall()
    cursor.close()
    conn.close()

    for row in rows:
        if needle_lower and (
            needle_lower == (row.get("username") or "").strip().lower()
            or needle_lower == (row.get("email") or "").strip().lower()
            or needle_phone and needle_phone == normalize_phone(row.get("telp") or "")
        ):
            return row
    return None



# ===== Membership inactive request helpers =====
MEMBERSHIP_PENDING = "pending"
MEMBERSHIP_APPROVED = "approved"
MEMBERSHIP_REJECTED = "rejected"
MEMBERSHIP_CANCELLED = "cancelled"

MEMBERSHIP_STATUS_LABELS = {
    MEMBERSHIP_PENDING: "Menunggu Review Admin",
    MEMBERSHIP_APPROVED: "Disetujui",
    MEMBERSHIP_REJECTED: "Ditolak",
    MEMBERSHIP_CANCELLED: "Dibatalkan",
}

def membership_status_label(value: str | None) -> str:
    return MEMBERSHIP_STATUS_LABELS.get(normalize_text(value).lower(), normalize_text(value) or "Menunggu Review Admin")

def membership_status_key(value: str | None) -> str:
    raw = normalize_text(value).lower()
    if raw in {"pending", "menunggu", "menunggu review admin"}:
        return MEMBERSHIP_PENDING
    if raw in {"approved", "approve", "disetujui", "acc"}:
        return MEMBERSHIP_APPROVED
    if raw in {"rejected", "reject", "ditolak"}:
        return MEMBERSHIP_REJECTED
    if raw in {"cancelled", "canceled", "dibatalkan", "cancel"}:
        return MEMBERSHIP_CANCELLED
    return raw or MEMBERSHIP_PENDING

def ensure_membership_request_schema(cursor) -> None:
    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS `membership_status_requests` (
          `id` varchar(100) NOT NULL,
          `member_id` int(11) NOT NULL,
          `member_name` varchar(255) DEFAULT NULL,
          `inactive_type` varchar(30) NOT NULL DEFAULT 'permanent',
          `reason` varchar(255) NOT NULL,
          `start_date` date NOT NULL,
          `return_date` date DEFAULT NULL,
          `note` text NOT NULL,
          `evidence_json` longtext DEFAULT NULL,
          `status` varchar(30) NOT NULL DEFAULT 'pending',
          `admin_id` int(11) DEFAULT NULL,
          `admin_name` varchar(255) DEFAULT NULL,
          `admin_note` text DEFAULT NULL,
          `decided_at` datetime DEFAULT NULL,
          `applied_at` datetime DEFAULT NULL,
          `manual_reactivated_at` datetime DEFAULT NULL,
          `created_at` datetime NOT NULL DEFAULT CURRENT_TIMESTAMP,
          `updated_at` datetime NOT NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
          PRIMARY KEY (`id`),
          KEY `idx_membership_member` (`member_id`),
          KEY `idx_membership_status` (`status`),
          KEY `idx_membership_start` (`start_date`)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_general_ci
        """
    )
    ensure_column(cursor, "membership_status_requests", "applied_at", "`applied_at` datetime DEFAULT NULL")
    ensure_column(cursor, "membership_status_requests", "manual_reactivated_at", "`manual_reactivated_at` datetime DEFAULT NULL")

def membership_evidence_payload(value) -> dict[str, object]:
    attachments = normalize_attachment_payload(value)
    return attachments[0] if attachments else {}

def membership_request_row_to_dict(row: dict[str, object]) -> dict[str, object]:
    if not row:
        return {}
    evidence = membership_evidence_payload(row.get("evidence_json"))
    status_key = membership_status_key(row.get("status"))
    start_date = row.get("start_date")
    return_date = row.get("return_date")
    created_at = row.get("created_at")
    updated_at = row.get("updated_at")
    decided_at = row.get("decided_at")
    applied_at = row.get("applied_at")
    manual_reactivated_at = row.get("manual_reactivated_at")
    return {
        "id": row.get("id"),
        "memberId": row.get("member_id"),
        "memberName": row.get("member_name") or "Anggota",
        "requestType": "Nonaktif",
        "inactiveType": "temporary" if normalize_text(row.get("inactive_type")).lower() == "temporary" else "permanent",
        "reason": row.get("reason") or "",
        "effectiveDate": start_date.isoformat() if hasattr(start_date, "isoformat") else normalize_text(start_date),
        "reactivateDate": return_date.isoformat() if hasattr(return_date, "isoformat") else normalize_text(return_date),
        "note": row.get("note") or "",
        "evidence": evidence,
        "evidenceName": evidence.get("name") or "",
        "evidenceType": evidence.get("mimeType") or "",
        "evidenceUrl": evidence.get("url") or "",
        "evidencePreviewable": bool(evidence.get("previewable")),
        "statusKey": status_key,
        "status": membership_status_label(status_key),
        "adminName": row.get("admin_name") or "",
        "adminNote": row.get("admin_note") or "",
        "submittedDate": created_at.date().isoformat() if hasattr(created_at, "date") else normalize_text(created_at)[:10],
        "createdAt": created_at.isoformat() if hasattr(created_at, "isoformat") else normalize_text(created_at),
        "updatedAt": updated_at.isoformat() if hasattr(updated_at, "isoformat") else normalize_text(updated_at),
        "decidedAt": decided_at.isoformat() if hasattr(decided_at, "isoformat") else normalize_text(decided_at),
        "appliedAt": applied_at.isoformat() if hasattr(applied_at, "isoformat") else normalize_text(applied_at),
        "manualReactivatedAt": manual_reactivated_at.isoformat() if hasattr(manual_reactivated_at, "isoformat") else normalize_text(manual_reactivated_at),
    }

def ensure_membership_columns(cursor) -> None:
    ensure_column(cursor, "anggota", "inactive_until", "`inactive_until` date DEFAULT NULL")
    ensure_column(cursor, "anggota", "inactive_from", "`inactive_from` date DEFAULT NULL")
    ensure_column(cursor, "anggota", "inactive_type", "`inactive_type` varchar(30) DEFAULT NULL")
    ensure_column(cursor, "anggota", "inactive_reason", "`inactive_reason` varchar(255) DEFAULT NULL")
    ensure_column(cursor, "anggota", "inactive_note", "`inactive_note` text DEFAULT NULL")
    ensure_membership_request_schema(cursor)

def apply_membership_status_transitions(cursor) -> None:
    """Aktifkan/nonaktifkan otomatis berdasarkan request yang sudah disetujui."""
    today = datetime.now().date()
    ensure_membership_columns(cursor)

    # Mulai masa nonaktif yang sudah sampai tanggal efektif.
    cursor.execute(
        """
        SELECT r.*, a.status_akun
        FROM `membership_status_requests` r
        JOIN `anggota` a ON a.id = r.member_id
        WHERE r.status = 'approved'
          AND r.start_date <= %s
          AND r.applied_at IS NULL
          AND r.manual_reactivated_at IS NULL
          AND COALESCE(a.status_akun, 'aktif') <> 'nonaktif'
          AND (r.return_date IS NULL OR r.return_date > %s)
        """,
        (today, today),
    )
    for row in cursor.fetchall() or []:
        cursor.execute(
            """
            UPDATE `anggota`
            SET status_akun = 'nonaktif', inactive_from = %s, inactive_until = %s,
                inactive_type = %s, inactive_reason = %s, inactive_note = %s
            WHERE id = %s
            """,
            (
                row.get("start_date"),
                row.get("return_date"),
                row.get("inactive_type"),
                row.get("reason"),
                row.get("note"),
                row.get("member_id"),
            ),
        )
        cursor.execute(
            """
            UPDATE `membership_status_requests`
            SET applied_at = COALESCE(applied_at, CURRENT_TIMESTAMP), updated_at = CURRENT_TIMESTAMP
            WHERE id = %s
            """,
            (row.get("id"),),
        )
        create_notification_once(
            cursor,
            "keanggotaan",
            "Status Keanggotaan Nonaktif",
            "Pengajuan nonaktif Anda sudah disetujui dan status akun menjadi nonaktif.",
            "/profil-anggota.html",
            {"target_user_id": row.get("member_id"), "request_id": row.get("id")},
            target_role="user",
            dedupe_key=f"membership-approved-active-{row.get('id')}",
        )

    # Aktif kembali otomatis untuk nonaktif berjangka.
    cursor.execute(
        """
        SELECT r.*
        FROM `membership_status_requests` r
        JOIN `anggota` a ON a.id = r.member_id
        WHERE r.status = 'approved'
          AND r.inactive_type = 'temporary'
          AND r.return_date IS NOT NULL
          AND r.return_date <= %s
          AND r.manual_reactivated_at IS NULL
          AND COALESCE(a.status_akun, 'aktif') = 'nonaktif'
        """,
        (today,),
    )
    for row in cursor.fetchall() or []:
        cursor.execute(
            """
            UPDATE `anggota`
            SET status_akun = 'aktif', inactive_until = NULL, inactive_from = NULL,
                inactive_type = NULL, inactive_reason = NULL, inactive_note = NULL
            WHERE id = %s
            """,
            (row.get("member_id"),),
        )
        create_notification_once(
            cursor,
            "keanggotaan",
            "Status Keanggotaan Aktif Kembali",
            "Masa nonaktif berjangka Anda sudah selesai. Status akun Anda aktif kembali.",
            "/profil-anggota.html",
            {"target_user_id": row.get("member_id"), "request_id": row.get("id")},
            target_role="user",
            dedupe_key=f"membership-auto-reactivate-{row.get('id')}",
        )


def _effective_membership_request_for_member(cursor, member_id):
    """Ambil pengajuan nonaktif approved yang masih aktif untuk 1 anggota.

    Helper ini dipakai untuk menjaga agar status anggota di Manajemen Anggota
    tidak tertimpa kembali menjadi aktif setelah request nonaktif disetujui.
    """
    try:
        member_id_int = int(member_id)
    except (TypeError, ValueError):
        return None

    today = datetime.now().date()
    ensure_membership_columns(cursor)
    cursor.execute(
        """
        SELECT *
        FROM membership_status_requests
        WHERE member_id = %s
          AND status = 'approved'
          AND start_date <= %s
          AND (return_date IS NULL OR return_date > %s)
          AND manual_reactivated_at IS NULL
        ORDER BY start_date DESC, decided_at DESC, updated_at DESC, created_at DESC
        LIMIT 1
        """,
        (member_id_int, today, today),
    )
    return cursor.fetchone()


def force_apply_effective_membership_status(cursor, member_id):
    """Paksa sinkron status_akun anggota dari request nonaktif yang approved.

    Kasus bug sebelumnya:
    - Approve dari Dashboard Admin aman.
    - Approve dari Manajemen Anggota bisa terlihat tetap Aktif karena data anggota
      yang dirender/sinkron masih stale.
    Fungsi ini memastikan sumber kebenaran tetap tabel membership_status_requests.
    """
    req = _effective_membership_request_for_member(cursor, member_id)
    if not req:
        return None

    cursor.execute(
        """
        UPDATE anggota
        SET status_akun = 'nonaktif',
            inactive_from = %s,
            inactive_until = %s,
            inactive_type = %s,
            inactive_reason = %s,
            inactive_note = %s,
            updated_at = CURRENT_TIMESTAMP
        WHERE id = %s
        """,
        (
            req.get("start_date"),
            req.get("return_date"),
            req.get("inactive_type") or "permanent",
            req.get("reason") or "",
            req.get("note") or "",
            req.get("member_id"),
        ),
    )
    cursor.execute(
        """
        UPDATE membership_status_requests
        SET applied_at = COALESCE(applied_at, CURRENT_TIMESTAMP),
            updated_at = CURRENT_TIMESTAMP
        WHERE id = %s
        """,
        (req.get("id"),),
    )
    return req


def repair_effective_membership_statuses_for_admin(cursor) -> None:
    """Sinkron ulang status anggota untuk halaman admin/manajemen anggota."""
    today = datetime.now().date()
    ensure_membership_columns(cursor)

    # Jalankan transisi otomatis lebih dulu untuk request yang belum applied
    # dan request temporary yang sudah melewati tanggal aktif kembali.
    apply_membership_status_transitions(cursor)

    # Perbaiki juga data lama yang sudah approved/applied tetapi status anggota
    # sempat tertimpa aktif oleh proses sync halaman Manajemen Anggota.
    cursor.execute(
        """
        SELECT DISTINCT member_id
        FROM membership_status_requests
        WHERE status = 'approved'
          AND start_date <= %s
          AND (return_date IS NULL OR return_date > %s)
          AND manual_reactivated_at IS NULL
        """,
        (today, today),
    )
    member_ids = [row.get("member_id") if isinstance(row, dict) else row[0] for row in (cursor.fetchall() or [])]
    for member_id in member_ids:
        force_apply_effective_membership_status(cursor, member_id)



def current_member_is_inactive() -> bool:
    return session.get("logged_in") and normalize_role_value(session.get("role") or "user") == "user" and normalize_status(session.get("status_akun") or "aktif") == "nonaktif"

def is_inactive_member_page_allowed(candidate: str) -> bool:
    allowed = {
        "dashboard-anggota.html",
        "profil-anggota.html",
        "profil.html",
        "unduh-sertifikat-anggota.html",
        "log-aktivitas-saya.html",
        "log-aktivitas.html",
        "home.html",
        "login.html",
    }
    return candidate in allowed

def can_manage_members() -> bool:
    if not session.get("logged_in"):
        return False
    role = normalize_role_value(session.get("role") or "")
    if role == "super_admin":
        return True
    if role == "admin":
        return admin_has_module_access("keanggotaan")
    return False


def can_manage_registration_forms() -> bool:
    return bool(session.get("logged_in")) and (session.get("role") or "") in {"admin", "super_admin"}


def registration_form_field_to_dict(item, index):
    return {
        "id": str(item and item.get("id") or f"field-{index + 1}"),
        "label": str(item and item.get("label") or f"Pertanyaan {index + 1}"),
        "type": str(item and item.get("type") or "text"),
        "required": bool(item and item.get("required")),
        "placeholder": str(item and item.get("placeholder") or ""),
        "options": [str(value) for value in (item and item.get("options") or []) if str(value).strip()],
    }


def normalize_registration_fields(value) -> list[dict[str, object]]:
    if value in (None, "", []):
        return []

    raw_items = value
    if isinstance(value, str):
        raw_text = value.strip()
        if not raw_text:
            return []
        try:
            parsed = json.loads(raw_text)
            raw_items = parsed if isinstance(parsed, list) else [parsed]
        except (TypeError, ValueError):
            raw_items = [raw_text]

    if isinstance(raw_items, dict):
        raw_items = [raw_items]

    normalized_items: list[dict[str, object]] = []
    for index, item in enumerate(raw_items or []):
        if not isinstance(item, dict):
            continue
        normalized_items.append(registration_form_field_to_dict(item, index))
    return normalized_items


def registration_form_row_to_dict(row: dict[str, object], submission_count: int = 0) -> dict[str, object]:
    fields_raw = row.get("fields_json") or row.get("fields") or "[]"
    try:
        fields = json.loads(fields_raw) if isinstance(fields_raw, str) else (fields_raw or [])
    except (TypeError, ValueError):
        fields = []

    # BACA DATA BARU imageUrl dan attachments
    attachments_raw = row.get("attachments") or "[]"
    try:
        attachments = json.loads(attachments_raw) if isinstance(attachments_raw, str) else (attachments_raw or [])
    except (TypeError, ValueError):
        attachments = []

    return {
        "id": row.get("id"),
        "title": row.get("title") or "Form Pendaftaran",
        "description": row.get("description") or "",
        "target": "internal" if (row.get("target") or "public") == "internal" else "public",
        "visibility": "draft" if (row.get("visibility") or "visible") == "draft" else "visible",
        "openDate": str(row.get("open_date") or ""),
        "closeDate": str(row.get("close_date") or ""),
        "quota": int(row.get("quota") or 0),
        "imageUrl": row.get("image_url") or "",
        "imageName": row.get("image_name") or "",
        "attachments": attachments if isinstance(attachments, list) else [],
        "fields": normalize_registration_fields(fields),
        "submissionCount": int(submission_count or 0),
        "createdAt": row.get("created_at") or "",
        "updatedAt": row.get("updated_at") or "",
    }


def registration_form_payload_to_db_values(data, existing=None):
    source = existing or {}
    fields = normalize_registration_fields(data.get("fields", source.get("fields_json", source.get("fields", []))))
    attachments = normalize_attachment_payload(data.get("attachments", source.get("attachments", [])))

    def value_for(*keys, default=""):
        for key in keys:
            current = data.get(key)
            if current not in (None, ""):
                return current
            current = source.get(key)
            if current not in (None, ""):
                return current
        return default

    quota_value = data.get("quota")
    if quota_value in (None, ""):
        quota_value = source.get("quota") or 0

    return {
        "title": str(value_for("title")).strip(),
        "description": str(value_for("description")).strip(),
        "target": "internal" if str(value_for("target", default="public")).strip().lower() == "internal" else "public",
        "visibility": "draft" if str(value_for("visibility", default="visible")).strip().lower() == "draft" else "visible",
        "open_date": str(value_for("openDate", "open_date")).strip() or None,
        "close_date": str(value_for("closeDate", "close_date")).strip() or None,
        "quota": int(quota_value or 0),
        "image_url": str(value_for("imageUrl", "image_url")).strip(),
        "image_name": str(value_for("imageName", "image_name")).strip(),
        "attachments": json.dumps(attachments, ensure_ascii=False),
        "fields_json": json.dumps(fields, ensure_ascii=False),
    }


def registration_form_status(form: dict[str, object], submission_count: int) -> dict[str, object]:
    now = datetime.now().date()
    if form.get("visibility") == "draft":
        return {"open": False, "reason": "draft", "text": "Draft"}

    open_date_value = str(form.get("openDate") or "").strip()
    close_date_value = str(form.get("closeDate") or "").strip()
    open_date = datetime.strptime(open_date_value, "%Y-%m-%d").date() if open_date_value else None
    close_date = datetime.strptime(close_date_value, "%Y-%m-%d").date() if close_date_value else None

    if open_date and now < open_date:
        return {"open": False, "reason": "not_open", "text": "Belum dibuka"}
    if close_date and now > close_date:
        return {"open": False, "reason": "closed", "text": "Sudah ditutup"}
    if int(form.get("quota") or 0) > 0 and int(submission_count or 0) >= int(form.get("quota") or 0):
        return {"open": False, "reason": "quota", "text": "Kuota penuh"}
    return {"open": True, "reason": "open", "text": "Aktif"}


def registration_form_submission_counts(cursor) -> dict[str, int]:
    cursor.execute("SELECT `form_id`, COUNT(*) AS `count` FROM `registration_form_submissions` GROUP BY `form_id`")
    counts: dict[str, int] = {}
    for row in cursor.fetchall() or []:
        counts[str(row.get("form_id") or "")] = int(row.get("count") or 0)
    return counts


def registration_submission_row_to_dict(row: dict[str, object]) -> dict[str, object]:
    answers_raw = row.get("answers_json") or "[]"
    try:
        answers = json.loads(answers_raw) if isinstance(answers_raw, str) else (answers_raw or [])
    except (TypeError, ValueError):
        answers = []
    return {
        "id": row.get("id"),
        "formId": row.get("form_id"),
        "submittedAt": row.get("submitted_at") or "",
        "submitter": {
            "key": row.get("submitter_key") or "",
            "identifier": row.get("submitter_identifier") or "",
            "role": row.get("submitter_role") or "public",
            "userId": row.get("submitter_user_id") or "",
            "source": row.get("submitter_source") or "public",
        },
        "answers": answers if isinstance(answers, list) else [],
    }


def registration_submitter_context(form: dict[str, object], payload_submitter) -> tuple[dict[str, object] | None, tuple[dict[str, object], int] | None]:
    viewer = current_user_context()
    submitter_data = payload_submitter if isinstance(payload_submitter, dict) else {}

    if form.get("target") == "internal":
        if not viewer.get("logged_in"):
            return None, ({"success": False, "error": "Form internal hanya bisa diisi setelah login."}, 401)
        submitter_key = f"member:{viewer.get('user_id') or viewer.get('username') or viewer.get('email') or 'internal'}"
        return {
            "key": submitter_key,
            "identifier": viewer.get("nama") or viewer.get("username") or viewer.get("email") or "Anggota",
            "role": viewer.get("role") or "user",
            "user_id": viewer.get("user_id") or "",
            "source": "internal",
        }, None

    if viewer.get("logged_in"):
        submitter_key = f"member:{viewer.get('user_id') or viewer.get('username') or viewer.get('email') or 'public'}"
        return {
            "key": submitter_key,
            "identifier": viewer.get("nama") or viewer.get("username") or viewer.get("email") or "Pengguna",
            "role": viewer.get("role") or "user",
            "user_id": viewer.get("user_id") or "",
            "source": "logged_in",
        }, None

    client_key = str(submitter_data.get("key") or submitter_data.get("clientKey") or request.headers.get("X-Registration-Client-Key") or "").strip()
    if not client_key:
        client_key = f"guest:{uuid.uuid4().hex}"
    return {
        "key": client_key,
        "identifier": str(submitter_data.get("identifier") or "Pengunjung").strip() or "Pengunjung",
        "role": str(submitter_data.get("role") or "public").strip() or "public",
        "user_id": "",
        "source": "public",
    }, None


def registration_submission_payload_to_rows(form: dict[str, object], answers_payload):
    fields = form.get("fields") if isinstance(form.get("fields"), list) else []
    answers_list = answers_payload if isinstance(answers_payload, list) else []
    answer_map = {}
    for item in answers_list:
        if not isinstance(item, dict):
            continue
        field_id = str(item.get("fieldId") or item.get("field_id") or "").strip()
        if not field_id:
            continue
        answer_map[field_id] = item.get("value")

    normalized_answers: list[dict[str, object]] = []
    for field in fields:
        if not isinstance(field, dict):
            continue
        field_id = str(field.get("id") or "").strip()
        field_label = str(field.get("label") or field_id or "Pertanyaan").strip()
        field_type = str(field.get("type") or "text").strip().lower()
        value = answer_map.get(field_id)

        # Upload Lampiran (Tambahan handling lampiran pada saat visitor isi form)
        if field_type == "file":
             # Value dari client adalah list url jika input type file (seperti attachment)
             if isinstance(value, list):
                 normalized_value = [str(item).strip() for item in value if str(item).strip()]
             elif value in (None, ""):
                 normalized_value = []
             else:
                 normalized_value = [str(value).strip()]
             if field.get("required") and not normalized_value:
                 return None, f'Mohon isi/upload: {field_label}'

        elif field_type == "checkbox":
            if isinstance(value, list):
                normalized_value = [str(item).strip() for item in value if str(item).strip()]
            elif value in (None, ""):
                normalized_value = []
            else:
                normalized_value = [str(value).strip()]
            if field.get("required") and not normalized_value:
                return None, f'Mohon isi: {field_label}'
        else:
            normalized_value = str(value or "").strip()
            if field.get("required") and not normalized_value:
                return None, f'Mohon isi: {field_label}'

        normalized_answers.append({
            "fieldId": field_id,
            "label": field_label,
            "type": field_type,
            "value": normalized_value,
        })

    return normalized_answers, None


def read_members_for_admin() -> list[dict[str, object]]:
    # Khusus halaman admin/manajemen anggota, sinkronkan dulu request nonaktif
    # yang sudah approved agar tabel anggota tidak menampilkan status stale.
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        repair_effective_membership_statuses_for_admin(cursor)
        conn.commit()
    except Exception as exc:
        conn.rollback()
        print(f"[WARN] Failed to repair membership status before admin read: {exc}")
    finally:
        cursor.close()
        conn.close()
    return read_member_rows()



def _session_actor_id() -> int | None:
    try:
        return int(session.get("user_id"))
    except (TypeError, ValueError):
        return None


def _member_action_allowed(actor_role: str, actor_id: int | None, target_role: str, target_id: int | None, action: str, requested_role: str | None = None) -> tuple[bool, str]:
    actor_role = normalize_role_value(actor_role)
    target_role = normalize_role_value(target_role)
    requested_role = normalize_role_value(requested_role or target_role)
    is_self = actor_id is not None and target_id is not None and int(actor_id) == int(target_id)

    if actor_role == "super_admin":
        if action == "delete" and is_self:
            return False, "Super Admin tidak dapat menghapus akunnya sendiri. Gunakan Super Admin lain jika akun ini perlu dihapus."
        return True, ""

    if actor_role == "admin":
        if action == "create":
            if requested_role == "super_admin":
                return False, "Admin biasa tidak dapat membuat akun Super Admin."
            return True, ""

        if action == "delete":
            if is_self:
                return False, "Admin tidak dapat menghapus akunnya sendiri."
            if target_role != "user":
                return False, "Admin biasa hanya dapat menghapus akun anggota biasa."
            return True, ""

        if action == "reset":
            if is_self:
                return True, ""
            if target_role != "user":
                return False, "Admin biasa tidak dapat reset password akun Admin atau Super Admin lain."
            return True, ""

        if action == "edit":
            if requested_role == "super_admin":
                return False, "Admin biasa tidak dapat menetapkan role Super Admin."
            if is_self:
                if requested_role != target_role:
                    return False, "Admin biasa tidak dapat mengubah role akunnya sendiri dari halaman ini."
                return True, ""
            if target_role != "user":
                return False, "Admin biasa tidak dapat mengedit akun Admin atau Super Admin lain."
            return True, ""

    return False, "Akses ditolak."


def _password_hash_from_payload(incoming_password: str, existing_hash: str | None, birth_date: str) -> str:
    incoming = normalize_text(incoming_password)
    current_hash = normalize_text(existing_hash)
    if current_hash and (not incoming or incoming == current_hash or incoming.startswith(("scrypt:", "pbkdf2:", "sha256$"))):
        return current_hash
    return hash_member_password(incoming, birth_date)


def sync_members_from_payload(payload: list[dict[str, object]]) -> None:
    ensure_notifications_schema()
    actor_role = normalize_role_value(session.get("role") or "")
    actor_id = _session_actor_id()
    if actor_role not in {"admin", "super_admin"}:
        raise PermissionError("Akses ditolak.")

    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        cursor.execute("START TRANSACTION")

        ensure_membership_columns(cursor)
        cursor.execute(
            """
            SELECT id, nama, username, telp, password, role, tgl_lahir, email, alamat, status_akun, inactive_until, inactive_from, inactive_type, inactive_reason, inactive_note, created_at, updated_at
            FROM anggota
            """
        )
        existing_rows = cursor.fetchall() or []
        existing_by_id: dict[int, dict[str, object]] = {}
        existing_ids: set[int] = set()
        for row in existing_rows:
            try:
                row_id = int(row.get("id"))
            except (TypeError, ValueError):
                continue
            existing_ids.add(row_id)
            existing_by_id[row_id] = row

        next_id = max(existing_ids) + 1 if existing_ids else 1
        kept_ids: set[int] = set()

        for index, item in enumerate(payload, start=1):
            if not isinstance(item, dict):
                continue
            birth_date = normalize_text(item.get("birthDate") or item.get("tanggalLahir"))
            status_value = normalize_status(item.get("status") or item.get("status_akun") or "aktif")
            inactive_until = parse_optional_date(item.get("inactiveUntil") or item.get("inactive_until"))
            inactive_from_value = parse_optional_date(item.get("inactiveFrom") or item.get("inactive_from"))
            inactive_type_value = normalize_text(item.get("inactiveType") or item.get("inactive_type"))
            inactive_reason_value = normalize_text(item.get("inactiveReason") or item.get("inactive_reason"))
            inactive_note_value = normalize_text(item.get("inactiveNote") or item.get("inactive_note"))
            if status_value == "aktif":
                inactive_until = None
                inactive_from_value = None
                inactive_type_value = ""
                inactive_reason_value = ""
                inactive_note_value = ""
            elif status_value == "nonaktif" and not inactive_from_value:
                inactive_from_value = datetime.now().date().isoformat()
            requested_role = normalize_role_value(item.get("role") or "user")
            explicit_status_change = bool(item.get("manualStatusChange") or item.get("_manualStatusChange"))

            raw_id = item.get("id")
            try:
                parsed_id = int(raw_id)
            except (TypeError, ValueError):
                parsed_id = None

            is_existing = parsed_id is not None and parsed_id in existing_by_id
            if is_existing:
                member_id = int(parsed_id)
                existing = existing_by_id[member_id]
                target_role = normalize_role_value(existing.get("role") or "user")
                allowed, message = _member_action_allowed(actor_role, actor_id, target_role, member_id, "edit", requested_role)
                if not allowed:
                    raise PermissionError(message)
                hashed_password = _password_hash_from_payload(
                    normalize_text(item.get("password")),
                    normalize_text(existing.get("password")),
                    birth_date,
                )
            else:
                member_id = parsed_id if parsed_id is not None and parsed_id > 0 else next_id
                while member_id in existing_ids or member_id in kept_ids:
                    member_id = next_id
                    next_id += 1
                allowed, message = _member_action_allowed(actor_role, actor_id, "user", member_id, "create", requested_role)
                if not allowed:
                    raise PermissionError(message)
                hashed_password = hash_member_password(normalize_text(item.get("password")), birth_date)

            # Jika ada pengajuan nonaktif approved yang sedang berlaku, jangan biarkan
            # payload stale dari halaman Manajemen Anggota menimpa status menjadi Aktif.
            # Status Aktif hanya boleh dipakai sebagai aktivasi manual bila perubahan
            # status memang berasal dari submit form edit anggota.
            if is_existing:
                effective_req = _effective_membership_request_for_member(cursor, member_id)
                if effective_req and status_value == "aktif" and not explicit_status_change:
                    status_value = "nonaktif"
                    inactive_from_value = _date_to_iso(effective_req.get("start_date")) or datetime.now().date().isoformat()
                    inactive_until = _date_to_iso(effective_req.get("return_date")) or None
                    inactive_type_value = normalize_text(effective_req.get("inactive_type")) or "permanent"
                    inactive_reason_value = normalize_text(effective_req.get("reason"))
                    inactive_note_value = normalize_text(effective_req.get("note"))

            kept_ids.add(member_id)

            cursor.execute(
                """
                INSERT INTO anggota
                (id, nama, username, telp, password, role, tgl_lahir, email, alamat, status_akun, inactive_until, inactive_from, inactive_type, inactive_reason, inactive_note, created_at, updated_at)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
                ON DUPLICATE KEY UPDATE
                  updated_at = IF(
                    NOT (
                      COALESCE(nama, '') = COALESCE(VALUES(nama), '') AND
                      COALESCE(username, '') = COALESCE(VALUES(username), '') AND
                      COALESCE(telp, '') = COALESCE(VALUES(telp), '') AND
                      COALESCE(password, '') = COALESCE(VALUES(password), '') AND
                      COALESCE(role, '') = COALESCE(VALUES(role), '') AND
                      COALESCE(tgl_lahir, '') = COALESCE(VALUES(tgl_lahir), '') AND
                      COALESCE(email, '') = COALESCE(VALUES(email), '') AND
                      COALESCE(alamat, '') = COALESCE(VALUES(alamat), '') AND
                      COALESCE(status_akun, '') = COALESCE(VALUES(status_akun), '') AND
                      COALESCE(inactive_until, '') = COALESCE(VALUES(inactive_until), '') AND
                      COALESCE(inactive_from, '') = COALESCE(VALUES(inactive_from), '') AND
                      COALESCE(inactive_type, '') = COALESCE(VALUES(inactive_type), '') AND
                      COALESCE(inactive_reason, '') = COALESCE(VALUES(inactive_reason), '') AND
                      COALESCE(inactive_note, '') = COALESCE(VALUES(inactive_note), '')
                    ),
                    CURRENT_TIMESTAMP,
                    updated_at
                  ),
                  nama = VALUES(nama),
                  username = VALUES(username),
                  telp = VALUES(telp),
                  password = VALUES(password),
                  role = VALUES(role),
                  tgl_lahir = VALUES(tgl_lahir),
                  email = VALUES(email),
                  alamat = VALUES(alamat),
                  status_akun = VALUES(status_akun),
                  inactive_until = VALUES(inactive_until),
                  inactive_from = VALUES(inactive_from),
                  inactive_type = VALUES(inactive_type),
                  inactive_reason = VALUES(inactive_reason),
                  inactive_note = VALUES(inactive_note)
                """,
                (
                    member_id,
                    normalize_text(item.get("name") or item.get("fullName") or "Anggota"),
                    normalize_text(item.get("username") or item.get("email") or item.get("phone") or f"anggota-{index}"),
                    normalize_text(item.get("phone")),
                    hashed_password,
                    requested_role,
                    birth_date,
                    normalize_text(item.get("email")),
                    normalize_text(item.get("address")),
                    status_value,
                    inactive_until,
                    inactive_from_value,
                    inactive_type_value,
                    inactive_reason_value,
                    inactive_note_value,
                ),
            )

            if requested_role == "admin":
                cursor.execute(
                    """
                    INSERT IGNORE INTO admin_module_permissions
                      (member_id, streaming, inventaris, publikasi, konten, keanggotaan)
                    VALUES (%s, 1, 1, 1, 1, 1)
                    """,
                    (member_id,),
                )
            else:
                cursor.execute("DELETE FROM admin_module_permissions WHERE member_id = %s", (member_id,))

            if is_existing:
                previous_status = normalize_status(existing.get("status_akun") or "aktif")
                if previous_status == "nonaktif" and status_value == "aktif" and explicit_status_change:
                    cursor.execute(
                        """
                        UPDATE membership_status_requests
                        SET manual_reactivated_at = COALESCE(manual_reactivated_at, CURRENT_TIMESTAMP),
                            updated_at = CURRENT_TIMESTAMP
                        WHERE member_id = %s
                          AND status = 'approved'
                          AND start_date <= CURDATE()
                          AND (return_date IS NULL OR return_date > CURDATE())
                          AND manual_reactivated_at IS NULL
                        """,
                        (member_id,),
                    )
                    create_notification_once(
                        cursor,
                        "keanggotaan",
                        "Status Keanggotaan Diaktifkan",
                        "Admin telah mengaktifkan kembali status keanggotaan Anda.",
                        "/profil-anggota.html",
                        {"target_user_id": member_id},
                        target_role="user",
                        dedupe_key=f"membership-manual-activate-{member_id}-{int(time.time())}",
                    )
                elif previous_status == "aktif" and status_value == "nonaktif":
                    create_notification_once(
                        cursor,
                        "keanggotaan",
                        "Status Keanggotaan Nonaktif",
                        "Admin telah mengubah status keanggotaan Anda menjadi nonaktif.",
                        "/profil-anggota.html",
                        {"target_user_id": member_id},
                        target_role="user",
                        dedupe_key=f"membership-manual-inactivate-{member_id}-{int(time.time())}",
                    )

        ids_to_delete = sorted(existing_ids - kept_ids)
        for delete_id in ids_to_delete:
            existing = existing_by_id.get(delete_id) or {}
            target_role = normalize_role_value(existing.get("role") or "user")
            allowed, message = _member_action_allowed(actor_role, actor_id, target_role, delete_id, "delete")
            if not allowed:
                raise PermissionError(message)

        if ids_to_delete:
            placeholders = ",".join(["%s"] * len(ids_to_delete))
            cursor.execute(f"DELETE FROM admin_module_permissions WHERE member_id IN ({placeholders})", tuple(ids_to_delete))
            cursor.execute(f"DELETE FROM anggota WHERE id IN ({placeholders})", tuple(ids_to_delete))

        conn.commit()
    except Exception as e:
        conn.rollback()
        raise e
    finally:
        cursor.close()
        conn.close()

def login_target_for_role(role: str) -> str:
    if role == "user":
        return url_for("dashboard_anggota")
    return url_for("dashboard")


def load_about_content_from_db() -> dict[str, object]:
    default_about = {
        "description": "Ringkasan profil organisasi, visi pelayanan multimedia, serta peran Crembo dalam mendukung kegiatan liturgi dan agenda komunitas. Konten ini nantinya diatur dari panel admin setelah login.",
        "buttonText": "Pelajari Lebih Lanjut",
        "buttonLink": "profil.html",
        "autoSeconds": 5,
        "images": [],
    }

    try:
        ensure_auth_schema()
        conn = mysql_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT description, button_text, button_link, auto_seconds FROM `tentang_crembo_config` WHERE `id`=1 LIMIT 1")
        cfg = cursor.fetchone() or {}
        cursor.execute("SELECT `id`, `type`, `url`, `order_index`, `is_visible` FROM `tentang_crembo_media` ORDER BY `order_index` ASC")
        media_rows = cursor.fetchall() or []
        cursor.close()
        conn.close()

        return {
            "description": cfg.get("description") or default_about["description"],
            "buttonText": cfg.get("button_text") or default_about["buttonText"],
            "buttonLink": cfg.get("button_link") or default_about["buttonLink"],
            "autoSeconds": int(cfg.get("auto_seconds") or default_about["autoSeconds"]),
            "images": [
                {
                    "id": row.get("id"),
                    "type": row.get("type") or "image",
                    "url": row.get("url") or "",
                    "order": row.get("order_index") or 0,
                    "active": bool(row.get("is_visible")),
                }
                for row in media_rows
            ],
        }
    except Exception:
        return default_about


def is_valid_instagram_url(url: str) -> bool:
    return bool(re.match(r"^https?://(www\.)?(instagram\.com|instagr\.am)/(p|reel|tv)/[A-Za-z0-9_\-\.]+/?", (url or "").strip(), re.IGNORECASE))


def normalize_instagram_url(url: str) -> str:
    raw = (url or "").strip()
    if raw and not re.match(r"^https?://", raw, re.IGNORECASE):
        raw = "https://" + raw
    return raw


def load_instagram_posts_from_db(limit: int | None = None, active_only: bool = True) -> list[dict[str, object]]:
    try:
        ensure_auth_schema()
        conn = mysql_connection()
        cursor = conn.cursor(dictionary=True)
        sql = "SELECT `id_instagram`, `judul_instagram`, `url_instagram`, `urutan`, `status`, `tgl_instagram` FROM `instagram_posts`"
        if active_only:
            sql += " WHERE `status` = 1"
        sql += " ORDER BY `urutan` ASC, `id_instagram` DESC"
        if limit is not None:
            sql += " LIMIT %s"
            cursor.execute(sql, (int(limit),))
        else:
            cursor.execute(sql)
        rows = cursor.fetchall() or []
        cursor.close()
        conn.close()
        return [
            {
                "id": row.get("id_instagram"),
                "title": row.get("judul_instagram") or "Instagram",
                "url": row.get("url_instagram") or "",
                "order": row.get("urutan") or 0,
                "active": bool(row.get("status")),
                "createdAt": row.get("tgl_instagram") or "",
            }
            for row in rows
        ]
    except Exception:
        return []


def save_instagram_posts_payload(payload: list[dict[str, object]]) -> int:
    ensure_auth_schema()
    conn = mysql_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("START TRANSACTION")
        cursor.execute("DELETE FROM `instagram_posts`")
        saved_count = 0
        for index, item in enumerate(payload, start=1):
            title = str(item.get("title") or item.get("judul") or "").strip()
            url = normalize_instagram_url(str(item.get("url") or item.get("url_instagram") or "").strip())
            if not title or not url or not is_valid_instagram_url(url):
                continue

            post_id = str(item.get("id") or item.get("id_instagram") or f"ig-{index}-{int(time.time())}")
            order_value = int(item.get("order") or item.get("urutan") or index)
            active_value = 1 if item.get("active", True) else 0

            cursor.execute(
                """
                INSERT INTO `instagram_posts`
                (`id_instagram`, `judul_instagram`, `url_instagram`, `urutan`, `tgl_instagram`, `ip`, `status`)
                VALUES (%s, %s, %s, %s, NOW(), %s, %s)
                """,
                (post_id, title, url, order_value, request.remote_addr or "127.0.0.1", active_value),
            )
            saved_count += 1

        conn.commit()
        return saved_count
    except Exception:
        conn.rollback()
        raise
    finally:
        cursor.close()
        conn.close()

def normalize_youtube_embed_type(value: str | None) -> str:
    kind = normalize_text(value).lower()
    return "playlist" if kind == "playlist" else "video"


def load_youtube_videos():
    try:
        conn = mysql_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT * FROM `youtube_embeds` WHERE `is_visible`=1 ORDER BY `order_index` ASC")
        rows = cursor.fetchall() or []
        conn.close()
        return [
            {
                "id": r["id"],
                "url": r["url"],
                "type": normalize_youtube_embed_type(r.get("embed_type")),
                "title": r.get("title") or "",
                "order": r["order_index"],
            }
            for r in rows
        ]
    except Exception:
        return []

def load_gmaps_url():
    try:
        conn = mysql_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT `url` FROM `google_maps_embed` WHERE `id`=1 LIMIT 1")
        row = cursor.fetchone()
        conn.close()
        return row["url"] if row and row["url"] else ""
    except Exception:
        return ""

def load_carousel_slides():
    try:
        conn = mysql_connection()
        cursor = conn.cursor(dictionary=True)
        # Ambil hanya yang aktif untuk di home
        cursor.execute("SELECT * FROM `carousel_slides` WHERE `is_visible`=1 ORDER BY `order_index` ASC")
        rows = cursor.fetchall() or []
        conn.close()
        return [{
            "id": r["id"],
            "title": r["title"],
            "slug": r["slug"],
            "description": r["description"],
            "buttonText": r["button_text"],
            "link": r["button_link"],
            "backgroundImage": r["background_image"],
            "order": r["order_index"],
            "active": bool(r["is_visible"])
        } for r in rows]
    except Exception:
        return []


def load_public_profile_menu_from_db() -> list[dict[str, str]]:
    default_profile_menu = [
        {"id": "sejarah", "label": "Sejarah"},
        {"id": "tentang-crembo", "label": "Tentang Crembo"},
        {"id": "struktur", "label": "Struktur"},
        {"id": "visi-misi", "label": "Visi & Misi"},
    ]

    try:
        ensure_auth_schema()
        conn = mysql_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute(
            """
            SELECT `id`, `title`
            FROM `organization_profiles`
            WHERE COALESCE(`is_visible`, 1) = 1
            ORDER BY `order_index` ASC, `title` ASC
            """
        )
        rows = cursor.fetchall() or []
        cursor.close()
        conn.close()

        profile_menu = []
        for row in rows:
            profile_id = (row.get("id") or "").strip()
            title = (row.get("title") or "").strip()
            if profile_id and title:
                profile_menu.append({"id": profile_id, "label": title})

        return profile_menu or default_profile_menu
    except Exception:
        return default_profile_menu

def build_home_page_data() -> dict[str, object]:
    # Ambil slide dari database
    db_slides = load_carousel_slides()
    
    # Jika database masih kosong, tampilkan default bawaan (agar layout tidak hancur)
    if not db_slides:
        db_slides = [
            {
                "id": "default-1",
                "title": "Jadwal Misa Mingguan",
                "slug": "jadwal-misa",
                "description": "Lihat jadwal pelayanan streaming mingguan terbaru dari tim multimedia.",
                "link": "jadwal-streaming.html",
                "buttonText": "Lihat Jadwal",
                "backgroundImage": "",
                "order": 1,
                "active": True,
            },
            {
                "id": "default-2",
                "title": "Open Recruitment Petugas",
                "slug": "open-recruitment",
                "description": "Bergabung sebagai petugas multimedia untuk mendukung pelayanan misa.",
                "link": "form-pendaftaran.html",
                "buttonText": "Daftar Sekarang",
                "backgroundImage": "",
                "order": 2,
                "active": True,
            }
        ]

    return {
        "carouselSlides": db_slides,
        "aboutContent": load_about_content_from_db(),
        "instagramPosts": load_instagram_posts_from_db(limit=12, active_only=True),
        "youtubeVideos": load_youtube_videos(),
        "gmapsUrl": load_gmaps_url(),
        "bigMassSchedules": [],
        "instagramAutoSeconds": 4,
        "profileMenu": load_public_profile_menu_from_db(),
    }


@app.errorhandler(403)
def forbidden_page(error):
    if request.path.startswith("/api/"):
        return jsonify({"ok": False, "message": "Akses modul ini tidak diizinkan."}), 403
    return "<h1>403 Forbidden</h1><p>Akses modul ini tidak diizinkan untuk akun Anda.</p>", 403

@app.route("/")
def index():
    home_page_data = build_home_page_data()
    return render_template("home.html", home_page_data=home_page_data, current_user=current_user_context())


@app.route("/login", methods=["GET", "POST"])
def login():
    ensure_auth_schema()

    if request.method == "POST":
        identifier = request.form.get("identifier", "").strip()
        password = request.form.get("password", "")

        if not identifier or not password:
            flash("Username, email, atau nomor WhatsApp dan kata sandi wajib diisi.", "error")
            return render_template("login.html")

        member = fetch_member(identifier)
        if not member:
            flash("Akun tidak ditemukan.", "error")
            return render_template("login.html")

        if not check_password_hash(member["password"], password):
            flash("Kata sandi salah.", "error")
            return render_template("login.html")

        session.clear()
        session["logged_in"] = True
        session["user_id"] = member["id"]
        session["username"] = member.get("username")
        session["nama"] = member.get("nama")
        session["role"] = member.get("role") or "user"
        session["telp"] = member.get("telp")
        session["email"] = member.get("email")
        session["alamat"] = member.get("alamat")
        session["tgl_lahir"] = member.get("tgl_lahir")
        session["status_akun"] = member.get("status_akun") or "aktif"

        return redirect(login_target_for_role(session["role"]))

    return render_template("login.html", current_user=current_user_context())



# ===== Password reset OTP via email =====
def ensure_password_reset_schema(cursor=None) -> None:
    """Buat/upgrade tabel OTP reset password."""
    own_conn = None
    own_cursor = cursor
    if own_cursor is None:
        own_conn = mysql_connection()
        own_cursor = own_conn.cursor()

    own_cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS `password_reset_otps` (
          `id` varchar(80) NOT NULL,
          `member_id` int(11) NOT NULL,
          `member_email` varchar(255) NOT NULL,
          `identifier` varchar(255) DEFAULT NULL,
          `otp_hash` varchar(255) NOT NULL,
          `reset_token_hash` varchar(255) DEFAULT NULL,
          `expires_at` datetime NOT NULL,
          `resend_available_at` datetime NOT NULL,
          `verified_at` datetime DEFAULT NULL,
          `used_at` datetime DEFAULT NULL,
          `attempts` int(11) NOT NULL DEFAULT 0,
          `request_ip` varchar(80) DEFAULT NULL,
          `user_agent` text DEFAULT NULL,
          `created_at` datetime NOT NULL DEFAULT CURRENT_TIMESTAMP,
          `updated_at` datetime NOT NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
          PRIMARY KEY (`id`),
          KEY `idx_password_reset_member` (`member_id`),
          KEY `idx_password_reset_email` (`member_email`),
          KEY `idx_password_reset_expires` (`expires_at`)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_general_ci
        """
    )
    ensure_column(own_cursor, "password_reset_otps", "identifier", "`identifier` varchar(255) DEFAULT NULL")
    ensure_column(own_cursor, "password_reset_otps", "otp_hash", "`otp_hash` varchar(255) NOT NULL")
    ensure_column(own_cursor, "password_reset_otps", "reset_token_hash", "`reset_token_hash` varchar(255) DEFAULT NULL")
    ensure_column(own_cursor, "password_reset_otps", "expires_at", "`expires_at` datetime NOT NULL")
    ensure_column(own_cursor, "password_reset_otps", "resend_available_at", "`resend_available_at` datetime NOT NULL")
    ensure_column(own_cursor, "password_reset_otps", "verified_at", "`verified_at` datetime DEFAULT NULL")
    ensure_column(own_cursor, "password_reset_otps", "used_at", "`used_at` datetime DEFAULT NULL")
    ensure_column(own_cursor, "password_reset_otps", "attempts", "`attempts` int(11) NOT NULL DEFAULT 0")
    ensure_column(own_cursor, "password_reset_otps", "request_ip", "`request_ip` varchar(80) DEFAULT NULL")
    ensure_column(own_cursor, "password_reset_otps", "user_agent", "`user_agent` text DEFAULT NULL")
    ensure_column(own_cursor, "password_reset_otps", "created_at", "`created_at` datetime NOT NULL DEFAULT CURRENT_TIMESTAMP")
    ensure_column(own_cursor, "password_reset_otps", "updated_at", "`updated_at` datetime NOT NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP")

    if own_conn is not None:
        own_conn.commit()
        own_cursor.close()
        own_conn.close()


def mask_email_address(email_value: str) -> str:
    email_value = normalize_text(email_value)
    if "@" not in email_value:
        return email_value
    local, domain = email_value.split("@", 1)
    if len(local) <= 2:
        masked_local = local[:1] + "*"
    else:
        masked_local = local[:2] + ("*" * max(2, len(local) - 2))
    return f"{masked_local}@{domain}"


def now_utc() -> datetime:
    return datetime.utcnow().replace(microsecond=0)


def seconds_until(dt_value) -> int:
    if not dt_value:
        return 0
    if isinstance(dt_value, str):
        try:
            dt_value = datetime.fromisoformat(dt_value)
        except ValueError:
            return 0
    return max(0, int((dt_value - now_utc()).total_seconds()))


def build_reset_otp_email(member: dict[str, object], otp_code: str, expires_at: datetime) -> tuple[str, str, str]:
    site_name = "Crembo Media"
    member_name = normalize_text(member.get("nama")) or normalize_text(member.get("username")) or "Anggota"
    expires_text = expires_at.strftime("%d/%m/%Y %H:%M:%S")
    subject = f"Kode OTP Reset Password - {site_name}"
    text_body = (
        f"Halo {member_name},\n\n"
        "Anda telah meminta reset password akun Crembo Media.\n"
        f"Kode OTP Anda: {otp_code}\n\n"
        f"Kode ini berlaku selama {PASSWORD_RESET_OTP_MINUTES} menit sampai {expires_text} WIB.\n"
        "Jangan berikan kode ini kepada siapa pun. Jika Anda tidak meminta reset password, abaikan email ini.\n\n"
        f"Salam,\nTim {site_name}"
    )
    spaced_otp = " ".join(otp_code)
    html_body = f"""<!doctype html>
<html>
<head>
  <meta charset="utf-8">
  <title>{html.escape(subject)}</title>
</head>
<body style="margin:0;padding:0;background:#f5f0ef;font-family:Inter,Segoe UI,Arial,sans-serif;color:#1a1a1a;">
  <div style="max-width:620px;margin:0 auto;padding:28px 14px;">
    <div style="background:linear-gradient(135deg,#3a0000,#800000 55%,#a52a2a);border-radius:18px 18px 0 0;padding:24px;color:#fff;">
      <table role="presentation" width="100%" cellspacing="0" cellpadding="0" style="border-collapse:collapse;">
        <tr>
          <td width="64" style="vertical-align:middle;padding-right:14px;">
            <img src="cid:crembo_logo" alt="Logo Crembo Media" width="58" height="58" style="display:block;width:58px;height:58px;object-fit:contain;border-radius:12px;background:rgba(255,255,255,.12);border:1px solid rgba(255,255,255,.22);padding:4px;">
          </td>
          <td style="vertical-align:middle;">
            <div style="font-size:22px;font-weight:900;letter-spacing:.4px;line-height:1.2;">CREMBO MEDIA</div>
            <div style="margin-top:4px;color:rgba(255,255,255,.82);font-size:13px;">Sistem Informasi Internal Komunitas</div>
          </td>
        </tr>
      </table>
    </div>
    <div style="background:#fff;border:1px solid rgba(128,0,0,.13);border-top:0;border-radius:0 0 18px 18px;padding:28px;box-shadow:0 12px 32px rgba(128,0,0,.13);">
      <h1 style="margin:0 0 8px;font-size:22px;color:#800000;">Kode OTP Reset Password</h1>
      <p style="margin:0 0 18px;line-height:1.7;color:#4a4a4a;">Halo <strong>{html.escape(member_name)}</strong>, gunakan kode OTP berikut untuk melanjutkan proses reset password akun Crembo Media Anda.</p>
      <div style="margin:22px 0;padding:18px;border-radius:14px;background:linear-gradient(135deg,#5c0000,#a00000);text-align:center;color:#fff;">
        <div style="font-size:30px;font-weight:900;letter-spacing:10px;font-family:Consolas,Monaco,monospace;">{html.escape(spaced_otp)}</div>
        <div style="margin-top:8px;font-size:12px;color:rgba(255,255,255,.82);">Berlaku {PASSWORD_RESET_OTP_MINUTES} menit sampai {html.escape(expires_text)} WIB</div>
      </div>
      <div style="border-left:4px solid #d4a017;background:#fff8e1;padding:12px 14px;border-radius:10px;color:#5c3b00;font-size:13px;line-height:1.7;">
        Jangan berikan kode ini kepada siapa pun. Jika Anda tidak meminta reset password, abaikan email ini atau hubungi admin.
      </div>
      <p style="margin:22px 0 0;color:#7a7a7a;font-size:12px;">Email ini dikirim otomatis oleh sistem {html.escape(site_name)}.</p>
    </div>
  </div>
</body>
</html>"""
    return subject, text_body, html_body


def send_email_message(recipient_email: str, recipient_name: str, subject: str, text_body: str, html_body: str) -> None:
    if not SMTP_USERNAME or not SMTP_PASSWORD:
        raise RuntimeError("SMTP belum dikonfigurasi. Isi SMTP_USERNAME dan SMTP_PASSWORD.")

    # Struktur related -> alternative agar logo bisa tampil inline via CID di Gmail.
    msg = MIMEMultipart("related")
    msg["Subject"] = subject
    msg["From"] = formataddr((SMTP_FROM_NAME, SMTP_FROM_EMAIL))
    msg["To"] = formataddr((recipient_name or recipient_email, recipient_email))

    alternative = MIMEMultipart("alternative")
    alternative.attach(MIMEText(text_body, "plain", "utf-8"))
    alternative.attach(MIMEText(html_body, "html", "utf-8"))
    msg.attach(alternative)

    logo_path = EMAIL_LOGO_PATH
    if logo_path and logo_path.exists() and logo_path.is_file():
        mime_type, _ = mimetypes.guess_type(str(logo_path))
        image_subtype = (mime_type or "image/png").split("/", 1)[-1]
        try:
            with logo_path.open("rb") as logo_file:
                logo_part = MIMEImage(logo_file.read(), _subtype=image_subtype)
            logo_part.add_header("Content-ID", "<crembo_logo>")
            logo_part.add_header("Content-Disposition", "inline", filename=logo_path.name)
            msg.attach(logo_part)
        except Exception as exc:
            # Email tetap dikirim meskipun file logo bermasalah.
            print(f"[WARN] Gagal melampirkan logo email: {exc}")
    else:
        print(f"[WARN] Logo email tidak ditemukan: {logo_path}")

    with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=30) as server:
        server.ehlo()
        server.starttls()
        server.ehlo()
        server.login(SMTP_USERNAME, SMTP_PASSWORD)
        server.sendmail(SMTP_FROM_EMAIL, [recipient_email], msg.as_string())

def create_password_reset_otp(cursor, member: dict[str, object], identifier: str) -> dict[str, object]:
    otp_code = f"{secrets.randbelow(1000000):06d}"
    reset_id = uuid.uuid4().hex
    issued_at = now_utc()
    expires_at = issued_at + timedelta(minutes=PASSWORD_RESET_OTP_MINUTES)
    resend_available_at = issued_at + timedelta(seconds=PASSWORD_RESET_RESEND_SECONDS)

    cursor.execute(
        """
        INSERT INTO `password_reset_otps`
          (`id`, `member_id`, `member_email`, `identifier`, `otp_hash`, `expires_at`, `resend_available_at`, `request_ip`, `user_agent`)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
        """,
        (
            reset_id,
            member.get("id"),
            normalize_text(member.get("email")),
            normalize_text(identifier),
            generate_password_hash(otp_code),
            expires_at,
            resend_available_at,
            request.headers.get("X-Forwarded-For", request.remote_addr or "")[:80],
            (request.headers.get("User-Agent") or "")[:1000],
        ),
    )
    return {
        "reset_id": reset_id,
        "otp": otp_code,
        "expires_at": expires_at,
        "resend_available_at": resend_available_at,
    }


def password_reset_response_payload(reset_id: str, email_value: str, expires_at, resend_available_at, extra: dict[str, object] | None = None) -> dict[str, object]:
    payload = {
        "ok": True,
        "resetId": reset_id,
        "maskedEmail": mask_email_address(email_value),
        "expiresIn": seconds_until(expires_at),
        "resendAfter": seconds_until(resend_available_at),
        "expiresAt": expires_at.isoformat() if hasattr(expires_at, "isoformat") else str(expires_at),
        "resendAvailableAt": resend_available_at.isoformat() if hasattr(resend_available_at, "isoformat") else str(resend_available_at),
    }
    if extra:
        payload.update(extra)
    return payload


@app.route("/api/password/forgot", methods=["POST"])
def api_password_forgot():
    ensure_auth_schema()
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        ensure_password_reset_schema(cursor)
        data = request.get_json(silent=True) or request.form
        identifier = normalize_text(data.get("identifier"))
        if not identifier:
            return jsonify({"ok": False, "message": "Username, email, atau nomor HP wajib diisi."}), 400

        member = fetch_member(identifier)
        if not member:
            return jsonify({"ok": False, "message": "Akun tidak ditemukan."}), 404

        email_value = normalize_text(member.get("email"))
        if not email_value or "@" not in email_value:
            return jsonify({"ok": False, "message": "Akun ini belum memiliki email terdaftar. Hubungi admin untuk reset password."}), 400

        # Batasi spam: jika OTP terakhir masih cooldown, minta user menunggu.
        cursor.execute(
            """
            SELECT `id`, `expires_at`, `resend_available_at`, `used_at`
            FROM `password_reset_otps`
            WHERE `member_id` = %s AND `used_at` IS NULL
            ORDER BY `created_at` DESC
            LIMIT 1
            """,
            (member.get("id"),),
        )
        last_reset = cursor.fetchone()
        if last_reset and seconds_until(last_reset.get("resend_available_at")) > 0 and seconds_until(last_reset.get("expires_at")) > 0:
            return jsonify({
                "ok": False,
                "message": f"Kode OTP sudah dikirim. Silakan tunggu {seconds_until(last_reset.get('resend_available_at'))} detik untuk kirim ulang.",
                "resetId": last_reset.get("id"),
                "maskedEmail": mask_email_address(email_value),
                "expiresIn": seconds_until(last_reset.get("expires_at")),
                "resendAfter": seconds_until(last_reset.get("resend_available_at")),
            }), 429

        reset_data = create_password_reset_otp(cursor, member, identifier)
        subject, text_body, html_body = build_reset_otp_email(member, reset_data["otp"], reset_data["expires_at"])
        send_email_message(email_value, normalize_text(member.get("nama")), subject, text_body, html_body)
        conn.commit()
        record_activity_for_member(
            member.get("id"),
            "REQUEST",
            "Autentikasi",
            "Meminta kode OTP reset password",
            "password.forgot",
            meta={"identifier": identifier, "email": mask_email_address(email_value)},
        )

        return jsonify(password_reset_response_payload(
            reset_data["reset_id"],
            email_value,
            reset_data["expires_at"],
            reset_data["resend_available_at"],
            {"message": f"Kode OTP berhasil dikirim ke {mask_email_address(email_value)}."},
        ))
    except Exception as exc:
        conn.rollback()
        return jsonify({"ok": False, "message": f"Gagal mengirim OTP: {exc}"}), 500
    finally:
        cursor.close()
        conn.close()


@app.route("/api/password/resend", methods=["POST"])
def api_password_resend():
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        ensure_password_reset_schema(cursor)
        data = request.get_json(silent=True) or request.form
        reset_id = normalize_text(data.get("resetId") or data.get("reset_id"))
        if not reset_id:
            return jsonify({"ok": False, "message": "Sesi reset password tidak valid."}), 400

        cursor.execute("SELECT * FROM `password_reset_otps` WHERE `id` = %s LIMIT 1", (reset_id,))
        reset_row = cursor.fetchone()
        if not reset_row or reset_row.get("used_at"):
            return jsonify({"ok": False, "message": "Sesi reset password tidak valid atau sudah digunakan."}), 400

        wait_seconds = seconds_until(reset_row.get("resend_available_at"))
        if wait_seconds > 0:
            return jsonify({"ok": False, "message": f"Tunggu {wait_seconds} detik sebelum kirim ulang OTP.", "resendAfter": wait_seconds}), 429

        cursor.execute("SELECT * FROM `anggota` WHERE `id` = %s LIMIT 1", (reset_row.get("member_id"),))
        member = cursor.fetchone()
        if not member:
            return jsonify({"ok": False, "message": "Akun tidak ditemukan."}), 404

        email_value = normalize_text(member.get("email") or reset_row.get("member_email"))
        otp_code = f"{secrets.randbelow(1000000):06d}"
        issued_at = now_utc()
        expires_at = issued_at + timedelta(minutes=PASSWORD_RESET_OTP_MINUTES)
        resend_available_at = issued_at + timedelta(seconds=PASSWORD_RESET_RESEND_SECONDS)

        subject, text_body, html_body = build_reset_otp_email(member, otp_code, expires_at)
        send_email_message(email_value, normalize_text(member.get("nama")), subject, text_body, html_body)

        cursor.execute(
            """
            UPDATE `password_reset_otps`
            SET `otp_hash` = %s,
                `expires_at` = %s,
                `resend_available_at` = %s,
                `verified_at` = NULL,
                `reset_token_hash` = NULL,
                `attempts` = 0,
                `updated_at` = CURRENT_TIMESTAMP
            WHERE `id` = %s
            """,
            (generate_password_hash(otp_code), expires_at, resend_available_at, reset_id),
        )
        conn.commit()

        return jsonify(password_reset_response_payload(
            reset_id,
            email_value,
            expires_at,
            resend_available_at,
            {"message": f"Kode OTP baru berhasil dikirim ke {mask_email_address(email_value)}."},
        ))
    except Exception as exc:
        conn.rollback()
        return jsonify({"ok": False, "message": f"Gagal mengirim ulang OTP: {exc}"}), 500
    finally:
        cursor.close()
        conn.close()


@app.route("/api/password/verify", methods=["POST"])
def api_password_verify():
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        ensure_password_reset_schema(cursor)
        data = request.get_json(silent=True) or request.form
        reset_id = normalize_text(data.get("resetId") or data.get("reset_id"))
        otp_code = normalize_text(data.get("otp"))
        if not reset_id or not re.fullmatch(r"\d{6}", otp_code or ""):
            return jsonify({"ok": False, "message": "Masukkan 6 digit kode OTP."}), 400

        cursor.execute("SELECT * FROM `password_reset_otps` WHERE `id` = %s LIMIT 1", (reset_id,))
        reset_row = cursor.fetchone()
        if not reset_row or reset_row.get("used_at"):
            return jsonify({"ok": False, "message": "Sesi reset password tidak valid atau sudah digunakan."}), 400
        if seconds_until(reset_row.get("expires_at")) <= 0:
            return jsonify({"ok": False, "message": "Kode OTP sudah kedaluwarsa. Silakan kirim ulang OTP."}), 400
        if int(reset_row.get("attempts") or 0) >= PASSWORD_RESET_MAX_ATTEMPTS:
            return jsonify({"ok": False, "message": "Percobaan OTP terlalu banyak. Silakan kirim ulang kode."}), 429

        if not check_password_hash(reset_row.get("otp_hash") or "", otp_code):
            cursor.execute("UPDATE `password_reset_otps` SET `attempts` = `attempts` + 1 WHERE `id` = %s", (reset_id,))
            conn.commit()
            return jsonify({"ok": False, "message": "Kode OTP salah. Periksa kembali."}), 400

        reset_token = secrets.token_urlsafe(32)
        reset_token_hash = generate_password_hash(reset_token)
        cursor.execute(
            """
            UPDATE `password_reset_otps`
            SET `verified_at` = %s,
                `reset_token_hash` = %s,
                `updated_at` = CURRENT_TIMESTAMP
            WHERE `id` = %s
            """,
            (now_utc(), reset_token_hash, reset_id),
        )
        conn.commit()

        return jsonify({
            "ok": True,
            "message": "OTP benar. Silakan buat password baru.",
            "resetId": reset_id,
            "resetToken": reset_token,
        })
    except Exception as exc:
        conn.rollback()
        return jsonify({"ok": False, "message": f"Gagal memverifikasi OTP: {exc}"}), 500
    finally:
        cursor.close()
        conn.close()


@app.route("/api/password/reset", methods=["POST"])
def api_password_reset():
    ensure_auth_schema()
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        ensure_password_reset_schema(cursor)
        data = request.get_json(silent=True) or request.form
        reset_id = normalize_text(data.get("resetId") or data.get("reset_id"))
        reset_token = normalize_text(data.get("resetToken") or data.get("reset_token"))
        new_password = str(data.get("password") or "")
        confirm_password = str(data.get("confirmPassword") or data.get("confirm_password") or "")

        if not reset_id or not reset_token:
            return jsonify({"ok": False, "message": "Sesi reset password tidak valid."}), 400
        if not new_password or not confirm_password:
            return jsonify({"ok": False, "message": "Password baru dan konfirmasi wajib diisi."}), 400
        if new_password != confirm_password:
            return jsonify({"ok": False, "message": "Konfirmasi password tidak cocok."}), 400

        rule_score = sum([
            len(new_password) >= 8,
            bool(re.search(r"[A-Z]", new_password)),
            bool(re.search(r"[0-9]", new_password)),
            bool(re.search(r"[^A-Za-z0-9]", new_password)),
        ])
        if rule_score < 3:
            return jsonify({"ok": False, "message": "Password terlalu lemah. Gunakan minimal 8 karakter dan kombinasi huruf besar, angka, atau simbol."}), 400

        cursor.execute("SELECT * FROM `password_reset_otps` WHERE `id` = %s LIMIT 1", (reset_id,))
        reset_row = cursor.fetchone()
        if not reset_row or reset_row.get("used_at"):
            return jsonify({"ok": False, "message": "Sesi reset password tidak valid atau sudah digunakan."}), 400
        if not reset_row.get("verified_at") or not reset_row.get("reset_token_hash"):
            return jsonify({"ok": False, "message": "OTP belum diverifikasi."}), 400
        if seconds_until(reset_row.get("expires_at")) <= 0:
            return jsonify({"ok": False, "message": "Sesi reset password sudah kedaluwarsa. Ulangi dari Lupa Sandi."}), 400
        if not check_password_hash(reset_row.get("reset_token_hash") or "", reset_token):
            return jsonify({"ok": False, "message": "Token reset password tidak valid."}), 400

        cursor.execute("SELECT `id`, `nama`, `email` FROM `anggota` WHERE `id` = %s LIMIT 1", (reset_row.get("member_id"),))
        member = cursor.fetchone()
        if not member:
            return jsonify({"ok": False, "message": "Akun tidak ditemukan."}), 404

        cursor.execute(
            """
            UPDATE `anggota`
            SET `password` = %s, `updated_at` = CURRENT_TIMESTAMP
            WHERE `id` = %s
            """,
            (generate_password_hash(new_password), reset_row.get("member_id")),
        )
        cursor.execute(
            "UPDATE `password_reset_otps` SET `used_at` = %s, `updated_at` = CURRENT_TIMESTAMP WHERE `id` = %s",
            (now_utc(), reset_id),
        )
        conn.commit()
        record_activity_for_member(
            reset_row.get("member_id"),
            "RESET",
            "Autentikasi",
            "Reset password berhasil melalui OTP",
            "password.reset",
        )

        return jsonify({"ok": True, "message": "Password berhasil diperbarui. Silakan login dengan password baru."})
    except Exception as exc:
        conn.rollback()
        return jsonify({"ok": False, "message": f"Gagal menyimpan password baru: {exc}"}), 500
    finally:
        cursor.close()
        conn.close()


@app.route("/logout")
def logout():
    try:
        record_activity("LOGOUT", "Autentikasi", "Logout dari sistem", "logout", method="GET")
    except Exception:
        pass
    session.clear()
    return redirect(url_for("login"))


@app.route("/dashboard")
def dashboard():
    if not session.get("logged_in"):
        return redirect(url_for("login"))
    if (session.get("role") or "user") == "user":
        return redirect(url_for("dashboard_anggota"))
    return render_template("dashboard.html", current_user=current_user_context())


@app.route("/dashboard-anggota")
def dashboard_anggota():
    if not session.get("logged_in"):
        return redirect(url_for("login"))
    if (session.get("role") or "user") != "user":
        return redirect(url_for("dashboard"))
    return render_template("dashboard-anggota.html", current_user=current_user_context())


@app.route("/profil")
def profil():
    if not session.get("logged_in"):
        return redirect(url_for("login"))
    if (session.get("role") or "user") == "user":
        return render_template("profil-anggota.html", current_user=current_user_context())
    return render_template("profil-admin.html", current_user=current_user_context())



@app.route("/api/membership/my", methods=["GET"])
def api_membership_my():
    """Ambil status keanggotaan user login sebagai JSON murni dan read-only.

    Endpoint ini dipanggil bersamaan dengan /api/profile/me dari halaman profil.
    Karena itu endpoint tidak menjalankan auto-transition/DDL agar tidak terjadi
    deadlock MySQL 1213 pada tabel anggota atau membership_status_requests.
    """
    if not session.get("logged_in"):
        return jsonify({"ok": False, "message": "Unauthorized"}), 401

    conn = None
    cursor = None
    try:
        conn = mysql_connection()
        cursor = conn.cursor(dictionary=True)
        user_id = session.get("user_id")

        cursor.execute("SELECT * FROM `anggota` WHERE id = %s LIMIT 1", (user_id,))
        member = cursor.fetchone()
        if not member:
            session.clear()
            return jsonify({"ok": False, "message": "Akun login tidak ditemukan. Silakan login ulang."}), 404

        sync_session_from_member_row(member)

        cursor.execute(
            """
            SELECT * FROM `membership_status_requests`
            WHERE member_id = %s
            ORDER BY created_at DESC, updated_at DESC
            LIMIT 1
            """,
            (user_id,),
        )
        latest = cursor.fetchone()

        return jsonify({
            "ok": True,
            "member": member_row_to_dict(member),
            "request": membership_request_row_to_dict(latest) if latest else None,
            "currentUser": current_user_context_from_row(member),
        })
    except mysql.connector.Error as exc:
        if conn:
            try:
                conn.rollback()
            except Exception:
                pass
        if getattr(exc, "errno", None) in {1205, 1213}:
            return jsonify({
                "ok": False,
                "message": "Database sedang sibuk memproses data keanggotaan. Silakan muat ulang halaman beberapa detik lagi."
            }), 503
        return jsonify({"ok": False, "message": f"Gagal memuat status keanggotaan: {exc}"}), 500
    except Exception as exc:
        if conn:
            try:
                conn.rollback()
            except Exception:
                pass
        return jsonify({"ok": False, "message": f"Gagal memuat status keanggotaan: {exc}"}), 500
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

@app.route("/api/membership/inactive-request", methods=["POST"])

def api_membership_submit_request():
    if not session.get("logged_in"):
        return jsonify({"ok": False, "message": "Unauthorized"}), 401
    if normalize_role_value(session.get("role") or "user") != "user":
        return jsonify({"ok": False, "message": "Pengajuan status nonaktif hanya untuk anggota biasa."}), 403

    ensure_auth_schema()
    ensure_notifications_schema()
    user_id = session.get("user_id")
    mode = "temporary" if normalize_text(request.form.get("inactiveType") or request.form.get("type")).lower() == "temporary" else "permanent"
    reason = normalize_text(request.form.get("reason"))
    start_date = parse_optional_date(request.form.get("effectiveDate") or request.form.get("startDate"))
    return_date = parse_optional_date(request.form.get("reactivateDate") or request.form.get("returnDate"))
    note = normalize_text(request.form.get("note"))
    request_id = normalize_text(request.form.get("id") or request.form.get("requestId"))

    if not reason or not start_date or not note:
        return jsonify({"ok": False, "message": "Alasan, tanggal mulai, dan catatan tambahan wajib diisi."}), 400
    if mode == "temporary":
        if not return_date:
            return jsonify({"ok": False, "message": "Tanggal aktif kembali wajib diisi untuk nonaktif berjangka."}), 400
        if return_date <= start_date:
            return jsonify({"ok": False, "message": "Tanggal aktif kembali harus sesudah tanggal mulai nonaktif."}), 400
    else:
        return_date = None

    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        cursor.execute("START TRANSACTION")
        ensure_membership_columns(cursor)
        cursor.execute("SELECT id, nama FROM anggota WHERE id = %s LIMIT 1", (user_id,))
        member = cursor.fetchone()
        if not member:
            conn.rollback()
            return jsonify({"ok": False, "message": "Data anggota tidak ditemukan."}), 404

        existing = None
        if request_id:
            cursor.execute("SELECT * FROM membership_status_requests WHERE id = %s AND member_id = %s LIMIT 1", (request_id, user_id))
            existing = cursor.fetchone()
            if existing and membership_status_key(existing.get("status")) != MEMBERSHIP_PENDING:
                conn.rollback()
                return jsonify({"ok": False, "message": "Pengajuan yang sudah diproses admin tidak bisa diubah."}), 400

        if not existing:
            cursor.execute("SELECT * FROM membership_status_requests WHERE member_id = %s AND status = 'pending' ORDER BY created_at DESC LIMIT 1", (user_id,))
            existing = cursor.fetchone()

        file_storage = request.files.get("evidence") or request.files.get("bukti") or request.files.get("file")
        evidence_payload = None
        if file_storage and file_storage.filename:
            evidence_payload = save_uploaded_attachment(file_storage)
        elif existing:
            evidence_payload = membership_evidence_payload(existing.get("evidence_json"))

        if not evidence_payload:
            conn.rollback()
            return jsonify({"ok": False, "message": "Bukti pendukung wajib diupload."}), 400

        now = datetime.now()
        if existing:
            req_id = existing.get("id")
            cursor.execute(
                """
                UPDATE membership_status_requests
                SET inactive_type=%s, reason=%s, start_date=%s, return_date=%s, note=%s,
                    evidence_json=%s, updated_at=CURRENT_TIMESTAMP
                WHERE id=%s AND member_id=%s AND status='pending'
                """,
                (mode, reason, start_date, return_date, note, json.dumps([evidence_payload], ensure_ascii=False), req_id, user_id),
            )
            action_label = "diperbarui"
        else:
            req_id = f"nonaktif-{int(time.time()*1000)}-{uuid.uuid4().hex[:8]}"
            cursor.execute(
                """
                INSERT INTO membership_status_requests
                (id, member_id, member_name, inactive_type, reason, start_date, return_date, note, evidence_json, status, created_at, updated_at)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,'pending',%s,%s)
                """,
                (req_id, user_id, member.get("nama") or session.get("nama") or "Anggota", mode, reason, start_date, return_date, note, json.dumps([evidence_payload], ensure_ascii=False), now, now),
            )
            action_label = "dikirim"

        create_notification(
            cursor,
            "keanggotaan",
            f"Pengajuan Nonaktif: {member.get('nama') or 'Anggota'}",
            f"{html.escape(member.get('nama') or 'Anggota')} mengajukan status nonaktif. Mohon tinjau di Manajemen Anggota.",
            "/manajemen-anggota.html",
            {"request_id": req_id},
            target_role="admin",
        )
        conn.commit()
        cursor.execute("SELECT * FROM membership_status_requests WHERE id = %s", (req_id,))
        row = cursor.fetchone()
        return jsonify({"ok": True, "message": f"Pengajuan nonaktif berhasil {action_label}.", "request": membership_request_row_to_dict(row)})
    except Exception as exc:
        conn.rollback()
        return jsonify({"ok": False, "message": str(exc)}), 400
    finally:
        cursor.close()
        conn.close()

@app.route("/api/membership/inactive-request/<request_id>", methods=["DELETE"])
def api_membership_cancel_request(request_id):
    if not session.get("logged_in"):
        return jsonify({"ok": False, "message": "Unauthorized"}), 401
    ensure_auth_schema()
    user_id = session.get("user_id")
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        ensure_membership_columns(cursor)
        cursor.execute("SELECT * FROM membership_status_requests WHERE id=%s AND member_id=%s LIMIT 1", (request_id, user_id))
        row = cursor.fetchone()
        if not row:
            return jsonify({"ok": False, "message": "Pengajuan tidak ditemukan."}), 404
        if membership_status_key(row.get("status")) != MEMBERSHIP_PENDING:
            return jsonify({"ok": False, "message": "Pengajuan yang sudah diproses admin tidak bisa dibatalkan."}), 400
        cursor.execute("UPDATE membership_status_requests SET status='cancelled', updated_at=CURRENT_TIMESTAMP WHERE id=%s", (request_id,))
        conn.commit()
        return jsonify({"ok": True, "message": "Pengajuan nonaktif berhasil dibatalkan."})
    except Exception as exc:
        conn.rollback()
        return jsonify({"ok": False, "message": str(exc)}), 400
    finally:
        cursor.close()
        conn.close()

@app.route("/api/membership/admin/requests", methods=["GET"])
def api_membership_admin_requests():
    if not can_manage_members():
        return jsonify({"ok": False, "message": "Unauthorized"}), 403
    ensure_auth_schema()
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        apply_membership_status_transitions(cursor)
        conn.commit()
        status_filter = membership_status_key(request.args.get("status") or "pending")
        search = normalize_text(request.args.get("search")).lower()
        sort_mode = normalize_text(request.args.get("sort")) or "newest"
        params = []
        where = "WHERE 1=1"
        if status_filter != "all":
            where += " AND r.status = %s"
            params.append(status_filter)
        if search:
            where += " AND (LOWER(r.member_name) LIKE %s OR LOWER(r.reason) LIKE %s OR LOWER(r.note) LIKE %s OR LOWER(r.status) LIKE %s)"
            like = f"%{search}%"
            params.extend([like, like, like, like])
        order = "ASC" if sort_mode == "oldest" else "DESC"
        cursor.execute(f"SELECT r.*, a.email, a.role, a.status_akun FROM membership_status_requests r LEFT JOIN anggota a ON a.id = r.member_id {where} ORDER BY r.created_at {order}, r.updated_at {order}", tuple(params))
        rows = cursor.fetchall() or []
        cursor.execute("SELECT status, COUNT(*) AS count FROM membership_status_requests GROUP BY status")
        counts = {membership_status_key(row.get("status")): int(row.get("count") or 0) for row in cursor.fetchall() or []}
        return jsonify({"ok": True, "items": [membership_request_row_to_dict(r) for r in rows], "counts": counts})
    finally:
        cursor.close()
        conn.close()

@app.route("/api/membership/admin/requests/<request_id>/respond", methods=["POST"])
def api_membership_admin_respond(request_id):
    if not can_manage_members():
        return jsonify({"ok": False, "message": "Unauthorized"}), 403
    payload = request.get_json(silent=True) or {}
    action = normalize_text(payload.get("action")).lower()
    admin_note = normalize_text(payload.get("adminNote") or payload.get("note"))
    if action not in {"approve", "reject"}:
        return jsonify({"ok": False, "message": "Aksi tidak valid."}), 400

    ensure_auth_schema()
    ensure_notifications_schema()
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        cursor.execute("START TRANSACTION")
        ensure_membership_columns(cursor)
        cursor.execute("SELECT * FROM membership_status_requests WHERE id=%s LIMIT 1", (request_id,))
        req = cursor.fetchone()
        if not req:
            conn.rollback()
            return jsonify({"ok": False, "message": "Pengajuan tidak ditemukan."}), 404
        if membership_status_key(req.get("status")) != MEMBERSHIP_PENDING:
            conn.rollback()
            return jsonify({"ok": False, "message": "Pengajuan sudah diproses."}), 400

        new_status = MEMBERSHIP_APPROVED if action == "approve" else MEMBERSHIP_REJECTED
        cursor.execute(
            """
            UPDATE membership_status_requests
            SET status=%s, admin_id=%s, admin_name=%s, admin_note=%s, decided_at=CURRENT_TIMESTAMP, updated_at=CURRENT_TIMESTAMP
            WHERE id=%s
            """,
            (new_status, session.get("user_id"), session.get("nama") or session.get("username"), admin_note, request_id),
        )

        member_id = req.get("member_id")
        if action == "approve":
            today = datetime.now().date()
            start_date = req.get("start_date")
            return_date = req.get("return_date")
            inactive_type = normalize_text(req.get("inactive_type")).lower() or "permanent"

            # Jika pengajuan sudah disetujui, status anggota harus sinkron dari endpoint yang sama
            # baik diproses lewat Dashboard Admin maupun lewat Manajemen Anggota.
            # Untuk request berjangka yang tanggal aktif kembalinya sudah lewat/sama hari ini,
            # akun langsung dianggap aktif kembali; selain itu akun menjadi nonaktif.
            already_finished = bool(inactive_type == "temporary" and return_date and return_date <= today)
            if already_finished:
                cursor.execute(
                    """
                    UPDATE anggota
                    SET status_akun='aktif', inactive_from=NULL, inactive_until=NULL, inactive_type=NULL,
                        inactive_reason=NULL, inactive_note=NULL
                    WHERE id=%s
                    """,
                    (member_id,),
                )
                cursor.execute(
                    """
                    UPDATE membership_status_requests
                    SET applied_at = COALESCE(applied_at, CURRENT_TIMESTAMP),
                        manual_reactivated_at = COALESCE(manual_reactivated_at, CURRENT_TIMESTAMP),
                        updated_at = CURRENT_TIMESTAMP
                    WHERE id = %s
                    """,
                    (request_id,),
                )
            else:
                cursor.execute(
                    """
                    UPDATE anggota
                    SET status_akun='nonaktif', inactive_from=%s, inactive_until=%s, inactive_type=%s,
                        inactive_reason=%s, inactive_note=%s
                    WHERE id=%s
                    """,
                    (start_date or today, return_date, inactive_type, req.get("reason"), req.get("note"), member_id),
                )
                cursor.execute(
                    """
                    UPDATE membership_status_requests
                    SET applied_at = COALESCE(applied_at, CURRENT_TIMESTAMP), updated_at = CURRENT_TIMESTAMP
                    WHERE id = %s
                    """,
                    (request_id,),
                )
                # Verifikasi ulang dari tabel request agar status anggota tidak tetap Aktif
                # ketika approve dilakukan dari halaman Manajemen Anggota.
                force_apply_effective_membership_status(cursor, member_id)
            create_notification_once(
                cursor,
                "keanggotaan",
                "Pengajuan Nonaktif Disetujui",
                "Pengajuan status nonaktif Anda telah disetujui oleh admin.",
                "/profil-anggota.html",
                {"target_user_id": member_id, "request_id": request_id},
                target_role="user",
                dedupe_key=f"membership-approved-{request_id}",
            )
        else:
            create_notification_once(
                cursor,
                "keanggotaan",
                "Pengajuan Nonaktif Ditolak",
                "Pengajuan status nonaktif Anda ditolak oleh admin." + (f" Catatan: {html.escape(admin_note)}" if admin_note else ""),
                "/profil-anggota.html",
                {"target_user_id": member_id, "request_id": request_id},
                target_role="user",
                dedupe_key=f"membership-rejected-{request_id}",
            )
        conn.commit()
        updated_member = None
        if member_id:
            fresh_conn = mysql_connection()
            fresh_cursor = fresh_conn.cursor(dictionary=True)
            try:
                fresh_cursor.execute(
                    """
                    SELECT id, nama, username, telp, password, role, tgl_lahir, email, alamat,
                           status_akun, inactive_until, inactive_from, inactive_type, inactive_reason, inactive_note,
                           created_at, updated_at
                    FROM anggota WHERE id = %s LIMIT 1
                    """,
                    (member_id,),
                )
                updated_row = fresh_cursor.fetchone()
                updated_member = member_row_to_dict(updated_row) if updated_row else None
            finally:
                fresh_cursor.close()
                fresh_conn.close()
        return jsonify({"ok": True, "message": "Pengajuan berhasil diproses.", "member": updated_member})
    except Exception as exc:
        conn.rollback()
        return jsonify({"ok": False, "message": str(exc)}), 400
    finally:
        cursor.close()
        conn.close()

@app.route("/api/membership/admin/pending-actions", methods=["GET"])
def api_membership_admin_pending_actions():
    if not can_manage_members():
        return jsonify({"ok": False, "items": []}), 403
    ensure_auth_schema()
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        ensure_membership_columns(cursor)
        cursor.execute(
            """
            SELECT * FROM membership_status_requests
            WHERE status = 'pending'
            ORDER BY created_at ASC
            LIMIT 10
            """
        )
        return jsonify({"ok": True, "items": [membership_request_row_to_dict(r) for r in cursor.fetchall() or []]})
    finally:
        cursor.close()
        conn.close()


def admin_user_row_to_dict(row: dict[str, object]) -> dict[str, object]:
    permissions = permission_row_to_dict(row)
    birth_date = _date_to_iso(row.get("tgl_lahir"))
    status_value = normalize_status(row.get("status_akun") or "aktif")
    return {
        "id": row.get("id"),
        "fullName": row.get("nama") or "Admin",
        "name": row.get("nama") or "Admin",
        "username": row.get("username") or "",
        "loginIdentifier": row.get("username") or "",
        "email": row.get("email") or "",
        "phone": row.get("telp") or "",
        "address": row.get("alamat") or "",
        "birthDate": birth_date,
        "role": "admin",
        "status": "Aktif" if status_value == "aktif" else "Nonaktif",
        "permissions": permissions,
        "createdAt": row.get("created_at").isoformat() if row.get("created_at") else "",
        "updatedAt": row.get("updated_at").isoformat() if row.get("updated_at") else "",
    }

def fetch_admin_users_for_super(cursor) -> list[dict[str, object]]:
    ensure_admin_permissions_schema(cursor)
    cursor.execute(
        """
        SELECT a.id, a.nama, a.username, a.telp, a.role, a.tgl_lahir, a.email, a.alamat,
               a.status_akun, a.created_at, a.updated_at,
               p.streaming, p.inventaris, p.publikasi, p.konten, p.keanggotaan
        FROM anggota a
        LEFT JOIN admin_module_permissions p ON p.member_id = a.id
        WHERE a.role = 'admin'
        ORDER BY a.nama ASC, a.id ASC
        """
    )
    return [admin_user_row_to_dict(row) for row in cursor.fetchall() or []]

def validate_admin_user_payload(payload: dict[str, object], *, existing_id: int | None = None) -> tuple[dict[str, object] | None, str | None]:
    full_name = normalize_text(payload.get("fullName") or payload.get("name"))
    username = normalize_text(payload.get("loginIdentifier") or payload.get("username"))
    email = normalize_text(payload.get("email"))
    phone = normalize_text(payload.get("phone") or payload.get("telp"))
    address = normalize_text(payload.get("address") or payload.get("alamat"))
    birth_date = parse_optional_date(payload.get("birthDate") or payload.get("tanggalLahir"))
    status_value = normalize_status(payload.get("status") or payload.get("status_akun") or "aktif")
    permissions = normalize_admin_permissions(payload.get("permissions"), default_all=False)

    if not full_name:
        return None, "Nama lengkap admin wajib diisi."
    if not username:
        return None, "Login identifier/username wajib diisi."
    if not email:
        return None, "Email admin wajib diisi."
    if not phone:
        return None, "Nomor WhatsApp admin wajib diisi."
    if not address:
        return None, "Alamat admin wajib diisi."
    if not birth_date:
        return None, "Tanggal lahir admin wajib diisi dengan format valid."

    return {
        "full_name": full_name,
        "username": username,
        "email": email,
        "phone": phone,
        "address": address,
        "birth_date": birth_date,
        "status": status_value,
        "permissions": permissions,
    }, None

def ensure_admin_user_unique(cursor, username: str, email: str, phone: str, *, except_id: int | None = None) -> str | None:
    params = [username, email, phone]
    sql = "SELECT id, username, email, telp FROM anggota WHERE (username = %s OR email = %s OR telp = %s)"
    if except_id is not None:
        sql += " AND id <> %s"
        params.append(except_id)
    sql += " LIMIT 1"
    cursor.execute(sql, tuple(params))
    existing = cursor.fetchone()
    if existing:
        return "Username, email, atau nomor WhatsApp sudah digunakan akun lain."
    return None

def upsert_admin_permissions(cursor, member_id: int, permissions: dict[str, bool]) -> None:
    normalized = normalize_admin_permissions(permissions, default_all=False)
    cursor.execute(
        """
        INSERT INTO admin_module_permissions
          (member_id, streaming, inventaris, publikasi, konten, keanggotaan)
        VALUES (%s, %s, %s, %s, %s, %s)
        ON DUPLICATE KEY UPDATE
          streaming = VALUES(streaming),
          inventaris = VALUES(inventaris),
          publikasi = VALUES(publikasi),
          konten = VALUES(konten),
          keanggotaan = VALUES(keanggotaan),
          updated_at = CURRENT_TIMESTAMP
        """,
        (
            member_id,
            1 if normalized.get("streaming") else 0,
            1 if normalized.get("inventaris") else 0,
            1 if normalized.get("publikasi") else 0,
            1 if normalized.get("konten") else 0,
            1 if normalized.get("keanggotaan") else 0,
        ),
    )

@app.route("/api/admin-users", methods=["GET"])
def api_admin_users_list():
    forbidden = require_super_admin_api()
    if forbidden:
        return forbidden
    ensure_auth_schema()
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        admins = fetch_admin_users_for_super(cursor)
        return jsonify({"ok": True, "admins": admins, "total": len(admins)})
    finally:
        cursor.close()
        conn.close()

@app.route("/api/admin-users", methods=["POST"])
def api_admin_users_create():
    forbidden = require_super_admin_api()
    if forbidden:
        return forbidden
    payload = request.get_json(silent=True) or {}
    values, error = validate_admin_user_payload(payload)
    if error:
        return jsonify({"ok": False, "message": error}), 400
    ensure_auth_schema()
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        cursor.execute("START TRANSACTION")
        unique_error = ensure_admin_user_unique(cursor, values["username"], values["email"], values["phone"])
        if unique_error:
            conn.rollback()
            return jsonify({"ok": False, "message": unique_error}), 400
        cursor.execute("SELECT COALESCE(MAX(id), 0) + 1 AS next_id FROM anggota")
        next_id = int((cursor.fetchone() or {}).get("next_id") or 1)
        password_hash = hash_member_password("", values["birth_date"])
        cursor.execute(
            """
            INSERT INTO anggota
              (id, nama, username, telp, password, role, tgl_lahir, email, alamat, status_akun, created_at, updated_at)
            VALUES (%s, %s, %s, %s, %s, 'admin', %s, %s, %s, %s, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
            """,
            (
                next_id,
                values["full_name"],
                values["username"],
                values["phone"],
                password_hash,
                values["birth_date"],
                values["email"],
                values["address"],
                values["status"],
            ),
        )
        upsert_admin_permissions(cursor, next_id, values["permissions"])
        conn.commit()
        admins = fetch_admin_users_for_super(cursor)
        return jsonify({"ok": True, "message": "Admin baru berhasil ditambahkan.", "admins": admins})
    except Exception as exc:
        conn.rollback()
        return jsonify({"ok": False, "message": str(exc)}), 400
    finally:
        cursor.close()
        conn.close()

@app.route("/api/admin-users/<int:admin_id>", methods=["PUT"])
def api_admin_users_update(admin_id: int):
    forbidden = require_super_admin_api()
    if forbidden:
        return forbidden
    payload = request.get_json(silent=True) or {}
    values, error = validate_admin_user_payload(payload, existing_id=admin_id)
    if error:
        return jsonify({"ok": False, "message": error}), 400
    ensure_auth_schema()
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        cursor.execute("START TRANSACTION")
        cursor.execute("SELECT id, role FROM anggota WHERE id = %s LIMIT 1", (admin_id,))
        row = cursor.fetchone()
        if not row or normalize_role_value(row.get("role") or "") != "admin":
            conn.rollback()
            return jsonify({"ok": False, "message": "Data admin tidak ditemukan."}), 404
        unique_error = ensure_admin_user_unique(cursor, values["username"], values["email"], values["phone"], except_id=admin_id)
        if unique_error:
            conn.rollback()
            return jsonify({"ok": False, "message": unique_error}), 400
        cursor.execute(
            """
            UPDATE anggota
            SET nama = %s, username = %s, telp = %s, tgl_lahir = %s, email = %s,
                alamat = %s, status_akun = %s, updated_at = CURRENT_TIMESTAMP
            WHERE id = %s AND role = 'admin'
            """,
            (
                values["full_name"],
                values["username"],
                values["phone"],
                values["birth_date"],
                values["email"],
                values["address"],
                values["status"],
                admin_id,
            ),
        )
        upsert_admin_permissions(cursor, admin_id, values["permissions"])
        conn.commit()
        admins = fetch_admin_users_for_super(cursor)
        return jsonify({"ok": True, "message": "Data admin berhasil diperbarui.", "admins": admins})
    except Exception as exc:
        conn.rollback()
        return jsonify({"ok": False, "message": str(exc)}), 400
    finally:
        cursor.close()
        conn.close()

@app.route("/api/admin-users/<int:admin_id>", methods=["DELETE"])
def api_admin_users_delete(admin_id: int):
    forbidden = require_super_admin_api()
    if forbidden:
        return forbidden
    if str(admin_id) == str(session.get("user_id") or ""):
        return jsonify({"ok": False, "message": "Anda tidak dapat menghapus akun sendiri dari halaman ini."}), 400
    ensure_auth_schema()
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        cursor.execute("START TRANSACTION")
        cursor.execute("SELECT id, role FROM anggota WHERE id = %s LIMIT 1", (admin_id,))
        row = cursor.fetchone()
        if not row or normalize_role_value(row.get("role") or "") != "admin":
            conn.rollback()
            return jsonify({"ok": False, "message": "Data admin tidak ditemukan."}), 404
        cursor.execute("DELETE FROM admin_module_permissions WHERE member_id = %s", (admin_id,))
        cursor.execute("DELETE FROM anggota WHERE id = %s AND role = 'admin'", (admin_id,))
        conn.commit()
        admins = fetch_admin_users_for_super(cursor)
        return jsonify({"ok": True, "message": "Data admin berhasil dihapus.", "admins": admins})
    except Exception as exc:
        conn.rollback()
        return jsonify({"ok": False, "message": str(exc)}), 400
    finally:
        cursor.close()
        conn.close()

@app.route("/api/anggota", methods=["GET"])
def api_anggota_list():
    if not can_manage_members():
        return jsonify({"ok": False, "message": "Unauthorized"}), 403
    response = jsonify({"ok": True, "members": read_members_for_admin()})
    response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    response.headers["Pragma"] = "no-cache"
    return response


@app.route("/api/anggota/sync", methods=["POST"])
def api_anggota_sync():
    if not can_manage_members():
        return jsonify({"ok": False, "message": "Unauthorized"}), 403

    payload = request.get_json(silent=True) or {}
    members = payload.get("members")
    if not isinstance(members, list):
        return jsonify({"ok": False, "message": "Payload members harus array."}), 400

    try:
        sync_members_from_payload(members)
    except PermissionError as exc:
        return jsonify({"ok": False, "message": str(exc)}), 403
    except Exception as exc:
        return jsonify({"ok": False, "message": str(exc)}), 400
    return jsonify({"ok": True, "members": read_members_for_admin()})


@app.route("/api/profile/me", methods=["GET"])
def api_profile_me():
    """Ambil profil anggota/admin yang sedang login langsung dari DB.

    Endpoint dibuat read-only supaya aman dipanggil paralel dengan
    /api/membership/my dari halaman profil anggota.
    """
    if not session.get("logged_in"):
        return jsonify({"ok": False, "message": "Unauthorized"}), 401

    conn = None
    cursor = None
    try:
        conn = mysql_connection()
        cursor = conn.cursor(dictionary=True)
        user_id = session.get("user_id")
        cursor.execute(
            """
            SELECT id, nama, username, telp, role, tgl_lahir, email, alamat,
                   status_akun, inactive_until, inactive_from, inactive_type,
                   inactive_reason, inactive_note, created_at, updated_at
            FROM anggota
            WHERE id = %s
            LIMIT 1
            """,
            (user_id,),
        )
        row = cursor.fetchone()
        if not row:
            session.clear()
            return jsonify({"ok": False, "message": "Akun login tidak ditemukan. Silakan login ulang."}), 404

        sync_session_from_member_row(row)
        user = current_user_context_from_row(
            row,
            include_admin_permissions=normalize_role_value(row.get("role") or "") in {"admin", "super_admin"},
        )
        return jsonify({"ok": True, "user": user, "member": member_row_to_dict(row)})
    except mysql.connector.Error as exc:
        if conn:
            try:
                conn.rollback()
            except Exception:
                pass
        if getattr(exc, "errno", None) in {1205, 1213}:
            return jsonify({
                "ok": False,
                "message": "Database sedang sibuk memproses data profil. Silakan muat ulang halaman beberapa detik lagi."
            }), 503
        return jsonify({"ok": False, "message": f"Gagal memuat profil login: {exc}"}), 500
    except Exception as exc:
        if conn:
            try:
                conn.rollback()
            except Exception:
                pass
        return jsonify({"ok": False, "message": f"Gagal memuat profil login: {exc}"}), 500
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()


@app.route("/api/profile/update", methods=["POST"])

def api_profile_update():
    if not session.get("logged_in"):
        return jsonify({"ok": False, "message": "Unauthorized"}), 401

    payload = request.get_json(silent=True) or {}
    # Extract fields (do not allow updating role, nama, etc. if not authorized)
    # The requirement is that full name is readonly, so we skip it.
    username = payload.get("username", "").strip()
    email = payload.get("email", "").strip()
    telp = payload.get("phone", "").strip()
    alamat = payload.get("address", "").strip()
    tgl_lahir = payload.get("birthDate", "").strip()

    if not username or not email:
        return jsonify({"ok": False, "message": "Username dan email wajib diisi."}), 400

    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        user_id = session.get("user_id")
        
        # Check uniqueness of username and email (excluding current user)
        cursor.execute("SELECT id FROM anggota WHERE (username = %s OR email = %s OR telp = %s) AND id != %s", (username, email, telp, user_id))
        if cursor.fetchone():
            return jsonify({"ok": False, "message": "Username, email, atau telepon sudah digunakan oleh akun lain."}), 400
            
        cursor.execute(
            """
            UPDATE anggota 
            SET username = %s, email = %s, telp = %s, alamat = %s, tgl_lahir = %s
            WHERE id = %s
            """,
            (username, email, telp, alamat, tgl_lahir, user_id)
        )
        conn.commit()
        
        # Update session dan kirim ulang user terbaru supaya frontend tidak perlu memakai localStorage lama.
        session["username"] = username
        session["email"] = email
        session["telp"] = telp
        session["alamat"] = alamat
        session["tgl_lahir"] = tgl_lahir
        refreshed_user = current_user_context()
        
        return jsonify({"ok": True, "message": "Profil berhasil diperbarui.", "user": refreshed_user})
    except Exception as e:
        conn.rollback()
        return jsonify({"ok": False, "message": str(e)}), 500
    finally:
        cursor.close()
        conn.close()


@app.route("/api/profile/password", methods=["POST"])
def api_profile_password():
    if not session.get("logged_in"):
        return jsonify({"ok": False, "message": "Unauthorized"}), 401

    payload = request.get_json(silent=True) or {}
    current_password = payload.get("currentPassword", "")
    new_password = payload.get("newPassword", "")

    if len(new_password) < 6:
        return jsonify({"ok": False, "message": "Password baru minimal 6 karakter."}), 400

    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        user_id = session.get("user_id")
        cursor.execute("SELECT password, tgl_lahir FROM anggota WHERE id = %s", (user_id,))
        row = cursor.fetchone()
        
        if not row or not check_password_hash(row["password"], current_password):
            return jsonify({"ok": False, "message": "Password saat ini salah."}), 400
            
        hashed_new_password = hash_member_password(new_password, row.get("tgl_lahir", ""))
        
        cursor.execute("UPDATE anggota SET password = %s WHERE id = %s", (hashed_new_password, user_id))
        conn.commit()
        
        return jsonify({"ok": True, "message": "Password berhasil diperbarui."})
    except Exception as e:
        conn.rollback()
        return jsonify({"ok": False, "message": str(e)}), 500
    finally:
        cursor.close()
        conn.close()


@app.route("/<path:page>")
def render_mockup_page(page: str):
    if page in {"favicon.ico"}:
        abort(404)

    asset_path = FRONTEND_DIR / page
    if asset_path.is_file() and not page.endswith(".html"):
        return send_from_directory(FRONTEND_DIR, page)

    candidate = page
    if not candidate.endswith(".html"):
        candidate = f"{candidate}.html"

    if candidate == "login.html":
        session.clear()

    # Role-aware dashboard guard for direct .html URLs.
    # Without this, /dashboard.html is served by the generic template router
    # and user/anggota accounts can accidentally see the admin dashboard.
    if candidate in {"dashboard.html", "dashboard-anggota.html"}:
        if not session.get("logged_in"):
            return redirect(url_for("login"))
        current_role = session.get("role") or "user"
        if candidate == "dashboard.html" and current_role == "user":
            return redirect(url_for("dashboard_anggota"))
        if candidate == "dashboard-anggota.html" and current_role != "user":
            return redirect(url_for("dashboard"))

    if candidate in {"log-aktivitas.html", "log-aktivitas-saya.html"}:
        if not session.get("logged_in"):
            return redirect(url_for("login"))

    if candidate == "kelola-data-admin.html":
        if not session.get("logged_in"):
            return redirect(url_for("login"))
        if normalize_role_value(session.get("role") or "") != "super_admin":
            abort(403)

    module_key = ADMIN_PAGE_MODULE_MAP.get(candidate)
    if module_key:
        if not session.get("logged_in"):
            return redirect(url_for("login"))
        if not admin_has_module_access(module_key):
            abort(403)

    if session.get("logged_in") and current_member_is_inactive() and not is_inactive_member_page_allowed(candidate):
        return redirect(url_for("dashboard_anggota", inactive="1"))

    if template_exists(candidate):
        extra_context = {"current_user": current_user_context()}
        if candidate == "home.html":
            extra_context["home_page_data"] = build_home_page_data()
        if candidate == "manajemen-anggota.html":
            extra_context["member_rows"] = read_members_for_admin() if session.get("logged_in") else []
        return render_template(candidate, **extra_context)

    abort(404)


def render_public_page(template_name: str, **context):
    if not template_exists(template_name):
        abort(404)
    return render_template(template_name, current_user=current_user_context(), **context)

@app.route("/pengumuman")
def public_news_page():
    return render_public_page("pengumuman.html")

@app.route("/pengumuman/kategori/<category_slug>")
def public_news_category_page(category_slug):
    # Menyertakan category_slug jika dibutuhkan oleh Jinja, walau JS di FE juga bisa parsing URL
    return render_public_page("pengumuman.html", category_slug=category_slug)

@app.route("/pengumuman/<news_id>")
def public_news_detail_page(news_id):
    return render_public_page("pengumuman-detail.html", news_id=news_id)


@app.route("/agenda")
def public_agenda_page():
    return render_public_page("agenda.html")

@app.route("/agenda/<agenda_id>")
def public_agenda_detail_page(agenda_id):
    # Mengirim parameter agenda_id ke template sehingga JS di client bisa membaca ID yang akan di load
    return render_public_page("agenda-detail.html", agenda_id=agenda_id)

@app.route("/form-pendaftaran")
def public_registration_forms_page():
    return render_public_page("form-pendaftaran.html")

@app.route("/form-pendaftaran/<form_id>")
def public_registration_form_detail_page(form_id):
    return render_public_page("form-pendaftaran-detail.html", form_id=form_id)

# --- TENTANG CREMBO ENDPOINTS ---

@app.route("/api/tentang/config", methods=["GET"])
def get_tentang_config():
    ensure_auth_schema()
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM `tentang_crembo_config` LIMIT 1")
    row = cursor.fetchone()
    cursor.close()
    conn.close()
    if row:
        return jsonify({
            "description": row["description"],
            "buttonText": row["button_text"],
            "buttonLink": row["button_link"],
            "autoSeconds": row["auto_seconds"]
        })
    return jsonify({})

@app.route("/api/tentang/config", methods=["POST"])
def set_tentang_config():
    ensure_auth_schema()
    data = request.json or {}
    description = data.get("description", "")
    button_text = data.get("buttonText", "")
    button_link = data.get("buttonLink", "")
    try:
        auto_seconds = int(data.get("autoSeconds", 5))
    except (ValueError, TypeError):
        auto_seconds = 5

    conn = mysql_connection()
    cursor = conn.cursor()
    cursor.execute(
        """
        INSERT INTO `tentang_crembo_config` (`id`, `description`, `button_text`, `button_link`, `auto_seconds`)
        VALUES (1, %s, %s, %s, %s)
        ON DUPLICATE KEY UPDATE
          `description` = VALUES(`description`),
          `button_text` = VALUES(`button_text`),
          `button_link` = VALUES(`button_link`),
          `auto_seconds` = VALUES(`auto_seconds`)
        """,
        (description, button_text, button_link, auto_seconds),
    )
    conn.commit()
    cursor.close()
    conn.close()
    return jsonify({"success": True})

@app.route("/api/tentang/media", methods=["GET"])
def get_tentang_media():
    ensure_auth_schema()
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT `id`, `type`, `url`, `order_index`, `is_visible` FROM `tentang_crembo_media` ORDER BY `order_index` ASC")
    rows = cursor.fetchall()
    cursor.close()
    conn.close()
    
    out = []
    for r in rows:
        out.append({
            "id": r["id"],
            "type": r["type"],
            "url": r["url"],
            "order": r["order_index"],
            "active": bool(r["is_visible"])
        })
    return jsonify(out)

@app.route("/api/tentang/media/sync", methods=["POST"])
def sync_tentang_media():
    ensure_auth_schema()
    payload = request.json
    if not isinstance(payload, list):
        return jsonify({"success": False, "error": "Invalid payload format"}), 400
        
    conn = mysql_connection()
    cursor = conn.cursor()
    
    # Simple sync: delete all and insert fresh order
    cursor.execute("DELETE FROM `tentang_crembo_media`")
    
    for item in payload:
        item_id = item.get("id", "")
        item_type = item.get("type", "image")
        url = item.get("url", "")
        order = int(item.get("order", 0))
        active = 1 if item.get("active", True) else 0
        
        if item_id:
            cursor.execute("""
                INSERT INTO `tentang_crembo_media` (`id`, `type`, `url`, `order_index`, `is_visible`)
                VALUES (%s, %s, %s, %s, %s)
            """, (item_id, item_type, url, order, active))
            
    conn.commit()
    cursor.close()
    conn.close()
    return jsonify({"success": True, "count": len(payload)})


@app.route("/api/instagram/posts", methods=["GET"])
def get_instagram_posts():
    ensure_auth_schema()
    return jsonify({"ok": True, "posts": load_instagram_posts_from_db(limit=None, active_only=False)})


@app.route("/api/instagram/posts/sync", methods=["POST"])
def sync_instagram_posts():
    ensure_auth_schema()
    payload = request.get_json(silent=True)
    if not isinstance(payload, list):
        return jsonify({"ok": False, "message": "Invalid payload format"}), 400

    saved_count = save_instagram_posts_payload(payload)
    return jsonify({"ok": True, "count": saved_count})

import werkzeug.utils

@app.route("/uploads/<path:filename>")
def serve_uploaded_file(filename: str):
    upload_folder = ensure_upload_folder()
    return send_from_directory(upload_folder, filename)


@app.route("/api/upload_image", methods=["POST"])
def upload_image():
    ensure_auth_schema()
    incoming_files = request.files.getlist("files")
    if not incoming_files:
        single_file = request.files.get("file")
        incoming_files = [single_file] if single_file and single_file.filename else []

    if not incoming_files:
        return jsonify({"success": False, "error": "No selected file"}), 400

    saved_files: list[dict[str, object]] = []
    try:
        for file in incoming_files:
            if not file or not file.filename:
                continue
            saved_files.append(save_uploaded_attachment(file))
    except ValueError as exc:
        return jsonify({"success": False, "error": str(exc)}), 400

    if not saved_files:
        return jsonify({"success": False, "error": "No selected file"}), 400

    return jsonify({
        "success": True,
        "files": saved_files,
        "url": saved_files[0]["url"],
    })

@app.route("/api/youtube", methods=["GET"])
def get_youtube():
    ensure_auth_schema()
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM `youtube_embeds` ORDER BY `order_index` ASC")
    rows = cursor.fetchall() or []
    conn.close()
    return jsonify([
        {
            "id": r["id"],
            "url": r["url"],
            "type": normalize_youtube_embed_type(r.get("embed_type")),
            "title": r.get("title") or "",
            "order": r["order_index"],
            "active": bool(r["is_visible"]),
        }
        for r in rows
    ])

@app.route("/api/youtube/sync", methods=["POST"])
def sync_youtube():
    ensure_auth_schema()
    payload = request.json or []
    if not isinstance(payload, list):
        return jsonify({"success": False, "error": "Payload YouTube tidak valid."}), 400

    conn = mysql_connection()
    cursor = conn.cursor()
    try:
        ensure_column(cursor, "youtube_embeds", "embed_type", "`embed_type` varchar(30) NOT NULL DEFAULT 'video'")
        ensure_column(cursor, "youtube_embeds", "title", "`title` varchar(255) DEFAULT NULL")
        try:
            cursor.execute("ALTER TABLE `youtube_embeds` MODIFY COLUMN `url` text NOT NULL")
        except Exception:
            pass

        cursor.execute("DELETE FROM `youtube_embeds`")
        for idx, item in enumerate(payload, start=1):
            if not isinstance(item, dict):
                continue
            item_id = normalize_text(item.get("id")) or f"yt-{uuid.uuid4().hex[:12]}"
            item_url = normalize_text(item.get("url"))
            if not item_url:
                continue
            item_type = normalize_youtube_embed_type(item.get("type") or item.get("embed_type"))
            item_title = normalize_text(item.get("title"))[:255]
            item_order = parse_required_int(item.get("order"), idx)
            item_active = 1 if item.get("active", True) else 0
            cursor.execute(
                """
                INSERT INTO `youtube_embeds`
                (`id`, `url`, `embed_type`, `title`, `order_index`, `is_visible`)
                VALUES (%s, %s, %s, %s, %s, %s)
                """,
                (item_id, item_url, item_type, item_title, item_order, item_active),
            )
        conn.commit()
        return jsonify({"success": True})
    except Exception as exc:
        conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 400
    finally:
        cursor.close()
        conn.close()

@app.route("/api/gmaps", methods=["GET"])
def api_get_gmaps():
    return jsonify({"url": load_gmaps_url()})

@app.route("/api/gmaps", methods=["POST"])
def api_set_gmaps():
    ensure_auth_schema()
    url = request.json.get("url", "")
    conn = mysql_connection()
    cursor = conn.cursor()
    cursor.execute("UPDATE `google_maps_embed` SET `url` = %s WHERE `id` = 1", (url,))
    conn.commit()
    conn.close()
    return jsonify({"success": True})

@app.route("/api/carousel", methods=["GET"])
def get_carousel():
    ensure_auth_schema()
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM `carousel_slides` ORDER BY `order_index` ASC")
    rows = cursor.fetchall() or []
    conn.close()
    return jsonify([{
        "id": r["id"],
        "title": r["title"],
        "slug": r["slug"],
        "description": r["description"],
        "buttonText": r["button_text"],
        "link": r["button_link"],
        "backgroundImage": r["background_image"],
        "order": r["order_index"],
        "active": bool(r["is_visible"])
    } for r in rows])

@app.route("/api/carousel/sync", methods=["POST"])
def sync_carousel():
    ensure_auth_schema()
    payload = request.json
    conn = mysql_connection()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM `carousel_slides`")
    for item in payload:
        cursor.execute("""
            INSERT INTO `carousel_slides` 
            (`id`, `title`, `slug`, `description`, `button_text`, `button_link`, `background_image`, `order_index`, `is_visible`) 
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            item.get("id"), item.get("title"), item.get("slug"), item.get("description"),
            item.get("buttonText"), item.get("link"), item.get("backgroundImage"),
            item.get("order"), 1 if item.get("active") else 0
        ))
    conn.commit()
    conn.close()
    return jsonify({"success": True})
@app.route("/api/sertifikat/config", methods=["GET"])
def get_sertifikat_config():
    ensure_auth_schema()
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM `sertifikat_config` LIMIT 1")
    row = cursor.fetchone()
    cursor.close()
    conn.close()
    if row:
        return jsonify({
            "ketuaName": row["ketua_name"],
            "pembinaName": row["pembina_name"],
            "ketuaSignUrl": row["ketua_sign_url"],
            "pembinaSignUrl": row["pembina_sign_url"],
            "updatedAt": row["updated_at"]
        })
    return jsonify({})

@app.route("/api/sertifikat/config", methods=["POST"])
def set_sertifikat_config():
    ensure_auth_schema()
    data = request.json or {}
    conn = mysql_connection()
    cursor = conn.cursor()
    cursor.execute("""
        UPDATE `sertifikat_config` 
        SET `ketua_name` = %s, `pembina_name` = %s, `ketua_sign_url` = %s, `pembina_sign_url` = %s
        WHERE `id` = 1
    """, (data.get("ketuaName", ""), data.get("pembinaName", ""), data.get("ketuaSignUrl", ""), data.get("pembinaSignUrl", "")))
    conn.commit()
    cursor.close()
    conn.close()
    return jsonify({"success": True})

# --- ORGANIZATION PROFILES ENDPOINTS ---

@app.route("/api/profiles", methods=["GET"])
def get_profiles():
    ensure_auth_schema()
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM `organization_profiles` ORDER BY `order_index` ASC")
    rows = cursor.fetchall() or []
    cursor.close()
    conn.close()
    profile_rows = []
    for row in rows:
        attachments = normalize_attachment_payload(row["attachment_url"])
        profile_rows.append({
            "id": row["id"],
            "title": row["title"],
            "description": row["description"],
            "attachmentUrl": attachments[0]["url"] if attachments else "",
            "attachments": attachments,
            "order": row["order_index"],
            "active": bool(row["is_visible"]),
            "createdAt": row["created_at"],
            "updatedAt": row["updated_at"],
        })
    return jsonify(profile_rows)

@app.route("/api/profiles", methods=["POST"])
def create_profile():
    ensure_auth_schema()
    data = request.json or {}
    profile_id = f"profile-{int(time.time() * 1000)}"
    attachments = normalize_attachment_payload(data.get("attachments"))
    if not attachments:
        attachments = normalize_attachment_payload(data.get("attachmentUrl"))

    conn = mysql_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("""
            INSERT INTO `organization_profiles` 
            (`id`, `title`, `description`, `attachment_url`, `order_index`, `is_visible`)
            VALUES (%s, %s, %s, %s, %s, %s)
        """, (
            profile_id,
            data.get("title", ""),
            data.get("description", ""),
            json.dumps(attachments, ensure_ascii=False),
            int(data.get("order", 0)),
            1 if data.get("active") else 0,
        ))
        conn.commit()
        return jsonify({"success": True, "id": profile_id})
    except Exception as e:
        conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 400
    finally:
        cursor.close()
        conn.close()

@app.route("/api/profiles/<profile_id>", methods=["PUT"])
def update_profile(profile_id):
    ensure_auth_schema()
    data = request.json or {}
    attachments = normalize_attachment_payload(data.get("attachments"))
    if not attachments:
        attachments = normalize_attachment_payload(data.get("attachmentUrl"))

    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        cursor.execute("SELECT `attachment_url` FROM `organization_profiles` WHERE `id` = %s LIMIT 1", (profile_id,))
        existing = cursor.fetchone()
        if not existing:
            return jsonify({"success": False, "error": "Profile not found"}), 404

        cursor.execute("""
            UPDATE `organization_profiles`
            SET `title` = %s, `description` = %s, `attachment_url` = %s, `order_index` = %s, `is_visible` = %s
            WHERE `id` = %s
        """, (
            data.get("title", ""),
            data.get("description", ""),
            json.dumps(attachments, ensure_ascii=False),
            int(data.get("order", 0)),
            1 if data.get("active") else 0,
            profile_id
        ))
        conn.commit()

        # Cleanup file lama yang sudah dihapus/diganti saat Edit
        if existing:
            try:
                old_atts_raw = existing.get("attachment_url") or "[]"
                old_atts = json.loads(old_atts_raw) if isinstance(old_atts_raw, str) else (old_atts_raw or [])
                new_att_urls = [a.get("url") for a in attachments if isinstance(a, dict) and a.get("url")]
                
                for oa in old_atts:
                    if isinstance(oa, dict) and oa.get("url") and oa.get("url") not in new_att_urls:
                        remove_physical_file(oa["url"])
            except Exception as e:
                print(f"[WARN] Failed to cleanup replaced files for profile {profile_id}: {e}")

        return jsonify({"success": True})
    except Exception as e:
        conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 400
    finally:
        cursor.close()
        conn.close()

@app.route("/api/profiles/<profile_id>", methods=["DELETE"])
def delete_profile(profile_id):
    ensure_auth_schema()
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        # 1. Fetch data profil untuk mendapatkan list file yang harus dihapus
        cursor.execute("SELECT `attachment_url` FROM `organization_profiles` WHERE `id` = %s LIMIT 1", (profile_id,))
        profile = cursor.fetchone()
        
        if not profile:
            return jsonify({"success": False, "error": "Profile not found"}), 404

        # 2. Hapus record dari database
        cursor.execute("DELETE FROM `organization_profiles` WHERE `id` = %s", (profile_id,))
        conn.commit()

        # 3. Clean up physical files di folder
        try:
            atts_raw = profile.get("attachment_url") or "[]"
            atts = json.loads(atts_raw) if isinstance(atts_raw, str) else (atts_raw or [])
            for att in atts:
                if isinstance(att, dict) and att.get("url"):
                    remove_physical_file(att["url"])
        except Exception as e:
            print(f"[WARN] Error during physical file cleanup for profile {profile_id}: {e}")

        return jsonify({"success": True})
    except Exception as e:
        conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 400
    finally:
        cursor.close()
        conn.close()


# --- REGISTRATION FORM ENDPOINTS ---


def registration_form_lookup(cursor, form_id: str, include_counts: bool = False):
    cursor.execute("SELECT * FROM `registration_forms` WHERE `id` = %s LIMIT 1", (form_id,))
    row = cursor.fetchone()
    if not row:
        return None
    submission_count = 0
    if include_counts:
        cursor.execute("SELECT COUNT(*) AS `count` FROM `registration_form_submissions` WHERE `form_id` = %s", (form_id,))
        count_row = cursor.fetchone() or {}
        submission_count = int(count_row.get("count") or 0)
    return registration_form_row_to_dict(row, submission_count)


@app.route("/api/registration/forms", methods=["GET"])
def get_registration_forms():
    ensure_registration_form_schema()
    scope = (request.args.get("scope") or "public").strip().lower()
    viewer = current_user_context()
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        cursor.execute("SELECT * FROM `registration_forms` ORDER BY `updated_at` DESC, `created_at` DESC")
        rows = cursor.fetchall() or []
        counts = registration_form_submission_counts(cursor)
        forms = [registration_form_row_to_dict(row, counts.get(str(row.get("id") or ""), 0)) for row in rows]

        if scope == "admin":
            if not can_manage_registration_forms():
                return jsonify({"success": False, "error": "Forbidden"}), 403
            return jsonify(forms)

        accessible_forms = []
        for form in forms:
            if form.get("visibility") == "draft":
                continue
            # if form.get("target") == "internal" and not viewer.get("logged_in"):
            #     continue
            accessible_forms.append(form)
        return jsonify(accessible_forms)
    finally:
        cursor.close()
        conn.close()


@app.route("/api/registration/forms/<form_id>", methods=["GET"])
def get_registration_form_detail(form_id):
    ensure_registration_form_schema()
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        form = registration_form_lookup(cursor, form_id, include_counts=True)
        if not form:
            return jsonify({"success": False, "error": "Form not found"}), 404

        viewer = current_user_context()
        if form.get("visibility") == "draft" and not can_manage_registration_forms():
            return jsonify({"success": False, "error": "Forbidden"}), 403
        # if form.get("target") == "internal" and not viewer.get("logged_in") and not can_manage_registration_forms():
        #     return jsonify({"success": False, "error": "Forbidden"}), 403

        return jsonify(form)
    finally:
        cursor.close()
        conn.close()


# Di dalam app.py, cari fungsi create_registration_form()
@app.route("/api/registration/forms", methods=["POST"])
def create_registration_form():
    ensure_registration_form_schema()
    if not can_manage_registration_forms():
        return jsonify({"success": False, "error": "Forbidden"}), 403

    data = request.json or {}
    values = registration_form_payload_to_db_values(data)
    if not values["title"] or not values["open_date"] or not values["close_date"]:
        return jsonify({"success": False, "error": "Judul, tanggal pembukaan, dan tanggal penutupan wajib diisi."}), 400
    if not values["fields_json"] or values["fields_json"] == "[]":
        return jsonify({"success": False, "error": "Tambahkan minimal 1 pertanyaan."}), 400

    try:
        open_date = datetime.strptime(values["open_date"], "%Y-%m-%d").date()
        close_date = datetime.strptime(values["close_date"], "%Y-%m-%d").date()
    except ValueError:
        return jsonify({"success": False, "error": "Format tanggal tidak valid."}), 400
    if close_date < open_date:
        return jsonify({"success": False, "error": "Tanggal penutupan harus sama atau setelah tanggal pembukaan."}), 400

    form_id = data.get("id") or f"registration-form-{int(time.time() * 1000)}"
    conn = mysql_connection()
    cursor = conn.cursor()
    try:
        cursor.execute(
            """
            INSERT INTO `registration_forms`
            (`id`, `title`, `description`, `target`, `visibility`, `open_date`, `close_date`, `quota`, `image_url`, `image_name`, `attachments`, `fields_json`, `created_by`, `created_by_name`, `created_by_role`)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """,
            (
                form_id,
                values["title"],
                values["description"],
                values["target"],
                values["visibility"],
                values["open_date"],
                values["close_date"],
                values["quota"],
                values["image_url"],
                values["image_name"],
                values["attachments"],
                values["fields_json"],
                str(session.get("user_id") or ""),
                str(session.get("nama") or session.get("username") or ""),
                str(session.get("role") or ""),
            ),
        )
        conn.commit()
        try:
            ensure_notifications_schema()
            nc = conn.cursor()
            try:
                # PERBAIKAN: ubah url_for agar mengarah ke halaman public FE, bukan ke API
                create_notification(nc, "form", f"Form Pendaftaran Baru: {values['title']}", values.get('description') or "Terdapat form pendaftaran baru.", url_for('public_registration_form_detail_page', form_id=form_id), {"form_id": form_id})
                conn.commit()
            finally:
                nc.close()
        except Exception:
            pass
        return jsonify({"success": True, "id": form_id})
    except mysql.connector.IntegrityError:
        conn.rollback()
        return jsonify({"success": False, "error": "ID form sudah digunakan."}), 409
    except Exception as error:
        conn.rollback()
        return jsonify({"success": False, "error": str(error)}), 400
    finally:
        cursor.close()
        conn.close()


@app.route("/api/registration/forms/<form_id>", methods=["PUT"])
def update_registration_form(form_id):
    ensure_registration_form_schema()
    if not can_manage_registration_forms():
        return jsonify({"success": False, "error": "Forbidden"}), 403

    data = request.json or {}
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        cursor.execute("SELECT * FROM `registration_forms` WHERE `id` = %s LIMIT 1", (form_id,))
        existing = cursor.fetchone()
        if not existing:
            return jsonify({"success": False, "error": "Form not found"}), 404

        values = registration_form_payload_to_db_values(data, existing)
        if not values["title"] or not values["open_date"] or not values["close_date"]:
            return jsonify({"success": False, "error": "Judul, tanggal pembukaan, dan tanggal penutupan wajib diisi."}), 400
        if not values["fields_json"] or values["fields_json"] == "[]":
            return jsonify({"success": False, "error": "Tambahkan minimal 1 pertanyaan."}), 400

        try:
            open_date = datetime.strptime(values["open_date"], "%Y-%m-%d").date()
            close_date = datetime.strptime(values["close_date"], "%Y-%m-%d").date()
        except ValueError:
            return jsonify({"success": False, "error": "Format tanggal tidak valid."}), 400
        if close_date < open_date:
            return jsonify({"success": False, "error": "Tanggal penutupan harus sama atau setelah tanggal pembukaan."}), 400

        cursor.execute(
            """
            UPDATE `registration_forms`
            SET `title` = %s,
                `description` = %s,
                `target` = %s,
                `visibility` = %s,
                `open_date` = %s,
                `close_date` = %s,
                `quota` = %s,
                `image_url` = %s,
                `image_name` = %s,
                `attachments` = %s,
                `fields_json` = %s
            WHERE `id` = %s
            """,
            (
                values["title"],
                values["description"],
                values["target"],
                values["visibility"],
                values["open_date"],
                values["close_date"],
                values["quota"],
                values["image_url"],
                values["image_name"],
                values["attachments"],
                values["fields_json"],
                form_id,
            ),
        )
        conn.commit()

        # Cleanup file lama yang sudah dihapus/diganti saat Edit
        if existing:
            try:
                # Cleanup Image
                if existing.get("image_url") and existing.get("image_url") != values["image_url"]:
                    remove_physical_file(existing["image_url"])
                
                # Cleanup Attachments
                old_atts_raw = existing.get("attachments") or "[]"
                old_atts = json.loads(old_atts_raw) if isinstance(old_atts_raw, str) else (old_atts_raw or [])
                new_atts_raw = values.get("attachments") or "[]"
                new_atts = json.loads(new_atts_raw) if isinstance(new_atts_raw, str) else (new_atts_raw or [])
                new_att_urls = [a.get("url") for a in new_atts if isinstance(a, dict) and a.get("url")]
                
                for oa in old_atts:
                    if isinstance(oa, dict) and oa.get("url") and oa.get("url") not in new_att_urls:
                        remove_physical_file(oa["url"])
            except Exception as e:
                print(f"[WARN] Failed to cleanup replaced files for registration form {form_id}: {e}")


        return jsonify({"success": True, "id": form_id})
    except Exception as error:
        conn.rollback()
        return jsonify({"success": False, "error": str(error)}), 400
    finally:
        cursor.close()
        conn.close()


@app.route("/api/registration/forms/<form_id>", methods=["DELETE"])
def delete_registration_form(form_id):
    ensure_registration_form_schema()
    if not can_manage_registration_forms():
        return jsonify({"success": False, "error": "Forbidden"}), 403

    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        # Fetch form data first to get the files we need to delete
        cursor.execute("SELECT `image_url`, `attachments` FROM `registration_forms` WHERE `id` = %s LIMIT 1", (form_id,))
        form = cursor.fetchone()
        
        if not form:
            return jsonify({"success": False, "error": "Form not found"}), 404

        # Delete the record from the database
        cursor.execute("DELETE FROM `registration_forms` WHERE `id` = %s", (form_id,))
        conn.commit()

        # Clean up physical files
        try:
            # Delete image
            if form.get("image_url"):
                remove_physical_file(form["image_url"])
            
            # Delete attachments
            attachments_raw = form.get("attachments") or "[]"
            attachments = json.loads(attachments_raw) if isinstance(attachments_raw, str) else (attachments_raw or [])
            for att in attachments:
                if isinstance(att, dict) and att.get("url"):
                    remove_physical_file(att["url"])
        except Exception as e:
            print(f"[WARN] Error during physical file cleanup for form {form_id}: {e}")


        return jsonify({"success": True})
    except Exception as error:
        conn.rollback()
        return jsonify({"success": False, "error": str(error)}), 400
    finally:
        cursor.close()
        conn.close()


@app.route("/api/registration/forms/<form_id>/submissions", methods=["GET"])
def get_registration_form_submissions(form_id):
    ensure_registration_form_schema()
    if not can_manage_registration_forms():
        return jsonify({"success": False, "error": "Forbidden"}), 403

    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        form = registration_form_lookup(cursor, form_id, include_counts=True)
        if not form:
            return jsonify({"success": False, "error": "Form not found"}), 404

        cursor.execute("SELECT * FROM `registration_form_submissions` WHERE `form_id` = %s ORDER BY `submitted_at` DESC", (form_id,))
        submissions = [registration_submission_row_to_dict(row) for row in cursor.fetchall() or []]
        return jsonify({"form": form, "submissions": submissions, "submissionCount": len(submissions)})
    finally:
        cursor.close()
        conn.close()


def registration_export_rows(form: dict[str, object], submissions: list[dict[str, object]]):
    fields = form.get("fields") if isinstance(form.get("fields"), list) else []
    headers = ["No", "Waktu Submit", "Identitas", "Role"] + [str(field.get("label") or "Pertanyaan") for field in fields if isinstance(field, dict)]
    rows = []

    # Format domain utama (Contoh untuk local) Jika dionline sesuaikan URL-nya jika ingin url full
    base_url = "http://127.0.0.1:5000" 

    for index, submission in enumerate(submissions, start=1):
        answer_map = {}
        for answer in submission.get("answers") if isinstance(submission.get("answers"), list) else []:
            if not isinstance(answer, dict):
                continue
            answer_map[str(answer.get("fieldId") or "")] = answer.get("value")

        row_values = [
            index,
            submission.get("submittedAt") or "",
            submission.get("submitter", {}).get("identifier") or "",
            submission.get("submitter", {}).get("role") or "public",
        ]
        for field in fields:
            if not isinstance(field, dict):
                continue
            
            value = answer_map.get(str(field.get("id") or ""))
            field_type = str(field.get("type") or "").strip().lower()

            if field_type == "file":
                # Convert path to full URL to make it clickable on PDF/Excel
                if isinstance(value, list):
                    urls = [f"{base_url}{item}" if str(item).startswith("/") else str(item) for item in value]
                    row_values.append(", ".join(urls))
                elif isinstance(value, str) and value.startswith("/"):
                    row_values.append(f"{base_url}{value}")
                else:
                    row_values.append(str(value or ""))
            else:
                if isinstance(value, list):
                    row_values.append(", ".join(str(item) for item in value))
                else:
                    row_values.append(str(value or ""))

        rows.append(row_values)

    return headers, rows


def registration_export_filename(form: dict[str, object], extension: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "-", str(form.get("title") or "form-pendaftaran").lower()).strip("-") or "form-pendaftaran"
    return f"pendaftar-{slug}.{extension}"


@app.route("/api/registration/forms/<form_id>/export.xlsx", methods=["GET"])
def export_registration_form_excel(form_id):
    ensure_registration_form_schema()
    if not can_manage_registration_forms():
        return jsonify({"success": False, "error": "Forbidden"}), 403

    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        form = registration_form_lookup(cursor, form_id, include_counts=True)
        if not form:
            return jsonify({"success": False, "error": "Form not found"}), 404
        cursor.execute("SELECT * FROM `registration_form_submissions` WHERE `form_id` = %s ORDER BY `submitted_at` DESC", (form_id,))
        submissions = [registration_submission_row_to_dict(row) for row in cursor.fetchall() or []]

        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill

        headers, rows = registration_export_rows(form, submissions)
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Pendaftar"
        worksheet.append(headers)
        for row in rows:
            worksheet.append(row)

        header_fill = PatternFill("solid", fgColor="7F0000")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for column_cells in worksheet.columns:
            max_length = 0
            column_letter = column_cells[0].column_letter
            for cell in column_cells:
                try:
                    cell_value = str(cell.value or "")
                    if len(cell_value) > max_length:
                        max_length = len(cell_value)
                except Exception:
                    continue
            worksheet.column_dimensions[column_letter].width = min(max_length + 4, 40)

        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return send_file(
            buffer,
            as_attachment=True,
            download_name=registration_export_filename(form, "xlsx"),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    finally:
        cursor.close()
        conn.close()


@app.route("/api/registration/forms/<form_id>/export.pdf", methods=["GET"])
def export_registration_form_pdf(form_id):
    ensure_registration_form_schema()
    if not can_manage_registration_forms():
        return jsonify({"success": False, "error": "Forbidden"}), 403

    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        form = registration_form_lookup(cursor, form_id, include_counts=True)
        if not form:
            return jsonify({"success": False, "error": "Form not found"}), 404
        cursor.execute("SELECT * FROM `registration_form_submissions` WHERE `form_id` = %s ORDER BY `submitted_at` DESC", (form_id,))
        submissions = [registration_submission_row_to_dict(row) for row in cursor.fetchall() or []]

        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Spacer, Table, TableStyle, Paragraph

        headers, rows = registration_export_rows(form, submissions)
        buffer = BytesIO()
        document = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=8 * mm, leftMargin=8 * mm, topMargin=10 * mm, bottomMargin=10 * mm)
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle("RegistrationTitle", parent=styles["Title"], fontName="Helvetica-Bold", fontSize=14, leading=16, textColor=colors.HexColor("#7F0000"))
        normal_style = ParagraphStyle("RegistrationNormal", parent=styles["BodyText"], fontName="Helvetica", fontSize=8, leading=10)

        elements = [
            Paragraph(f"Data Pendaftar: {form.get('title')}", title_style),
            Spacer(1, 4 * mm),
            Paragraph(f"Total pendaftar: {len(submissions)}", normal_style),
            Spacer(1, 4 * mm),
        ]

        table_data = [headers] + rows if rows else [headers, ["-", "-", "-", "-"] + ["" for _ in headers[4:]]]
        table = Table(table_data, repeatRows=1)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#7F0000")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 7),
                    ("ALIGN", (0, 0), (0, -1), "CENTER"),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D1D5DB")),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.HexColor("#FFFFFF")]),
                ]
            )
        )
        elements.append(table)
        document.build(elements)
        buffer.seek(0)
        return send_file(
            buffer,
            as_attachment=True,
            download_name=registration_export_filename(form, "pdf"),
            mimetype="application/pdf",
        )
    finally:
        cursor.close()
        conn.close()


@app.route("/api/registration/forms/<form_id>/submit", methods=["POST"])
def submit_registration_form(form_id):
    ensure_registration_form_schema()
    data = request.json or {}
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        form = registration_form_lookup(cursor, form_id, include_counts=True)
        if not form:
            return jsonify({"success": False, "error": "Form not found"}), 404

        submission_count = int(form.get("submissionCount") or 0)
        state = registration_form_status(form, submission_count)
        if not state["open"]:
            return jsonify({"success": False, "error": state["text"]}), 400

        submitter, error_payload = registration_submitter_context(form, data.get("submitter"))
        if error_payload:
            error_body, status_code = error_payload
            return jsonify(error_body), status_code

        answers, error_message = registration_submission_payload_to_rows(form, data.get("answers"))
        if error_message:
            return jsonify({"success": False, "error": error_message}), 400

        cursor.execute(
            "SELECT 1 FROM `registration_form_submissions` WHERE `form_id` = %s AND `submitter_key` = %s LIMIT 1",
            (form_id, submitter["key"]),
        )
        if cursor.fetchone():
            return jsonify({"success": False, "error": "Anda sudah pernah mengirim jawaban untuk form ini."}), 409

        submission_id = f"submission-{int(time.time() * 1000)}-{uuid.uuid4().hex[:6]}"
        cursor.execute(
            """
            INSERT INTO `registration_form_submissions`
            (`id`, `form_id`, `submitter_key`, `submitter_identifier`, `submitter_role`, `submitter_user_id`, `submitter_source`, `answers_json`)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            """,
            (
                submission_id,
                form_id,
                submitter["key"],
                submitter["identifier"],
                submitter["role"],
                submitter["user_id"],
                submitter["source"],
                json.dumps(answers, ensure_ascii=False),
            ),
        )
        conn.commit()
        return jsonify({"success": True, "id": submission_id})
    except Exception as error:
        conn.rollback()
        return jsonify({"success": False, "error": str(error)}), 400
    finally:
        cursor.close()
        conn.close()


@app.route("/api/registration/submissions/me", methods=["GET"])
def get_my_registration_submissions():
    ensure_registration_form_schema()
    viewer = current_user_context()
    client_key = str(request.args.get("clientKey") or request.headers.get("X-Registration-Client-Key") or "").strip()

    if viewer.get("logged_in"):
        submitter_key = f"member:{viewer.get('user_id') or viewer.get('username') or viewer.get('email') or 'public'}"
    else:
        if not client_key:
            return jsonify([])
        submitter_key = client_key

    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        cursor.execute(
            "SELECT * FROM `registration_form_submissions` WHERE `submitter_key` = %s ORDER BY `submitted_at` DESC",
            (submitter_key,),
        )
        submissions = [registration_submission_row_to_dict(row) for row in cursor.fetchall() or []]
        return jsonify(submissions)
    finally:
        cursor.close()
        conn.close()


@app.route("/api/session", methods=["GET"])
def get_session_context():
    response = jsonify(current_user_context())
    # Hindari data profil/session lama terbaca ulang oleh browser saat berganti akun.
    response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    response.headers["Pragma"] = "no-cache"
    return response


# --- SEARCH ENDPOINT ---

@app.route("/api/search", methods=["GET"])
def api_search():
    """
    Global search endpoint. Mencari di: pengumuman (news), agenda, jadwal streaming, profil.
    Query params:
        q      : kata kunci pencarian
        type   : filter tipe (Pengumuman | Agenda | Jadwal | Profil) — kosong = semua
        sort   : newest | oldest | az | za
        page   : halaman (mulai 1)
        per_page: jumlah per halaman (default 10)
    """
    query   = (request.args.get("q") or "").strip().lower()
    ftype   = (request.args.get("type") or "").strip()
    sort    = (request.args.get("sort") or "newest").strip()
    page    = max(1, int(request.args.get("page") or 1))
    per_page = min(50, max(1, int(request.args.get("per_page") or 10)))

    results = []

    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        # ── 1. Pengumuman (news) ─────────────────────────────────────────────────
        if not ftype or ftype == "Pengumuman":
            try:
                cursor.execute("""
                    SELECT id, title, summary, content, thumbnails, published_at, created_at, status
                    FROM `news`
                    WHERE status = 'published'
                    ORDER BY published_at DESC, created_at DESC
                """)
                for row in (cursor.fetchall() or []):
                    title_val   = (row.get("title") or "")
                    summary_val = (row.get("summary") or "")
                    content_raw = re.sub(r'<[^>]+>', ' ', row.get("content") or "")
                    haystack    = (title_val + " " + summary_val + " " + content_raw).lower()
                    if query and query not in haystack:
                        continue
                    snippet = (summary_val or content_raw[:200]).strip()
                    date_val = row.get("published_at") or row.get("created_at")
                    date_str = date_val.strftime("%d %b %Y") if date_val else ""
                    # Ambil URL thumbnail pertama dari JSON
                    thumb_url = ""
                    try:
                        thumbs = json.loads(row.get("thumbnails") or "[]")
                        if thumbs and isinstance(thumbs, list) and isinstance(thumbs[0], dict):
                            thumb_url = thumbs[0].get("url") or ""
                    except Exception:
                        pass
                    results.append({
                        "type":    "Pengumuman",
                        "title":   title_val,
                        "snippet": snippet[:220] if snippet else "",
                        "date":    date_str,
                        "date_ts": date_val.timestamp() if date_val else 0,
                        "url":     f"/pengumuman/{row['id']}",
                        "thumb":   thumb_url,
                    })
            except Exception as e:
                pass  # Tabel belum ada atau error, skip

        # ── 2. Agenda ────────────────────────────────────────────────────────────
        if not ftype or ftype == "Agenda":
            try:
                cursor.execute("""
                    SELECT id, title, description, start_date, location, status, image_url
                    FROM `agendas`
                    WHERE status = 'active'
                    ORDER BY start_date DESC
                """)
                for row in (cursor.fetchall() or []):
                    title_val   = (row.get("title") or "")
                    desc_val    = re.sub(r'<[^>]+>', ' ', row.get("description") or "")
                    loc_val     = (row.get("location") or "")
                    haystack    = (title_val + " " + desc_val + " " + loc_val).lower()
                    if query and query not in haystack:
                        continue
                    date_val = row.get("start_date")
                    date_str = date_val.strftime("%d %b %Y") if hasattr(date_val, "strftime") else str(date_val or "")
                    import datetime as _dt
                    date_ts = _dt.datetime.combine(date_val, _dt.time.min).timestamp() if isinstance(date_val, _dt.date) else 0
                    snippet = desc_val[:220].strip() if desc_val.strip() else loc_val
                    results.append({
                        "type":    "Agenda",
                        "title":   title_val,
                        "snippet": snippet,
                        "date":    date_str,
                        "date_ts": date_ts,
                        "url":     f"/agenda/{row['id']}",
                        "thumb":   row.get("image_url") or "",
                    })
            except Exception:
                pass

        # ── 3. Form Pendaftaran ───────────────────────────────────────────────────
        if not ftype or ftype == "FormPendaftaran":
            try:
                cursor.execute("""
                    SELECT id, title, description, visibility, image_url, updated_at, created_at
                    FROM `registration_forms`
                    WHERE visibility != 'draft'
                    ORDER BY updated_at DESC, created_at DESC
                """)
                for row in (cursor.fetchall() or []):
                    title_val = (row.get("title") or "")
                    desc_raw  = re.sub(r'<[^>]+>', ' ', row.get("description") or "")
                    haystack  = (title_val + " " + desc_raw).lower()
                    if query and query not in haystack:
                        continue
                    upd = row.get("updated_at") or row.get("created_at")
                    date_ts = upd.timestamp() if hasattr(upd, "timestamp") else 0
                    date_str = upd.strftime("%d %b %Y") if upd else ""
                    results.append({
                        "type":    "FormPendaftaran",
                        "title":   title_val,
                        "snippet": desc_raw[:220].strip(),
                        "date":    date_str,
                        "date_ts": date_ts,
                        "url":     f"/form-pendaftaran-detail.html?id={row['id']}",
                        "thumb":   row.get("image_url") or "",
                    })
            except Exception:
                pass

        # ── 4. Profil Organisasi ─────────────────────────────────────────────────
        if not ftype or ftype == "Profil":
            try:
                cursor.execute("""
                    SELECT id, title, description, updated_at
                    FROM `organization_profiles`
                    WHERE is_visible = 1
                    ORDER BY order_index ASC
                """)
                for row in (cursor.fetchall() or []):
                    title_val = (row.get("title") or "")
                    desc_raw  = re.sub(r'<[^>]+>', ' ', row.get("description") or "")
                    haystack  = (title_val + " " + desc_raw).lower()
                    if query and query not in haystack:
                        continue
                    upd = row.get("updated_at")
                    date_ts = upd.timestamp() if hasattr(upd, "timestamp") else 0
                    slug_id = str(row.get("id") or "").replace("profile-", "").lower()
                    safe_id = re.sub(r'[^a-z0-9\-]', '-', title_val.lower())
                    results.append({
                        "type":    "Profil",
                        "title":   title_val,
                        "snippet": desc_raw[:220].strip(),
                        "date":    "–",
                        "date_ts": date_ts,
                        "url":     f"/profil.html?profil={row['id']}",
                        "thumb":   "",
                    })
            except Exception:
                pass

    finally:
        cursor.close()
        conn.close()

    # ── Sorting ──────────────────────────────────────────────────────────────────
    if sort == "az":
        results.sort(key=lambda r: (r["title"] or "").lower())
    elif sort == "za":
        results.sort(key=lambda r: (r["title"] or "").lower(), reverse=True)
    elif sort == "oldest":
        results.sort(key=lambda r: r["date_ts"])
    else:  # newest (default)
        results.sort(key=lambda r: r["date_ts"], reverse=True)

    # Hitung count per tipe (SEBELUM pagination)
    counts = {"Pengumuman": 0, "Agenda": 0, "FormPendaftaran": 0, "Profil": 0}
    for r in results:
        t = r.get("type", "")
        if t in counts:
            counts[t] += 1
    total = len(results)

    # ── Pagination ───────────────────────────────────────────────────────────────
    total_pages = max(1, (total + per_page - 1) // per_page)
    page = min(page, total_pages)
    start = (page - 1) * per_page
    paged = results[start: start + per_page]

    # Bersihkan date_ts dari output
    for r in paged:
        r.pop("date_ts", None)

    return jsonify({
        "query":       (request.args.get("q") or "").strip(),
        "type":        ftype,
        "sort":        sort,
        "page":        page,
        "per_page":    per_page,
        "total":       total,
        "total_pages": total_pages,
        "counts":      counts,
        "results":     paged,
    })


# --- AGENDA ENDPOINTS ---

def agenda_row_to_dict(row):
    attachments_raw = row.get("attachments") or "[]"
    try:
        attachments = json.loads(attachments_raw) if isinstance(attachments_raw, str) else (attachments_raw or [])
    except (TypeError, ValueError):
        attachments = []

    return {
        "id": row.get("id"),
        "title": row.get("title") or "Agenda",
        "description": row.get("description") or "",
        "startDate": str(row.get("start_date") or ""),
        "startTime": str(row.get("start_time") or "")[:5],
        "endDate": str(row.get("end_date") or ""),
        "endTime": str(row.get("end_time") or "")[:5],
        "location": row.get("location") or "",
        "registrationLink": row.get("registration_link") or "",
        "imageUrl": row.get("image_url") or "",
        "imageName": row.get("image_name") or "",
        "attachments": attachments if isinstance(attachments, list) else [],
        "status": "active" if (row.get("status") or "active") == "active" else "inactive",
        "order": int(row.get("order_index") or 0),
        "createdAt": row.get("created_at") or "",
        "updatedAt": row.get("updated_at") or "",
    }


def agenda_payload_to_db_values(data, existing=None):
    source = existing or {}
    attachments = normalize_attachment_payload(data.get("attachments", source.get("attachments", [])))

    def value_for(*keys, default=""):
        for key in keys:
            current = data.get(key)
            if current not in (None, ""):
                return current
            current = source.get(key)
            if current not in (None, ""):
                return current
        return default

    order_value = data.get("order")
    if order_value in (None, ""):
        order_value = data.get("orderIndex")
    if order_value in (None, ""):
        order_value = source.get("order_index") or 0

    status_value = str(value_for("status", default="active")).lower()

    return {
        "title": str(value_for("title")).strip(),
        "description": str(value_for("description")).strip(),
        "start_date": str(value_for("startDate", "start_date")).strip(),
        "start_time": str(value_for("startTime", "start_time")).strip(),
        "end_date": str(value_for("endDate", "end_date")).strip() or None,
        "end_time": str(value_for("endTime", "end_time")).strip() or None,
        "location": str(value_for("location")).strip(),
        "registration_link": str(value_for("registrationLink", "registration_link")).strip(),
        "image_url": str(value_for("imageUrl", "image_url")).strip(),
        "image_name": str(value_for("imageName", "image_name")).strip(),
        "attachments": json.dumps(attachments, ensure_ascii=False),
        "status": "inactive" if status_value == "inactive" else "active",
        "order_index": int(order_value or 0),
    }


@app.route("/api/agendas", methods=["GET"])
def get_agendas():
    ensure_agenda_schema()
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        cursor.execute("SELECT * FROM `agendas` ORDER BY `order_index` ASC, `updated_at` DESC")
        rows = cursor.fetchall() or []
        return jsonify([agenda_row_to_dict(row) for row in rows])
    finally:
        cursor.close()
        conn.close()


@app.route("/api/agendas/<agenda_id>", methods=["GET"])
def get_agenda_detail(agenda_id):
    ensure_agenda_schema()
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        cursor.execute("SELECT * FROM `agendas` WHERE `id` = %s LIMIT 1", (agenda_id,))
        row = cursor.fetchone()
        if not row:
            return jsonify({"error": "Agenda not found"}), 404
        return jsonify(agenda_row_to_dict(row))
    finally:
        cursor.close()
        conn.close()


@app.route("/api/agendas", methods=["POST"])
def create_agenda():
    ensure_agenda_schema()
    data = request.json or {}
    values = agenda_payload_to_db_values(data)
    if not values["title"] or not values["start_date"] or not values["start_time"]:
        return jsonify({"success": False, "error": "Judul, tanggal mulai, dan waktu mulai wajib diisi."}), 400

    agenda_id = f"agenda-{int(time.time() * 1000)}"
    conn = mysql_connection()
    cursor = conn.cursor()
    try:
        cursor.execute(
            """
            INSERT INTO `agendas`
            (`id`, `title`, `description`, `start_date`, `start_time`, `end_date`, `end_time`, `location`, `registration_link`, `image_url`, `image_name`, `attachments`, `status`, `order_index`)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """,
            (
                agenda_id,
                values["title"],
                values["description"],
                values["start_date"],
                values["start_time"],
                values["end_date"],
                values["end_time"],
                values["location"],
                values["registration_link"],
                values["image_url"],
                values["image_name"],
                values["attachments"],
                values["status"],
                values["order_index"],
            ),
        )
        conn.commit()
        try:
            ensure_notifications_schema()
            nc = conn.cursor()
            try:
                create_notification(nc, "agenda", f"Agenda Baru: {values['title']}", values.get('description') or "Terdapat agenda kegiatan baru.", url_for('get_agendas') if False else f"/agenda/{agenda_id}", {"agenda_id": agenda_id})
                conn.commit()
            finally:
                nc.close()
        except Exception:
            pass
        return jsonify({"success": True, "id": agenda_id})
    except Exception as e:
        conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 400
    finally:
        cursor.close()
        conn.close()


@app.route("/api/agendas/<agenda_id>", methods=["PUT"])
def update_agenda(agenda_id):
    ensure_agenda_schema()
    data = request.json or {}
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        cursor.execute("SELECT * FROM `agendas` WHERE `id` = %s LIMIT 1", (agenda_id,))
        existing = cursor.fetchone()
        if not existing:
            return jsonify({"success": False, "error": "Agenda not found"}), 404

        values = agenda_payload_to_db_values(data, existing)
        if not values["title"] or not values["start_date"] or not values["start_time"]:
            return jsonify({"success": False, "error": "Judul, tanggal mulai, dan waktu mulai wajib diisi."}), 400

        cursor.execute(
            """
            UPDATE `agendas`
            SET `title` = %s,
                `description` = %s,
                `start_date` = %s,
                `start_time` = %s,
                `end_date` = %s,
                `end_time` = %s,
                `location` = %s,
                `registration_link` = %s,
                `image_url` = %s,
                `image_name` = %s,
                `attachments` = %s,
                `status` = %s,
                `order_index` = %s
            WHERE `id` = %s
            """,
            (
                values["title"],
                values["description"],
                values["start_date"],
                values["start_time"],
                values["end_date"],
                values["end_time"],
                values["location"],
                values["registration_link"],
                values["image_url"],
                values["image_name"],
                values["attachments"],
                values["status"],
                values["order_index"],
                agenda_id,
            ),
        )
        conn.commit()
        
        # Cleanup file lama yang sudah dihapus/diganti saat Edit
        if existing:
            try:
                # Cleanup Image
                if existing.get("image_url") and existing.get("image_url") != values["image_url"]:
                    remove_physical_file(existing["image_url"])
                
                # Cleanup Attachments
                old_atts_raw = existing.get("attachments") or "[]"
                old_atts = json.loads(old_atts_raw) if isinstance(old_atts_raw, str) else (old_atts_raw or [])
                new_atts_raw = values.get("attachments") or "[]"
                new_atts = json.loads(new_atts_raw) if isinstance(new_atts_raw, str) else (new_atts_raw or [])
                new_att_urls = [a.get("url") for a in new_atts if isinstance(a, dict) and a.get("url")]
                
                for oa in old_atts:
                    if isinstance(oa, dict) and oa.get("url") and oa.get("url") not in new_att_urls:
                        remove_physical_file(oa["url"])
            except Exception as e:
                print(f"[WARN] Failed to cleanup replaced files for agenda {agenda_id}: {e}")

        return jsonify({"success": True})
    except Exception as e:
        conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 400
    finally:
        cursor.close()
        conn.close()


@app.route("/api/agendas/<agenda_id>", methods=["DELETE"])
def delete_agenda(agenda_id):
    ensure_agenda_schema()
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        # 1. Fetch agenda data first to get the files we need to delete
        cursor.execute("SELECT `image_url`, `attachments` FROM `agendas` WHERE `id` = %s LIMIT 1", (agenda_id,))
        agenda = cursor.fetchone()
        
        if not agenda:
            return jsonify({"success": False, "error": "Agenda not found"}), 404

        # 2. Delete the record from the database
        cursor.execute("DELETE FROM `agendas` WHERE `id` = %s", (agenda_id,))
        conn.commit()

        # 3. Clean up physical files
        try:
            # Delete image
            if agenda.get("image_url"):
                remove_physical_file(agenda["image_url"])
            
            # Delete attachments
            attachments_raw = agenda.get("attachments") or "[]"
            attachments = json.loads(attachments_raw) if isinstance(attachments_raw, str) else (attachments_raw or [])
            for att in attachments:
                if isinstance(att, dict) and att.get("url"):
                    remove_physical_file(att["url"])
        except Exception as e:
            print(f"[WARN] Error during physical file cleanup for agenda {agenda_id}: {e}")

        return jsonify({"success": True})
    except Exception as e:
        conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 400
    finally:
        cursor.close()
        conn.close()

# --- NEWS & CATEGORIES ENDPOINTS ---

@app.route("/api/news-categories", methods=["GET"])
def get_news_categories():
    ensure_news_schema()
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        cursor.execute("SELECT * FROM `news_categories` ORDER BY `order_index`, `name`")
        categories = cursor.fetchall()
        return jsonify([{
            "id": cat["id"],
            "name": cat["name"],
            "slug": cat["slug"],
            "description": cat["description"],
            "order": cat["order_index"]
        } for cat in categories])
    finally:
        cursor.close()
        conn.close()

@app.route("/api/news-categories", methods=["POST"])
def create_news_category():
    ensure_news_schema()
    data = request.json or {}
    name = (data.get("name") or "").strip()
    slug = (data.get("slug") or "").strip() or re.sub(r'[^a-z0-9]+', '-', name.lower()).strip('-')
    category_id = f"cat-{int(time.time() * 1000)}"

    conn = mysql_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("""
            INSERT INTO `news_categories` (`id`, `name`, `slug`, `description`, `order_index`)
            VALUES (%s, %s, %s, %s, %s)
        """, (
            category_id,
            name,
            slug,
            data.get("description", ""),
            int(data.get("order", 0))
        ))
        conn.commit()
        return jsonify({"success": True, "id": category_id})
    except Exception as e:
        conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 400
    finally:
        cursor.close()
        conn.close()

@app.route("/api/news-categories/<category_id>", methods=["PUT"])
def update_news_category(category_id):
    ensure_news_schema()
    data = request.json or {}
    name = (data.get("name") or "").strip()
    slug = (data.get("slug") or "").strip() or re.sub(r'[^a-z0-9]+', '-', name.lower()).strip('-')

    conn = mysql_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("""
            UPDATE `news_categories`
            SET `name` = %s, `slug` = %s, `description` = %s, `order_index` = %s
            WHERE `id` = %s
        """, (
            name,
            slug,
            data.get("description", ""),
            int(data.get("order", 0)),
            category_id
        ))
        conn.commit()
        return jsonify({"success": True})
    except Exception as e:
        conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 400
    finally:
        cursor.close()
        conn.close()

@app.route("/api/news-categories/<category_id>", methods=["DELETE"])
def delete_news_category(category_id):
    ensure_news_schema()
    conn = mysql_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("DELETE FROM `news_category_mapping` WHERE `category_id` = %s", (category_id,))
        cursor.execute("DELETE FROM `news_categories` WHERE `id` = %s", (category_id,))
        conn.commit()
        return jsonify({"success": True})
    except Exception as e:
        conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 400
    finally:
        cursor.close()
        conn.close()

@app.route("/api/news", methods=["GET"])
def get_news():
    ensure_news_schema()
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        cursor.execute("""
            SELECT * FROM `news` ORDER BY `published_at` DESC, `created_at` DESC
        """)
        news_items = cursor.fetchall()
        
        result = []
        for item in news_items:
            # Get categories for this news
            cursor.execute("""
                SELECT c.id, c.name, c.slug FROM `news_category_mapping` m
                JOIN `news_categories` c ON m.category_id = c.id
                WHERE m.news_id = %s
            """, (item["id"],))
            categories = cursor.fetchall()
            
            result.append({
                "id": item["id"],
                "title": item["title"],
                "slug": item["slug"],
                "content": item["content"],
                "summary": item["summary"],
                "thumbnails": json.loads(item["thumbnails"] or "[]"),
                "attachments": json.loads(item["attachments"] or "[]"),
                "status": item["status"],
                "publishedAt": item["published_at"].isoformat() if item["published_at"] else None,
                "createdAt": item["created_at"].isoformat() if item["created_at"] else None,
                "categories": [{"id": c["id"], "name": c["name"], "slug": c["slug"]} for c in categories]
            })
        return jsonify(result)
    finally:
        cursor.close()
        conn.close()

@app.route("/api/news/<news_id>", methods=["GET"])
def get_news_detail(news_id):
    ensure_news_schema()
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        cursor.execute("SELECT * FROM `news` WHERE `id` = %s", (news_id,))
        item = cursor.fetchone()
        if not item:
            return jsonify({"error": "News not found"}), 404

        # Get categories for this news
        cursor.execute("""
            SELECT c.id, c.name, c.slug FROM `news_category_mapping` m
            JOIN `news_categories` c ON m.category_id = c.id
            WHERE m.news_id = %s
        """, (news_id,))
        categories = cursor.fetchall()

        return jsonify({
            "id": item["id"],
            "title": item["title"],
            "slug": item["slug"],
            "content": item["content"],
            "summary": item["summary"],
            "thumbnails": json.loads(item["thumbnails"] or "[]"),
            "attachments": json.loads(item["attachments"] or "[]"),
            "status": item["status"],
            "publishedAt": item["published_at"].isoformat() if item["published_at"] else None,
            "createdAt": item["created_at"].isoformat() if item["created_at"] else None,
            "categories": [{"id": c["id"], "name": c["name"], "slug": c["slug"]} for c in categories]
        })
    finally:
        cursor.close()
        conn.close()

@app.route("/api/news", methods=["POST"])
def create_news():
    ensure_news_schema()
    data = request.json or {}
    news_id = f"news-{int(time.time() * 1000)}"
    
    title = (data.get("title") or "").strip()
    slug = (data.get("slug") or "").strip() or re.sub(r'[^a-z0-9]+', '-', title.lower()).strip('-')
    content = data.get("content", "")
    summary = data.get("summary", "")
    thumbnails = normalize_attachment_payload(data.get("thumbnails", []))
    attachments = normalize_attachment_payload(data.get("attachments", []))
    status = data.get("status", "draft")
    category_ids = data.get("categoryIds", [])

    conn = mysql_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("""
            INSERT INTO `news` (`id`, `title`, `slug`, `content`, `summary`, `thumbnails`, `attachments`, `status`)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            news_id,
            title,
            slug,
            content,
            summary,
            json.dumps(thumbnails, ensure_ascii=False),
            json.dumps(attachments, ensure_ascii=False),
            status
        ))

        # Add categories
        for cat_id in category_ids:
            cursor.execute("""
                INSERT IGNORE INTO `news_category_mapping` (`news_id`, `category_id`)
                VALUES (%s, %s)
            """, (news_id, cat_id))

        conn.commit()
        try:
            ensure_notifications_schema()
            nc = conn.cursor()
            try:
                create_notification(nc, "news", f"Pengumuman Baru: {title}", summary or "Terdapat pengumuman baru.", f"/pengumuman/{news_id}", {"news_id": news_id})
                conn.commit()
            finally:
                nc.close()
        except Exception:
            pass
        return jsonify({"success": True, "id": news_id})
    except Exception as e:
        conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 400
    finally:
        cursor.close()
        conn.close()

@app.route("/api/news/<news_id>", methods=["PUT"])
def update_news(news_id):
    ensure_news_schema()
    data = request.json or {}
    
    title = (data.get("title") or "").strip()
    slug = (data.get("slug") or "").strip() or re.sub(r'[^a-z0-9]+', '-', title.lower()).strip('-')
    content = data.get("content", "")
    summary = data.get("summary", "")
    thumbnails = normalize_attachment_payload(data.get("thumbnails", []))
    attachments = normalize_attachment_payload(data.get("attachments", []))
    status = data.get("status", "draft")
    category_ids = data.get("categoryIds", [])

    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        cursor.execute("SELECT * FROM `news` WHERE `id` = %s LIMIT 1", (news_id,))
        existing = cursor.fetchone()
        if not existing:
            return jsonify({"success": False, "error": "News not found"}), 404

        cursor.execute("""
            UPDATE `news`
            SET `title` = %s, `slug` = %s, `content` = %s, `summary` = %s, 
                `thumbnails` = %s, `attachments` = %s, `status` = %s
            WHERE `id` = %s
        """, (
            title,
            slug,
            content,
            summary,
            json.dumps(thumbnails, ensure_ascii=False),
            json.dumps(attachments, ensure_ascii=False),
            status,
            news_id
        ))

        # Update categories
        cursor.execute("DELETE FROM `news_category_mapping` WHERE `news_id` = %s", (news_id,))
        for cat_id in category_ids:
            cursor.execute("""
                INSERT IGNORE INTO `news_category_mapping` (`news_id`, `category_id`)
                VALUES (%s, %s)
            """, (news_id, cat_id))

        conn.commit()

        # Clean up old files if they are replaced
        try:
            old_thumbs_raw = existing.get("thumbnails") or "[]"
            old_thumbs = json.loads(old_thumbs_raw) if isinstance(old_thumbs_raw, str) else (old_thumbs_raw or [])
            new_thumb_urls = [t.get("url") for t in thumbnails if isinstance(t, dict) and t.get("url")]
            for ot in old_thumbs:
                if isinstance(ot, dict) and ot.get("url") and ot.get("url") not in new_thumb_urls:
                    remove_physical_file(ot["url"])
                    
            old_atts_raw = existing.get("attachments") or "[]"
            old_atts = json.loads(old_atts_raw) if isinstance(old_atts_raw, str) else (old_atts_raw or [])
            new_att_urls = [a.get("url") for a in attachments if isinstance(a, dict) and a.get("url")]
            for oa in old_atts:
                if isinstance(oa, dict) and oa.get("url") and oa.get("url") not in new_att_urls:
                    remove_physical_file(oa["url"])
        except Exception as e:
            print(f"[WARN] Failed to cleanup replaced files for news {news_id}: {e}")

        return jsonify({"success": True})
    except Exception as e:
        conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 400
    finally:
        cursor.close()
        conn.close()

@app.route("/api/news/<news_id>", methods=["DELETE"])
def delete_news(news_id):
    ensure_news_schema()
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        # 1. Fetch news data first to get the files we need to delete
        cursor.execute("SELECT `thumbnails`, `attachments` FROM `news` WHERE `id` = %s LIMIT 1", (news_id,))
        news = cursor.fetchone()
        
        if not news:
            return jsonify({"success": False, "error": "News not found"}), 404
            
        # 2. Delete mappings and record
        cursor.execute("DELETE FROM `news_category_mapping` WHERE `news_id` = %s", (news_id,))
        cursor.execute("DELETE FROM `news` WHERE `id` = %s", (news_id,))
        conn.commit()
        
        # 3. Clean up physical files
        try:
            # Clean thumbnails
            thumbnails_raw = news.get("thumbnails") or "[]"
            thumbnails = json.loads(thumbnails_raw) if isinstance(thumbnails_raw, str) else (thumbnails_raw or [])
            for thumb in thumbnails:
                if isinstance(thumb, dict) and thumb.get("url"):
                    remove_physical_file(thumb["url"])
                    
            # Clean attachments
            attachments_raw = news.get("attachments") or "[]"
            attachments = json.loads(attachments_raw) if isinstance(attachments_raw, str) else (attachments_raw or [])
            for att in attachments:
                if isinstance(att, dict) and att.get("url"):
                    remove_physical_file(att["url"])
        except Exception as e:
            print(f"[WARN] Error during physical file cleanup for news {news_id}: {e}")

        return jsonify({"success": True})
    except Exception as e:
        conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 400
    finally:
        cursor.close()
        conn.close()

# --- EXPORT EXCEL & PDF RIWAYAT PEMINJAMAN ---

@app.route("/api/pengajuan/export.xlsx", methods=["GET"])
def export_pengajuan_excel():
    ensure_auth_schema()
    role = session.get("role", "")
    if role not in ["admin", "super_admin"]:
        abort(403)
        
    status_filter = request.args.get('status')
    from_date = request.args.get('from_date')
    to_date = request.args.get('to_date')
    
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        ensure_loan_schema(cursor)
        query = """
            SELECT l.*, a.nama as member_name 
            FROM `loan_requests` l
            LEFT JOIN `anggota` a ON l.member_id = a.id
            WHERE 1=1
        """
        params = []
        if status_filter and status_filter != 'all':
            query += " AND l.`status` = %s"
            params.append(status_filter)
        if from_date:
            query += " AND l.`tanggal_mulai` >= %s"
            params.append(from_date)
        if to_date:
            query += " AND l.`tanggal_mulai` <= %s"
            params.append(to_date)
            
        query += " ORDER BY l.`created_at` DESC"
        cursor.execute(query, tuple(params))
        items = cursor.fetchall() or []
        
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        
        headers = [
            "ID Laporan", "Peminjam", "Barang", "Qty", 
            "Mulai Pinjam", "Target Kembali", "Status",
            "Waktu Ambil", "Kondisi Ambil", "Catatan Ambil", "Bukti Ambil",
            "Waktu Kembali", "Kondisi Kembali", "Catatan Kembali", "Bukti Kembali", "Catatan Admin"
        ]
        
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Riwayat Peminjaman"
        worksheet.append(headers)
        
        base_url = "http://127.0.0.1:5000"
        
        for item in items:
            p_info = safe_json_loads(item.get("pickup_info"), {})
            r_info = safe_json_loads(item.get("return_info"), {})
            
            p_kondisi = ", ".join([u.get("status", "") for u in p_info.get("units", [])]) if p_info.get("units") else "-"
            r_kondisi = ", ".join([u.get("status", "") for u in r_info.get("units", [])]) if r_info.get("units") else "-"
            
            p_catatan = " | ".join([u.get("reason", "") for u in p_info.get("units", []) if u.get("reason")]) if p_info.get("units") else "-"
            r_catatan = " | ".join([u.get("reason", "") for u in r_info.get("units", []) if u.get("reason")]) if r_info.get("units") else "-"
            
            p_foto = f"{base_url}{p_info.get('photo')}" if p_info.get("photo") and str(p_info.get("photo")).startswith("/") else (p_info.get("photo") or "-")
            r_foto = f"{base_url}{r_info.get('photo')}" if r_info.get("photo") and str(r_info.get("photo")).startswith("/") else (r_info.get("photo") or "-")

            mulai_str = f"{item.get('tanggal_mulai')} {str(item.get('waktu_mulai', ''))[:5]}"
            selesai_str = f"{item.get('tanggal_selesai')} {str(item.get('waktu_selesai', ''))[:5]}"
            p_waktu = f"{p_info.get('date', '')} {p_info.get('time', '')}" if p_info.get("date") else "-"
            r_waktu = f"{r_info.get('date', '')} {r_info.get('time', '')}" if r_info.get("date") else "-"
            
            status_map = {"pending": "Pending", "approved": "Disetujui", "taken": "Dipinjam", "returned": "Selesai", "rejected": "Ditolak", "cancelled": "Dibatalkan"}
            status_indo = status_map.get(item.get("status"), item.get("status"))
            
            row = [
                item.get("id"),
                item.get("member_name") or "-",
                f"{item.get('barang_name')} ({item.get('barang_code')})",
                item.get("jumlah", 1),
                mulai_str,
                selesai_str,
                status_indo,
                p_waktu,
                p_kondisi,
                p_catatan,
                p_foto,
                r_waktu,
                r_kondisi,
                r_catatan,
                r_foto,
                item.get("admin_note") or "-"
            ]
            worksheet.append(row)
            
        header_fill = PatternFill("solid", fgColor="7F0000")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2), 2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrapText=True)

        for column_cells in worksheet.columns:
            max_length = 0
            column_letter = column_cells[0].column_letter
            for cell in column_cells:
                try:
                    cell_value = str(cell.value or "")
                    if len(cell_value) > max_length:
                        max_length = len(cell_value)
                except Exception:
                    continue
            worksheet.column_dimensions[column_letter].width = min(max_length + 2, 40)

        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        
        filename = f"riwayat-peminjaman-{datetime.now().strftime('%Y%m%d%H%M')}.xlsx"
        
        return send_file(
            buffer,
            as_attachment=False, # Karena file excel browser cenderung akan download langsung
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    finally:
        cursor.close()
        conn.close()

@app.route("/api/pengajuan/export.pdf", methods=["GET"])
def export_pengajuan_pdf():
    ensure_auth_schema()
    role = session.get("role", "")
    if role not in ["admin", "super_admin"]:
        abort(403)
        
    status_filter = request.args.get('status')
    from_date = request.args.get('from_date')
    to_date = request.args.get('to_date')
    
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        ensure_loan_schema(cursor)
        query = """
            SELECT l.*, a.nama as member_name 
            FROM `loan_requests` l
            LEFT JOIN `anggota` a ON l.member_id = a.id
            WHERE 1=1
        """
        params = []
        if status_filter and status_filter != 'all':
            query += " AND l.`status` = %s"
            params.append(status_filter)
        if from_date:
            query += " AND l.`tanggal_mulai` >= %s"
            params.append(from_date)
        if to_date:
            query += " AND l.`tanggal_mulai` <= %s"
            params.append(to_date)
            
        query += " ORDER BY l.`created_at` DESC"
        cursor.execute(query, tuple(params))
        items = cursor.fetchall() or []
        
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Spacer, Table, TableStyle, Paragraph

        headers = ["Peminjam", "Barang", "Tgl Ambil", "Kondisi/Bukti Ambil", "Tgl Kembali", "Kondisi/Bukti Kembali", "Status"]
        rows = []
        base_url = "http://127.0.0.1:5000"
        
        for item in items:
            p_info = safe_json_loads(item.get("pickup_info"), {})
            r_info = safe_json_loads(item.get("return_info"), {})
            
            p_waktu = f"{p_info.get('date', '')} {p_info.get('time', '')}" if p_info.get("date") else "-"
            r_waktu = f"{r_info.get('date', '')} {r_info.get('time', '')}" if r_info.get("date") else "-"
            
            p_kondisi = ", ".join([u.get("status", "") for u in p_info.get("units", [])]) if p_info.get("units") else "-"
            r_kondisi = ", ".join([u.get("status", "") for u in r_info.get("units", [])]) if r_info.get("units") else "-"
            
            p_foto = f"{base_url}{p_info.get('photo')}" if p_info.get("photo") and str(p_info.get("photo")).startswith("/") else (p_info.get("photo") or "")
            r_foto = f"{base_url}{r_info.get('photo')}" if r_info.get("photo") and str(r_info.get("photo")).startswith("/") else (r_info.get("photo") or "")
            
            p_cell = f"{p_kondisi}<br/><a href='{p_foto}' color='blue'>Lihat Foto Ambil</a>" if p_foto else p_kondisi
            r_cell = f"{r_kondisi}<br/><a href='{r_foto}' color='blue'>Lihat Foto Kembali</a>" if r_foto else r_kondisi
            
            status_map = {"pending": "Pending", "approved": "Disetujui", "taken": "Dipinjam", "returned": "Selesai", "rejected": "Ditolak", "cancelled": "Dibatalkan"}
            
            rows.append([
                item.get("member_name") or "-",
                item.get("barang_name") or "-",
                p_waktu,
                p_cell,
                r_waktu,
                r_cell,
                status_map.get(item.get("status"), item.get("status"))
            ])
            
        buffer = BytesIO()
        document = SimpleDocTemplate(
            buffer, pagesize=landscape(A4),
            rightMargin=8 * mm, leftMargin=8 * mm, topMargin=10 * mm, bottomMargin=10 * mm,
        )
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle("Title", parent=styles["Title"], fontName="Helvetica-Bold", fontSize=14, leading=16, textColor=colors.HexColor("#7F0000"))
        normal_style = ParagraphStyle("Normal", parent=styles["BodyText"], fontName="Helvetica", fontSize=8, leading=10)
        cell_style = ParagraphStyle("Cell", parent=styles["BodyText"], fontName="Helvetica", fontSize=7, leading=9)

        elements = [
            Paragraph("Rekap Data Riwayat Peminjaman dan Pengembalian", title_style),
            Spacer(1, 4 * mm),
            Paragraph(f"Dicetak pada: {datetime.now().strftime('%d %B %Y %H:%M')} WIB", normal_style),
            Spacer(1, 4 * mm),
        ]

        wrapped_rows = []
        for row in rows:
            wrapped_row = []
            for cell_data in row:
                wrapped_row.append(Paragraph(str(cell_data), cell_style))
            wrapped_rows.append(wrapped_row)

        table_data = [headers] + wrapped_rows if rows else [headers, ["-" for _ in headers]]
        page_width = landscape(A4)[0] - (16 * mm) 
        col_widths = [0.15*page_width, 0.20*page_width, 0.12*page_width, 0.18*page_width, 0.12*page_width, 0.18*page_width, 0.05*page_width]
        
        table = Table(table_data, colWidths=col_widths, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#7F0000")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("ALIGN", (0, 0), (0, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D1D5DB")),
        ]))
        elements.append(table)
        document.build(elements)
        buffer.seek(0)
        
        filename = f"riwayat-peminjaman-{datetime.now().strftime('%Y%m%d%H%M')}.pdf"
        
        return send_file(
            buffer,
            as_attachment=False,  # Buka di tab baru (preview)
            download_name=filename,
            mimetype="application/pdf",
        )
    finally:
        cursor.close()
        conn.close()

@app.route("/api/form-kerusakan/export.xlsx", methods=["GET"])
def export_kerusakan_excel():
    ensure_auth_schema()
    user_context = current_user_context()
    role = user_context.get("role", "")
    
    if role not in ["admin", "super_admin"]:
        abort(403)
        
    search_query = request.args.get('search', '').strip()
    status_filter = request.args.get('status')
    severity_filter = request.args.get('severity')
    
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        ensure_damage_schema(cursor)
        
        query = "SELECT d.*, a.nama as member_name FROM `form_kerusakan_barang` d LEFT JOIN `anggota` a ON d.member_id = a.id WHERE 1=1"
        params = []
        
        if status_filter and status_filter != 'all':
            query += " AND d.`status` = %s"
            params.append(status_filter)
        
        if severity_filter and severity_filter != 'all':
            query += " AND d.`tingkat_kerusakan` = %s"
            params.append(severity_filter)
        
        if search_query:
            query += " AND (d.`barang_name` LIKE %s OR d.`deskripsi_kerusakan` LIKE %s OR a.`nama` LIKE %s)"
            search_param = f"%{search_query}%"
            params.extend([search_param, search_param, search_param])
        
        query += " ORDER BY d.`created_at` DESC"
        
        cursor.execute(query, tuple(params))
        items = cursor.fetchall() or []
        
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        
        # Susunan Header baru
        headers = [
            "ID Laporan", "Barang", "Kode Barang", "Pelapor", 
            "Tingkat Kerusakan", "Deskripsi Kerusakan", "Waktu Kejadian", "Diinput", "Foto Bukti", "Status"
        ]
        
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Laporan Kerusakan"
        worksheet.append(headers)
        
        base_url = "http://127.0.0.1:5000"
        
        for item in items:
            waktu_kej = item.get("waktu_kejadian").strftime("%Y-%m-%d %H:%M:%S") if item.get("waktu_kejadian") else "-"
            diinput = item.get("created_at").strftime("%Y-%m-%d %H:%M:%S") if item.get("created_at") else "-"
            
            # Format foto ke full URL
            foto_links = []
            foto_raw = safe_json_loads(item.get("foto_kerusakan"), [])
            if foto_raw:
                for f in foto_raw:
                    if isinstance(f, dict) and f.get("url"):
                        url = str(f.get("url"))
                        if url.startswith("/"): url = f"{base_url}{url}"
                        foto_links.append(url)
            foto_label = "\n".join(foto_links) if foto_links else "-"
            
            row = [
                item.get("id"),
                item.get("barang_name") or "-",
                item.get("barang_code") or "-",
                item.get("member_name") or "-",
                item.get("tingkat_kerusakan") or "-",
                item.get("deskripsi_kerusakan") or "-",
                waktu_kej,
                diinput,
                foto_label,
                item.get("status") or "-",
            ]
            worksheet.append(row)
            
        header_fill = PatternFill("solid", fgColor="7F0000")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2), 2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrapText=True)

        for column_cells in worksheet.columns:
            max_length = 0
            column_letter = column_cells[0].column_letter
            for cell in column_cells:
                try:
                    cell_value = str(cell.value or "")
                    if "\n" in cell_value:
                        lines = cell_value.split("\n")
                        line_lengths = [len(l) for l in lines]
                        if line_lengths and max(line_lengths) > max_length:
                            max_length = max(line_lengths)
                    elif len(cell_value) > max_length:
                        max_length = len(cell_value)
                except Exception:
                    continue
            worksheet.column_dimensions[column_letter].width = min(max_length + 4, 60)

        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        
        filename = f"laporan-kerusakan-{datetime.now().strftime('%Y%m%d%H%M')}.xlsx"
        
        return send_file(
            buffer,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    finally:
        cursor.close()
        conn.close()

@app.route("/api/form-kerusakan/export.pdf", methods=["GET"])
def export_kerusakan_pdf():
    ensure_auth_schema()
    user_context = current_user_context()
    role = user_context.get("role", "")
    
    if role not in ["admin", "super_admin"]:
        abort(403)
        
    search_query = request.args.get('search', '').strip()
    status_filter = request.args.get('status')
    severity_filter = request.args.get('severity')
    
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        ensure_damage_schema(cursor)
        
        query = "SELECT d.*, a.nama as member_name FROM `form_kerusakan_barang` d LEFT JOIN `anggota` a ON d.member_id = a.id WHERE 1=1"
        params = []
        
        if status_filter and status_filter != 'all':
            query += " AND d.`status` = %s"
            params.append(status_filter)
        
        if severity_filter and severity_filter != 'all':
            query += " AND d.`tingkat_kerusakan` = %s"
            params.append(severity_filter)
        
        if search_query:
            query += " AND (d.`barang_name` LIKE %s OR d.`deskripsi_kerusakan` LIKE %s OR a.`nama` LIKE %s)"
            search_param = f"%{search_query}%"
            params.extend([search_param, search_param, search_param])
        
        query += " ORDER BY d.`created_at` DESC"
        
        cursor.execute(query, tuple(params))
        items = cursor.fetchall() or []
        
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Spacer, Table, TableStyle, Paragraph

        # Susunan header baru di PDF
        headers = ["Barang", "Pelapor", "Tingkat", "Deskripsi", "Waktu", "Diinput", "Foto Bukti", "Status"]
        rows = []
        base_url = "http://127.0.0.1:5000"
        
        for item in items:
            waktu_kej = item.get("waktu_kejadian").strftime("%Y-%m-%d %H:%M") if item.get("waktu_kejadian") else "-"
            diinput = item.get("created_at").strftime("%Y-%m-%d %H:%M") if item.get("created_at") else "-"
            
            foto_links = []
            foto_raw = safe_json_loads(item.get("foto_kerusakan"), [])
            if foto_raw:
                for f in foto_raw:
                    if isinstance(f, dict) and f.get("url"):
                        url = str(f.get("url"))
                        if url.startswith("/"): url = f"{base_url}{url}"
                        # Kita bungkus url foto menggunakan tag link agar bisa di klik di PDF
                        foto_links.append(f"<a href='{url}' color='blue'>{url}</a>")
            foto_label = "<br/>".join(foto_links) if foto_links else "-"

            rows.append([
                item.get("barang_name") or "-",
                item.get("member_name") or "-",
                item.get("tingkat_kerusakan") or "-",
                item.get("deskripsi_kerusakan") or "-",
                waktu_kej,
                diinput,
                foto_label,
                item.get("status") or "-"
            ])
            
        buffer = BytesIO()
        document = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            rightMargin=8 * mm,
            leftMargin=8 * mm,
            topMargin=10 * mm,
            bottomMargin=10 * mm,
        )
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle("DamageTitle", parent=styles["Title"], fontName="Helvetica-Bold", fontSize=14, leading=16, textColor=colors.HexColor("#7F0000"))
        normal_style = ParagraphStyle("DamageNormal", parent=styles["BodyText"], fontName="Helvetica", fontSize=8, leading=10)
        cell_style = ParagraphStyle("CellNormal", parent=styles["BodyText"], fontName="Helvetica", fontSize=7, leading=9, wordWrap='CJK')

        elements = [
            Paragraph("Rekap Data Laporan Kerusakan Barang", title_style),
            Spacer(1, 4 * mm),
            Paragraph(f"Dicetak pada: {datetime.now().strftime('%d %B %Y %H:%M')} WIB", normal_style),
            Spacer(1, 4 * mm),
        ]

        wrapped_rows = []
        for row in rows:
            wrapped_row = []
            for cell_data in row:
                text = str(cell_data).replace('\n', '<br/>')
                wrapped_row.append(Paragraph(text, cell_style))
            wrapped_rows.append(wrapped_row)

        table_data = [headers] + wrapped_rows if rows else [headers, ["-" for _ in headers]]
        
        page_width = landscape(A4)[0] - (16 * mm) 
        # Sesuaikan proporsi persentase ke-8 kolom 
        col_widths = [
            0.15 * page_width,
            0.12 * page_width,
            0.08 * page_width,
            0.20 * page_width,
            0.10 * page_width,
            0.10 * page_width,
            0.15 * page_width,
            0.10 * page_width
        ]
        
        table = Table(table_data, colWidths=col_widths, repeatRows=1)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#7F0000")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 7),
                    ("ALIGN", (0, 0), (0, -1), "CENTER"),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D1D5DB")),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.HexColor("#FFFFFF")]),
                ]
            )
        )
        elements.append(table)
        document.build(elements)
        buffer.seek(0)
        
        filename = f"laporan-kerusakan-{datetime.now().strftime('%Y%m%d%H%M')}.pdf"
        
        return send_file(
            buffer,
            as_attachment=False,  # FALSE agar terbuka di tab browser dulu (preview) sebelum didownload
            download_name=filename,
            mimetype="application/pdf",
        )
    finally:
        cursor.close()
        conn.close()

@app.route("/api/streaming/roles", methods=["GET", "POST"])
def manage_streaming_roles():
    ensure_streaming_schema()
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        if request.method == "POST":
            payload = request.json or []
            cursor.execute("DELETE FROM `streaming_roles`")
            for idx, r in enumerate(payload):
                cursor.execute("INSERT INTO `streaming_roles` (role_name, order_index) VALUES (%s, %s)", (r['name'], idx+1))
            conn.commit()
            return jsonify({"success": True})
        
        cursor.execute("SELECT role_name as name FROM `streaming_roles` ORDER BY order_index ASC")
        return jsonify(cursor.fetchall())
    finally:
        cursor.close()
        conn.close()

@app.route("/api/streaming/config/weekly", methods=["GET", "POST"])
def manage_weekly_config():
    ensure_streaming_schema()
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        if request.method == "POST":
            payload = request.json or {}
            cursor.execute("DELETE FROM `streaming_weekly_config`")
            for day, times in payload.items():
                for t in times:
                    parts = t.split(' - ')
                    jam = parts[0].strip()
                    nama = parts[1].strip() if len(parts) > 1 else "Misa"
                    cursor.execute("INSERT INTO `streaming_weekly_config` (day_name, start_time, mass_name) VALUES (%s, %s, %s)", (day, jam, nama))
            conn.commit()
            return jsonify({"success": True})
        
        cursor.execute("SELECT * FROM `streaming_weekly_config` ORDER BY FIELD(day_name, 'Senin', 'Selasa', 'Rabu', 'Kamis', 'Jumat', 'Sabtu', 'Minggu'), start_time ASC")
        rows = cursor.fetchall()
        result = {}
        for r in rows:
            day = r['day_name']
            if day not in result: result[day] = []
            jam_str = format_time_hhmm(r['start_time'])
            result[day].append(f"{jam_str} - {r['mass_name']}")
        return jsonify(result)
    finally:
        cursor.close()
        conn.close()

@app.route("/api/streaming/cancelled", methods=["GET", "POST", "DELETE"])
def manage_cancelled_mass():
    ensure_streaming_schema()
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        if request.method == "POST":
            data = request.json
            cursor.execute("INSERT INTO `streaming_cancelled` (mass_date, mass_time) VALUES (%s, %s)", (data['date'], data['time']))
            conn.commit()
            return jsonify({"success": True})
        
        if request.method == "DELETE":
            data = request.json
            # Perbaikan: Langsung bandingkan mass_time tanpa DATE_FORMAT (MySQL otomatis konversi '18:00' ke TIME)
            cursor.execute("DELETE FROM `streaming_cancelled` WHERE mass_date = %s AND mass_time = %s", (data['date'], data['time'][:5]))
            conn.commit()
            return jsonify({"success": True})

        # Perbaikan: Ambil string dengan DATE_FORMAT otomatis dari DB, hindari format GMT 
        cursor.execute("SELECT mass_date as date, DATE_FORMAT(mass_time, '%H:%i') as time FROM `streaming_cancelled` ORDER BY mass_date DESC")
        rows = cursor.fetchall()
        # Flask serializes datetime.date as HTTP-date ("Thu, 01 May ..."), harus dikonversi ke ISO
        for r in rows:
            if hasattr(r['date'], 'isoformat'):
                r['date'] = r['date'].isoformat()
        return jsonify(rows)
    finally:
        cursor.close()
        conn.close()

@app.route("/api/streaming/schedule", methods=["GET"])
def get_streaming_schedule():
    ensure_streaming_schema()
    ensure_misa_besar_schema()
    month = int(request.args.get("month", datetime.now().month))
    year = int(request.args.get("year", datetime.now().year))
    
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        cursor.execute("SELECT role_name AS name FROM `streaming_roles` ORDER BY order_index ASC, id ASC")
        roles = [row["name"] for row in (cursor.fetchall() or []) if row.get("name")]

        cursor.execute("SELECT * FROM `streaming_weekly_config` ORDER BY start_time ASC")
        weekly_configs = cursor.fetchall() or []
        
        cursor.execute(
            """
            SELECT DATE_FORMAT(mass_date, '%Y-%m-%d') AS mass_date,
                   DATE_FORMAT(mass_time, '%H:%i') AS mass_time
            FROM `streaming_cancelled`
            WHERE MONTH(mass_date) = %s AND YEAR(mass_date) = %s
            """,
            (month, year),
        )
        cancelled_list = cursor.fetchall() or []
        cancelled_set = {f"{c['mass_date']}_{c['mass_time']}" for c in cancelled_list}

        cursor.execute(
            """
            SELECT DATE_FORMAT(misa_date, '%Y-%m-%d') AS misa_date,
                   DATE_FORMAT(misa_time, '%H:%i') AS misa_time,
                   misa_name
            FROM `misa_besar`
            WHERE status = 'published' AND MONTH(misa_date) = %s AND YEAR(misa_date) = %s
            """,
            (month, year),
        )
        big_mass_rows = cursor.fetchall() or []
        big_mass_conflict_set = {f"{row['misa_date']}_{row['misa_time']}" for row in big_mass_rows}

        cursor.execute(
            """
            SELECT DATE_FORMAT(sa.schedule_date, '%Y-%m-%d') AS schedule_date,
                   DATE_FORMAT(sa.schedule_time, '%H:%i') AS schedule_time,
                   sa.role_name,
                   sa.member_id,
                   COALESCE(a.nama, CONCAT('ID ', sa.member_id)) AS member_name
            FROM `streaming_assignments` sa
            LEFT JOIN `anggota` a ON a.id = sa.member_id
            WHERE MONTH(sa.schedule_date) = %s AND YEAR(sa.schedule_date) = %s
            """,
            (month, year),
        )
        assignment_rows = cursor.fetchall() or []
        assignment_map: dict[tuple[str, str], dict[str, str]] = {}
        assignment_id_map: dict[tuple[str, str], dict[str, str]] = {}
        for row in assignment_rows:
            key = (row["schedule_date"], row["schedule_time"])
            assignment_map.setdefault(key, {})[row["role_name"]] = row.get("member_name") or ""
            assignment_id_map.setdefault(key, {})[row["role_name"]] = str(row.get("member_id") or "")
        
        day_map = {0: 'Senin', 1: 'Selasa', 2: 'Rabu', 3: 'Kamis', 4: 'Jumat', 5: 'Sabtu', 6: 'Minggu'}
        
        schedule = []
        num_days = calendar.monthrange(year, month)[1]
        
        for day in range(1, num_days + 1):
            date_obj = datetime(year, month, day)
            day_name = day_map[date_obj.weekday()]
            date_str = date_obj.strftime("%Y-%m-%d")
            
            for cfg in weekly_configs:
                if cfg['day_name'] == day_name:
                    jam_str = format_time_hhmm(cfg['start_time'])
                    key = f"{date_str}_{jam_str}"
                    if key in cancelled_set or key in big_mass_conflict_set:
                        continue
                    assignment_key = (date_str, jam_str)
                    assignments = {role_name: assignment_map.get(assignment_key, {}).get(role_name, "") for role_name in roles}
                    assignment_ids = {role_name: assignment_id_map.get(assignment_key, {}).get(role_name, "") for role_name in roles}
                    schedule.append({
                        "date": date_str,
                        "time": jam_str,
                        "massName": cfg['mass_name'],
                        "dayName": day_name,
                        "roles": roles,
                        "assignments": assignments,
                        "assignmentIds": assignment_ids,
                    })
        
        schedule.sort(key=lambda x: (x['date'], x['time']))
        return jsonify({"success": True, "schedule": schedule, "roles": roles})
    finally:
        cursor.close()
        conn.close()

@app.route("/api/streaming/active-members", methods=["GET"])
def get_active_members():
    ensure_auth_schema()
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        cursor.execute("SELECT id, nama as name, role FROM anggota WHERE status_akun = 'aktif' ORDER BY nama ASC")
        return jsonify(cursor.fetchall())
    finally:
        cursor.close()
        conn.close()

@app.route("/api/streaming/assignments", methods=["GET"])
def get_current_assignments():
    ensure_streaming_schema()
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        cursor.execute("SELECT DATE_FORMAT(schedule_date, '%Y-%m-%d') as schedule_date, DATE_FORMAT(schedule_time, '%H:%i') as schedule_time, role_name, member_id FROM streaming_assignments")
        return jsonify(cursor.fetchall())
    finally:
        cursor.close()
        conn.close()

@app.route("/api/streaming/assignments/save", methods=["POST"])
def save_assignments():
    ensure_streaming_schema()
    payload = request.json or [] 
    
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True, buffered=True)
    try:
        cursor.execute("SELECT schedule_date, schedule_time, role_name, member_id FROM streaming_assignments")
        existing_assignments = {}
        for r in cursor.fetchall():
            t_str = format_time_hhmm(r['schedule_time'])
            d_str = str(r['schedule_date'])
            existing_assignments[(d_str, t_str, r['role_name'])] = str(r['member_id'])
            
        ensure_notifications_schema()
        DAYS_INDO = ["Senin", "Selasa", "Rabu", "Kamis", "Jumat", "Sabtu", "Minggu"]
        
        for item in payload:
            d_str = item.get('date')
            t_str = item.get('time')
            r_str = item.get('role')
            m_id = str(item.get('memberId')) if item.get('memberId') else None
            m_name = item.get('massName') or 'Misa'
            
            key = (d_str, t_str, r_str)
            
            if m_id:
                cursor.execute("""
                    INSERT INTO streaming_assignments (schedule_date, schedule_time, role_name, member_id, request_source, created_at)
                    VALUES (%s, %s, %s, %s, 'admin', CURRENT_TIMESTAMP)
                    ON DUPLICATE KEY UPDATE
                        created_at = IF(member_id <> VALUES(member_id), CURRENT_TIMESTAMP, created_at),
                        request_source = IF(member_id <> VALUES(member_id), 'admin', request_source),
                        member_id = VALUES(member_id)
                """, (d_str, t_str, r_str, m_id))
                
                if existing_assignments.get(key) != m_id:
                    date_obj = datetime.strptime(d_str, "%Y-%m-%d")
                    day_name = DAYS_INDO[date_obj.weekday()]
                    date_formatted = date_obj.strftime("%d/%m/%Y")
                    
                    title = f"Tugas Baru: {r_str}"
                    body = f"Anda ditugaskan sebagai <b>{r_str}</b> untuk <b>{m_name}</b> pada hari {day_name}, {date_formatted} jam {t_str} WIB."
                    
                    create_notification(cursor, "tugas", title, body, "/jadwal-tugas-misa-anggota.html", {"target_user_id": m_id}, target_role=None)
            else:
                cursor.execute("""
                    DELETE FROM streaming_assignments 
                    WHERE schedule_date = %s AND schedule_time = %s AND role_name = %s
                """, (d_str, t_str, r_str))
        
        conn.commit()
        return jsonify({"success": True, "message": "Penugasan berhasil disimpan"})
    except Exception as e:
        conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 400
    finally:
        cursor.close()
        conn.close()


@app.route("/api/misa-besar/public", methods=["GET"])
def api_misa_besar_public():
    """Daftar Misa Besar untuk halaman Jadwal Streaming anggota.

    Endpoint ini sengaja hanya mengembalikan Misa Besar berstatus published,
    sehingga anggota biasa tidak menerima data draft. Response juga menyertakan
    penanda jika user sesi saat ini sedang bertugas pada misa tersebut.
    """
    ensure_misa_besar_schema()
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True, buffered=True)
    try:
        month = parse_optional_int(request.args.get("month"))
        year = parse_optional_int(request.args.get("year"))
        current_user_id = normalize_text(session.get("user_id"))

        query = """
            SELECT id,
                   misa_name AS misaName,
                   DATE_FORMAT(misa_date, '%Y-%m-%d') AS misaDate,
                   DATE_FORMAT(misa_time, '%H:%i') AS misaTime,
                   misa_note AS misaNote,
                   allow_member_request AS allowMemberRequest,
                   status,
                   created_at AS updatedAt
            FROM misa_besar
            WHERE status = 'published'
        """
        params: list[object] = []
        if month:
            query += " AND MONTH(misa_date) = %s"
            params.append(month)
        if year:
            query += " AND YEAR(misa_date) = %s"
            params.append(year)
        query += " ORDER BY misa_date ASC, misa_time ASC, id ASC"

        cursor.execute(query, tuple(params))
        events = cursor.fetchall() or []

        for ev in events:
            cursor.execute(
                """
                SELECT id, role_name AS role, required_count AS count
                FROM misa_besar_names
                WHERE misa_id = %s
                ORDER BY id ASC
                """,
                (ev["id"],),
            )
            roles = cursor.fetchall() or []
            current_user_roles: list[str] = []
            for role in roles:
                cursor.execute(
                    """
                    SELECT a.id, a.nama AS name
                    FROM misa_besar_assignments bma
                    JOIN anggota a ON bma.member_id = a.id
                    WHERE bma.role_id = %s
                    ORDER BY a.nama ASC
                    """,
                    (role["id"],),
                )
                members_data = cursor.fetchall() or []
                role["members"] = [str(member.get("id")) for member in members_data if member.get("id") is not None]
                role["memberIds"] = role["members"]
                role["memberNames"] = [normalize_text(member.get("name")) for member in members_data if normalize_text(member.get("name"))]
                role["count"] = parse_required_int(role.get("count"), 1)
                if current_user_id and current_user_id in set(role["memberIds"]):
                    current_user_roles.append(normalize_text(role.get("role")) or "Role")

            ev["allowMemberRequest"] = bool(ev.get("allowMemberRequest"))
            ev["status"] = "published"
            ev["roles"] = roles
            ev["currentUserRoles"] = current_user_roles
            ev["isCurrentUserAssigned"] = bool(current_user_roles)

        return jsonify({"success": True, "items": events})
    finally:
        cursor.close()
        conn.close()


@app.route("/api/misa-besar", methods=["GET", "POST"])
def api_misa_besar():
    ensure_misa_besar_schema()
    ensure_streaming_schema()
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True, buffered=True)
    try:
        if request.method == "POST":
            data = request.json
            allow_req = 1 if data.get('allowMemberRequest') else 0
            cursor.execute("""
                INSERT INTO misa_besar (misa_name, misa_date, misa_time, misa_note, allow_member_request, status)
                VALUES (%s, %s, %s, %s, %s, %s)
            """, (data['misaName'], data['misaDate'], data['misaTime'], data.get('misaNote',''), allow_req, data['status']))
            misa_id = cursor.lastrowid
            
            for r in data.get('roles', []):
                cursor.execute("INSERT INTO misa_besar_names (misa_id, role_name, required_count) VALUES (%s, %s, %s)",
                               (misa_id, r['role'], r['count']))
                role_id = cursor.lastrowid
                
                # Simpan list member ID langsung (yang tidak kosong)
                for m_id in r.get('members', []):
                    if str(m_id).strip():
                        cursor.execute("INSERT IGNORE INTO misa_besar_assignments (role_id, member_id) VALUES (%s, %s)", (role_id, m_id))

            removed_regular_assignments = clear_streaming_assignments_for_published_misa_besar(
                cursor, data.get('misaDate'), data.get('misaTime'), data.get('status')
            )
            new_snapshot = fetch_misa_besar_notification_snapshot(cursor, misa_id)
            notified_count = notify_misa_besar_assignment_changes(cursor, None, new_snapshot)
            open_role_notified_count = notify_misa_besar_open_roles_if_newly_published(cursor, None, new_snapshot)
            
            conn.commit()
            return jsonify({
                "success": True,
                "id": misa_id,
                "notifiedCount": notified_count,
                "openRoleNotifiedCount": open_role_notified_count,
                "removedRegularAssignments": removed_regular_assignments,
            })

        # GET DATA
        cursor.execute("SELECT id, misa_name as misaName, DATE_FORMAT(misa_date, '%Y-%m-%d') as misaDate, DATE_FORMAT(misa_time, '%H:%i') as misaTime, misa_note as misaNote, allow_member_request as allowMemberRequest, status, created_at as updatedAt FROM misa_besar ORDER BY misa_date DESC")
        events = cursor.fetchall()
        
        for ev in events:
            cursor.execute("SELECT id, role_name as role, required_count as count FROM misa_besar_names WHERE misa_id = %s", (ev['id'],))
            roles = cursor.fetchall()
            for r in roles:
                cursor.execute("""
                    SELECT a.id, a.nama as name 
                    FROM misa_besar_assignments bma 
                    JOIN anggota a ON bma.member_id = a.id 
                    WHERE bma.role_id = %s
                """, (r['id'],))
                members_data = cursor.fetchall()
                # Sesuaikan response dengan format Frontend
                r['members'] = [str(m['id']) for m in members_data]
                r['memberNames'] = [m['name'] for m in members_data]
                r['multi'] = r['count'] > 1
            ev['roles'] = roles
            
        return jsonify(events)
    finally:
        cursor.close()
        conn.close()

@app.route("/api/misa-besar/<int:misa_id>", methods=["PUT", "DELETE"])
def api_misa_besar_detail(misa_id):
    ensure_misa_besar_schema()
    ensure_streaming_schema()
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True, buffered=True)
    try:
        if request.method == "DELETE":
            cursor.execute("DELETE FROM misa_besar WHERE id = %s", (misa_id,))
            conn.commit()
            return jsonify({"success": True})
        
        if request.method == "PUT":
            data = request.json
            old_snapshot = fetch_misa_besar_notification_snapshot(cursor, misa_id)
            allow_req = 1 if data.get('allowMemberRequest') else 0
            
            # Update header Misa
            cursor.execute("""
                UPDATE misa_besar SET misa_name=%s, misa_date=%s, misa_time=%s, misa_note=%s, allow_member_request=%s, status=%s
                WHERE id=%s
            """, (data['misaName'], data['misaDate'], data['misaTime'], data.get('misaNote',''), allow_req, data['status'], misa_id))
            
            # Reset roles untuk mempermudah handling list members tanpa error FK constraint
            # Karena tabel misa_besar_names punya constraint ON DELETE CASCADE, assignments otomatis terhapus
            cursor.execute("DELETE FROM misa_besar_names WHERE misa_id = %s", (misa_id,))
            
            # Insert Ulang Roles dan Assignments
            for r in data.get('roles', []):
                cursor.execute("INSERT INTO misa_besar_names (misa_id, role_name, required_count) VALUES (%s, %s, %s)",
                               (misa_id, r['role'], r['count']))
                role_id = cursor.lastrowid
                
                for m_id in r.get('members', []):
                    if str(m_id).strip():
                        cursor.execute("INSERT IGNORE INTO misa_besar_assignments (role_id, member_id) VALUES (%s, %s)", (role_id, m_id))

            removed_regular_assignments = clear_streaming_assignments_for_published_misa_besar(
                cursor, data.get('misaDate'), data.get('misaTime'), data.get('status')
            )
            new_snapshot = fetch_misa_besar_notification_snapshot(cursor, misa_id)
            notified_count = notify_misa_besar_assignment_changes(cursor, old_snapshot, new_snapshot)
            open_role_notified_count = notify_misa_besar_open_roles_if_newly_published(cursor, old_snapshot, new_snapshot)
            
            conn.commit()
            return jsonify({
                "success": True,
                "notifiedCount": notified_count,
                "openRoleNotifiedCount": open_role_notified_count,
                "removedRegularAssignments": removed_regular_assignments,
            })
    finally:
        cursor.close()
        conn.close()

@app.route("/api/misa-besar/<int:misa_id>/status", methods=["PUT"])
def api_misa_besar_status(misa_id):
    """Endpoint khusus untuk mengubah status ke Draft atau Published dengan cepat"""
    ensure_misa_besar_schema()
    ensure_streaming_schema()
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True, buffered=True)
    try:
        data = request.json or {}
        old_snapshot = fetch_misa_besar_notification_snapshot(cursor, misa_id)
        new_status = normalize_misa_besar_status(data.get('status', 'draft'))
        
        cursor.execute("UPDATE misa_besar SET status = %s WHERE id = %s", (new_status, misa_id))
        removed_regular_assignments = 0
        if new_status == "published":
            cursor.execute(
                """
                SELECT DATE_FORMAT(misa_date, '%Y-%m-%d') AS misaDate,
                       DATE_FORMAT(misa_time, '%H:%i') AS misaTime
                FROM misa_besar
                WHERE id = %s
                LIMIT 1
                """,
                (misa_id,),
            )
            event_row = cursor.fetchone()
            if event_row:
                removed_regular_assignments = clear_streaming_assignments_for_published_misa_besar(
                    cursor, event_row.get("misaDate"), event_row.get("misaTime"), new_status
                )
        new_snapshot = fetch_misa_besar_notification_snapshot(cursor, misa_id)
        notified_count = notify_misa_besar_assignment_changes(cursor, old_snapshot, new_snapshot)
        open_role_notified_count = notify_misa_besar_open_roles_if_newly_published(cursor, old_snapshot, new_snapshot)
        conn.commit()
        return jsonify({
            "success": True,
            "notifiedCount": notified_count,
            "openRoleNotifiedCount": open_role_notified_count,
            "removedRegularAssignments": removed_regular_assignments,
        })
    except Exception as e:
        conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        cursor.close()
        conn.close()

@app.route("/api/misa-besar/assign", methods=["POST"])
def api_misa_besar_assign():
    ensure_misa_besar_schema()
    ensure_streaming_schema()
    data = request.json or {}
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True, buffered=True)
    try:
        old_snapshot = fetch_misa_besar_notification_snapshot(cursor, int(data['misaId']))
        cursor.execute("""
            SELECT n.role_name FROM misa_besar_assignments a 
            JOIN misa_besar_names n ON a.role_id = n.id 
            WHERE n.misa_id = %s AND a.member_id = %s
        """, (data['misaId'], data['memberId']))
        existing = cursor.fetchone()
        if existing:
            return jsonify({"success": False, "message": f"Gagal: Orang tersebut sudah bertugas sebagai {existing['role_name']} di jadwal ini."}), 400

        cursor.execute("INSERT INTO misa_besar_assignments (role_id, member_id) VALUES (%s, %s)", (data['roleId'], data['memberId']))
        removed_regular_assignments = 0
        new_snapshot = fetch_misa_besar_notification_snapshot(cursor, int(data['misaId']))
        if new_snapshot:
            removed_regular_assignments = clear_streaming_assignments_for_published_misa_besar(
                cursor, new_snapshot.get("misaDate"), new_snapshot.get("misaTime"), new_snapshot.get("status")
            )
        notified_count = notify_misa_besar_assignment_changes(cursor, old_snapshot, new_snapshot)
        conn.commit()
        return jsonify({
            "success": True,
            "notifiedCount": notified_count,
            "removedRegularAssignments": removed_regular_assignments,
        })
    finally:
        cursor.close()
        conn.close()

@app.route("/api/misa-besar/unassign", methods=["POST"])
def api_misa_besar_unassign():
    data = request.json
    conn = mysql_connection()
    cursor = conn.cursor(buffered=True)
    try:
        cursor.execute("DELETE FROM misa_besar_assignments WHERE role_id = %s AND member_id = %s", (data['roleId'], data['memberId']))
        conn.commit()
        return jsonify({"success": True})
    finally:
        cursor.close()
        conn.close()


# -----------------------------------------------------------------------------
# Request Tugas Anggota: self-service assignment for Misa Biasa & Misa Besar
# -----------------------------------------------------------------------------

def ensure_task_request_schema(cursor=None) -> None:
    """Pastikan tabel penugasan memiliki metadata untuk request anggota."""
    ensure_streaming_schema()
    ensure_misa_besar_schema()

    owns_connection = cursor is None
    conn = None
    if owns_connection:
        conn = mysql_connection()
        cursor = conn.cursor(buffered=True)

    try:
        ensure_column(cursor, "streaming_assignments", "created_at", "`created_at` datetime NOT NULL DEFAULT CURRENT_TIMESTAMP")
        ensure_column(cursor, "streaming_assignments", "request_source", "`request_source` varchar(30) NOT NULL DEFAULT 'admin'")
        ensure_column(cursor, "misa_besar_assignments", "created_at", "`created_at` datetime NOT NULL DEFAULT CURRENT_TIMESTAMP")
        ensure_column(cursor, "misa_besar_assignments", "request_source", "`request_source` varchar(30) NOT NULL DEFAULT 'admin'")
        if conn:
            conn.commit()
    finally:
        if owns_connection:
            cursor.close()
            conn.close()


def require_active_request_member(cursor):
    if not session.get("logged_in"):
        return None, (jsonify({"success": False, "error": "Anda harus login terlebih dahulu."}), 401)

    user_id = session.get("user_id")
    if not user_id:
        return None, (jsonify({"success": False, "error": "Sesi tidak valid."}), 401)

    cursor.execute(
        """
        SELECT id, nama, role, status_akun
        FROM anggota
        WHERE id = %s
        LIMIT 1
        """,
        (user_id,),
    )
    member = cursor.fetchone()
    if not member:
        return None, (jsonify({"success": False, "error": "Akun anggota tidak ditemukan."}), 404)

    if normalize_role_value(member.get("role") or "user") != "user":
        return None, (jsonify({"success": False, "error": "Halaman request tugas hanya untuk anggota biasa."}), 403)

    if normalize_status(member.get("status_akun") or "aktif") != "aktif":
        return None, (jsonify({"success": False, "error": "Akun Anda sedang nonaktif."}), 403)

    member["id"] = str(member.get("id"))
    member["name"] = normalize_text(member.get("nama")) or "Anggota"
    return member, None


def request_task_is_past(date_text: str, time_text: str = "00:00") -> bool:
    try:
        schedule_dt = datetime.strptime(f"{date_text} {format_time_hhmm(time_text)}", "%Y-%m-%d %H:%M")
        return schedule_dt < datetime.now()
    except Exception:
        try:
            return datetime.strptime(date_text, "%Y-%m-%d").date() < datetime.now().date()
        except Exception:
            return False


def request_task_day_name(date_text: str) -> str:
    try:
        return DAYS_INDO[datetime.strptime(date_text, "%Y-%m-%d").weekday()]
    except Exception:
        return "-"


def request_task_format_date(date_text: str) -> str:
    try:
        return datetime.strptime(date_text, "%Y-%m-%d").strftime("%d/%m/%Y")
    except Exception:
        return normalize_text(date_text) or "-"


def request_task_schedule_key(kind: str, date_text: str, time_text: str, misa_id: object | None = None) -> str:
    if kind == "besar":
        return f"besar:{misa_id}"
    return f"biasa:{date_text}:{format_time_hhmm(time_text)}"


def request_task_get_regular_cfg(cursor, date_text: str, time_text: str):
    day_name = request_task_day_name(date_text)
    cursor.execute(
        """
        SELECT mass_name, day_name, DATE_FORMAT(start_time, '%H:%i') AS start_time
        FROM streaming_weekly_config
        WHERE day_name = %s AND DATE_FORMAT(start_time, '%H:%i') = %s
        LIMIT 1
        """,
        (day_name, format_time_hhmm(time_text)),
    )
    return cursor.fetchone()


def request_task_regular_slot_blocked(cursor, date_text: str, time_text: str) -> str:
    time_text = format_time_hhmm(time_text)
    cursor.execute(
        """
        SELECT 1 FROM streaming_cancelled
        WHERE mass_date = %s AND DATE_FORMAT(mass_time, '%H:%i') = %s
        LIMIT 1
        """,
        (date_text, time_text),
    )
    if cursor.fetchone():
        return "Jadwal ditiadakan."

    cursor.execute(
        """
        SELECT misa_name FROM misa_besar
        WHERE status = 'published'
          AND misa_date = %s
          AND DATE_FORMAT(misa_time, '%H:%i') = %s
        LIMIT 1
        """,
        (date_text, time_text),
    )
    row = cursor.fetchone()
    if row:
        return f"Bentrok dengan Misa Besar: {normalize_text(row.get('misa_name')) or 'Misa Besar'}."
    return ""


def request_task_fetch_roles(cursor) -> list[str]:
    cursor.execute("SELECT role_name AS name FROM streaming_roles ORDER BY order_index ASC, id ASC")
    return [normalize_text(row.get("name")) for row in (cursor.fetchall() or []) if normalize_text(row.get("name"))]


def request_task_regular_items(cursor, month: int, year: int, member_id: str) -> list[dict[str, object]]:
    roles = request_task_fetch_roles(cursor)
    cursor.execute("SELECT day_name, mass_name, DATE_FORMAT(start_time, '%H:%i') AS start_time FROM streaming_weekly_config ORDER BY start_time ASC, id ASC")
    weekly_configs = cursor.fetchall() or []

    cursor.execute(
        """
        SELECT DATE_FORMAT(schedule_date, '%Y-%m-%d') AS schedule_date,
               DATE_FORMAT(schedule_time, '%H:%i') AS schedule_time,
               role_name, member_id, COALESCE(a.nama, CONCAT('ID ', sa.member_id)) AS member_name
        FROM streaming_assignments sa
        LEFT JOIN anggota a ON a.id = sa.member_id
        WHERE MONTH(schedule_date) = %s AND YEAR(schedule_date) = %s
        """,
        (month, year),
    )
    assignments: dict[tuple[str, str], dict[str, dict[str, str]]] = {}
    for row in cursor.fetchall() or []:
        key = (row.get("schedule_date"), format_time_hhmm(row.get("schedule_time")))
        assignments.setdefault(key, {})[normalize_text(row.get("role_name"))] = {
            "memberId": str(row.get("member_id") or ""),
            "memberName": normalize_text(row.get("member_name")),
        }

    items: list[dict[str, object]] = []
    num_days = calendar.monthrange(year, month)[1]
    for day in range(1, num_days + 1):
        date_obj = datetime(year, month, day)
        date_text = date_obj.strftime("%Y-%m-%d")
        day_name = DAYS_INDO[date_obj.weekday()]
        for cfg in weekly_configs:
            if normalize_text(cfg.get("day_name")) != day_name:
                continue
            time_text = format_time_hhmm(cfg.get("start_time"))
            blocked_reason = request_task_regular_slot_blocked(cursor, date_text, time_text)
            if blocked_reason:
                continue

            key = (date_text, time_text)
            slot_assignments = assignments.get(key, {})
            role_details = []
            open_roles = []
            current_user_roles = []
            for role_name in roles:
                assigned = slot_assignments.get(role_name) or {}
                member_id_text = normalize_text(assigned.get("memberId"))
                member_name = normalize_text(assigned.get("memberName"))
                filled = bool(member_id_text)
                if not filled:
                    open_roles.append({"role": role_name})
                elif member_id and str(member_id) == member_id_text:
                    current_user_roles.append(role_name)
                role_details.append({
                    "role": role_name,
                    "filled": filled,
                    "memberId": member_id_text,
                    "memberName": member_name,
                })

            is_past = request_task_is_past(date_text, time_text)
            is_current_user_assigned = bool(current_user_roles)
            can_request = bool(open_roles) and not is_current_user_assigned and not is_past
            status_label = "Terbuka" if can_request else "Tertutup"
            if is_current_user_assigned:
                reason = "Anda sudah bertugas di jadwal ini."
            elif is_past:
                reason = "Jadwal sudah lewat."
            elif not open_roles:
                reason = "Semua role sudah terisi."
            else:
                reason = "Masih ada role kosong."

            items.append({
                "id": request_task_schedule_key("biasa", date_text, time_text),
                "type": "biasa",
                "typeLabel": "Misa Biasa",
                "misaName": normalize_text(cfg.get("mass_name")) or "Misa Biasa",
                "date": date_text,
                "dateLabel": request_task_format_date(date_text),
                "dayName": day_name,
                "time": time_text,
                "roles": role_details,
                "openRoles": open_roles,
                "currentUserRoles": current_user_roles,
                "isCurrentUserAssigned": is_current_user_assigned,
                "canRequest": can_request,
                "status": status_label,
                "statusReason": reason,
            })

    items.sort(key=lambda item: (item.get("date"), item.get("time")))
    return items


def request_task_big_items(cursor, month: int, year: int, member_id: str) -> list[dict[str, object]]:
    cursor.execute(
        """
        SELECT id, misa_name, DATE_FORMAT(misa_date, '%Y-%m-%d') AS misa_date,
               DATE_FORMAT(misa_time, '%H:%i') AS misa_time, misa_note,
               allow_member_request, status
        FROM misa_besar
        WHERE status = 'published' AND MONTH(misa_date) = %s AND YEAR(misa_date) = %s
        ORDER BY misa_date ASC, misa_time ASC, id ASC
        """,
        (month, year),
    )
    events = cursor.fetchall() or []
    items: list[dict[str, object]] = []
    for event in events:
        misa_id = event.get("id")
        cursor.execute(
            """
            SELECT n.id AS role_id, n.role_name, n.required_count,
                   a.member_id, COALESCE(ag.nama, CONCAT('ID ', a.member_id)) AS member_name
            FROM misa_besar_names n
            LEFT JOIN misa_besar_assignments a ON a.role_id = n.id
            LEFT JOIN anggota ag ON ag.id = a.member_id
            WHERE n.misa_id = %s
            ORDER BY n.id ASC, a.id ASC
            """,
            (misa_id,),
        )
        grouped: dict[str, dict[str, object]] = {}
        for row in cursor.fetchall() or []:
            role_id = str(row.get("role_id") or "")
            if not role_id:
                continue
            if role_id not in grouped:
                grouped[role_id] = {
                    "roleId": role_id,
                    "role": normalize_text(row.get("role_name")) or "Role",
                    "requiredCount": max(1, parse_required_int(row.get("required_count"), 1)),
                    "members": [],
                }
            if row.get("member_id") not in (None, ""):
                grouped[role_id]["members"].append({
                    "memberId": str(row.get("member_id")),
                    "memberName": normalize_text(row.get("member_name")),
                })

        role_details = []
        open_roles = []
        current_user_roles = []
        for role in grouped.values():
            members = role.get("members") or []
            filled_count = len(members)
            required_count = max(1, parse_required_int(role.get("requiredCount"), 1))
            empty_count = max(required_count - filled_count, 0)
            for member_row in members:
                if member_id and str(member_row.get("memberId")) == str(member_id):
                    current_user_roles.append(normalize_text(role.get("role")) or "Role")
            if empty_count > 0:
                open_roles.append({
                    "role": normalize_text(role.get("role")) or "Role",
                    "roleId": role.get("roleId"),
                    "emptyCount": empty_count,
                })
            role_details.append({
                "roleId": role.get("roleId"),
                "role": normalize_text(role.get("role")) or "Role",
                "requiredCount": required_count,
                "filledCount": filled_count,
                "emptyCount": empty_count,
                "members": members,
            })

        date_text = normalize_text(event.get("misa_date"))
        time_text = format_time_hhmm(event.get("misa_time"))
        allow_request = bool(event.get("allow_member_request"))
        is_past = request_task_is_past(date_text, time_text)
        is_current_user_assigned = bool(current_user_roles)
        can_request = allow_request and bool(open_roles) and not is_current_user_assigned and not is_past
        if is_current_user_assigned:
            reason = "Anda sudah bertugas di misa ini."
        elif is_past:
            reason = "Jadwal sudah lewat."
        elif not allow_request:
            reason = "Request ditutup admin."
        elif not open_roles:
            reason = "Semua role sudah terisi."
        else:
            reason = "Masih ada role kosong."

        items.append({
            "id": request_task_schedule_key("besar", date_text, time_text, misa_id),
            "type": "besar",
            "typeLabel": "Misa Besar",
            "misaId": misa_id,
            "misaName": normalize_text(event.get("misa_name")) or "Misa Besar",
            "date": date_text,
            "dateLabel": request_task_format_date(date_text),
            "dayName": request_task_day_name(date_text),
            "time": time_text,
            "note": normalize_text(event.get("misa_note")),
            "allowMemberRequest": allow_request,
            "roles": role_details,
            "openRoles": open_roles,
            "currentUserRoles": current_user_roles,
            "isCurrentUserAssigned": is_current_user_assigned,
            "canRequest": can_request,
            "status": "Terbuka" if can_request else "Tertutup",
            "statusReason": reason,
        })
    return items


@app.route("/api/request-tugas/me", methods=["GET"])
def api_request_tugas_me():
    ensure_auth_schema()
    viewer = current_user_context()
    return jsonify({"success": True, "currentUser": viewer})


@app.route("/api/request-tugas/schedules", methods=["GET"])
def api_request_tugas_schedules():
    ensure_task_request_schema()
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True, buffered=True)
    try:
        member, error = require_active_request_member(cursor)
        if error:
            return error

        month = parse_required_int(request.args.get("month"), datetime.now().month)
        year = parse_required_int(request.args.get("year"), datetime.now().year)
        kind = normalize_text(request.args.get("type")) or "biasa"
        status_filter = (normalize_text(request.args.get("slot")) or "all").lower()
        search_text = normalize_text(request.args.get("search")).lower()

        if kind == "besar":
            items = request_task_big_items(cursor, month, year, member["id"])
        elif kind == "all":
            items = request_task_regular_items(cursor, month, year, member["id"]) + request_task_big_items(cursor, month, year, member["id"])
        else:
            kind = "biasa"
            items = request_task_regular_items(cursor, month, year, member["id"])

        if status_filter in {"open", "terbuka"}:
            items = [item for item in items if item.get("canRequest")]
        elif status_filter in {"closed", "tertutup"}:
            items = [item for item in items if not item.get("canRequest")]

        if search_text:
            def haystack(item):
                role_texts = []
                for role in item.get("roles") or []:
                    role_texts.append(normalize_text(role.get("role")))
                    for member_row in role.get("members") or []:
                        role_texts.append(normalize_text(member_row.get("memberName")))
                    if role.get("memberName"):
                        role_texts.append(normalize_text(role.get("memberName")))
                return " ".join([
                    normalize_text(item.get("typeLabel")),
                    normalize_text(item.get("misaName")),
                    normalize_text(item.get("dateLabel")),
                    normalize_text(item.get("dayName")),
                    normalize_text(item.get("time")),
                    normalize_text(item.get("statusReason")),
                    " ".join(role_texts),
                ]).lower()
            items = [item for item in items if search_text in haystack(item)]

        items.sort(key=lambda item: (item.get("date") or "", item.get("time") or "", item.get("misaName") or ""))
        return jsonify({
            "success": True,
            "currentUser": {"id": member["id"], "name": member["name"]},
            "items": items,
        })
    finally:
        cursor.close()
        conn.close()


@app.route("/api/request-tugas/claim", methods=["POST"])
def api_request_tugas_claim():
    ensure_task_request_schema()
    ensure_notifications_schema()
    data = request.get_json(silent=True) or {}
    kind = normalize_text(data.get("type"))
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True, buffered=True)
    try:
        member, error = require_active_request_member(cursor)
        if error:
            return error
        member_id = member["id"]

        if kind == "besar":
            misa_id = parse_required_int(data.get("misaId"), 0)
            role_id = parse_required_int(data.get("roleId"), 0)
            if not misa_id or not role_id:
                return jsonify({"success": False, "error": "Jadwal Misa Besar dan role wajib dipilih."}), 400

            cursor.execute(
                """
                SELECT mb.id, mb.misa_name, DATE_FORMAT(mb.misa_date, '%Y-%m-%d') AS misa_date,
                       DATE_FORMAT(mb.misa_time, '%H:%i') AS misa_time, mb.allow_member_request, mb.status,
                       n.id AS role_id, n.role_name, n.required_count
                FROM misa_besar mb
                JOIN misa_besar_names n ON n.misa_id = mb.id
                WHERE mb.id = %s AND n.id = %s
                LIMIT 1
                """,
                (misa_id, role_id),
            )
            event = cursor.fetchone()
            if not event:
                return jsonify({"success": False, "error": "Jadwal atau role Misa Besar tidak ditemukan."}), 404
            if normalize_misa_besar_status(event.get("status")) != "published":
                return jsonify({"success": False, "error": "Misa Besar belum dipublish."}), 400
            if not bool(event.get("allow_member_request")):
                return jsonify({"success": False, "error": "Request anggota untuk Misa Besar ini sedang ditutup."}), 400
            if request_task_is_past(event.get("misa_date"), event.get("misa_time")):
                return jsonify({"success": False, "error": "Jadwal Misa Besar sudah lewat."}), 400

            cursor.execute(
                """
                SELECT n.role_name
                FROM misa_besar_assignments a
                JOIN misa_besar_names n ON n.id = a.role_id
                WHERE n.misa_id = %s AND a.member_id = %s
                LIMIT 1
                """,
                (misa_id, member_id),
            )
            existing = cursor.fetchone()
            if existing:
                return jsonify({"success": False, "error": f"Anda sudah bertugas sebagai {existing.get('role_name')} di Misa Besar ini."}), 409

            cursor.execute("SELECT COUNT(*) AS filled FROM misa_besar_assignments WHERE role_id = %s", (role_id,))
            filled = parse_required_int((cursor.fetchone() or {}).get("filled"), 0)
            required_count = max(1, parse_required_int(event.get("required_count"), 1))
            if filled >= required_count:
                return jsonify({"success": False, "error": "Slot role ini sudah penuh."}), 409

            cursor.execute(
                """
                INSERT INTO misa_besar_assignments (role_id, member_id, request_source, created_at)
                VALUES (%s, %s, 'member_request', CURRENT_TIMESTAMP)
                """,
                (role_id, member_id),
            )
            message = f"Anda berhasil terdaftar sebagai {event.get('role_name')} untuk {event.get('misa_name')}."
            create_task_success_notification(
                cursor,
                member_id=member_id,
                member_role=member.get("role") or "user",
                misa_type="misa_besar",
                misa_name=event.get("misa_name") or "Misa Besar",
                role_name=event.get("role_name") or "Role",
                date_text=event.get("misa_date"),
                time_text=event.get("misa_time"),
                source="member_request",
                misa_besar_id=misa_id,
            )

        else:
            kind = "biasa"
            date_text = parse_optional_date(data.get("date"))
            time_text = format_time_hhmm(data.get("time"))
            role_name = normalize_text(data.get("role"))
            if not date_text or not time_text or not role_name:
                return jsonify({"success": False, "error": "Jadwal dan role wajib dipilih."}), 400

            cfg = request_task_get_regular_cfg(cursor, date_text, time_text)
            if not cfg:
                return jsonify({"success": False, "error": "Jadwal Misa Biasa tidak ditemukan."}), 404
            blocked_reason = request_task_regular_slot_blocked(cursor, date_text, time_text)
            if blocked_reason:
                return jsonify({"success": False, "error": blocked_reason}), 400
            if request_task_is_past(date_text, time_text):
                return jsonify({"success": False, "error": "Jadwal Misa Biasa sudah lewat."}), 400
            cursor.execute("SELECT 1 FROM streaming_roles WHERE role_name = %s LIMIT 1", (role_name,))
            if not cursor.fetchone():
                return jsonify({"success": False, "error": "Role tidak ditemukan di konfigurasi jadwal streaming."}), 404

            cursor.execute(
                """
                SELECT role_name FROM streaming_assignments
                WHERE schedule_date = %s AND DATE_FORMAT(schedule_time, '%H:%i') = %s AND member_id = %s
                LIMIT 1
                """,
                (date_text, time_text, member_id),
            )
            existing = cursor.fetchone()
            if existing:
                return jsonify({"success": False, "error": f"Anda sudah bertugas sebagai {existing.get('role_name')} di jadwal ini."}), 409

            cursor.execute(
                """
                SELECT member_id FROM streaming_assignments
                WHERE schedule_date = %s AND DATE_FORMAT(schedule_time, '%H:%i') = %s AND role_name = %s
                LIMIT 1
                """,
                (date_text, time_text, role_name),
            )
            if cursor.fetchone():
                return jsonify({"success": False, "error": "Slot role ini sudah terisi."}), 409

            cursor.execute(
                """
                INSERT INTO streaming_assignments (schedule_date, schedule_time, role_name, member_id, request_source, created_at)
                VALUES (%s, %s, %s, %s, 'member_request', CURRENT_TIMESTAMP)
                """,
                (date_text, time_text, role_name, member_id),
            )
            message = f"Anda berhasil terdaftar sebagai {role_name} untuk {cfg.get('mass_name') or 'Misa Biasa'}."
            create_task_success_notification(
                cursor,
                member_id=member_id,
                member_role=member.get("role") or "user",
                misa_type="misa_biasa",
                misa_name=cfg.get("mass_name") or "Misa Biasa",
                role_name=role_name,
                date_text=date_text,
                time_text=time_text,
                source="member_request",
            )

        conn.commit()
        return jsonify({"success": True, "message": message})
    except Exception as exc:
        conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 400
    finally:
        cursor.close()
        conn.close()




# -----------------------------------------------------------------------------
# Riwayat Tugas Saya: histori tugas anggota dari jadwal Misa Biasa & Misa Besar
# -----------------------------------------------------------------------------

@app.route("/api/riwayat-tugas-saya", methods=["GET"])
def api_riwayat_tugas_saya():
    ensure_task_request_schema()
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True, buffered=True)
    try:
        member, error = require_active_request_member(cursor)
        if error:
            return error
        member_id = member["id"]

        page = max(1, parse_required_int(request.args.get("page"), 1))
        page_size = max(1, min(25, parse_required_int(request.args.get("pageSize"), 5)))
        kind_filter = (normalize_text(request.args.get("type")) or "all").lower()
        month_filter = normalize_text(request.args.get("month")) or "all"
        year_filter = normalize_text(request.args.get("year")) or "all"
        search_text = normalize_text(request.args.get("search")).lower()
        sort_mode = (normalize_text(request.args.get("sort")) or "date_desc").lower().replace("-", "_")
        stats_range = (normalize_text(request.args.get("statsRange")) or "month").lower()

        def source_label(source_value: object) -> str:
            source = normalize_text(source_value).lower()
            if source in {"member_request", "request", "mandiri", "self"}:
                return "Request Mandiri"
            if source in {"exchange", "swap", "tukar", "replacement", "replace", "ganti", "pengganti"}:
                return "Penukaran / Pengganti"
            return "Ditugaskan Admin"

        def status_for(date_text: str, time_text: str, cancelled: bool = False) -> str:
            if cancelled:
                return "Dibatalkan"
            try:
                schedule_dt = datetime.strptime(f"{date_text} {format_time_hhmm(time_text)}", "%Y-%m-%d %H:%M")
                return "Selesai" if schedule_dt < datetime.now() else "Terdaftar"
            except Exception:
                return "Terdaftar"

        def build_item(kind: str, type_label: str, misa_name: str, date_text: str, time_text: str, role_name: str, request_source: object, created_at: object = None, *, cancelled: bool = False, note: str = "") -> dict[str, object]:
            time_clean = format_time_hhmm(time_text)
            source = normalize_text(request_source) or "admin"
            status = status_for(date_text, time_clean, cancelled)
            day_name = request_task_day_name(date_text)
            created_iso = created_at.isoformat() if hasattr(created_at, "isoformat") else (normalize_text(created_at) or None)
            return {
                "type": kind,
                "typeLabel": type_label,
                "misaName": normalize_text(misa_name) or type_label,
                "date": normalize_text(date_text),
                "dateLabel": request_task_format_date(date_text),
                "dayName": day_name,
                "time": time_clean,
                "role": normalize_text(role_name) or "-",
                "status": status,
                "source": source,
                "sourceLabel": source_label(source),
                "isExchange": source_label(source) == "Penukaran / Pengganti",
                "createdAt": created_iso,
                "note": normalize_text(note) or "Nama Anda tercatat pada jadwal tugas streaming.",
                "scheduleLabel": f"{day_name}, {request_task_format_date(date_text)} jam {time_clean} WIB",
            }

        rows: list[dict[str, object]] = []

        cursor.execute(
            """
            SELECT DATE_FORMAT(sa.schedule_date, '%Y-%m-%d') AS date,
                   DATE_FORMAT(sa.schedule_time, '%H:%i') AS time,
                   sa.role_name AS role,
                   COALESCE(sa.request_source, 'admin') AS request_source,
                   sa.created_at,
                   cfg.mass_name
            FROM streaming_assignments sa
            LEFT JOIN streaming_weekly_config cfg
              ON cfg.day_name = CASE WEEKDAY(sa.schedule_date)
                WHEN 0 THEN 'Senin' WHEN 1 THEN 'Selasa' WHEN 2 THEN 'Rabu'
                WHEN 3 THEN 'Kamis' WHEN 4 THEN 'Jumat' WHEN 5 THEN 'Sabtu'
                ELSE 'Minggu' END
              AND DATE_FORMAT(cfg.start_time, '%H:%i') = DATE_FORMAT(sa.schedule_time, '%H:%i')
            WHERE sa.member_id = %s
            """,
            (member_id,),
        )
        for row in cursor.fetchall() or []:
            rows.append(build_item(
                "biasa",
                "Misa Biasa",
                row.get("mass_name") or "Misa Biasa",
                normalize_text(row.get("date")),
                format_time_hhmm(row.get("time")),
                row.get("role"),
                row.get("request_source"),
                row.get("created_at"),
            ))

        cursor.execute(
            """
            SELECT mb.misa_name, DATE_FORMAT(mb.misa_date, '%Y-%m-%d') AS date,
                   DATE_FORMAT(mb.misa_time, '%H:%i') AS time,
                   n.role_name AS role,
                   COALESCE(a.request_source, 'admin') AS request_source,
                   a.created_at
            FROM misa_besar_assignments a
            JOIN misa_besar_names n ON n.id = a.role_id
            JOIN misa_besar mb ON mb.id = n.misa_id
            WHERE a.member_id = %s AND mb.status = 'published'
            """,
            (member_id,),
        )
        for row in cursor.fetchall() or []:
            rows.append(build_item(
                "besar",
                "Misa Besar",
                row.get("misa_name") or "Misa Besar",
                normalize_text(row.get("date")),
                format_time_hhmm(row.get("time")),
                row.get("role"),
                row.get("request_source"),
                row.get("created_at"),
            ))

        # Riwayat pembatalan ikut ditampilkan sebagai status Dibatalkan agar histori tugas tetap lengkap.
        try:
            ensure_task_cancellation_schema(cursor)
            cursor.execute(
                """
                SELECT kind, COALESCE(type_label, IF(kind='besar','Misa Besar','Misa Biasa')) AS type_label,
                       DATE_FORMAT(schedule_date, '%Y-%m-%d') AS date,
                       DATE_FORMAT(schedule_time, '%H:%i') AS time,
                       misa_name, role_name, request_source, cancelled_at, note
                FROM task_cancellations
                WHERE member_id = %s
                """,
                (member_id,),
            )
            for row in cursor.fetchall() or []:
                kind = "besar" if normalize_text(row.get("kind")).lower() == "besar" else "biasa"
                rows.append(build_item(
                    kind,
                    "Misa Besar" if kind == "besar" else "Misa Biasa",
                    row.get("misa_name") or row.get("type_label") or "Misa",
                    normalize_text(row.get("date")),
                    format_time_hhmm(row.get("time")),
                    row.get("role_name"),
                    row.get("request_source"),
                    row.get("cancelled_at"),
                    cancelled=True,
                    note=row.get("note") or "Tugas ini sudah dibatalkan.",
                ))
        except Exception as exc:
            print(f"[WARN] Gagal memuat riwayat pembatalan tugas: {exc}")

        if kind_filter in {"biasa", "besar"}:
            rows = [item for item in rows if item.get("type") == kind_filter]

        if month_filter.lower() != "all":
            month_int = min(12, max(1, parse_required_int(month_filter, datetime.now().month)))
            rows = [item for item in rows if normalize_text(item.get("date"))[5:7] == f"{month_int:02d}"]

        if year_filter.lower() != "all":
            year_int = parse_required_int(year_filter, datetime.now().year)
            rows = [item for item in rows if normalize_text(item.get("date"))[:4] == str(year_int)]

        if search_text:
            def matches(item):
                haystack = " ".join([
                    normalize_text(item.get("misaName")),
                    normalize_text(item.get("typeLabel")),
                    normalize_text(item.get("role")),
                    normalize_text(item.get("status")),
                    normalize_text(item.get("sourceLabel")),
                    normalize_text(item.get("date")),
                    normalize_text(item.get("time")),
                    normalize_text(item.get("note")),
                ]).lower()
                return search_text in haystack
            rows = [item for item in rows if matches(item)]

        def schedule_dt(item):
            try:
                return datetime.strptime(f"{item.get('date')} {format_time_hhmm(item.get('time'))}", "%Y-%m-%d %H:%M")
            except Exception:
                return datetime(1970, 1, 1)

        if sort_mode in {"date_asc", "tanggal_terlama"}:
            rows.sort(key=schedule_dt)
        elif sort_mode in {"role_asc"}:
            rows.sort(key=lambda item: normalize_text(item.get("role")).lower())
        elif sort_mode in {"status_asc"}:
            rows.sort(key=lambda item: normalize_text(item.get("status")).lower())
        elif sort_mode in {"type_asc", "jenis_asc"}:
            rows.sort(key=lambda item: (normalize_text(item.get("typeLabel")).lower(), schedule_dt(item)))
        else:
            rows.sort(key=schedule_dt, reverse=True)

        total = len(rows)
        total_pages = max(1, (total + page_size - 1) // page_size)
        page = min(page, total_pages)
        start = (page - 1) * page_size
        paged_rows = rows[start:start + page_size]

        now = datetime.now()
        def in_stats_range(item):
            try:
                item_date = datetime.strptime(normalize_text(item.get("date")), "%Y-%m-%d")
            except Exception:
                return False
            if stats_range == "year":
                return item_date.year == now.year
            if stats_range == "week":
                week_start = now - timedelta(days=now.weekday())
                week_start = datetime(week_start.year, week_start.month, week_start.day)
                week_end = week_start + timedelta(days=7)
                return week_start <= item_date < week_end
            return item_date.year == now.year and item_date.month == now.month

        summary_rows = [item for item in rows if in_stats_range(item)]
        summary = {
            "total": len(summary_rows),
            "completed": sum(1 for item in summary_rows if item.get("status") == "Selesai"),
            "exchange": sum(1 for item in summary_rows if item.get("isExchange")),
            "upcoming": sum(1 for item in summary_rows if item.get("status") == "Terdaftar"),
            "cancelled": sum(1 for item in summary_rows if item.get("status") == "Dibatalkan"),
        }

        return jsonify({
            "success": True,
            "items": paged_rows,
            "summary": summary,
            "pagination": {
                "page": page,
                "pageSize": page_size,
                "total": total,
                "totalPages": total_pages,
            },
            "filters": {
                "type": kind_filter,
                "month": month_filter,
                "year": year_filter,
                "sort": sort_mode,
                "search": search_text,
            },
        })
    finally:
        cursor.close()
        conn.close()

@app.route("/api/request-tugas/history", methods=["GET"])
def api_request_tugas_history():
    ensure_task_request_schema()
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True, buffered=True)
    try:
        member, error = require_active_request_member(cursor)
        if error:
            return error
        member_id = member["id"]
        page = max(1, parse_required_int(request.args.get("page"), 1))
        page_size = max(1, min(25, parse_required_int(request.args.get("pageSize"), 5)))
        kind_filter = (normalize_text(request.args.get("type")) or "all").lower()
        role_filter = normalize_text(request.args.get("role")).lower()
        search_text = normalize_text(request.args.get("search")).lower()
        month_filter = normalize_text(request.args.get("month")) or str(datetime.now().month)
        year_filter = normalize_text(request.args.get("year")) or str(datetime.now().year)
        period_filter = (normalize_text(request.args.get("period")) or "all").lower()
        sort_mode = (normalize_text(request.args.get("sort")) or "date_desc").lower()

        rows: list[dict[str, object]] = []
        cursor.execute(
            """
            SELECT DATE_FORMAT(sa.schedule_date, '%Y-%m-%d') AS date,
                   DATE_FORMAT(sa.schedule_time, '%H:%i') AS time,
                   sa.role_name AS role,
                   COALESCE(sa.request_source, 'admin') AS request_source,
                   sa.created_at,
                   cfg.mass_name
            FROM streaming_assignments sa
            LEFT JOIN streaming_weekly_config cfg
              ON cfg.day_name = CASE WEEKDAY(sa.schedule_date)
                WHEN 0 THEN 'Senin' WHEN 1 THEN 'Selasa' WHEN 2 THEN 'Rabu'
                WHEN 3 THEN 'Kamis' WHEN 4 THEN 'Jumat' WHEN 5 THEN 'Sabtu'
                ELSE 'Minggu' END
              AND DATE_FORMAT(cfg.start_time, '%H:%i') = DATE_FORMAT(sa.schedule_time, '%H:%i')
            WHERE sa.member_id = %s
            """,
            (member_id,),
        )
        for row in cursor.fetchall() or []:
            date_text = normalize_text(row.get("date"))
            time_text = format_time_hhmm(row.get("time"))
            rows.append({
                "type": "biasa",
                "typeLabel": "Misa Biasa",
                "misaName": normalize_text(row.get("mass_name")) or "Misa Biasa",
                "date": date_text,
                "dateLabel": request_task_format_date(date_text),
                "dayName": request_task_day_name(date_text),
                "time": time_text,
                "role": normalize_text(row.get("role")),
                "status": "Terdaftar",
                "source": normalize_text(row.get("request_source")) or "admin",
                "createdAt": row.get("created_at").isoformat() if row.get("created_at") else None,
            })

        cursor.execute(
            """
            SELECT mb.misa_name, DATE_FORMAT(mb.misa_date, '%Y-%m-%d') AS date,
                   DATE_FORMAT(mb.misa_time, '%H:%i') AS time,
                   n.role_name AS role,
                   COALESCE(a.request_source, 'admin') AS request_source,
                   a.created_at
            FROM misa_besar_assignments a
            JOIN misa_besar_names n ON n.id = a.role_id
            JOIN misa_besar mb ON mb.id = n.misa_id
            WHERE a.member_id = %s
            """,
            (member_id,),
        )
        for row in cursor.fetchall() or []:
            date_text = normalize_text(row.get("date"))
            time_text = format_time_hhmm(row.get("time"))
            rows.append({
                "type": "besar",
                "typeLabel": "Misa Besar",
                "misaName": normalize_text(row.get("misa_name")) or "Misa Besar",
                "date": date_text,
                "dateLabel": request_task_format_date(date_text),
                "dayName": request_task_day_name(date_text),
                "time": time_text,
                "role": normalize_text(row.get("role")),
                "status": "Terdaftar",
                "source": normalize_text(row.get("request_source")) or "admin",
                "createdAt": row.get("created_at").isoformat() if row.get("created_at") else None,
            })

        if kind_filter in {"biasa", "besar"}:
            rows = [item for item in rows if item.get("type") == kind_filter]

        if month_filter.lower() != "all":
            month_int = min(12, max(1, parse_required_int(month_filter, datetime.now().month)))
            rows = [item for item in rows if normalize_text(item.get("date"))[5:7] == f"{month_int:02d}"]

        if year_filter.lower() != "all":
            year_int = parse_required_int(year_filter, datetime.now().year)
            rows = [item for item in rows if normalize_text(item.get("date"))[:4] == str(year_int)]

        today_text = datetime.now().strftime("%Y-%m-%d")
        if period_filter in {"upcoming", "akan_datang", "future"}:
            rows = [item for item in rows if normalize_text(item.get("date")) >= today_text]
        elif period_filter in {"past", "lewat", "passed"}:
            rows = [item for item in rows if normalize_text(item.get("date")) < today_text]

        if role_filter:
            rows = [item for item in rows if role_filter in normalize_text(item.get("role")).lower()]
        if search_text:
            rows = [item for item in rows if search_text in " ".join([
                normalize_text(item.get("typeLabel")), normalize_text(item.get("misaName")),
                normalize_text(item.get("dateLabel")), normalize_text(item.get("dayName")),
                normalize_text(item.get("time")), normalize_text(item.get("role")),
                normalize_text(item.get("source")),
                normalize_text(item.get("sourceLabel")), normalize_text(item.get("description")),
            ]).lower()]

        def request_history_schedule_key(item):
            return (normalize_text(item.get("date")), normalize_text(item.get("time")), normalize_text(item.get("misaName")), normalize_text(item.get("role")))

        if sort_mode == "date_asc":
            rows.sort(key=request_history_schedule_key)
        elif sort_mode == "created_desc":
            rows.sort(key=lambda item: (normalize_text(item.get("createdAt")), normalize_text(item.get("date")), normalize_text(item.get("time"))), reverse=True)
        elif sort_mode == "created_asc":
            rows.sort(key=lambda item: (normalize_text(item.get("createdAt")), normalize_text(item.get("date")), normalize_text(item.get("time"))))
        else:
            rows.sort(key=request_history_schedule_key, reverse=True)
        total = len(rows)
        start = (page - 1) * page_size
        end = start + page_size
        page_items = rows[start:end]
        for idx, item in enumerate(page_items, start=start + 1):
            item["no"] = idx
            item["sourceLabel"] = "Request Mandiri" if item.get("source") == "member_request" else "Ditugaskan Admin"
            item["description"] = "Request otomatis masuk ke daftar petugas." if item.get("source") == "member_request" else "Nama Anda terdaftar pada jadwal tugas."

        return jsonify({
            "success": True,
            "items": page_items,
            "pagination": {
                "page": page,
                "pageSize": page_size,
                "total": total,
                "totalPages": max(1, (total + page_size - 1) // page_size),
            },
            "filters": {
                "type": kind_filter,
                "month": month_filter,
                "year": year_filter,
                "period": period_filter,
                "sort": sort_mode,
            },
        })
    finally:
        cursor.close()
        conn.close()


# -----------------------------------------------------------------------------
# Pembatalan Tugas Anggota: self-service cancel assignment H-3
# -----------------------------------------------------------------------------

def ensure_task_cancellation_schema(cursor=None) -> None:
    """Pastikan tabel riwayat pembatalan tugas tersedia."""
    ensure_task_request_schema(cursor)
    ensure_notifications_schema()

    owns_connection = cursor is None
    conn = None
    if owns_connection:
        conn = mysql_connection()
        cursor = conn.cursor(buffered=True)

    try:
        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS `task_cancellations` (
              `id` int(11) NOT NULL AUTO_INCREMENT,
              `member_id` int(11) NOT NULL,
              `member_name` varchar(255) DEFAULT NULL,
              `kind` varchar(20) NOT NULL,
              `type_label` varchar(50) DEFAULT NULL,
              `misa_id` int(11) DEFAULT NULL,
              `role_id` int(11) DEFAULT NULL,
              `assignment_id` int(11) DEFAULT NULL,
              `schedule_date` date NOT NULL,
              `schedule_time` time NOT NULL,
              `misa_name` varchar(255) DEFAULT NULL,
              `role_name` varchar(100) DEFAULT NULL,
              `request_source` varchar(30) DEFAULT NULL,
              `cancelled_at` datetime NOT NULL DEFAULT CURRENT_TIMESTAMP,
              `status` varchar(30) NOT NULL DEFAULT 'batal',
              `note` text DEFAULT NULL,
              PRIMARY KEY (`id`),
              KEY `idx_task_cancel_member` (`member_id`),
              KEY `idx_task_cancel_date` (`schedule_date`),
              KEY `idx_task_cancel_kind` (`kind`)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_general_ci
            """
        )
        # Aman untuk database lama yang sudah pernah dibuat sebelum struktur lengkap.
        ensure_column(cursor, "task_cancellations", "member_name", "`member_name` varchar(255) DEFAULT NULL")
        ensure_column(cursor, "task_cancellations", "type_label", "`type_label` varchar(50) DEFAULT NULL")
        ensure_column(cursor, "task_cancellations", "misa_id", "`misa_id` int(11) DEFAULT NULL")
        ensure_column(cursor, "task_cancellations", "role_id", "`role_id` int(11) DEFAULT NULL")
        ensure_column(cursor, "task_cancellations", "assignment_id", "`assignment_id` int(11) DEFAULT NULL")
        ensure_column(cursor, "task_cancellations", "request_source", "`request_source` varchar(30) DEFAULT NULL")
        ensure_column(cursor, "task_cancellations", "note", "`note` text DEFAULT NULL")
        if conn:
            conn.commit()
    finally:
        if owns_connection:
            cursor.close()
            conn.close()


def cancel_task_can_cancel(date_text: str) -> bool:
    """Boleh dibatalkan maksimal H-3: hari ini <= tanggal tugas - 3 hari."""
    try:
        schedule_date = datetime.strptime(normalize_text(date_text), "%Y-%m-%d").date()
        return schedule_date >= (datetime.now().date() + timedelta(days=3))
    except Exception:
        return False


def cancel_task_rule_label(date_text: str) -> str:
    try:
        schedule_date = datetime.strptime(normalize_text(date_text), "%Y-%m-%d").date()
        deadline = schedule_date - timedelta(days=3)
        return request_task_format_date(deadline.strftime("%Y-%m-%d"))
    except Exception:
        return "-"


def cancel_task_build_item(row: dict[str, object], *, kind: str) -> dict[str, object]:
    date_text = normalize_text(row.get("date") or row.get("schedule_date") or row.get("misa_date"))
    time_text = format_time_hhmm(row.get("time") or row.get("schedule_time") or row.get("misa_time"))
    type_label = "Misa Besar" if kind == "besar" else "Misa Biasa"
    misa_name = normalize_text(row.get("misa_name") or row.get("mass_name")) or type_label
    role_name = normalize_text(row.get("role") or row.get("role_name")) or "Role"
    can_cancel = cancel_task_can_cancel(date_text)
    return {
        "id": f"{kind}:{row.get('assignment_id') or row.get('id') or row.get('role_id') or ''}:{date_text}:{time_text}:{role_name}",
        "type": kind,
        "typeLabel": type_label,
        "assignmentId": row.get("assignment_id") or row.get("id"),
        "misaId": row.get("misa_id"),
        "roleId": row.get("role_id"),
        "misaName": misa_name,
        "date": date_text,
        "dateLabel": request_task_format_date(date_text),
        "dayName": request_task_day_name(date_text),
        "time": time_text,
        "role": role_name,
        "status": "Terdaftar",
        "source": normalize_text(row.get("request_source")) or "admin",
        "sourceLabel": "Request Mandiri" if normalize_text(row.get("request_source")) == "member_request" else "Ditugaskan Admin",
        "canCancel": can_cancel,
        "deadlineLabel": cancel_task_rule_label(date_text),
    }


def cancel_task_filter_items(items: list[dict[str, object]], *, search_text: str = "", kind_filter: str = "all") -> list[dict[str, object]]:
    keyword = normalize_text(search_text).lower()
    kind = (normalize_text(kind_filter) or "all").lower()
    filtered = []
    for item in items:
        if kind in {"biasa", "besar"} and item.get("type") != kind:
            continue
        haystack = " ".join([
            normalize_text(item.get("typeLabel")), normalize_text(item.get("misaName")),
            normalize_text(item.get("dayName")), normalize_text(item.get("dateLabel")),
            normalize_text(item.get("date")), normalize_text(item.get("time")),
            normalize_text(item.get("role")), normalize_text(item.get("status")),
            normalize_text(item.get("sourceLabel")),
        ]).lower()
        if keyword and keyword not in haystack:
            continue
        filtered.append(item)
    return filtered


def cancel_task_sort_items(items: list[dict[str, object]], sort_mode: str, *, history: bool = False) -> list[dict[str, object]]:
    mode = normalize_text(sort_mode) or ("cancelled_desc" if history else "date_asc")

    def schedule_key(item):
        return (normalize_text(item.get("date")), normalize_text(item.get("time")), normalize_text(item.get("misaName")), normalize_text(item.get("role")))

    def cancelled_key(item):
        return normalize_text(item.get("cancelledAt"))

    if mode == "date_desc":
        return sorted(items, key=schedule_key, reverse=True)
    if mode == "cancelled_asc":
        return sorted(items, key=cancelled_key)
    if mode == "cancelled_desc":
        return sorted(items, key=cancelled_key, reverse=True)
    return sorted(items, key=schedule_key)


def cancel_task_paginate(items: list[dict[str, object]], page: int, page_size: int) -> tuple[list[dict[str, object]], dict[str, int]]:
    page = max(1, page)
    page_size = max(1, min(25, page_size))
    total = len(items)
    total_pages = max(1, (total + page_size - 1) // page_size)
    if page > total_pages:
        page = total_pages
    start = (page - 1) * page_size
    end = start + page_size
    page_items = items[start:end]
    for no, item in enumerate(page_items, start=start + 1):
        item["no"] = no
    return page_items, {"page": page, "pageSize": page_size, "total": total, "totalPages": total_pages}


def cancel_task_notify(cursor, *, member: dict[str, object], item: dict[str, object]) -> None:
    member_id = normalize_text(member.get("id"))
    member_name = normalize_text(member.get("name") or member.get("nama")) or "Anggota"
    type_label = normalize_text(item.get("typeLabel")) or "Jadwal"
    misa_name = normalize_text(item.get("misaName")) or type_label
    role_name = normalize_text(item.get("role")) or "Role"
    day_label = normalize_text(item.get("dayName")) or request_task_day_name(item.get("date"))
    date_label = normalize_text(item.get("dateLabel")) or request_task_format_date(item.get("date"))
    time_text = format_time_hhmm(item.get("time"))
    safe_member = html.escape(member_name)
    safe_type = html.escape(type_label)
    safe_misa = html.escape(misa_name)
    safe_role = html.escape(role_name)
    safe_day = html.escape(day_label)
    safe_date = html.escape(date_label)
    safe_time = html.escape(time_text)

    admin_body = (
        f"<b>{safe_member}</b> membatalkan tugas sebagai <b>{safe_role}</b> pada "
        f"<b>{safe_type} - {safe_misa}</b>, hari {safe_day}, {safe_date} jam {safe_time} WIB."
    )
    create_notification(
        cursor,
        "tugas",
        f"Pembatalan Tugas: {member_name}",
        admin_body,
        "/dashboard.html",
        {
            "notification_kind": "task_cancelled_by_member",
            "member_id": member_id,
            "member_name": member_name,
            "misa_type": item.get("type"),
            "misa_name": misa_name,
            "role": role_name,
            "misa_date": item.get("date"),
            "misa_time": time_text,
        },
        target_role="admin",
    )

    user_body = (
        f"Tugas Anda sebagai <b>{safe_role}</b> untuk <b>{safe_type} - {safe_misa}</b> "
        f"pada hari {safe_day}, {safe_date} jam {safe_time} WIB telah berhasil dibatalkan."
    )
    create_notification(
        cursor,
        "tugas",
        f"Tugas Dibatalkan: {role_name}",
        user_body,
        "/pembatalan-tugas-anggota.html",
        {
            "target_user_id": member_id,
            "notification_kind": "task_cancelled_success",
            "misa_type": item.get("type"),
            "misa_name": misa_name,
            "role": role_name,
            "misa_date": item.get("date"),
            "misa_time": time_text,
        },
        target_role="user",
    )


def cancel_task_insert_history(cursor, *, member: dict[str, object], item: dict[str, object]) -> None:
    cursor.execute(
        """
        INSERT INTO task_cancellations
          (member_id, member_name, kind, type_label, misa_id, role_id, assignment_id,
           schedule_date, schedule_time, misa_name, role_name, request_source, status)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 'batal')
        """,
        (
            member.get("id"), member.get("name") or member.get("nama"), item.get("type"), item.get("typeLabel"),
            item.get("misaId"), item.get("roleId"), item.get("assignmentId"), item.get("date"),
            format_time_hhmm(item.get("time")), item.get("misaName"), item.get("role"), item.get("source"),
        ),
    )


@app.route("/api/cancel-tugas/me", methods=["GET"])
def api_cancel_tugas_me():
    ensure_task_cancellation_schema()
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True, buffered=True)
    try:
        member, error = require_active_request_member(cursor)
        if error:
            return error
        return jsonify({"success": True, "member": {"id": member.get("id"), "name": member.get("name"), "role": "user"}})
    finally:
        cursor.close()
        conn.close()


@app.route("/api/cancel-tugas/active", methods=["GET"])
def api_cancel_tugas_active():
    ensure_task_cancellation_schema()
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True, buffered=True)
    try:
        member, error = require_active_request_member(cursor)
        if error:
            return error
        member_id = member["id"]
        today = datetime.now().date()
        min_cancel_date = (today + timedelta(days=3)).strftime("%Y-%m-%d")
        month = request.args.get("month")
        year = request.args.get("year")
        if month in (None, "", "current"):
            month = str(today.month)
        if year in (None, "", "current"):
            year = str(today.year)
        month_int = parse_required_int(month, today.month)
        year_int = parse_required_int(year, today.year)
        month_int = min(12, max(1, month_int))
        start_date = f"{year_int:04d}-{month_int:02d}-01"
        last_day = calendar.monthrange(year_int, month_int)[1]
        end_date = f"{year_int:04d}-{month_int:02d}-{last_day:02d}"
        kind_filter = (normalize_text(request.args.get("type")) or "all").lower()
        search_text = normalize_text(request.args.get("search"))
        sort_mode = normalize_text(request.args.get("sort")) or "date_asc"
        page = max(1, parse_required_int(request.args.get("page"), 1))
        page_size = max(1, min(25, parse_required_int(request.args.get("pageSize"), 5)))

        items: list[dict[str, object]] = []
        if kind_filter in {"all", "biasa"}:
            cursor.execute(
                """
                SELECT sa.id AS assignment_id, DATE_FORMAT(sa.schedule_date, '%Y-%m-%d') AS date,
                       DATE_FORMAT(sa.schedule_time, '%H:%i') AS time,
                       sa.role_name AS role, COALESCE(sa.request_source, 'admin') AS request_source,
                       COALESCE(cfg.mass_name, 'Misa Biasa') AS mass_name
                FROM streaming_assignments sa
                LEFT JOIN streaming_weekly_config cfg
                  ON cfg.day_name = CASE WEEKDAY(sa.schedule_date)
                    WHEN 0 THEN 'Senin' WHEN 1 THEN 'Selasa' WHEN 2 THEN 'Rabu'
                    WHEN 3 THEN 'Kamis' WHEN 4 THEN 'Jumat' WHEN 5 THEN 'Sabtu'
                    ELSE 'Minggu' END
                  AND DATE_FORMAT(cfg.start_time, '%H:%i') = DATE_FORMAT(sa.schedule_time, '%H:%i')
                WHERE sa.member_id = %s
                  AND sa.schedule_date BETWEEN %s AND %s
                  AND sa.schedule_date >= %s
                """,
                (member_id, start_date, end_date, min_cancel_date),
            )
            for row in cursor.fetchall() or []:
                items.append(cancel_task_build_item(row, kind="biasa"))

        if kind_filter in {"all", "besar"}:
            cursor.execute(
                """
                SELECT a.id AS assignment_id, mb.id AS misa_id, n.id AS role_id,
                       mb.misa_name, DATE_FORMAT(mb.misa_date, '%Y-%m-%d') AS date,
                       DATE_FORMAT(mb.misa_time, '%H:%i') AS time,
                       n.role_name AS role, COALESCE(a.request_source, 'admin') AS request_source
                FROM misa_besar_assignments a
                JOIN misa_besar_names n ON n.id = a.role_id
                JOIN misa_besar mb ON mb.id = n.misa_id
                WHERE a.member_id = %s
                  AND mb.status = 'published'
                  AND mb.misa_date BETWEEN %s AND %s
                  AND mb.misa_date >= %s
                """,
                (member_id, start_date, end_date, min_cancel_date),
            )
            for row in cursor.fetchall() or []:
                items.append(cancel_task_build_item(row, kind="besar"))

        items = cancel_task_filter_items(items, search_text=search_text, kind_filter=kind_filter)
        items = cancel_task_sort_items(items, sort_mode, history=False)
        page_items, pagination = cancel_task_paginate(items, page, page_size)
        return jsonify({
            "success": True,
            "items": page_items,
            "pagination": pagination,
            "filters": {"month": month_int, "year": year_int, "type": kind_filter},
            "rule": {"minCancelDate": min_cancel_date, "description": "Pembatalan hanya bisa dilakukan paling lambat H-3 sebelum jadwal."},
        })
    finally:
        cursor.close()
        conn.close()


@app.route("/api/cancel-tugas/history", methods=["GET"])
def api_cancel_tugas_history():
    ensure_task_cancellation_schema()
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True, buffered=True)
    try:
        member, error = require_active_request_member(cursor)
        if error:
            return error
        member_id = member["id"]
        kind_filter = (normalize_text(request.args.get("type")) or "all").lower()
        month = normalize_text(request.args.get("month")) or "all"
        year = normalize_text(request.args.get("year")) or "all"
        search_text = normalize_text(request.args.get("search"))
        sort_mode = normalize_text(request.args.get("sort")) or "cancelled_desc"
        page = max(1, parse_required_int(request.args.get("page"), 1))
        page_size = max(1, min(25, parse_required_int(request.args.get("pageSize"), 5)))

        conditions = ["member_id = %s"]
        params: list[object] = [member_id]
        if kind_filter in {"biasa", "besar"}:
            conditions.append("kind = %s")
            params.append(kind_filter)
        if month != "all":
            month_int = min(12, max(1, parse_required_int(month, 1)))
            conditions.append("MONTH(schedule_date) = %s")
            params.append(month_int)
        if year != "all":
            year_int = parse_required_int(year, datetime.now().year)
            conditions.append("YEAR(schedule_date) = %s")
            params.append(year_int)
        where_clause = " AND ".join(conditions)
        cursor.execute(
            f"""
            SELECT id, kind, type_label, misa_id, role_id, assignment_id,
                   DATE_FORMAT(schedule_date, '%Y-%m-%d') AS date,
                   DATE_FORMAT(schedule_time, '%H:%i') AS time,
                   misa_name, role_name AS role, request_source,
                   cancelled_at, status
            FROM task_cancellations
            WHERE {where_clause}
            """,
            tuple(params),
        )
        items = []
        for row in cursor.fetchall() or []:
            date_text = normalize_text(row.get("date"))
            time_text = format_time_hhmm(row.get("time"))
            kind = normalize_text(row.get("kind")) or "biasa"
            type_label = normalize_text(row.get("type_label")) or ("Misa Besar" if kind == "besar" else "Misa Biasa")
            cancelled_at = row.get("cancelled_at")
            items.append({
                "id": row.get("id"),
                "type": kind,
                "typeLabel": type_label,
                "misaId": row.get("misa_id"),
                "roleId": row.get("role_id"),
                "assignmentId": row.get("assignment_id"),
                "misaName": normalize_text(row.get("misa_name")) or type_label,
                "date": date_text,
                "dateLabel": request_task_format_date(date_text),
                "dayName": request_task_day_name(date_text),
                "time": time_text,
                "role": normalize_text(row.get("role")) or "Role",
                "source": normalize_text(row.get("request_source")) or "admin",
                "status": "Batal",
                "cancelledAt": cancelled_at.isoformat() if hasattr(cancelled_at, "isoformat") else normalize_text(cancelled_at),
                "cancelledAtLabel": cancelled_at.strftime("%d/%m/%Y %H:%M:%S") if hasattr(cancelled_at, "strftime") else normalize_text(cancelled_at),
            })
        items = cancel_task_filter_items(items, search_text=search_text, kind_filter=kind_filter)
        items = cancel_task_sort_items(items, sort_mode, history=True)
        page_items, pagination = cancel_task_paginate(items, page, page_size)
        return jsonify({"success": True, "items": page_items, "pagination": pagination})
    finally:
        cursor.close()
        conn.close()


@app.route("/api/cancel-tugas/cancel", methods=["POST"])
def api_cancel_tugas_cancel():
    ensure_task_cancellation_schema()
    data = request.get_json(silent=True) or {}
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True, buffered=True)
    try:
        member, error = require_active_request_member(cursor)
        if error:
            return error
        member_id = member["id"]
        kind = (normalize_text(data.get("type")) or "biasa").lower()
        assignment_id = parse_optional_int(data.get("assignmentId"))

        if kind == "besar":
            misa_id = parse_optional_int(data.get("misaId"))
            role_id = parse_optional_int(data.get("roleId"))
            if assignment_id is not None:
                cursor.execute(
                    """
                    SELECT a.id AS assignment_id, mb.id AS misa_id, n.id AS role_id,
                           mb.misa_name, DATE_FORMAT(mb.misa_date, '%Y-%m-%d') AS date,
                           DATE_FORMAT(mb.misa_time, '%H:%i') AS time,
                           n.role_name AS role, COALESCE(a.request_source, 'admin') AS request_source
                    FROM misa_besar_assignments a
                    JOIN misa_besar_names n ON n.id = a.role_id
                    JOIN misa_besar mb ON mb.id = n.misa_id
                    WHERE a.id = %s AND a.member_id = %s
                    LIMIT 1
                    """,
                    (assignment_id, member_id),
                )
            else:
                cursor.execute(
                    """
                    SELECT a.id AS assignment_id, mb.id AS misa_id, n.id AS role_id,
                           mb.misa_name, DATE_FORMAT(mb.misa_date, '%Y-%m-%d') AS date,
                           DATE_FORMAT(mb.misa_time, '%H:%i') AS time,
                           n.role_name AS role, COALESCE(a.request_source, 'admin') AS request_source
                    FROM misa_besar_assignments a
                    JOIN misa_besar_names n ON n.id = a.role_id
                    JOIN misa_besar mb ON mb.id = n.misa_id
                    WHERE mb.id = %s AND n.id = %s AND a.member_id = %s
                    LIMIT 1
                    """,
                    (misa_id, role_id, member_id),
                )
            row = cursor.fetchone()
            if not row:
                return jsonify({"success": False, "error": "Tugas Misa Besar tidak ditemukan atau bukan milik Anda."}), 404
            item = cancel_task_build_item(row, kind="besar")
            if not cancel_task_can_cancel(item.get("date")):
                return jsonify({"success": False, "error": "Tugas ini sudah melewati batas pembatalan H-3."}), 400
            cancel_task_insert_history(cursor, member=member, item=item)
            cursor.execute("DELETE FROM misa_besar_assignments WHERE id = %s AND member_id = %s", (item.get("assignmentId"), member_id))
        else:
            date_text = parse_optional_date(data.get("date"))
            time_text = format_time_hhmm(data.get("time"))
            role_name = normalize_text(data.get("role"))
            if assignment_id is not None:
                cursor.execute(
                    """
                    SELECT sa.id AS assignment_id, DATE_FORMAT(sa.schedule_date, '%Y-%m-%d') AS date,
                           DATE_FORMAT(sa.schedule_time, '%H:%i') AS time,
                           sa.role_name AS role, COALESCE(sa.request_source, 'admin') AS request_source,
                           COALESCE(cfg.mass_name, 'Misa Biasa') AS mass_name
                    FROM streaming_assignments sa
                    LEFT JOIN streaming_weekly_config cfg
                      ON cfg.day_name = CASE WEEKDAY(sa.schedule_date)
                        WHEN 0 THEN 'Senin' WHEN 1 THEN 'Selasa' WHEN 2 THEN 'Rabu'
                        WHEN 3 THEN 'Kamis' WHEN 4 THEN 'Jumat' WHEN 5 THEN 'Sabtu'
                        ELSE 'Minggu' END
                      AND DATE_FORMAT(cfg.start_time, '%H:%i') = DATE_FORMAT(sa.schedule_time, '%H:%i')
                    WHERE sa.id = %s AND sa.member_id = %s
                    LIMIT 1
                    """,
                    (assignment_id, member_id),
                )
            else:
                cursor.execute(
                    """
                    SELECT sa.id AS assignment_id, DATE_FORMAT(sa.schedule_date, '%Y-%m-%d') AS date,
                           DATE_FORMAT(sa.schedule_time, '%H:%i') AS time,
                           sa.role_name AS role, COALESCE(sa.request_source, 'admin') AS request_source,
                           COALESCE(cfg.mass_name, 'Misa Biasa') AS mass_name
                    FROM streaming_assignments sa
                    LEFT JOIN streaming_weekly_config cfg
                      ON cfg.day_name = CASE WEEKDAY(sa.schedule_date)
                        WHEN 0 THEN 'Senin' WHEN 1 THEN 'Selasa' WHEN 2 THEN 'Rabu'
                        WHEN 3 THEN 'Kamis' WHEN 4 THEN 'Jumat' WHEN 5 THEN 'Sabtu'
                        ELSE 'Minggu' END
                      AND DATE_FORMAT(cfg.start_time, '%H:%i') = DATE_FORMAT(sa.schedule_time, '%H:%i')
                    WHERE sa.schedule_date = %s AND DATE_FORMAT(sa.schedule_time, '%H:%i') = %s
                      AND sa.role_name = %s AND sa.member_id = %s
                    LIMIT 1
                    """,
                    (date_text, time_text, role_name, member_id),
                )
            row = cursor.fetchone()
            if not row:
                return jsonify({"success": False, "error": "Tugas Misa Biasa tidak ditemukan atau bukan milik Anda."}), 404
            item = cancel_task_build_item(row, kind="biasa")
            if not cancel_task_can_cancel(item.get("date")):
                return jsonify({"success": False, "error": "Tugas ini sudah melewati batas pembatalan H-3."}), 400
            cancel_task_insert_history(cursor, member=member, item=item)
            cursor.execute("DELETE FROM streaming_assignments WHERE id = %s AND member_id = %s", (item.get("assignmentId"), member_id))

        cancel_task_notify(cursor, member=member, item=item)
        conn.commit()
        return jsonify({"success": True, "message": "Tugas berhasil dibatalkan.", "item": item})
    except Exception as exc:
        conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 400
    finally:
        cursor.close()
        conn.close()

# -----------------------------------------------------------------------------
# Penukaran Jadwal Tugas Anggota: tukeran / pengganti tugas
# -----------------------------------------------------------------------------

def ensure_task_exchange_schema(cursor=None) -> None:
    ensure_task_request_schema(cursor)
    ensure_notifications_schema()
    owns_connection = cursor is None
    conn = None
    if owns_connection:
        conn = mysql_connection()
        cursor = conn.cursor(buffered=True)
    try:
        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS `task_exchange_requests` (
              `id` int(11) NOT NULL AUTO_INCREMENT,
              `requester_id` varchar(50) NOT NULL,
              `target_user_id` varchar(50) NOT NULL,
              `kind` varchar(20) NOT NULL DEFAULT 'biasa',
              `request_mode` varchar(30) NOT NULL DEFAULT 'swap',
              `my_assignment_id` int(11) DEFAULT NULL,
              `my_misa_id` int(11) DEFAULT NULL,
              `my_role_id` int(11) DEFAULT NULL,
              `my_type_label` varchar(80) DEFAULT NULL,
              `my_misa_name` varchar(255) DEFAULT NULL,
              `my_role_name` varchar(255) DEFAULT NULL,
              `my_schedule_date` date DEFAULT NULL,
              `my_schedule_time` time DEFAULT NULL,
              `target_assignment_id` int(11) DEFAULT NULL,
              `target_misa_id` int(11) DEFAULT NULL,
              `target_role_id` int(11) DEFAULT NULL,
              `target_type_label` varchar(80) DEFAULT NULL,
              `target_misa_name` varchar(255) DEFAULT NULL,
              `target_role_name` varchar(255) DEFAULT NULL,
              `target_schedule_date` date DEFAULT NULL,
              `target_schedule_time` time DEFAULT NULL,
              `reason` text DEFAULT NULL,
              `status` varchar(30) NOT NULL DEFAULT 'pending',
              `response_note` text DEFAULT NULL,
              `created_at` datetime NOT NULL DEFAULT CURRENT_TIMESTAMP,
              `updated_at` datetime NOT NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
              `responded_at` datetime DEFAULT NULL,
              `cancelled_at` datetime DEFAULT NULL,
              `auto_cancelled_at` datetime DEFAULT NULL,
              PRIMARY KEY (`id`),
              KEY `idx_exchange_requester` (`requester_id`),
              KEY `idx_exchange_target` (`target_user_id`),
              KEY `idx_exchange_status` (`status`),
              KEY `idx_exchange_my_date` (`my_schedule_date`)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_general_ci
            """
        )
        for col, definition in {
            "request_mode": "`request_mode` varchar(30) NOT NULL DEFAULT 'swap'",
            "my_assignment_id": "`my_assignment_id` int(11) DEFAULT NULL",
            "my_misa_id": "`my_misa_id` int(11) DEFAULT NULL",
            "my_role_id": "`my_role_id` int(11) DEFAULT NULL",
            "my_type_label": "`my_type_label` varchar(80) DEFAULT NULL",
            "my_misa_name": "`my_misa_name` varchar(255) DEFAULT NULL",
            "my_role_name": "`my_role_name` varchar(255) DEFAULT NULL",
            "my_schedule_date": "`my_schedule_date` date DEFAULT NULL",
            "my_schedule_time": "`my_schedule_time` time DEFAULT NULL",
            "target_assignment_id": "`target_assignment_id` int(11) DEFAULT NULL",
            "target_misa_id": "`target_misa_id` int(11) DEFAULT NULL",
            "target_role_id": "`target_role_id` int(11) DEFAULT NULL",
            "target_type_label": "`target_type_label` varchar(80) DEFAULT NULL",
            "target_misa_name": "`target_misa_name` varchar(255) DEFAULT NULL",
            "target_role_name": "`target_role_name` varchar(255) DEFAULT NULL",
            "target_schedule_date": "`target_schedule_date` date DEFAULT NULL",
            "target_schedule_time": "`target_schedule_time` time DEFAULT NULL",
            "response_note": "`response_note` text DEFAULT NULL",
            "responded_at": "`responded_at` datetime DEFAULT NULL",
            "cancelled_at": "`cancelled_at` datetime DEFAULT NULL",
            "auto_cancelled_at": "`auto_cancelled_at` datetime DEFAULT NULL",
        }.items():
            ensure_column(cursor, "task_exchange_requests", col, definition)
        if conn:
            conn.commit()
    finally:
        if owns_connection:
            cursor.close()
            conn.close()


def exchange_current_member(cursor, *, require_active: bool = True):
    if not session.get("logged_in"):
        return None, (jsonify({"success": False, "error": "Anda harus login terlebih dahulu."}), 401)
    user_id = session.get("user_id")
    if not user_id:
        return None, (jsonify({"success": False, "error": "Sesi tidak valid."}), 401)
    cursor.execute(
        """
        SELECT id, nama, role, status_akun
        FROM anggota
        WHERE id = %s
        LIMIT 1
        """,
        (user_id,),
    )
    member = cursor.fetchone()
    if not member:
        return None, (jsonify({"success": False, "error": "Akun tidak ditemukan."}), 404)
    if require_active and normalize_status(member.get("status_akun") or "aktif") != "aktif":
        return None, (jsonify({"success": False, "error": "Akun Anda sedang nonaktif."}), 403)
    member["id"] = str(member.get("id"))
    member["name"] = normalize_text(member.get("nama")) or "Anggota"
    member["roleNormalized"] = normalize_role_value(member.get("role") or "user")
    return member, None


def exchange_date_is_eligible(date_text: str) -> bool:
    """Penukaran hanya bisa untuk jadwal besok atau setelahnya; hari-H tidak bisa."""
    try:
        return datetime.strptime(normalize_text(date_text), "%Y-%m-%d").date() > datetime.now().date()
    except Exception:
        return False


def exchange_status_label(status: str) -> str:
    return {
        "pending": "Menunggu",
        "accepted": "Diterima",
        "rejected": "Ditolak",
        "cancelled": "Dibatalkan",
        "auto_cancelled": "Batal Otomatis",
    }.get(normalize_text(status).lower(), normalize_text(status) or "Menunggu")


def exchange_mode_label(mode: str) -> str:
    return "Menggantikan" if normalize_text(mode).lower() in {"substitute", "menggantikan"} else "Tukeran"


def exchange_kind_label(kind: str) -> str:
    return "Misa Besar" if normalize_text(kind).lower() == "besar" else "Misa Biasa"


def exchange_member_role_for_target(cursor, member_id: object) -> str:
    cursor.execute("SELECT role FROM anggota WHERE id = %s LIMIT 1", (member_id,))
    row = cursor.fetchone()
    return normalize_role_value((row or {}).get("role") or "user") if isinstance(row, dict) else "user"


def exchange_task_label_from_parts(name: str, date_text: str, time_text: str, role: str) -> str:
    return f"{normalize_text(name) or 'Misa'} - {request_task_day_name(date_text)}, {request_task_format_date(date_text)} {format_time_hhmm(time_text)} WIB ({normalize_text(role) or 'Role'})"


def exchange_format_row(row: dict[str, object]) -> dict[str, object]:
    date_text = normalize_text(row.get("date") or row.get("schedule_date") or row.get("my_schedule_date"))
    time_text = format_time_hhmm(row.get("time") or row.get("schedule_time") or row.get("my_schedule_time"))
    kind = normalize_text(row.get("kind") or row.get("type") or "biasa").lower()
    type_label = normalize_text(row.get("type_label") or row.get("typeLabel")) or exchange_kind_label(kind)
    misa_name = normalize_text(row.get("misa_name") or row.get("mass_name") or row.get("misaName")) or type_label
    role = normalize_text(row.get("role") or row.get("role_name") or row.get("roleName")) or "Role"
    member_id = normalize_text(row.get("member_id") or row.get("memberId"))
    member_name = normalize_text(row.get("member_name") or row.get("memberName")) or "Anggota"
    assignment_id = row.get("assignment_id") or row.get("assignmentId")
    return {
        "assignmentId": assignment_id,
        "id": f"{kind}:{assignment_id}",
        "type": kind,
        "typeLabel": type_label,
        "misaId": row.get("misa_id") or row.get("misaId"),
        "roleId": row.get("role_id") or row.get("roleId"),
        "misaName": misa_name,
        "date": date_text,
        "dateLabel": request_task_format_date(date_text),
        "dayName": request_task_day_name(date_text),
        "time": time_text,
        "role": role,
        "memberId": member_id,
        "memberName": member_name,
        "label": exchange_task_label_from_parts(misa_name, date_text, time_text, role),
    }


def exchange_fetch_assignment(cursor, *, kind: str, assignment_id: object | None = None, member_id: object | None = None, for_update: bool = False) -> dict[str, object] | None:
    clean_kind = normalize_text(kind).lower()
    lock_clause = " FOR UPDATE" if for_update else ""
    params: list[object] = []
    if clean_kind == "besar":
        where_parts = []
        if assignment_id is not None:
            where_parts.append("a.id = %s")
            params.append(assignment_id)
        if member_id is not None:
            where_parts.append("a.member_id = %s")
            params.append(member_id)
        if not where_parts:
            return None
        cursor.execute(
            f"""
            SELECT a.id AS assignment_id, a.member_id, COALESCE(mem.nama, CONCAT('ID ', a.member_id)) AS member_name,
                   mb.id AS misa_id, n.id AS role_id, mb.misa_name,
                   DATE_FORMAT(mb.misa_date, '%Y-%m-%d') AS date,
                   DATE_FORMAT(mb.misa_time, '%H:%i') AS time,
                   n.role_name AS role, 'besar' AS kind, 'Misa Besar' AS type_label
            FROM misa_besar_assignments a
            JOIN misa_besar_names n ON n.id = a.role_id
            JOIN misa_besar mb ON mb.id = n.misa_id
            LEFT JOIN anggota mem ON mem.id = a.member_id
            WHERE {' AND '.join(where_parts)} AND mb.status = 'published'
            LIMIT 1{lock_clause}
            """,
            tuple(params),
        )
    else:
        where_parts = []
        if assignment_id is not None:
            where_parts.append("sa.id = %s")
            params.append(assignment_id)
        if member_id is not None:
            where_parts.append("sa.member_id = %s")
            params.append(member_id)
        if not where_parts:
            return None
        cursor.execute(
            f"""
            SELECT sa.id AS assignment_id, sa.member_id, COALESCE(mem.nama, CONCAT('ID ', sa.member_id)) AS member_name,
                   NULL AS misa_id, NULL AS role_id,
                   COALESCE(cfg.mass_name, 'Misa Biasa') AS misa_name,
                   DATE_FORMAT(sa.schedule_date, '%Y-%m-%d') AS date,
                   DATE_FORMAT(sa.schedule_time, '%H:%i') AS time,
                   sa.role_name AS role, 'biasa' AS kind, 'Misa Biasa' AS type_label
            FROM streaming_assignments sa
            LEFT JOIN anggota mem ON mem.id = sa.member_id
            LEFT JOIN streaming_weekly_config cfg
              ON cfg.day_name = CASE WEEKDAY(sa.schedule_date)
                WHEN 0 THEN 'Senin' WHEN 1 THEN 'Selasa' WHEN 2 THEN 'Rabu'
                WHEN 3 THEN 'Kamis' WHEN 4 THEN 'Jumat' WHEN 5 THEN 'Sabtu'
                ELSE 'Minggu' END
              AND DATE_FORMAT(cfg.start_time, '%H:%i') = DATE_FORMAT(sa.schedule_time, '%H:%i')
            WHERE {' AND '.join(where_parts)}
            LIMIT 1{lock_clause}
            """,
            tuple(params),
        )
    row = cursor.fetchone()
    return exchange_format_row(row) if row else None


def exchange_fetch_user_tasks(cursor, *, member_id: object, kind: str, month: int, year: int) -> list[dict[str, object]]:
    today = datetime.now().date()
    start_date = f"{year:04d}-{month:02d}-01"
    end_date = f"{year:04d}-{month:02d}-{calendar.monthrange(year, month)[1]:02d}"
    items: list[dict[str, object]] = []
    clean_kind = normalize_text(kind).lower()
    if clean_kind in {"all", "biasa"}:
        cursor.execute(
            """
            SELECT sa.id AS assignment_id, sa.member_id, COALESCE(mem.nama, CONCAT('ID ', sa.member_id)) AS member_name,
                   COALESCE(cfg.mass_name, 'Misa Biasa') AS misa_name,
                   DATE_FORMAT(sa.schedule_date, '%Y-%m-%d') AS date,
                   DATE_FORMAT(sa.schedule_time, '%H:%i') AS time,
                   sa.role_name AS role, 'biasa' AS kind, 'Misa Biasa' AS type_label
            FROM streaming_assignments sa
            LEFT JOIN anggota mem ON mem.id = sa.member_id
            LEFT JOIN streaming_weekly_config cfg
              ON cfg.day_name = CASE WEEKDAY(sa.schedule_date)
                WHEN 0 THEN 'Senin' WHEN 1 THEN 'Selasa' WHEN 2 THEN 'Rabu'
                WHEN 3 THEN 'Kamis' WHEN 4 THEN 'Jumat' WHEN 5 THEN 'Sabtu'
                ELSE 'Minggu' END
              AND DATE_FORMAT(cfg.start_time, '%H:%i') = DATE_FORMAT(sa.schedule_time, '%H:%i')
            WHERE sa.member_id = %s AND sa.schedule_date BETWEEN %s AND %s AND sa.schedule_date > %s
            ORDER BY sa.schedule_date ASC, sa.schedule_time ASC
            """,
            (member_id, start_date, end_date, today),
        )
        items.extend(exchange_format_row(row) for row in (cursor.fetchall() or []))
    if clean_kind in {"all", "besar"}:
        cursor.execute(
            """
            SELECT a.id AS assignment_id, a.member_id, COALESCE(mem.nama, CONCAT('ID ', a.member_id)) AS member_name,
                   mb.id AS misa_id, n.id AS role_id, mb.misa_name,
                   DATE_FORMAT(mb.misa_date, '%Y-%m-%d') AS date,
                   DATE_FORMAT(mb.misa_time, '%H:%i') AS time,
                   n.role_name AS role, 'besar' AS kind, 'Misa Besar' AS type_label
            FROM misa_besar_assignments a
            JOIN misa_besar_names n ON n.id = a.role_id
            JOIN misa_besar mb ON mb.id = n.misa_id
            LEFT JOIN anggota mem ON mem.id = a.member_id
            WHERE a.member_id = %s AND mb.status = 'published' AND mb.misa_date BETWEEN %s AND %s AND mb.misa_date > %s
            ORDER BY mb.misa_date ASC, mb.misa_time ASC
            """,
            (member_id, start_date, end_date, today),
        )
        items.extend(exchange_format_row(row) for row in (cursor.fetchall() or []))
    items.sort(key=lambda item: (item.get("date") or "", item.get("time") or ""))
    return items


def exchange_fetch_target_tasks(cursor, *, current_member_id: object, kind: str, month: int, year: int) -> list[dict[str, object]]:
    today = datetime.now().date()
    start_date = f"{year:04d}-{month:02d}-01"
    end_date = f"{year:04d}-{month:02d}-{calendar.monthrange(year, month)[1]:02d}"
    items: list[dict[str, object]] = []
    clean_kind = normalize_text(kind).lower()
    if clean_kind == "biasa":
        cursor.execute(
            """
            SELECT sa.id AS assignment_id, sa.member_id, COALESCE(mem.nama, CONCAT('ID ', sa.member_id)) AS member_name,
                   COALESCE(cfg.mass_name, 'Misa Biasa') AS misa_name,
                   DATE_FORMAT(sa.schedule_date, '%Y-%m-%d') AS date,
                   DATE_FORMAT(sa.schedule_time, '%H:%i') AS time,
                   sa.role_name AS role, 'biasa' AS kind, 'Misa Biasa' AS type_label
            FROM streaming_assignments sa
            JOIN anggota mem ON mem.id = sa.member_id
            LEFT JOIN streaming_weekly_config cfg
              ON cfg.day_name = CASE WEEKDAY(sa.schedule_date)
                WHEN 0 THEN 'Senin' WHEN 1 THEN 'Selasa' WHEN 2 THEN 'Rabu'
                WHEN 3 THEN 'Kamis' WHEN 4 THEN 'Jumat' WHEN 5 THEN 'Sabtu'
                ELSE 'Minggu' END
              AND DATE_FORMAT(cfg.start_time, '%H:%i') = DATE_FORMAT(sa.schedule_time, '%H:%i')
            WHERE sa.member_id <> %s AND mem.status_akun = 'aktif'
              AND sa.schedule_date BETWEEN %s AND %s AND sa.schedule_date > %s
              AND NOT EXISTS (
                SELECT 1
                FROM streaming_assignments mine
                WHERE mine.member_id = %s
                  AND mine.schedule_date = sa.schedule_date
                  AND DATE_FORMAT(mine.schedule_time, '%H:%i') = DATE_FORMAT(sa.schedule_time, '%H:%i')
              )
            ORDER BY sa.schedule_date ASC, sa.schedule_time ASC
            """,
            (current_member_id, start_date, end_date, today, current_member_id),
        )
        items.extend(exchange_format_row(row) for row in (cursor.fetchall() or []))
    elif clean_kind == "besar":
        cursor.execute(
            """
            SELECT a.id AS assignment_id, a.member_id, COALESCE(mem.nama, CONCAT('ID ', a.member_id)) AS member_name,
                   mb.id AS misa_id, n.id AS role_id, mb.misa_name,
                   DATE_FORMAT(mb.misa_date, '%Y-%m-%d') AS date,
                   DATE_FORMAT(mb.misa_time, '%H:%i') AS time,
                   n.role_name AS role, 'besar' AS kind, 'Misa Besar' AS type_label
            FROM misa_besar_assignments a
            JOIN misa_besar_names n ON n.id = a.role_id
            JOIN misa_besar mb ON mb.id = n.misa_id
            JOIN anggota mem ON mem.id = a.member_id
            WHERE a.member_id <> %s AND mem.status_akun = 'aktif' AND mb.status = 'published'
              AND mb.misa_date BETWEEN %s AND %s AND mb.misa_date > %s
              AND NOT EXISTS (
                SELECT 1
                FROM misa_besar_assignments mine
                JOIN misa_besar_names mine_role ON mine_role.id = mine.role_id
                WHERE mine.member_id = %s AND mine_role.misa_id = mb.id
              )
            ORDER BY mb.misa_date ASC, mb.misa_time ASC
            """,
            (current_member_id, start_date, end_date, today, current_member_id),
        )
        items.extend(exchange_format_row(row) for row in (cursor.fetchall() or []))
    return items


def exchange_fetch_active_members(cursor, *, exclude_member_id: object | None = None) -> list[dict[str, object]]:
    params: list[object] = []
    where = "WHERE status_akun = 'aktif'"
    if exclude_member_id is not None:
        where += " AND id <> %s"
        params.append(exclude_member_id)
    cursor.execute(
        f"""
        SELECT id, nama, role
        FROM anggota
        {where}
        ORDER BY FIELD(role, 'user', 'admin', 'super_admin'), nama ASC
        """,
        tuple(params),
    )
    results = []
    for row in cursor.fetchall() or []:
        role_value = normalize_role_value(row.get("role") or "user")
        results.append({
            "id": str(row.get("id")),
            "name": normalize_text(row.get("nama")) or f"Anggota {row.get('id')}",
            "role": role_value,
            "roleLabel": "Super Admin" if role_value == "super_admin" else ("Admin" if role_value == "admin" else "Anggota"),
        })
    return results


def exchange_has_pending_for_assignment(cursor, requester_id: object, kind: str, assignment_id: object, exclude_id: object | None = None) -> bool:
    params: list[object] = [str(requester_id), normalize_text(kind).lower(), assignment_id]
    extra = ""
    if exclude_id is not None:
        extra = " AND id <> %s"
        params.append(exclude_id)
    cursor.execute(
        f"""
        SELECT id FROM task_exchange_requests
        WHERE requester_id = %s AND kind = %s AND my_assignment_id = %s AND status = 'pending'{extra}
        LIMIT 1
        """,
        tuple(params),
    )
    return cursor.fetchone() is not None


def exchange_member_already_in_schedule(cursor, *, kind: str, member_id: object, date_text: str, time_text: str, misa_id: object | None = None, exclude_assignment_id: object | None = None) -> bool:
    clean_kind = normalize_text(kind).lower()
    params: list[object] = []
    if clean_kind == "besar":
        where = "a.member_id = %s AND mb.id = %s"
        params = [member_id, misa_id]
        if exclude_assignment_id is not None:
            where += " AND a.id <> %s"
            params.append(exclude_assignment_id)
        cursor.execute(
            f"""
            SELECT a.id FROM misa_besar_assignments a
            JOIN misa_besar_names n ON n.id = a.role_id
            JOIN misa_besar mb ON mb.id = n.misa_id
            WHERE {where}
            LIMIT 1
            """,
            tuple(params),
        )
    else:
        where = "member_id = %s AND schedule_date = %s AND DATE_FORMAT(schedule_time, '%H:%i') = %s"
        params = [member_id, date_text, format_time_hhmm(time_text)]
        if exclude_assignment_id is not None:
            where += " AND id <> %s"
            params.append(exclude_assignment_id)
        cursor.execute(f"SELECT id FROM streaming_assignments WHERE {where} LIMIT 1", tuple(params))
    return cursor.fetchone() is not None




def exchange_schedule_member_ids(cursor, *, kind: str, date_text: str, time_text: str, misa_id: object | None = None) -> list[str]:
    # Return all member IDs assigned in a single misa session, so candidates never create double roles.
    clean_kind = normalize_text(kind).lower()
    if clean_kind == "besar":
        if not misa_id:
            return []
        cursor.execute(
            """
            SELECT DISTINCT a.member_id
            FROM misa_besar_assignments a
            JOIN misa_besar_names n ON n.id = a.role_id
            WHERE n.misa_id = %s AND a.member_id IS NOT NULL
            """,
            (misa_id,),
        )
    else:
        cursor.execute(
            """
            SELECT DISTINCT member_id
            FROM streaming_assignments
            WHERE schedule_date = %s
              AND DATE_FORMAT(schedule_time, '%H:%i') = %s
              AND member_id IS NOT NULL
            """,
            (date_text, format_time_hhmm(time_text)),
        )
    return [str(row.get("member_id")) for row in (cursor.fetchall() or []) if row.get("member_id") is not None]

def exchange_insert_notification(cursor, *, target_user_id: object, target_role: str | None, title: str, body: str, request_id: object, status: str = "pending", url: str | None = None):
    role_norm = normalize_role_value(target_role or exchange_member_role_for_target(cursor, target_user_id))
    return create_notification(
        cursor,
        "tukar",
        title,
        body,
        url,
        {
            "target_user_id": str(target_user_id),
            "exchange_request_id": int(request_id) if str(request_id).isdigit() else request_id,
            "exchange_status": status,
        },
        target_role=role_norm,
    )


def exchange_notify_new_request(cursor, request_row: dict[str, object]):
    target_role = exchange_member_role_for_target(cursor, request_row.get("target_user_id"))
    requester_name = normalize_text(request_row.get("requester_name")) or "Teman"
    mode_label = exchange_mode_label(request_row.get("request_mode"))
    my_label = exchange_task_label_from_parts(request_row.get("my_misa_name"), normalize_text(request_row.get("my_schedule_date")), request_row.get("my_schedule_time"), request_row.get("my_role_name"))
    if normalize_text(request_row.get("request_mode")).lower() in {"substitute", "menggantikan"}:
        title = f"Permintaan Pengganti dari {requester_name}"
        body = f"<b>{html.escape(requester_name)}</b> meminta Anda menggantikan tugas <b>{html.escape(str(request_row.get('my_role_name') or 'Role'))}</b> pada <b>{html.escape(str(request_row.get('my_misa_name') or 'Misa'))}</b>, {html.escape(request_task_day_name(normalize_text(request_row.get('my_schedule_date'))))}, {html.escape(request_task_format_date(normalize_text(request_row.get('my_schedule_date'))))} jam {html.escape(format_time_hhmm(request_row.get('my_schedule_time')))} WIB."
    else:
        target_label = exchange_task_label_from_parts(request_row.get("target_misa_name"), normalize_text(request_row.get("target_schedule_date")), request_row.get("target_schedule_time"), request_row.get("target_role_name"))
        title = f"Permintaan Tukar Jadwal dari {requester_name}"
        body = f"<b>{html.escape(requester_name)}</b> mengajak tukeran jadwal. Jadwal dia: {html.escape(my_label)}. Jadwal Anda: {html.escape(target_label)}."
    url = None if target_role in {"admin", "super_admin"} else "/penukaran-jadwal-tugas-anggota.html"
    exchange_insert_notification(cursor, target_user_id=request_row.get("target_user_id"), target_role=target_role, title=title, body=body, request_id=request_row.get("id"), status="pending", url=url)


def exchange_notify_requester_result(cursor, request_row: dict[str, object], status: str):
    status_label = exchange_status_label(status)
    target_name = normalize_text(request_row.get("target_name")) or "Teman"
    title_map = {
        "accepted": "Penukaran Jadwal Diterima",
        "rejected": "Penukaran Jadwal Ditolak",
        "cancelled": "Penukaran Jadwal Dibatalkan",
        "auto_cancelled": "Penukaran Jadwal Batal Otomatis",
    }
    title = title_map.get(status, f"Status Penukaran: {status_label}")
    my_label = exchange_task_label_from_parts(request_row.get("my_misa_name"), normalize_text(request_row.get("my_schedule_date")), request_row.get("my_schedule_time"), request_row.get("my_role_name"))
    if status == "accepted":
        body = f"Permintaan Anda sudah <b>diterima</b> oleh <b>{html.escape(target_name)}</b>. Jadwal terkait: {html.escape(my_label)}."
    elif status == "rejected":
        body = f"Permintaan Anda <b>ditolak</b> oleh <b>{html.escape(target_name)}</b>. Jadwal terkait: {html.escape(my_label)}."
    elif status == "cancelled":
        body = f"Permintaan penukaran/pengganti untuk jadwal {html.escape(my_label)} sudah dibatalkan."
    else:
        body = f"Permintaan penukaran/pengganti untuk jadwal {html.escape(my_label)} otomatis dibatalkan karena belum direspons sampai hari-H jadwal."
    exchange_insert_notification(cursor, target_user_id=request_row.get("requester_id"), target_role="user", title=title, body=body, request_id=request_row.get("id"), status=status, url="/penukaran-jadwal-tugas-anggota.html")


def exchange_fetch_request_row(cursor, request_id: object, for_update: bool = False) -> dict[str, object] | None:
    lock_clause = " FOR UPDATE" if for_update else ""
    cursor.execute(
        f"""
        SELECT er.*, req.nama AS requester_name, tgt.nama AS target_name,
               req.role AS requester_role, tgt.role AS target_role
        FROM task_exchange_requests er
        LEFT JOIN anggota req ON req.id = er.requester_id
        LEFT JOIN anggota tgt ON tgt.id = er.target_user_id
        WHERE er.id = %s
        LIMIT 1{lock_clause}
        """,
        (request_id,),
    )
    return cursor.fetchone()


def exchange_expire_pending_requests(cursor) -> int:
    today = datetime.now().date()
    cursor.execute(
        """
        SELECT er.*, req.nama AS requester_name, tgt.nama AS target_name,
               req.role AS requester_role, tgt.role AS target_role
        FROM task_exchange_requests er
        LEFT JOIN anggota req ON req.id = er.requester_id
        LEFT JOIN anggota tgt ON tgt.id = er.target_user_id
        WHERE er.status = 'pending' AND er.my_schedule_date <= %s
        """,
        (today,),
    )
    rows = cursor.fetchall() or []
    for row in rows:
        cursor.execute(
            """
            UPDATE task_exchange_requests
            SET status = 'auto_cancelled', auto_cancelled_at = NOW(), updated_at = NOW()
            WHERE id = %s AND status = 'pending'
            """,
            (row.get("id"),),
        )
        row["status"] = "auto_cancelled"
        exchange_notify_requester_result(cursor, row, "auto_cancelled")
        exchange_insert_notification(
            cursor,
            target_user_id=row.get("target_user_id"),
            target_role=row.get("target_role"),
            title="Permintaan Tukar Jadwal Batal Otomatis",
            body="Permintaan tukar/ganti tugas otomatis dibatalkan karena belum direspons sampai hari-H jadwal.",
            request_id=row.get("id"),
            status="auto_cancelled",
            url=None if normalize_role_value(row.get("target_role") or "user") in {"admin", "super_admin"} else "/penukaran-jadwal-tugas-anggota.html",
        )
    return len(rows)


def exchange_request_to_dict(row: dict[str, object], *, direction: str) -> dict[str, object]:
    status = normalize_text(row.get("status") or "pending").lower()
    my_date = normalize_text(row.get("my_schedule_date"))
    target_date = normalize_text(row.get("target_schedule_date"))
    created_at = row.get("created_at")
    responded_at = row.get("responded_at")
    return {
        "id": row.get("id"),
        "direction": direction,
        "type": normalize_text(row.get("kind")) or "biasa",
        "typeLabel": exchange_kind_label(row.get("kind")),
        "mode": normalize_text(row.get("request_mode")) or "swap",
        "modeLabel": exchange_mode_label(row.get("request_mode")),
        "requesterId": str(row.get("requester_id") or ""),
        "requesterName": normalize_text(row.get("requester_name")) or "Anggota",
        "targetUserId": str(row.get("target_user_id") or ""),
        "targetName": normalize_text(row.get("target_name")) or "Teman",
        "myAssignmentId": row.get("my_assignment_id"),
        "targetAssignmentId": row.get("target_assignment_id"),
        "myMisaName": normalize_text(row.get("my_misa_name")),
        "myRole": normalize_text(row.get("my_role_name")),
        "myDate": my_date,
        "myDateLabel": request_task_format_date(my_date),
        "myDayName": request_task_day_name(my_date),
        "myTime": format_time_hhmm(row.get("my_schedule_time")),
        "myLabel": exchange_task_label_from_parts(row.get("my_misa_name"), my_date, row.get("my_schedule_time"), row.get("my_role_name")),
        "targetMisaName": normalize_text(row.get("target_misa_name")),
        "targetRole": normalize_text(row.get("target_role_name")),
        "targetDate": target_date,
        "targetDateLabel": request_task_format_date(target_date),
        "targetDayName": request_task_day_name(target_date),
        "targetTime": format_time_hhmm(row.get("target_schedule_time")),
        "targetLabel": exchange_task_label_from_parts(row.get("target_misa_name"), target_date, row.get("target_schedule_time"), row.get("target_role_name")) if target_date else "-",
        "reason": normalize_text(row.get("reason")),
        "status": status,
        "statusLabel": exchange_status_label(status),
        "canAct": status == "pending" and direction == "incoming" and exchange_date_is_eligible(my_date),
        "canCancel": status == "pending" and direction == "outgoing" and exchange_date_is_eligible(my_date),
        "createdAt": created_at.isoformat() if hasattr(created_at, "isoformat") else normalize_text(created_at),
        "createdAtLabel": created_at.strftime("%d/%m/%Y %H:%M") if hasattr(created_at, "strftime") else normalize_text(created_at),
        "respondedAt": responded_at.isoformat() if hasattr(responded_at, "isoformat") else normalize_text(responded_at),
    }


def exchange_filter_requests(items: list[dict[str, object]], search_text: str = "") -> list[dict[str, object]]:
    keyword = normalize_text(search_text).lower()
    if not keyword:
        return items
    filtered = []
    for item in items:
        haystack = " ".join([
            normalize_text(item.get("requesterName")), normalize_text(item.get("targetName")),
            normalize_text(item.get("myLabel")), normalize_text(item.get("targetLabel")),
            normalize_text(item.get("myRole")), normalize_text(item.get("targetRole")),
            normalize_text(item.get("reason")), normalize_text(item.get("statusLabel")),
            normalize_text(item.get("typeLabel")), normalize_text(item.get("modeLabel")),
        ]).lower()
        if keyword in haystack:
            filtered.append(item)
    return filtered


def exchange_paginate(items: list[dict[str, object]], page: int, page_size: int = 5):
    total = len(items)
    total_pages = max(1, (total + page_size - 1) // page_size)
    page = max(1, min(page, total_pages))
    start = (page - 1) * page_size
    return items[start:start + page_size], {"page": page, "pageSize": page_size, "total": total, "totalPages": total_pages, "hasPrev": page > 1, "hasNext": page < total_pages}


@app.route("/api/task-exchanges/me", methods=["GET"])
def api_task_exchanges_me():
    ensure_task_exchange_schema()
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True, buffered=True)
    try:
        member, error = exchange_current_member(cursor)
        if error:
            return error
        return jsonify({"success": True, "member": {"id": member["id"], "name": member["name"], "role": member["roleNormalized"]}})
    finally:
        cursor.close()
        conn.close()


@app.route("/api/task-exchanges/options", methods=["GET"])
def api_task_exchanges_options():
    ensure_task_exchange_schema()
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True, buffered=True)
    try:
        member, error = require_active_request_member(cursor)
        if error:
            return error
        exchange_expire_pending_requests(cursor)
        conn.commit()
        now = datetime.now()
        month = min(12, max(1, parse_required_int(request.args.get("month"), now.month)))
        year = parse_required_int(request.args.get("year"), now.year)
        kind = normalize_text(request.args.get("type")) or "biasa"
        if kind not in {"biasa", "besar"}:
            kind = "biasa"
        my_tasks = exchange_fetch_user_tasks(cursor, member_id=member["id"], kind=kind, month=month, year=year)
        target_tasks = exchange_fetch_target_tasks(cursor, current_member_id=member["id"], kind=kind, month=month, year=year)
        active_members = exchange_fetch_active_members(cursor, exclude_member_id=member["id"])
        for task in my_tasks:
            task["hasPendingRequest"] = exchange_has_pending_for_assignment(cursor, member["id"], task.get("type"), task.get("assignmentId"))
            task["assignedMemberIds"] = exchange_schedule_member_ids(
                cursor,
                kind=task.get("type"),
                date_text=task.get("date"),
                time_text=task.get("time"),
                misa_id=task.get("misaId"),
            )
        return jsonify({"success": True, "myTasks": my_tasks, "targetTasks": target_tasks, "members": active_members, "filters": {"month": month, "year": year, "type": kind}})
    finally:
        cursor.close()
        conn.close()


@app.route("/api/task-exchanges", methods=["POST"])
def api_task_exchanges_create():
    ensure_task_exchange_schema()
    data = request.get_json(silent=True) or {}
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True, buffered=True)
    try:
        member, error = require_active_request_member(cursor)
        if error:
            return error
        requester_id = member["id"]
        mode = normalize_text(data.get("mode") or data.get("requestMode") or "swap").lower()
        if mode in {"tukeran", "swap"}:
            mode = "swap"
        elif mode in {"menggantikan", "substitute"}:
            mode = "substitute"
        else:
            return jsonify({"success": False, "error": "Tipe request tidak valid."}), 400
        kind = normalize_text(data.get("type") or data.get("kind") or "biasa").lower()
        if kind not in {"biasa", "besar"}:
            return jsonify({"success": False, "error": "Jenis misa tidak valid."}), 400
        my_assignment_id = parse_optional_int(data.get("myAssignmentId"))
        if my_assignment_id is None:
            return jsonify({"success": False, "error": "Pilih jadwal Anda terlebih dahulu."}), 400
        reason = normalize_text(data.get("reason"))
        if not reason:
            return jsonify({"success": False, "error": "Alasan wajib diisi."}), 400
        my_task = exchange_fetch_assignment(cursor, kind=kind, assignment_id=my_assignment_id, member_id=requester_id, for_update=True)
        if not my_task:
            return jsonify({"success": False, "error": "Jadwal Anda tidak ditemukan atau bukan milik Anda."}), 404
        if not exchange_date_is_eligible(my_task.get("date")):
            return jsonify({"success": False, "error": "Penukaran hanya bisa diajukan minimal H-1. Jadwal hari ini atau yang sudah lewat tidak bisa diajukan."}), 400
        if exchange_has_pending_for_assignment(cursor, requester_id, kind, my_assignment_id):
            return jsonify({"success": False, "error": "Jadwal ini sudah memiliki request aktif. Batalkan request sebelumnya terlebih dahulu."}), 409

        target_task = None
        target_user_id = None
        if mode == "swap":
            target_assignment_id = parse_optional_int(data.get("targetAssignmentId"))
            if target_assignment_id is None:
                return jsonify({"success": False, "error": "Pilih jadwal teman yang ingin diajak tukeran."}), 400
            target_task = exchange_fetch_assignment(cursor, kind=kind, assignment_id=target_assignment_id, for_update=True)
            if not target_task:
                return jsonify({"success": False, "error": "Jadwal teman tidak ditemukan."}), 404
            if str(target_task.get("memberId")) == str(requester_id):
                return jsonify({"success": False, "error": "Anda tidak bisa memilih jadwal sendiri sebagai jadwal tukar."}), 400
            if not exchange_date_is_eligible(target_task.get("date")):
                return jsonify({"success": False, "error": "Jadwal teman harus jadwal yang akan datang minimal H-1."}), 400
            if exchange_member_already_in_schedule(cursor, kind=kind, member_id=requester_id, date_text=target_task.get("date"), time_text=target_task.get("time"), misa_id=target_task.get("misaId")):
                return jsonify({"success": False, "error": "Jadwal teman tidak bisa dipilih karena Anda sudah bertugas pada sesi tersebut."}), 409
            if exchange_member_already_in_schedule(cursor, kind=kind, member_id=target_task.get("memberId"), date_text=my_task.get("date"), time_text=my_task.get("time"), misa_id=my_task.get("misaId")):
                return jsonify({"success": False, "error": "Teman tersebut sudah bertugas pada sesi jadwal Anda, sehingga tidak bisa ditukar."}), 409
            target_user_id = str(target_task.get("memberId"))
        else:
            target_user_id = normalize_text(data.get("targetUserId"))
            if not target_user_id:
                return jsonify({"success": False, "error": "Pilih teman pengganti."}), 400
            if str(target_user_id) == str(requester_id):
                return jsonify({"success": False, "error": "Teman pengganti tidak boleh diri sendiri."}), 400
            cursor.execute("SELECT id, nama, role, status_akun FROM anggota WHERE id = %s LIMIT 1", (target_user_id,))
            target_member = cursor.fetchone()
            if not target_member or normalize_status(target_member.get("status_akun") or "aktif") != "aktif":
                return jsonify({"success": False, "error": "Teman pengganti tidak ditemukan atau tidak aktif."}), 404
            if exchange_member_already_in_schedule(cursor, kind=kind, member_id=target_user_id, date_text=my_task.get("date"), time_text=my_task.get("time"), misa_id=my_task.get("misaId")):
                return jsonify({"success": False, "error": "Teman pengganti sudah bertugas pada sesi tersebut, sehingga tidak bisa memegang 2 role."}), 409
            target_task = {"memberId": target_user_id, "memberName": normalize_text(target_member.get("nama")) or "Teman"}

        cursor.execute(
            """
            INSERT INTO task_exchange_requests
            (requester_id, target_user_id, kind, request_mode,
             my_assignment_id, my_misa_id, my_role_id, my_type_label, my_misa_name, my_role_name, my_schedule_date, my_schedule_time,
             target_assignment_id, target_misa_id, target_role_id, target_type_label, target_misa_name, target_role_name, target_schedule_date, target_schedule_time,
             reason, status)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 'pending')
            """,
            (
                requester_id, target_user_id, kind, mode,
                my_task.get("assignmentId"), my_task.get("misaId"), my_task.get("roleId"), my_task.get("typeLabel"), my_task.get("misaName"), my_task.get("role"), my_task.get("date"), format_time_hhmm(my_task.get("time")),
                target_task.get("assignmentId") if mode == "swap" else None,
                target_task.get("misaId") if mode == "swap" else None,
                target_task.get("roleId") if mode == "swap" else None,
                target_task.get("typeLabel") if mode == "swap" else None,
                target_task.get("misaName") if mode == "swap" else None,
                target_task.get("role") if mode == "swap" else None,
                target_task.get("date") if mode == "swap" else None,
                format_time_hhmm(target_task.get("time")) if mode == "swap" else None,
                reason,
            ),
        )
        req_id = cursor.lastrowid
        row = exchange_fetch_request_row(cursor, req_id)
        exchange_notify_new_request(cursor, row)
        conn.commit()
        return jsonify({"success": True, "message": "Request berhasil dikirim.", "request": exchange_request_to_dict(row, direction="outgoing")})
    except Exception as exc:
        conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 400
    finally:
        cursor.close()
        conn.close()


def exchange_list_requests(*, direction: str):
    ensure_task_exchange_schema()
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True, buffered=True)
    try:
        member, error = exchange_current_member(cursor)
        if error:
            return error
        exchange_expire_pending_requests(cursor)
        conn.commit()
        member_id = member["id"]
        search_text = normalize_text(request.args.get("search"))
        sort_mode = normalize_text(request.args.get("sort")) or "created_desc"
        page = max(1, parse_required_int(request.args.get("page"), 1))
        page_size = max(1, min(25, parse_required_int(request.args.get("pageSize"), 5)))
        status_filter = normalize_text(request.args.get("status")) or "all"
        kind_filter = normalize_text(request.args.get("type")) or "all"
        conditions = ["er.target_user_id = %s" if direction == "incoming" else "er.requester_id = %s"]
        params: list[object] = [member_id]
        if status_filter != "all":
            conditions.append("er.status = %s")
            params.append(status_filter)
        if kind_filter in {"biasa", "besar"}:
            conditions.append("er.kind = %s")
            params.append(kind_filter)
        where_clause = " AND ".join(conditions)
        cursor.execute(
            f"""
            SELECT er.*, req.nama AS requester_name, tgt.nama AS target_name,
                   req.role AS requester_role, tgt.role AS target_role
            FROM task_exchange_requests er
            LEFT JOIN anggota req ON req.id = er.requester_id
            LEFT JOIN anggota tgt ON tgt.id = er.target_user_id
            WHERE {where_clause}
            """,
            tuple(params),
        )
        items = [exchange_request_to_dict(row, direction=direction) for row in (cursor.fetchall() or [])]
        items = exchange_filter_requests(items, search_text)
        if sort_mode == "created_asc":
            items.sort(key=lambda item: item.get("createdAt") or "")
        elif sort_mode == "date_asc":
            items.sort(key=lambda item: (item.get("myDate") or "", item.get("myTime") or ""))
        elif sort_mode == "date_desc":
            items.sort(key=lambda item: (item.get("myDate") or "", item.get("myTime") or ""), reverse=True)
        else:
            items.sort(key=lambda item: item.get("createdAt") or "", reverse=True)
        page_items, pagination = exchange_paginate(items, page, page_size)
        return jsonify({"success": True, "items": page_items, "pagination": pagination})
    finally:
        cursor.close()
        conn.close()


@app.route("/api/task-exchanges/incoming", methods=["GET"])
def api_task_exchanges_incoming():
    return exchange_list_requests(direction="incoming")


@app.route("/api/task-exchanges/outgoing", methods=["GET"])
def api_task_exchanges_outgoing():
    return exchange_list_requests(direction="outgoing")


@app.route("/api/task-exchanges/pending-actions", methods=["GET"])
def api_task_exchanges_pending_actions():
    ensure_task_exchange_schema()
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True, buffered=True)
    try:
        member, error = exchange_current_member(cursor)
        if error:
            return error
        exchange_expire_pending_requests(cursor)
        conn.commit()
        limit = max(1, min(20, parse_required_int(request.args.get("limit"), 5)))
        cursor.execute(
            """
            SELECT er.*, req.nama AS requester_name, tgt.nama AS target_name,
                   req.role AS requester_role, tgt.role AS target_role
            FROM task_exchange_requests er
            LEFT JOIN anggota req ON req.id = er.requester_id
            LEFT JOIN anggota tgt ON tgt.id = er.target_user_id
            WHERE er.target_user_id = %s AND er.status = 'pending' AND er.my_schedule_date > CURDATE()
            ORDER BY er.created_at DESC
            LIMIT %s
            """,
            (member["id"], limit),
        )
        items = [exchange_request_to_dict(row, direction="incoming") for row in (cursor.fetchall() or [])]
        return jsonify({"success": True, "items": items})
    finally:
        cursor.close()
        conn.close()


@app.route("/api/task-exchanges/<int:request_id>/respond", methods=["POST"])
def api_task_exchanges_respond(request_id):
    ensure_task_exchange_schema()
    data = request.get_json(silent=True) or {}
    action = normalize_text(data.get("action") or "").lower()
    if action not in {"accept", "reject"}:
        return jsonify({"success": False, "error": "Aksi tidak valid."}), 400
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True, buffered=True)
    try:
        member, error = exchange_current_member(cursor)
        if error:
            return error
        exchange_expire_pending_requests(cursor)
        row = exchange_fetch_request_row(cursor, request_id, for_update=True)
        if not row:
            return jsonify({"success": False, "error": "Request tidak ditemukan."}), 404
        if str(row.get("target_user_id")) != str(member["id"]):
            return jsonify({"success": False, "error": "Request ini bukan untuk Anda."}), 403
        if normalize_text(row.get("status")) != "pending":
            return jsonify({"success": False, "error": f"Request sudah berstatus {exchange_status_label(row.get('status'))}."}), 400
        if not exchange_date_is_eligible(normalize_text(row.get("my_schedule_date"))):
            cursor.execute("UPDATE task_exchange_requests SET status = 'auto_cancelled', auto_cancelled_at = NOW(), updated_at = NOW() WHERE id = %s", (request_id,))
            row["status"] = "auto_cancelled"
            exchange_notify_requester_result(cursor, row, "auto_cancelled")
            conn.commit()
            return jsonify({"success": False, "error": "Request otomatis batal karena sudah masuk hari-H atau jadwal sudah lewat."}), 400
        if action == "reject":
            cursor.execute("UPDATE task_exchange_requests SET status = 'rejected', responded_at = NOW(), updated_at = NOW() WHERE id = %s", (request_id,))
            row["status"] = "rejected"
            exchange_notify_requester_result(cursor, row, "rejected")
            conn.commit()
            return jsonify({"success": True, "message": "Request berhasil ditolak."})

        kind = normalize_text(row.get("kind") or "biasa")
        mode = normalize_text(row.get("request_mode") or "swap")
        my_task = exchange_fetch_assignment(cursor, kind=kind, assignment_id=row.get("my_assignment_id"), member_id=row.get("requester_id"), for_update=True)
        if not my_task:
            return jsonify({"success": False, "error": "Jadwal pengaju sudah tidak valid."}), 409
        if not exchange_date_is_eligible(my_task.get("date")):
            return jsonify({"success": False, "error": "Jadwal pengaju sudah masuk hari-H atau lewat."}), 400
        if mode == "swap":
            target_task = exchange_fetch_assignment(cursor, kind=kind, assignment_id=row.get("target_assignment_id"), member_id=member["id"], for_update=True)
            if not target_task:
                return jsonify({"success": False, "error": "Jadwal Anda untuk ditukar sudah tidak valid."}), 409
            if not exchange_date_is_eligible(target_task.get("date")):
                return jsonify({"success": False, "error": "Jadwal Anda sudah masuk hari-H atau lewat."}), 400
            if exchange_member_already_in_schedule(cursor, kind=kind, member_id=row.get("requester_id"), date_text=target_task.get("date"), time_text=target_task.get("time"), misa_id=target_task.get("misaId")):
                return jsonify({"success": False, "error": "Pengaju sudah bertugas pada sesi jadwal Anda. Request tidak bisa diterima karena akan membuat double role."}), 409
            if exchange_member_already_in_schedule(cursor, kind=kind, member_id=member["id"], date_text=my_task.get("date"), time_text=my_task.get("time"), misa_id=my_task.get("misaId")):
                return jsonify({"success": False, "error": "Anda sudah bertugas pada sesi jadwal pengaju. Request tidak bisa diterima karena akan membuat double role."}), 409
            if kind == "besar":
                cursor.execute("UPDATE misa_besar_assignments SET member_id = %s, request_source = 'exchange', created_at = NOW() WHERE id = %s", (member["id"], my_task.get("assignmentId")))
                cursor.execute("UPDATE misa_besar_assignments SET member_id = %s, request_source = 'exchange', created_at = NOW() WHERE id = %s", (row.get("requester_id"), target_task.get("assignmentId")))
            else:
                cursor.execute("UPDATE streaming_assignments SET member_id = %s, request_source = 'exchange', created_at = NOW() WHERE id = %s", (member["id"], my_task.get("assignmentId")))
                cursor.execute("UPDATE streaming_assignments SET member_id = %s, request_source = 'exchange', created_at = NOW() WHERE id = %s", (row.get("requester_id"), target_task.get("assignmentId")))
        else:
            if exchange_member_already_in_schedule(cursor, kind=kind, member_id=member["id"], date_text=my_task.get("date"), time_text=my_task.get("time"), misa_id=my_task.get("misaId"), exclude_assignment_id=my_task.get("assignmentId")):
                return jsonify({"success": False, "error": "Anda sudah bertugas pada jadwal/misa tersebut."}), 409
            if kind == "besar":
                cursor.execute("UPDATE misa_besar_assignments SET member_id = %s, request_source = 'exchange', created_at = NOW() WHERE id = %s", (member["id"], my_task.get("assignmentId")))
            else:
                cursor.execute("UPDATE streaming_assignments SET member_id = %s, request_source = 'exchange', created_at = NOW() WHERE id = %s", (member["id"], my_task.get("assignmentId")))
        cursor.execute("UPDATE task_exchange_requests SET status = 'accepted', responded_at = NOW(), updated_at = NOW() WHERE id = %s", (request_id,))
        row["status"] = "accepted"
        exchange_notify_requester_result(cursor, row, "accepted")
        conn.commit()
        return jsonify({"success": True, "message": "Request berhasil diterima dan jadwal sudah diperbarui."})
    except Exception as exc:
        conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 400
    finally:
        cursor.close()
        conn.close()


@app.route("/api/task-exchanges/<int:request_id>/cancel", methods=["POST"])
def api_task_exchanges_cancel(request_id):
    ensure_task_exchange_schema()
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True, buffered=True)
    try:
        member, error = require_active_request_member(cursor)
        if error:
            return error
        row = exchange_fetch_request_row(cursor, request_id, for_update=True)
        if not row:
            return jsonify({"success": False, "error": "Request tidak ditemukan."}), 404
        if str(row.get("requester_id")) != str(member["id"]):
            return jsonify({"success": False, "error": "Anda hanya bisa membatalkan request keluar milik Anda."}), 403
        if normalize_text(row.get("status")) != "pending":
            return jsonify({"success": False, "error": "Request ini sudah tidak bisa dibatalkan."}), 400
        cursor.execute("UPDATE task_exchange_requests SET status = 'cancelled', cancelled_at = NOW(), updated_at = NOW() WHERE id = %s", (request_id,))
        row["status"] = "cancelled"
        target_role = row.get("target_role")
        exchange_insert_notification(
            cursor,
            target_user_id=row.get("target_user_id"),
            target_role=target_role,
            title="Permintaan Tukar Jadwal Dibatalkan",
            body=f"Permintaan tukar/ganti tugas dari <b>{html.escape(member['name'])}</b> telah dibatalkan oleh pengaju.",
            request_id=request_id,
            status="cancelled",
            url=None if normalize_role_value(target_role or "user") in {"admin", "super_admin"} else "/penukaran-jadwal-tugas-anggota.html",
        )
        exchange_notify_requester_result(cursor, row, "cancelled")
        conn.commit()
        return jsonify({"success": True, "message": "Request berhasil dibatalkan."})
    except Exception as exc:
        conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 400
    finally:
        cursor.close()
        conn.close()

# ---------------------------------------------------------------------------
# Monitoring Tugas Anggota - kewajiban bulanan Misa Biasa saja
# ---------------------------------------------------------------------------

def ensure_monthly_monitoring_schema(cursor) -> None:
    """Schema kecil untuk menyimpan target minimum tugas bulanan.

    Default target minimum disamakan dengan kebutuhan sistem saat ini: 3 tugas
    Misa Biasa per bulan. Jika row default sudah ada, nilainya tidak ditimpa;
    admin tetap bisa mengubahnya dari halaman monitoring admin.
    """
    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS `monthly_task_settings` (
          `id` varchar(50) NOT NULL,
          `target_minimum` int NOT NULL DEFAULT 3,
          `updated_at` datetime NOT NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
          PRIMARY KEY (`id`)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_general_ci
        """
    )
    try:
        ensure_column(cursor, "monthly_task_settings", "target_minimum", "`target_minimum` int NOT NULL DEFAULT 3")
    except Exception:
        pass
    cursor.execute(
        """
        INSERT IGNORE INTO `monthly_task_settings` (`id`, `target_minimum`)
        VALUES ('default', 3)
        """
    )
    try:
        ensure_column(cursor, "anggota", "inactive_until", "`inactive_until` date DEFAULT NULL")
    except Exception:
        pass
def monitoring_require_admin():
    if not session.get("logged_in"):
        return jsonify({"success": False, "error": "Anda harus login terlebih dahulu."}), 401
    if normalize_role_value(session.get("role") or "user") not in {"admin", "super_admin"}:
        return jsonify({"success": False, "error": "Akses hanya untuk admin/super admin."}), 403
    return None


def monitoring_get_target_minimum(cursor) -> int:
    """Ambil target minimum bulanan dengan fallback aman.

    Target default sekarang 3. Jika tabel/row belum ada, backend membuat row
    default bernilai 3 agar halaman anggota tidak menampilkan 0/2 lagi.
    """
    try:
        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS `monthly_task_settings` (
              `id` varchar(50) NOT NULL,
              `target_minimum` int NOT NULL DEFAULT 3,
              `updated_at` datetime NOT NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
              PRIMARY KEY (`id`)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_general_ci
            """
        )
        try:
            ensure_column(cursor, "monthly_task_settings", "target_minimum", "`target_minimum` int NOT NULL DEFAULT 3")
        except Exception:
            pass
        cursor.execute("SELECT target_minimum FROM monthly_task_settings WHERE id = 'default' LIMIT 1")
        row = cursor.fetchone()
        if not row:
            cursor.execute("INSERT INTO monthly_task_settings (id, target_minimum) VALUES ('default', 3)")
            return 3
        value = fetch_scalar_value(row, 3)
        return max(1, min(99, parse_required_int(value, 3)))
    except Exception as exc:
        print(f"[WARN] Gagal membaca target minimum monitoring: {exc}")
        return 3
def monitoring_account_role_label(role_value: str) -> str:
    role = normalize_role_value(role_value or "user")
    if role == "super_admin":
        return "Super Admin"
    if role == "admin":
        return "Admin"
    return "Anggota"


def monitoring_is_member_active_for_period(row: dict[str, object], year: int, month: int) -> bool:
    # Saat ini status aktif/nonaktif di database menjadi sumber utama.
    # Jika kolom inactive_until tersedia dan tanggalnya sudah lewat sebelum periode, anggota dianggap aktif lagi.
    status = normalize_status(row.get("status_akun") or "aktif")
    if status == "aktif":
        return True
    inactive_until = row.get("inactive_until")
    if not inactive_until:
        return False
    try:
        if hasattr(inactive_until, "year"):
            until_date = inactive_until
        else:
            until_date = datetime.strptime(str(inactive_until)[:10], "%Y-%m-%d").date()
        period_start = datetime(year, month, 1).date()
        return until_date <= period_start
    except Exception:
        return False


def monitoring_status(total_tasks: int, target_minimum: int) -> tuple[str, str, int]:
    shortage = max(0, int(target_minimum) - int(total_tasks or 0))
    if shortage <= 0:
        return "Aman", "ok", shortage
    if shortage == 1:
        return "Perlu Tambahan", "warn", shortage
    return "Kritis", "danger", shortage


def monitoring_count_regular_assignments(cursor, year: int, month: int) -> dict[int, int]:
    cursor.execute(
        """
        SELECT member_id, COUNT(*) AS total
        FROM streaming_assignments
        WHERE YEAR(schedule_date) = %s AND MONTH(schedule_date) = %s
        GROUP BY member_id
        """,
        (year, month),
    )
    result: dict[int, int] = {}
    for row in cursor.fetchall() or []:
        try:
            result[int(row.get("member_id") if isinstance(row, dict) else row[0])] = int(row.get("total") if isinstance(row, dict) else row[1])
        except Exception:
            continue
    return result


def monitoring_fetch_members(cursor, year: int, month: int) -> list[dict[str, object]]:
    try:
        ensure_column(cursor, "anggota", "inactive_until", "`inactive_until` date DEFAULT NULL")
    except Exception:
        pass
    cursor.execute(
        """
        SELECT id, nama, username, role, status_akun, inactive_until
        FROM anggota
        ORDER BY nama ASC, id ASC
        """
    )
    rows = cursor.fetchall() or []
    active_members: list[dict[str, object]] = []
    for row in rows:
        if not monitoring_is_member_active_for_period(row, year, month):
            continue
        role = normalize_role_value(row.get("role") or "user")
        active_members.append({
            "id": row.get("id"),
            "name": normalize_text(row.get("nama") or row.get("username")) or f"Anggota {row.get('id')}",
            "accountRole": role,
            "accountRoleLabel": monitoring_account_role_label(role),
        })
    return active_members


def monitoring_month_range(year: int, month: int) -> tuple[str, str]:
    last_day = calendar.monthrange(year, month)[1]
    return f"{year}-{month:02d}-01", f"{year}-{month:02d}-{last_day:02d}"


def monitoring_regular_slot_conflict(cursor, date_text: str, time_text: str) -> bool:
    time_clean = format_time_hhmm(time_text)
    cursor.execute(
        """
        SELECT 1 FROM streaming_cancelled
        WHERE mass_date = %s AND DATE_FORMAT(mass_time, '%H:%i') = %s
        LIMIT 1
        """,
        (date_text, time_clean),
    )
    if cursor.fetchone():
        return True
    cursor.execute(
        """
        SELECT 1 FROM misa_besar
        WHERE status = 'published' AND misa_date = %s AND DATE_FORMAT(misa_time, '%H:%i') = %s
        LIMIT 1
        """,
        (date_text, time_clean),
    )
    return cursor.fetchone() is not None


def monitoring_open_regular_slots(cursor, year: int, month: int, member_id: object | None = None) -> list[dict[str, object]]:
    """Ambil slot misa biasa yang masih punya role kosong, hanya tanggal setelah hari ini."""
    roles = request_task_fetch_roles(cursor)
    if not roles:
        return []
    cursor.execute(
        """
        SELECT day_name, mass_name, DATE_FORMAT(start_time, '%H:%i') AS start_time
        FROM streaming_weekly_config
        ORDER BY FIELD(day_name, 'Senin','Selasa','Rabu','Kamis','Jumat','Sabtu','Minggu'), start_time ASC, id ASC
        """
    )
    configs = cursor.fetchall() or []
    cursor.execute(
        """
        SELECT DATE_FORMAT(schedule_date, '%Y-%m-%d') AS schedule_date,
               DATE_FORMAT(schedule_time, '%H:%i') AS schedule_time,
               role_name, member_id
        FROM streaming_assignments
        WHERE YEAR(schedule_date) = %s AND MONTH(schedule_date) = %s
        """,
        (year, month),
    )
    assignment_map: dict[tuple[str, str], dict[str, str]] = {}
    assigned_members_by_slot: dict[tuple[str, str], set[str]] = {}
    for row in cursor.fetchall() or []:
        key = (normalize_text(row.get("schedule_date")), format_time_hhmm(row.get("schedule_time")))
        assignment_map.setdefault(key, {})[normalize_text(row.get("role_name"))] = str(row.get("member_id") or "")
        assigned_members_by_slot.setdefault(key, set()).add(str(row.get("member_id") or ""))

    today = datetime.now().date()
    items: list[dict[str, object]] = []
    for day in range(1, calendar.monthrange(year, month)[1] + 1):
        date_obj = datetime(year, month, day).date()
        if date_obj <= today:
            continue
        date_text = date_obj.strftime("%Y-%m-%d")
        day_name = DAYS_INDO[date_obj.weekday()]
        for cfg in configs:
            if normalize_text(cfg.get("day_name")) != day_name:
                continue
            time_text = format_time_hhmm(cfg.get("start_time"))
            if monitoring_regular_slot_conflict(cursor, date_text, time_text):
                continue
            key = (date_text, time_text)
            if member_id and str(member_id) in assigned_members_by_slot.get(key, set()):
                # 1 orang hanya boleh 1 role dalam 1 misa/sesi.
                continue
            filled_by_role = assignment_map.get(key, {})
            open_roles = [role for role in roles if not filled_by_role.get(role)]
            if not open_roles:
                continue
            items.append({
                "id": request_task_schedule_key("biasa", date_text, time_text),
                "date": date_text,
                "dateLabel": request_task_format_date(date_text),
                "dayName": day_name,
                "time": time_text,
                "misaName": normalize_text(cfg.get("mass_name")) or "Misa Biasa",
                "openRoles": open_roles,
            })
    items.sort(key=lambda item: (item.get("date"), item.get("time")))
    return items


def monitoring_has_open_future_slots(cursor, year: int, month: int) -> bool:
    return bool(monitoring_open_regular_slots(cursor, year, month, member_id=None))


def monitoring_notification_url_for_role(role_value: str) -> str:
    role = normalize_role_value(role_value or "user")
    if role == "admin":
        return "/jadwal-tugas-streaming-admin.html"
    return "/request-tugas-anggota.html"


def create_monthly_requirement_notifications() -> int:
    """Buat notifikasi target bulanan untuk user/admin biasa, bukan super_admin.

    - H-7 dan H-1 sebelum bulan berikutnya.
    - Harian pada bulan berjalan jika masih kurang dan masih ada slot masa depan yang kosong.
    """
    ensure_auth_schema()
    ensure_streaming_schema()
    ensure_notifications_schema()
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True, buffered=True)
    sent = 0
    try:
        ensure_monthly_monitoring_schema(cursor)
        target = monitoring_get_target_minimum(cursor)
        today = datetime.now().date()

        periods: list[tuple[int, int, str, str]] = []
        # Notifikasi harian bulan berjalan.
        periods.append((today.year, today.month, "daily_current", "Pengingat target tugas bulan ini"))
        # H-7 / H-1 untuk bulan berikutnya.
        first_next_month = (datetime(today.year, today.month, 28).date() + timedelta(days=4)).replace(day=1)
        h7_date = first_next_month - timedelta(days=7)
        h1_date = first_next_month - timedelta(days=1)
        if today == h7_date:
            periods.append((first_next_month.year, first_next_month.month, "h7_next_month", "Pengingat H-7 target tugas bulan depan"))
        if today == h1_date:
            periods.append((first_next_month.year, first_next_month.month, "h1_next_month", "Pengingat H-1 target tugas bulan depan"))

        for year, month, code, title in periods:
            if not monitoring_has_open_future_slots(cursor, year, month):
                continue
            counts = monitoring_count_regular_assignments(cursor, year, month)
            members = monitoring_fetch_members(cursor, year, month)
            for member in members:
                role_value = normalize_role_value(member.get("accountRole") or "user")
                if role_value == "super_admin":
                    continue
                total = counts.get(int(member.get("id") or 0), 0)
                shortage = max(0, target - total)
                if shortage <= 0:
                    continue
                month_name = calendar.month_name[month]
                month_label = ["", "Januari", "Februari", "Maret", "April", "Mei", "Juni", "Juli", "Agustus", "September", "Oktober", "November", "Desember"][month]
                body = (
                    f"Target tugas Misa Biasa bulan <b>{html.escape(month_label)} {year}</b> belum terpenuhi. "
                    f"Saat ini Anda memiliki <b>{total}</b> tugas dari target minimum <b>{target}</b>; "
                    f"masih kurang <b>{shortage}</b> tugas. Silakan mengambil/request jadwal yang masih kosong."
                )
                dedupe_day = today.strftime("%Y-%m-%d") if code == "daily_current" else code
                dedupe_key = f"monthly-monitoring:{year}-{month:02d}:{member.get('id')}:{code}:{dedupe_day}:target{target}"
                create_notification_once(
                    cursor,
                    "tugas",
                    title,
                    body,
                    monitoring_notification_url_for_role(role_value),
                    {
                        "target_user_id": str(member.get("id")),
                        "notification_kind": "monthly_task_requirement",
                        "period_year": year,
                        "period_month": month,
                        "target_minimum": target,
                        "current_total": total,
                        "shortage": shortage,
                    },
                    target_role="admin" if role_value == "admin" else "user",
                    dedupe_key=dedupe_key,
                )
                sent += 1
        conn.commit()
        return sent
    except Exception:
        conn.rollback()
        raise
    finally:
        cursor.close()
        conn.close()


@app.route("/api/monitoring-tugas", methods=["GET"])
def api_monitoring_tugas():
    auth_error = monitoring_require_admin()
    if auth_error:
        return auth_error
    ensure_auth_schema()
    ensure_streaming_schema()
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True, buffered=True)
    try:
        ensure_monthly_monitoring_schema(cursor)
        now = datetime.now()
        month = max(1, min(12, parse_required_int(request.args.get("month"), now.month)))
        year = parse_required_int(request.args.get("year"), now.year)
        search_text = normalize_text(request.args.get("search")).lower()
        role_filter = normalize_text(request.args.get("role") or "all")
        status_filter = normalize_text(request.args.get("status") or "all")
        sort_mode = normalize_text(request.args.get("sort") or "name_asc")
        target = monitoring_get_target_minimum(cursor)
        counts = monitoring_count_regular_assignments(cursor, year, month)
        members = monitoring_fetch_members(cursor, year, month)

        items: list[dict[str, object]] = []
        for member in members:
            total = counts.get(int(member.get("id") or 0), 0)
            label, class_name, shortage = monitoring_status(total, target)
            item = {
                "id": member.get("id"),
                "name": member.get("name"),
                "accountRole": member.get("accountRole"),
                "accountRoleLabel": member.get("accountRoleLabel"),
                "totalTasks": total,
                "shortage": shortage,
                "status": label,
                "statusClass": class_name,
                "canSchedule": shortage > 0,
            }
            items.append(item)

        if role_filter != "all":
            items = [i for i in items if normalize_text(i.get("accountRole")) == role_filter]
        if status_filter != "all":
            items = [i for i in items if normalize_text(i.get("status")) == status_filter]
        if search_text:
            items = [i for i in items if search_text in (normalize_text(i.get("name")) + " " + normalize_text(i.get("accountRoleLabel"))).lower()]

        if sort_mode == "name_desc":
            items.sort(key=lambda i: normalize_text(i.get("name")).lower(), reverse=True)
        elif sort_mode == "total_asc":
            items.sort(key=lambda i: (i.get("totalTasks") or 0, normalize_text(i.get("name")).lower()))
        elif sort_mode == "total_desc":
            items.sort(key=lambda i: (i.get("totalTasks") or 0, normalize_text(i.get("name")).lower()), reverse=True)
        elif sort_mode == "shortage_asc":
            items.sort(key=lambda i: (i.get("shortage") or 0, normalize_text(i.get("name")).lower()))
        elif sort_mode == "shortage_desc":
            items.sort(key=lambda i: (i.get("shortage") or 0, normalize_text(i.get("name")).lower()), reverse=True)
        elif sort_mode == "status_asc":
            order = {"Aman": 0, "Perlu Tambahan": 1, "Kritis": 2}
            items.sort(key=lambda i: (order.get(normalize_text(i.get("status")), 9), normalize_text(i.get("name")).lower()))
        elif sort_mode == "status_desc":
            order = {"Aman": 0, "Perlu Tambahan": 1, "Kritis": 2}
            items.sort(key=lambda i: (order.get(normalize_text(i.get("status")), 9), normalize_text(i.get("name")).lower()), reverse=True)
        else:
            items.sort(key=lambda i: normalize_text(i.get("name")).lower())

        # Ringkasan dihitung dari seluruh anggota aktif pada periode tersebut, bukan hanya hasil search.
        all_statuses = []
        for member in members:
            total = counts.get(int(member.get("id") or 0), 0)
            label, class_name, shortage = monitoring_status(total, target)
            all_statuses.append((label, shortage))
        summary = {
            "totalMembers": len(members),
            "completedMembers": sum(1 for label, shortage in all_statuses if shortage == 0),
            "shortage1Members": sum(1 for label, shortage in all_statuses if shortage == 1),
            "shortageMultiMembers": sum(1 for label, shortage in all_statuses if shortage > 1),
            "targetMinimum": target,
        }
        return jsonify({
            "success": True,
            "items": items,
            "summary": summary,
            "period": {"month": month, "year": year},
            "roles": [
                {"value": "user", "label": "Anggota"},
                {"value": "admin", "label": "Admin"},
                {"value": "super_admin", "label": "Super Admin"},
            ],
        })
    finally:
        cursor.close()
        conn.close()


@app.route("/api/monitoring-tugas/target", methods=["POST"])
def api_monitoring_tugas_target():
    auth_error = monitoring_require_admin()
    if auth_error:
        return auth_error
    payload = request.get_json(silent=True) or {}
    target = max(1, min(99, parse_required_int(payload.get("targetMinimum"), 2)))
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        ensure_monthly_monitoring_schema(cursor)
        cursor.execute(
            """
            INSERT INTO monthly_task_settings (`id`, `target_minimum`)
            VALUES ('default', %s)
            ON DUPLICATE KEY UPDATE target_minimum = VALUES(target_minimum), updated_at = CURRENT_TIMESTAMP
            """,
            (target,),
        )
        conn.commit()
        return jsonify({"success": True, "targetMinimum": target})
    finally:
        cursor.close()
        conn.close()


@app.route("/api/monitoring-tugas/slots", methods=["GET"])
def api_monitoring_tugas_slots():
    auth_error = monitoring_require_admin()
    if auth_error:
        return auth_error
    ensure_auth_schema()
    ensure_streaming_schema()
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True, buffered=True)
    try:
        ensure_monthly_monitoring_schema(cursor)
        now = datetime.now()
        month = max(1, min(12, parse_required_int(request.args.get("month"), now.month)))
        year = parse_required_int(request.args.get("year"), now.year)
        member_id = request.args.get("memberId")
        if not member_id:
            return jsonify({"success": False, "error": "Member wajib dipilih."}), 400
        cursor.execute("SELECT id, nama, role, status_akun FROM anggota WHERE id = %s LIMIT 1", (member_id,))
        member = cursor.fetchone()
        if not member:
            return jsonify({"success": False, "error": "Anggota tidak ditemukan."}), 404
        if normalize_status(member.get("status_akun") or "aktif") != "aktif":
            return jsonify({"success": False, "error": "Anggota sedang nonaktif."}), 400
        slots = monitoring_open_regular_slots(cursor, year, month, member_id=member_id)
        return jsonify({"success": True, "slots": slots, "member": {"id": member.get("id"), "name": member.get("nama")}})
    finally:
        cursor.close()
        conn.close()


@app.route("/api/monitoring-tugas/schedule", methods=["POST"])
def api_monitoring_tugas_schedule():
    auth_error = monitoring_require_admin()
    if auth_error:
        return auth_error
    payload = request.get_json(silent=True) or {}
    member_id = normalize_text(payload.get("memberId"))
    date_text = normalize_text(payload.get("date"))
    time_text = format_time_hhmm(payload.get("time"))
    role_name = normalize_text(payload.get("role"))
    if not member_id or not date_text or not time_text or not role_name:
        return jsonify({"success": False, "error": "Anggota, tanggal, jam, dan role wajib dipilih."}), 400
    try:
        date_obj = datetime.strptime(date_text, "%Y-%m-%d").date()
    except Exception:
        return jsonify({"success": False, "error": "Tanggal tidak valid."}), 400
    if date_obj <= datetime.now().date():
        return jsonify({"success": False, "error": "Hanya jadwal mulai besok dan seterusnya yang bisa dipilih."}), 400

    ensure_auth_schema()
    ensure_streaming_schema()
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True, buffered=True)
    try:
        cursor.execute("SELECT id, nama, role, status_akun FROM anggota WHERE id = %s LIMIT 1", (member_id,))
        member = cursor.fetchone()
        if not member:
            return jsonify({"success": False, "error": "Anggota tidak ditemukan."}), 404
        if normalize_status(member.get("status_akun") or "aktif") != "aktif":
            return jsonify({"success": False, "error": "Anggota sedang nonaktif."}), 400
        cursor.execute("SELECT role_name FROM streaming_roles WHERE role_name = %s LIMIT 1", (role_name,))
        if not cursor.fetchone():
            return jsonify({"success": False, "error": "Role tidak valid."}), 400
        cfg = request_task_get_regular_cfg(cursor, date_text, time_text)
        if not cfg:
            return jsonify({"success": False, "error": "Jadwal Misa Biasa tidak ditemukan pada tanggal/jam tersebut."}), 400
        if monitoring_regular_slot_conflict(cursor, date_text, time_text):
            return jsonify({"success": False, "error": "Jadwal diblokir/ditiadakan atau bentrok Misa Besar."}), 400
        cursor.execute(
            """
            SELECT role_name FROM streaming_assignments
            WHERE schedule_date = %s AND DATE_FORMAT(schedule_time, '%H:%i') = %s AND member_id = %s
            LIMIT 1
            """,
            (date_text, time_text, member_id),
        )
        if cursor.fetchone():
            return jsonify({"success": False, "error": "Anggota ini sudah bertugas di sesi yang sama."}), 400
        cursor.execute(
            """
            SELECT member_id FROM streaming_assignments
            WHERE schedule_date = %s AND DATE_FORMAT(schedule_time, '%H:%i') = %s AND role_name = %s
            LIMIT 1
            """,
            (date_text, time_text, role_name),
        )
        if cursor.fetchone():
            return jsonify({"success": False, "error": "Role ini sudah terisi."}), 409
        cursor.execute(
            """
            INSERT INTO streaming_assignments (schedule_date, schedule_time, role_name, member_id, request_source, created_at)
            VALUES (%s, %s, %s, %s, 'admin', CURRENT_TIMESTAMP)
            """,
            (date_text, time_text, role_name, member_id),
        )
        ensure_notifications_schema()
        day_name = request_task_day_name(date_text)
        date_label = request_task_format_date(date_text)
        misa_name = normalize_text(cfg.get("mass_name")) or "Misa Biasa"
        title = f"Tugas Baru: {role_name}"
        body = (
            f"Anda dijadwalkan sebagai <b>{html.escape(role_name)}</b> untuk <b>{html.escape(misa_name)}</b> "
            f"pada hari {html.escape(day_name)}, {html.escape(date_label)} jam {html.escape(time_text)} WIB."
        )
        create_notification(
            cursor,
            "tugas",
            title,
            body,
            notification_target_url_for_member_role(member.get("role"), default_user_url="/jadwal-tugas-misa-anggota.html"),
            {
                "target_user_id": str(member_id),
                "notification_kind": "monitoring_admin_schedule",
                "misa_type": "misa_biasa",
                "misa_name": misa_name,
                "misa_date": date_text,
                "misa_time": time_text,
                "role": role_name,
            },
            target_role=None,
        )
        conn.commit()
        return jsonify({"success": True, "message": f"{member.get('nama')} berhasil dijadwalkan sebagai {role_name}."})
    except Exception as exc:
        conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 400
    finally:
        cursor.close()
        conn.close()



# ---------------------------------------------------------------------------
# Monitoring Kewajiban Tugas Anggota - statistik personal anggota
# ---------------------------------------------------------------------------

def member_monitoring_current_user(cursor):
    """Ambil user login terbaru dari database untuk halaman monitoring anggota."""
    if not session.get("logged_in"):
        return None, (jsonify({"success": False, "error": "Anda harus login terlebih dahulu."}), 401)
    user_id = session.get("user_id")
    if not user_id:
        return None, (jsonify({"success": False, "error": "Sesi login tidak valid."}), 401)
    cursor.execute(
        """
        SELECT id, nama, username, role, status_akun
        FROM anggota
        WHERE id = %s
        LIMIT 1
        """,
        (user_id,),
    )
    user = cursor.fetchone()
    if not user:
        return None, (jsonify({"success": False, "error": "Akun tidak ditemukan."}), 404)
    if normalize_status(user.get("status_akun") or "aktif") != "aktif":
        return None, (jsonify({"success": False, "error": "Akun Anda sedang nonaktif."}), 403)
    user["id"] = str(user.get("id"))
    user["name"] = normalize_text(user.get("nama") or user.get("username") or "Anggota")
    user["roleNormalized"] = normalize_role_value(user.get("role") or "user")
    user["roleLabel"] = monitoring_account_role_label(user.get("role") or "user")
    return user, None


def member_monitoring_month_name(month: int) -> str:
    month_names = [
        "Januari", "Februari", "Maret", "April", "Mei", "Juni",
        "Juli", "Agustus", "September", "Oktober", "November", "Desember",
    ]
    try:
        return month_names[int(month) - 1]
    except Exception:
        return "-"


def member_monitoring_period_label(month: str | int, year: str | int) -> str:
    month_text = normalize_text(month)
    year_text = normalize_text(year)
    if month_text.lower() == "all" and year_text.lower() == "all":
        return "Semua Bulan & Semua Tahun"
    if month_text.lower() == "all":
        return f"Semua Bulan {year_text}"
    if year_text.lower() == "all":
        return f"{member_monitoring_month_name(parse_required_int(month_text, 1))} Semua Tahun"
    return f"{member_monitoring_month_name(parse_required_int(month_text, 1))} {year_text}"


def member_monitoring_parse_month(value, default_month: int | None = None) -> str:
    raw = normalize_text(value)
    if raw.lower() == "all":
        return "all"
    if raw == "":
        return str(default_month or datetime.now().month)
    return str(min(12, max(1, parse_required_int(raw, default_month or datetime.now().month))))


def member_monitoring_parse_year(value, default_year: int | None = None) -> str:
    raw = normalize_text(value)
    if raw.lower() == "all":
        return "all"
    if raw == "":
        return str(default_year or datetime.now().year)
    return str(parse_required_int(raw, default_year or datetime.now().year))


def member_monitoring_schedule_dt(date_text: str, time_text: str) -> datetime:
    try:
        return datetime.strptime(f"{normalize_text(date_text)} {format_time_hhmm(time_text)}", "%Y-%m-%d %H:%M")
    except Exception:
        try:
            return datetime.strptime(normalize_text(date_text), "%Y-%m-%d")
        except Exception:
            return datetime(1970, 1, 1)


def member_monitoring_task_item(kind: str, type_label: str, misa_name: str, date_text: str, time_text: str, role_name: str, created_at=None) -> dict[str, object]:
    schedule_dt = member_monitoring_schedule_dt(date_text, time_text)
    is_completed = schedule_dt < datetime.now()
    created_text = created_at.isoformat() if hasattr(created_at, "isoformat") else normalize_text(created_at)
    return {
        "type": kind,
        "typeLabel": type_label,
        "misaName": normalize_text(misa_name) or type_label,
        "date": normalize_text(date_text),
        "dateLabel": request_task_format_date(normalize_text(date_text)),
        "dayName": request_task_day_name(normalize_text(date_text)),
        "time": format_time_hhmm(time_text),
        "role": normalize_text(role_name) or "-",
        "status": "Selesai" if is_completed else "Terjadwal",
        "completed": is_completed,
        "createdAt": created_text,
        "display": f"{normalize_text(misa_name) or type_label} - {request_task_day_name(normalize_text(date_text))}, {request_task_format_date(normalize_text(date_text))} jam {format_time_hhmm(time_text)} WIB",
    }


def member_monitoring_fetch_regular_tasks(cursor, member_id: object, *, month: str = "all", year: str = "all") -> list[dict[str, object]]:
    """Ambil semua tugas Misa Biasa milik anggota secara defensif."""
    try:
        ensure_streaming_schema()
    except Exception as exc:
        print(f"[WARN] ensure_streaming_schema monitoring anggota gagal: {exc}")

    def build_filters(alias: str = "sa"):
        where_parts = [f"{alias}.member_id = %s"]
        params_local: list[object] = [member_id]
        if normalize_text(month).lower() != "all":
            where_parts.append(f"MONTH({alias}.schedule_date) = %s")
            params_local.append(parse_required_int(month, datetime.now().month))
        if normalize_text(year).lower() != "all":
            where_parts.append(f"YEAR({alias}.schedule_date) = %s")
            params_local.append(parse_required_int(year, datetime.now().year))
        return " AND ".join(where_parts), params_local

    rows = []
    where_sql, params = build_filters("sa")
    try:
        cursor.execute(
            f"""
            SELECT DATE_FORMAT(sa.schedule_date, '%Y-%m-%d') AS date,
                   TIME_FORMAT(sa.schedule_time, '%H:%i') AS time,
                   sa.role_name,
                   sa.created_at,
                   COALESCE(cfg.mass_name, 'Misa Biasa') AS mass_name
            FROM streaming_assignments sa
            LEFT JOIN streaming_weekly_config cfg
              ON cfg.day_name = CASE WEEKDAY(sa.schedule_date)
                WHEN 0 THEN 'Senin' WHEN 1 THEN 'Selasa' WHEN 2 THEN 'Rabu'
                WHEN 3 THEN 'Kamis' WHEN 4 THEN 'Jumat' WHEN 5 THEN 'Sabtu'
                ELSE 'Minggu' END
              AND TIME_FORMAT(cfg.start_time, '%H:%i') = TIME_FORMAT(sa.schedule_time, '%H:%i')
            LEFT JOIN streaming_cancelled sc
              ON sc.mass_date = sa.schedule_date
             AND TIME_FORMAT(sc.mass_time, '%H:%i') = TIME_FORMAT(sa.schedule_time, '%H:%i')
            LEFT JOIN misa_besar mb
              ON mb.status = 'published'
             AND mb.misa_date = sa.schedule_date
             AND TIME_FORMAT(mb.misa_time, '%H:%i') = TIME_FORMAT(sa.schedule_time, '%H:%i')
            WHERE {where_sql}
              AND sc.id IS NULL
              AND mb.id IS NULL
            ORDER BY sa.schedule_date ASC, sa.schedule_time ASC, sa.role_name ASC
            """,
            tuple(params),
        )
        rows = cursor.fetchall() or []
    except Exception as exc:
        print(f"[WARN] Query lengkap tugas Misa Biasa gagal, pakai fallback: {exc}")
        try:
            where_sql, params = build_filters("sa")
            cursor.execute(
                f"""
                SELECT DATE_FORMAT(sa.schedule_date, '%Y-%m-%d') AS date,
                       TIME_FORMAT(sa.schedule_time, '%H:%i') AS time,
                       sa.role_name,
                       sa.created_at,
                       'Misa Biasa' AS mass_name
                FROM streaming_assignments sa
                WHERE {where_sql}
                ORDER BY sa.schedule_date ASC, sa.schedule_time ASC, sa.role_name ASC
                """,
                tuple(params),
            )
            rows = cursor.fetchall() or []
        except Exception as fallback_exc:
            print(f"[ERROR] Fallback tugas Misa Biasa anggota gagal: {fallback_exc}")
            rows = []

    tasks: list[dict[str, object]] = []
    for row in rows:
        try:
            tasks.append(member_monitoring_task_item(
                "biasa",
                "Misa Biasa",
                row.get("mass_name") or "Misa Biasa",
                row.get("date"),
                row.get("time"),
                row.get("role_name"),
                row.get("created_at"),
            ))
        except Exception as exc:
            print(f"[WARN] Skip row tugas Misa Biasa yang tidak valid: {exc}")
    return tasks
def member_monitoring_fetch_big_tasks(cursor, member_id: object, *, month: str = "all", year: str = "all") -> list[dict[str, object]]:
    """Ambil tugas Misa Besar milik anggota dengan fallback aman."""
    try:
        ensure_misa_besar_schema()
    except Exception as exc:
        print(f"[WARN] ensure_misa_besar_schema monitoring anggota gagal: {exc}")

    where = ["a.member_id = %s", "COALESCE(mb.status, 'published') = 'published'"]
    params: list[object] = [member_id]
    if normalize_text(month).lower() != "all":
        where.append("MONTH(mb.misa_date) = %s")
        params.append(parse_required_int(month, datetime.now().month))
    if normalize_text(year).lower() != "all":
        where.append("YEAR(mb.misa_date) = %s")
        params.append(parse_required_int(year, datetime.now().year))
    where_sql = " AND ".join(where)

    try:
        cursor.execute(
            f"""
            SELECT mb.misa_name,
                   DATE_FORMAT(mb.misa_date, '%Y-%m-%d') AS date,
                   TIME_FORMAT(mb.misa_time, '%H:%i') AS time,
                   n.role_name,
                   a.created_at
            FROM misa_besar_assignments a
            JOIN misa_besar_names n ON n.id = a.role_id
            JOIN misa_besar mb ON mb.id = n.misa_id
            WHERE {where_sql}
            ORDER BY mb.misa_date ASC, mb.misa_time ASC, n.role_name ASC
            """,
            tuple(params),
        )
        rows = cursor.fetchall() or []
    except Exception as exc:
        print(f"[ERROR] Gagal mengambil tugas Misa Besar anggota: {exc}")
        rows = []

    tasks: list[dict[str, object]] = []
    for row in rows:
        try:
            tasks.append(member_monitoring_task_item(
                "besar",
                "Misa Besar",
                row.get("misa_name") or "Misa Besar",
                row.get("date"),
                row.get("time"),
                row.get("role_name"),
                row.get("created_at"),
            ))
        except Exception as exc:
            print(f"[WARN] Skip row tugas Misa Besar yang tidak valid: {exc}")
    return tasks
def member_monitoring_fetch_all_tasks(cursor, member_id: object, *, month: str = "all", year: str = "all", include_big: bool = True) -> list[dict[str, object]]:
    tasks: list[dict[str, object]] = []
    try:
        tasks.extend(member_monitoring_fetch_regular_tasks(cursor, member_id, month=month, year=year))
    except Exception as exc:
        print(f"[ERROR] Ambil tugas Misa Biasa gagal: {exc}")
    if include_big:
        try:
            tasks.extend(member_monitoring_fetch_big_tasks(cursor, member_id, month=month, year=year))
        except Exception as exc:
            print(f"[ERROR] Ambil tugas Misa Besar gagal: {exc}")
    try:
        tasks.sort(key=lambda item: member_monitoring_schedule_dt(item.get("date"), item.get("time")))
    except Exception:
        pass
    return tasks
def member_monitoring_available_years(cursor, member_id: object) -> list[int]:
    years = {datetime.now().year}
    try:
        cursor.execute(
            """
            SELECT DISTINCT YEAR(schedule_date) AS year_value
            FROM streaming_assignments
            WHERE member_id = %s
            UNION
            SELECT DISTINCT YEAR(mb.misa_date) AS year_value
            FROM misa_besar_assignments a
            JOIN misa_besar_names n ON n.id = a.role_id
            JOIN misa_besar mb ON mb.id = n.misa_id
            WHERE a.member_id = %s
            """,
            (member_id, member_id),
        )
        for row in cursor.fetchall() or []:
            year_val = fetch_scalar_value(row, None)
            if year_val:
                years.add(int(year_val))
    except Exception:
        pass
    current = datetime.now().year
    years.update({current - 1, current, current + 1})
    return sorted(years)


def member_monitoring_period_list(cursor, member_id: object, *, month: str, year: str, default_all_months: bool = False) -> list[tuple[int, int]]:
    month_value = normalize_text(month).lower() or "all"
    year_value = normalize_text(year).lower() or str(datetime.now().year)
    years = member_monitoring_available_years(cursor, member_id) if year_value == "all" else [parse_required_int(year_value, datetime.now().year)]
    periods: list[tuple[int, int]] = []
    if month_value == "all" or default_all_months:
        for y in years:
            for m in range(1, 13):
                periods.append((y, m))
    else:
        m = min(12, max(1, parse_required_int(month_value, datetime.now().month)))
        for y in years:
            periods.append((y, m))
    return periods


def member_monitoring_progress_group(tasks: list[dict[str, object]]) -> list[dict[str, object]]:
    groups: dict[str, dict[str, object]] = {}
    for task in tasks:
        type_key = normalize_text(task.get("type")) or "biasa"
        if type_key not in groups:
            groups[type_key] = {
                "type": type_key,
                "title": normalize_text(task.get("typeLabel")) or "Misa",
                "count": 0,
                "completed": 0,
                "schedules": [],
                "roles": [],
            }
        group = groups[type_key]
        group["count"] += 1
        if task.get("completed"):
            group["completed"] += 1
        group["schedules"].append(task.get("display"))
        group["roles"].append(task.get("role"))
    result: list[dict[str, object]] = []
    for group in groups.values():
        total = max(0, int(group.get("count") or 0))
        completed = max(0, int(group.get("completed") or 0))
        percentage = min(100, round((completed / total) * 100)) if total else 0
        if total == 0:
            status_label = "Belum Ada"
        elif completed >= total:
            status_label = "Selesai"
        elif completed > 0:
            status_label = "Sebagian Selesai"
        else:
            status_label = "Terjadwal"
        result.append({
            "type": group.get("type"),
            "title": group.get("title"),
            "count": total,
            "completed": completed,
            "schedules": group.get("schedules") or [],
            "roles": sorted(set([normalize_text(role) for role in group.get("roles") or [] if normalize_text(role)])),
            "status": status_label,
            "percentage": percentage,
            "progressClass": "good" if percentage >= 80 else ("warn" if percentage >= 60 else "low"),
        })
    result.sort(key=lambda item: 0 if item.get("type") == "biasa" else 1)
    return result


def member_monitoring_regular_count_for_period(cursor, member_id: object, year: int, month: int) -> int:
    return len(member_monitoring_fetch_regular_tasks(cursor, member_id, month=str(month), year=str(year)))


def member_monitoring_api_error(exc: Exception, message: str = "Gagal mengambil data monitoring."):
    """Return JSON error agar frontend tidak menerima HTML 500/teks kosong."""
    detail = str(exc)
    print(f"[ERROR] {message} {detail}")
    shown = message if not detail else f"{message} Detail: {detail}"
    return jsonify({"success": False, "error": shown, "detail": detail}), 500
@app.route("/api/monitoring-kewajiban-tugas/profile", methods=["GET"])
def api_member_monitoring_profile():
    ensure_auth_schema()
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True, buffered=True)
    try:
        member, error = member_monitoring_current_user(cursor)
        if error:
            return error
        now = datetime.now()
        return jsonify({
            "success": True,
            "member": {
                "id": member.get("id"),
                "name": member.get("name"),
                "username": member.get("username") or "",
                "role": member.get("roleNormalized"),
                "roleLabel": member.get("roleLabel"),
                "initial": (member.get("name") or "A")[:1].upper(),
            },
            "current": {"month": now.month, "year": now.year},
            "years": member_monitoring_available_years(cursor, member.get("id")),
        })
    except Exception as exc:
        return member_monitoring_api_error(exc)
    finally:
        cursor.close()
        conn.close()

@app.route("/api/monitoring-kewajiban-tugas/summary", methods=["GET"])
def api_member_monitoring_summary():
    ensure_auth_schema()
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True, buffered=True)
    try:
        ensure_monthly_monitoring_schema(cursor)
        member, error = member_monitoring_current_user(cursor)
        if error:
            return error
        now = datetime.now()
        member_id = member.get("id")
        target = monitoring_get_target_minimum(cursor)
        progress_month = member_monitoring_parse_month(request.args.get("progressMonth"), now.month)
        progress_year = member_monitoring_parse_year(request.args.get("progressYear"), now.year)
        shortage_month = member_monitoring_parse_month(request.args.get("shortageMonth"), now.month)
        shortage_year = member_monitoring_parse_year(request.args.get("shortageYear"), now.year)
        total_month = member_monitoring_parse_month(request.args.get("totalMonth"), now.month)
        total_year = member_monitoring_parse_year(request.args.get("totalYear"), now.year)

        progress_tasks = member_monitoring_fetch_all_tasks(cursor, member_id, month=progress_month, year=progress_year, include_big=True)
        progress_total = len(progress_tasks)
        progress_completed = sum(1 for task in progress_tasks if task.get("completed"))
        progress_percentage = min(100, round((progress_completed / progress_total) * 100)) if progress_total else 0

        shortage_periods = member_monitoring_period_list(cursor, member_id, month=shortage_month, year=shortage_year)
        shortage_total_value = 0
        shortage_regular_total = 0
        for y, m in shortage_periods:
            total_regular = member_monitoring_regular_count_for_period(cursor, member_id, y, m)
            shortage_regular_total += total_regular
            shortage_total_value += max(0, target - total_regular)

        total_tasks = member_monitoring_fetch_all_tasks(cursor, member_id, month=total_month, year=total_year, include_big=True)
        total_scheduled = len(total_tasks)
        total_completed = sum(1 for task in total_tasks if task.get("completed"))

        return jsonify({
            "success": True,
            "targetMinimum": target,
            "progress": {
                "completed": progress_completed,
                "total": progress_total,
                "percentage": progress_percentage,
                "periodLabel": member_monitoring_period_label(progress_month, progress_year),
            },
            "shortage": {
                "shortage": shortage_total_value,
                "regularTotal": shortage_regular_total,
                "periodLabel": member_monitoring_period_label(shortage_month, shortage_year),
            },
            "total": {
                "scheduled": total_scheduled,
                "completed": total_completed,
                "periodLabel": member_monitoring_period_label(total_month, total_year),
            },
        })
    except Exception as exc:
        return member_monitoring_api_error(exc)
    finally:
        cursor.close()
        conn.close()

@app.route("/api/monitoring-kewajiban-tugas/progress", methods=["GET"])
def api_member_monitoring_progress():
    ensure_auth_schema()
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True, buffered=True)
    try:
        member, error = member_monitoring_current_user(cursor)
        if error:
            return error
        now = datetime.now()
        month = member_monitoring_parse_month(request.args.get("month"), now.month)
        year = member_monitoring_parse_year(request.args.get("year"), now.year)
        search_text = normalize_text(request.args.get("search")).lower()
        sort_mode = normalize_text(request.args.get("sort")) or "date-desc"
        tasks = member_monitoring_fetch_all_tasks(cursor, member.get("id"), month=month, year=year, include_big=True)
        if search_text:
            tasks = [task for task in tasks if search_text in " ".join([
                normalize_text(task.get("typeLabel")),
                normalize_text(task.get("misaName")),
                normalize_text(task.get("role")),
                normalize_text(task.get("date")),
                normalize_text(task.get("time")),
                normalize_text(task.get("status")),
            ]).lower()]
        if sort_mode == "date-asc":
            tasks.sort(key=lambda task: member_monitoring_schedule_dt(task.get("date"), task.get("time")))
        elif sort_mode == "status":
            tasks.sort(key=lambda task: (normalize_text(task.get("status")), member_monitoring_schedule_dt(task.get("date"), task.get("time"))))
        else:
            tasks.sort(key=lambda task: member_monitoring_schedule_dt(task.get("date"), task.get("time")), reverse=True)
        groups = member_monitoring_progress_group(tasks)
        return jsonify({
            "success": True,
            "items": groups,
            "tasks": tasks,
            "period": {"month": month, "year": year, "label": member_monitoring_period_label(month, year)},
        })
    except Exception as exc:
        return member_monitoring_api_error(exc)
    finally:
        cursor.close()
        conn.close()

@app.route("/api/monitoring-kewajiban-tugas/shortage", methods=["GET"])
def api_member_monitoring_shortage():
    ensure_auth_schema()
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True, buffered=True)
    try:
        ensure_monthly_monitoring_schema(cursor)
        member, error = member_monitoring_current_user(cursor)
        if error:
            return error
        now = datetime.now()
        month = member_monitoring_parse_month(request.args.get("month"), None)
        if not normalize_text(request.args.get("month")):
            month = "all"
        year = member_monitoring_parse_year(request.args.get("year"), now.year)
        status_filter = normalize_text(request.args.get("status") or "all").lower()
        sort_mode = normalize_text(request.args.get("sort") or "month_asc").lower()
        target = monitoring_get_target_minimum(cursor)
        periods = member_monitoring_period_list(cursor, member.get("id"), month=month, year=year, default_all_months=(month == "all"))
        rows = []
        for y, m in periods:
            total = member_monitoring_regular_count_for_period(cursor, member.get("id"), y, m)
            shortage = max(0, target - total)
            fulfilled = shortage == 0
            rows.append({
                "month": m,
                "year": y,
                "period": f"{y}-{m:02d}",
                "monthText": f"{member_monitoring_month_name(m)} {y}",
                "targetMinimum": target,
                "totalTasks": total,
                "shortage": shortage,
                "status": "Terpenuhi" if fulfilled else f"Kurang {shortage} tugas",
                "statusClass": "ok" if fulfilled else "danger",
            })
        if status_filter == "fulfilled":
            rows = [row for row in rows if row.get("shortage") == 0]
        elif status_filter == "unfulfilled":
            rows = [row for row in rows if row.get("shortage") > 0]
        if sort_mode == "month_desc":
            rows.sort(key=lambda row: (row.get("year"), row.get("month")), reverse=True)
        elif sort_mode == "shortage_desc":
            rows.sort(key=lambda row: (row.get("shortage"), row.get("year"), row.get("month")), reverse=True)
        elif sort_mode == "shortage_asc":
            rows.sort(key=lambda row: (row.get("shortage"), row.get("year"), row.get("month")))
        else:
            rows.sort(key=lambda row: (row.get("year"), row.get("month")))
        return jsonify({"success": True, "items": rows, "targetMinimum": target})
    except Exception as exc:
        return member_monitoring_api_error(exc)
    finally:
        cursor.close()
        conn.close()

@app.route("/api/monitoring-kewajiban-tugas/stats", methods=["GET"])
def api_member_monitoring_stats():
    ensure_auth_schema()
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True, buffered=True)
    try:
        ensure_monthly_monitoring_schema(cursor)
        member, error = member_monitoring_current_user(cursor)
        if error:
            return error
        now = datetime.now()
        month = member_monitoring_parse_month(request.args.get("month"), now.month)
        year = member_monitoring_parse_year(request.args.get("year"), now.year)
        target = monitoring_get_target_minimum(cursor)
        periods = member_monitoring_period_list(cursor, member.get("id"), month=month, year=year)
        rows = []
        for y, m in periods:
            regular_tasks = member_monitoring_fetch_regular_tasks(cursor, member.get("id"), month=str(m), year=str(y))
            total_tasks = len(regular_tasks)
            completed = sum(1 for task in regular_tasks if task.get("completed"))
            shortage = max(0, target - total_tasks)
            percentage = 100 if total_tasks >= target else (round((total_tasks / target) * 100) if target else 0)
            rows.append({
                "month": m,
                "year": y,
                "period": f"{y}-{m:02d}",
                "monthText": f"{member_monitoring_month_name(m)} {y}",
                "targetMinimum": target,
                "targetTasks": total_tasks,
                "completedTasks": completed,
                "shortage": shortage,
                "percentage": min(100, percentage),
                "statusClass": "ok" if shortage == 0 else ("warn" if shortage == 1 else "danger"),
            })
        rows.sort(key=lambda row: (row.get("year"), row.get("month")), reverse=True)
        return jsonify({"success": True, "items": rows, "targetMinimum": target})
    except Exception as exc:
        return member_monitoring_api_error(exc)
    finally:
        cursor.close()
        conn.close()
# -----------------------------------------------------------------------------
# Evaluasi Streaming: Misa Biasa & Misa Besar
# -----------------------------------------------------------------------------

EVAL_DEFAULT_QUESTIONS = [
    {
        "id": "kesan-umum",
        "label": "Apa kesan umum pelayanan streaming pada misa ini?",
        "type": "single_choice",
        "required": True,
        "helpText": "Pilih satu penilaian umum.",
        "options": ["Sangat Baik", "Baik", "Cukup", "Perlu Pendampingan"],
    },
    {
        "id": "bagian-terbaik",
        "label": "Bagian mana yang sudah paling baik?",
        "type": "long_text",
        "required": False,
        "helpText": "Contoh: alur kamera, audio, koordinasi.",
        "options": [],
    },
]

EVAL_GENERAL_OPTIONS = {"aman", "kendala ringan", "kendala serius", "urgent", "kendala serius (urgent)"}
EVAL_REQUIRED_CHECKS = ["streaming_lancar", "koordinasi_baik"]


def eval_safe_json(value, fallback=None):
    if fallback is None:
        fallback = []
    return safe_json_loads(value, fallback)


def eval_date_to_iso(value) -> str:
    if not value:
        return ""
    if hasattr(value, "strftime"):
        return value.strftime("%Y-%m-%d")
    raw = normalize_text(value)
    return raw[:10]


def eval_time_to_hhmm(value) -> str:
    return format_time_hhmm(value)


def eval_datetime_from_parts(date_text: str, time_text: str) -> datetime | None:
    try:
        return datetime.strptime(f"{date_text} {eval_time_to_hhmm(time_text)}", "%Y-%m-%d %H:%M")
    except Exception:
        return None


def eval_day_name(date_text: str) -> str:
    try:
        return DAYS_INDO[datetime.strptime(date_text, "%Y-%m-%d").weekday()]
    except Exception:
        return "-"


def eval_format_date_id(date_text: str) -> str:
    try:
        return datetime.strptime(date_text, "%Y-%m-%d").strftime("%d/%m/%Y")
    except Exception:
        return normalize_text(date_text) or "-"


def eval_format_period_id(date_text: str, time_text: str = "") -> str:
    time_label = eval_time_to_hhmm(time_text) if time_text else ""
    base = f"{eval_day_name(date_text)}, {eval_format_date_id(date_text)}"
    return f"{base} jam {time_label} WIB" if time_label else base


def ensure_streaming_evaluation_schema(cursor=None) -> None:
    ensure_streaming_schema()
    ensure_misa_besar_schema()
    ensure_notifications_schema()
    owns_connection = cursor is None
    conn = None
    if owns_connection:
        conn = mysql_connection()
        cursor = conn.cursor(buffered=True)
    try:
        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS `streaming_evaluation_settings` (
              `id` int NOT NULL DEFAULT 1,
              `start_month` int NOT NULL DEFAULT 5,
              `start_year` int NOT NULL DEFAULT 2026,
              `updated_at` datetime NOT NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
              PRIMARY KEY (`id`)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_general_ci
            """
        )
        cursor.execute(
            """
            INSERT IGNORE INTO `streaming_evaluation_settings` (`id`, `start_month`, `start_year`)
            VALUES (1, 5, 2026)
            """
        )
        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS `streaming_evaluation_questions` (
              `id` varchar(100) NOT NULL,
              `label` varchar(500) NOT NULL,
              `question_type` varchar(40) NOT NULL DEFAULT 'short_text',
              `required` tinyint(1) NOT NULL DEFAULT 0,
              `help_text` text DEFAULT NULL,
              `options_json` longtext DEFAULT NULL,
              `order_index` int NOT NULL DEFAULT 0,
              `created_at` datetime NOT NULL DEFAULT CURRENT_TIMESTAMP,
              `updated_at` datetime NOT NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
              PRIMARY KEY (`id`)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_general_ci
            """
        )
        # Pertanyaan tambahan evaluasi bersifat opsional.
        # Jangan auto-seed ulang ketika tabel kosong, supaya admin bisa menyimpan konfigurasi tanpa pertanyaan tambahan.
        # Tombol "Reset ke Default" tetap tersedia untuk mengisi kembali pertanyaan default.
        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS `streaming_evaluations` (
              `id` int NOT NULL AUTO_INCREMENT,
              `schedule_kind` varchar(30) NOT NULL,
              `schedule_key` varchar(120) NOT NULL,
              `schedule_date` date NOT NULL,
              `schedule_time` time NOT NULL,
              `misa_name` varchar(255) NOT NULL,
              `misa_type_label` varchar(50) NOT NULL,
              `evaluator_id` int DEFAULT NULL,
              `evaluator_name` varchar(255) NOT NULL,
              `evaluator_role` varchar(100) DEFAULT NULL,
              `staff_json` longtext DEFAULT NULL,
              `extra_staff_json` longtext DEFAULT NULL,
              `staff_evaluations_json` longtext DEFAULT NULL,
              `technical_issue` longtext DEFAULT NULL,
              `nontechnical_issue` longtext DEFAULT NULL,
              `checklist_json` longtext DEFAULT NULL,
              `final_note` longtext DEFAULT NULL,
              `dynamic_answers_json` longtext DEFAULT NULL,
              `general_assessment` varchar(100) NOT NULL,
              `submitted_at` datetime NOT NULL DEFAULT CURRENT_TIMESTAMP,
              PRIMARY KEY (`id`),
              UNIQUE KEY `uniq_streaming_evaluation_schedule` (`schedule_kind`, `schedule_key`),
              KEY `idx_streaming_eval_date` (`schedule_date`, `schedule_time`),
              KEY `idx_streaming_eval_kind` (`schedule_kind`)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_general_ci
            """
        )
        if owns_connection:
            conn.commit()
    finally:
        if owns_connection:
            cursor.close()
            conn.close()


def eval_get_settings(cursor) -> dict[str, int]:
    ensure_streaming_evaluation_schema(cursor)
    cursor.execute("SELECT `start_month`, `start_year` FROM `streaming_evaluation_settings` WHERE `id` = 1 LIMIT 1")
    row = cursor.fetchone() or {}
    if not isinstance(row, dict):
        return {"startMonth": 5, "startYear": 2026}
    return {"startMonth": parse_required_int(row.get("start_month"), 5), "startYear": parse_required_int(row.get("start_year"), 2026)}


def eval_start_date_from_settings(settings: dict[str, int]) -> datetime.date:
    month = max(1, min(12, parse_required_int(settings.get("startMonth"), 5)))
    year = parse_required_int(settings.get("startYear"), 2026)
    return datetime(year, month, 1).date()


def eval_question_row_to_dict(row) -> dict[str, object]:
    if not isinstance(row, dict):
        return {}
    return {
        "id": normalize_text(row.get("id")),
        "label": normalize_text(row.get("label")),
        "type": normalize_text(row.get("question_type")) or "short_text",
        "required": bool(row.get("required")),
        "helpText": normalize_text(row.get("help_text")),
        "options": eval_safe_json(row.get("options_json"), []),
        "orderIndex": parse_required_int(row.get("order_index"), 0),
    }


def eval_fetch_questions(cursor) -> list[dict[str, object]]:
    ensure_streaming_evaluation_schema(cursor)
    cursor.execute(
        """
        SELECT `id`, `label`, `question_type`, `required`, `help_text`, `options_json`, `order_index`
        FROM `streaming_evaluation_questions`
        ORDER BY `order_index` ASC, `created_at` ASC
        """
    )
    rows = [eval_question_row_to_dict(row) for row in cursor.fetchall() or []]
    return [row for row in rows if row.get("id") and row.get("label")]


def eval_normalize_question_payload(items) -> list[dict[str, object]]:
    normalized = []
    if not isinstance(items, list):
        items = []
    allowed_types = {"short_text", "long_text", "single_choice", "multi_choice"}
    seen = set()
    for idx, item in enumerate(items, start=1):
        if not isinstance(item, dict):
            continue
        label = normalize_text(item.get("label"))
        if not label:
            continue
        raw_id = normalize_text(item.get("id")) or re.sub(r"[^a-z0-9]+", "-", label.lower()).strip("-") or f"q-{idx}"
        qid = raw_id[:90]
        if qid in seen:
            qid = f"{qid[:75]}-{idx}"
        seen.add(qid)
        qtype = normalize_text(item.get("type") or item.get("question_type")) or "short_text"
        if qtype not in allowed_types:
            qtype = "short_text"
        options = item.get("options") if isinstance(item.get("options"), list) else []
        clean_options = []
        for opt in options:
            opt_text = normalize_text(opt)
            if opt_text and opt_text not in clean_options:
                clean_options.append(opt_text)
        if qtype in {"single_choice", "multi_choice"} and not clean_options:
            clean_options = ["Ya", "Tidak"]
        normalized.append({
            "id": qid,
            "label": label,
            "type": qtype,
            "required": bool(item.get("required")),
            "helpText": normalize_text(item.get("helpText") or item.get("help_text")),
            "options": clean_options,
            "orderIndex": idx,
        })
    return normalized


def eval_evaluation_map(cursor, start_date: str | None = None, end_date: str | None = None) -> dict[tuple[str, str], dict[str, object]]:
    params = []
    where = []
    if start_date:
        where.append("schedule_date >= %s")
        params.append(start_date)
    if end_date:
        where.append("schedule_date <= %s")
        params.append(end_date)
    sql = "SELECT * FROM streaming_evaluations"
    if where:
        sql += " WHERE " + " AND ".join(where)
    cursor.execute(sql, tuple(params))
    result = {}
    for row in cursor.fetchall() or []:
        if isinstance(row, dict):
            result[(normalize_text(row.get("schedule_kind")), normalize_text(row.get("schedule_key")))] = row
    return result


def eval_fetch_active_members(cursor) -> list[dict[str, object]]:
    cursor.execute(
        """
        SELECT id, nama, username, role, status_akun
        FROM anggota
        WHERE LOWER(COALESCE(status_akun, 'aktif')) IN ('', 'aktif', 'active')
        ORDER BY nama ASC, username ASC
        """
    )
    members = []
    for row in cursor.fetchall() or []:
        if not isinstance(row, dict):
            continue
        members.append({
            "id": str(row.get("id")),
            "name": normalize_text(row.get("nama") or row.get("username")),
            "username": normalize_text(row.get("username")),
            "accountRole": normalize_role_value(row.get("role") or "user"),
        })
    return members


def eval_fetch_regular_roles(cursor) -> list[str]:
    cursor.execute("SELECT role_name FROM streaming_roles ORDER BY order_index ASC, id ASC")
    roles = [normalize_text(row.get("role_name") if isinstance(row, dict) else row[0]) for row in cursor.fetchall() or []]
    return [r for r in roles if r]


def eval_fetch_regular_config(cursor) -> list[dict[str, object]]:
    cursor.execute(
        """
        SELECT day_name, mass_name, DATE_FORMAT(start_time, '%H:%i') AS start_time
        FROM streaming_weekly_config
        ORDER BY FIELD(day_name, 'Senin','Selasa','Rabu','Kamis','Jumat','Sabtu','Minggu'), start_time ASC
        """
    )
    return cursor.fetchall() or []


def eval_fetch_regular_assignments(cursor, start_date: str, end_date: str) -> dict[tuple[str, str], dict[str, dict[str, object]]]:
    cursor.execute(
        """
        SELECT sa.id AS assignment_id, DATE_FORMAT(sa.schedule_date, '%Y-%m-%d') AS schedule_date,
               DATE_FORMAT(sa.schedule_time, '%H:%i') AS schedule_time, sa.role_name, sa.member_id,
               COALESCE(a.nama, CONCAT('ID ', sa.member_id)) AS member_name,
               COALESCE(a.role, 'user') AS account_role
        FROM streaming_assignments sa
        LEFT JOIN anggota a ON a.id = sa.member_id
        WHERE sa.schedule_date BETWEEN %s AND %s
        """,
        (start_date, end_date),
    )
    result: dict[tuple[str, str], dict[str, dict[str, object]]] = {}
    for row in cursor.fetchall() or []:
        date_text = normalize_text(row.get("schedule_date"))
        time_text = eval_time_to_hhmm(row.get("schedule_time"))
        key = (date_text, time_text)
        result.setdefault(key, {})[normalize_text(row.get("role_name"))] = {
            "assignmentId": row.get("assignment_id"),
            "role": normalize_text(row.get("role_name")),
            "memberId": str(row.get("member_id")),
            "memberName": normalize_text(row.get("member_name")),
            "accountRole": normalize_role_value(row.get("account_role") or "user"),
        }
    return result


def eval_fetch_regular_blocked(cursor, start_date: str, end_date: str) -> set[tuple[str, str]]:
    blocked = set()
    cursor.execute(
        """
        SELECT DATE_FORMAT(mass_date, '%Y-%m-%d') AS mass_date, DATE_FORMAT(mass_time, '%H:%i') AS mass_time
        FROM streaming_cancelled
        WHERE mass_date BETWEEN %s AND %s
        """,
        (start_date, end_date),
    )
    for row in cursor.fetchall() or []:
        blocked.add((normalize_text(row.get("mass_date")), eval_time_to_hhmm(row.get("mass_time"))))
    cursor.execute(
        """
        SELECT DATE_FORMAT(misa_date, '%Y-%m-%d') AS misa_date, DATE_FORMAT(misa_time, '%H:%i') AS misa_time
        FROM misa_besar
        WHERE status = 'published' AND misa_date BETWEEN %s AND %s
        """,
        (start_date, end_date),
    )
    for row in cursor.fetchall() or []:
        blocked.add((normalize_text(row.get("misa_date")), eval_time_to_hhmm(row.get("misa_time"))))
    return blocked


def eval_schedule_member_ids(schedule: dict[str, object]) -> set[str]:
    result = set()
    for staff in schedule.get("staff") or []:
        mid = normalize_text(staff.get("memberId") if isinstance(staff, dict) else "")
        if mid:
            result.add(mid)
    return result


def eval_schedule_has_staff(schedule: dict[str, object]) -> bool:
    return any(normalize_text(s.get("memberId")) for s in (schedule.get("staff") or []) if isinstance(s, dict))


def eval_build_regular_schedules(cursor, start_date: datetime.date, end_date: datetime.date) -> list[dict[str, object]]:
    roles = eval_fetch_regular_roles(cursor)
    cfg_rows = eval_fetch_regular_config(cursor)
    assignments = eval_fetch_regular_assignments(cursor, start_date.isoformat(), end_date.isoformat())
    blocked = eval_fetch_regular_blocked(cursor, start_date.isoformat(), end_date.isoformat())
    schedules = []
    current = start_date
    while current <= end_date:
        day_name = DAYS_INDO[current.weekday()]
        for cfg in cfg_rows:
            if normalize_text(cfg.get("day_name")) != day_name:
                continue
            date_text = current.isoformat()
            time_text = eval_time_to_hhmm(cfg.get("start_time"))
            if (date_text, time_text) in blocked:
                continue
            staff = []
            assigned_by_role = assignments.get((date_text, time_text), {})
            for role in roles:
                assigned = assigned_by_role.get(role) or {}
                staff.append({
                    "slotId": f"biasa:{date_text}:{time_text}:{role}",
                    "assignmentId": assigned.get("assignmentId"),
                    "role": role,
                    "roleId": role,
                    "memberId": assigned.get("memberId") or "",
                    "memberName": assigned.get("memberName") or "Belum terisi",
                    "accountRole": assigned.get("accountRole") or "",
                    "attendance": "present" if assigned.get("memberId") else "empty",
                })
            schedules.append({
                "kind": "misa_biasa",
                "kindLabel": "Misa Biasa",
                "scheduleKey": f"biasa:{date_text}:{time_text}",
                "id": f"misa_biasa|biasa:{date_text}:{time_text}",
                "misaId": None,
                "misaName": normalize_text(cfg.get("mass_name")) or "Misa Biasa",
                "date": date_text,
                "time": time_text,
                "dayName": day_name,
                "dateText": eval_format_date_id(date_text),
                "displayDateTime": eval_format_period_id(date_text, time_text),
                "staff": staff,
            })
        current += timedelta(days=1)
    return schedules


def eval_build_misa_besar_schedules(cursor, start_date: datetime.date, end_date: datetime.date) -> list[dict[str, object]]:
    cursor.execute(
        """
        SELECT id, misa_name, DATE_FORMAT(misa_date, '%Y-%m-%d') AS misa_date,
               DATE_FORMAT(misa_time, '%H:%i') AS misa_time, misa_note, allow_member_request, status
        FROM misa_besar
        WHERE status = 'published' AND misa_date BETWEEN %s AND %s
        ORDER BY misa_date ASC, misa_time ASC, id ASC
        """,
        (start_date.isoformat(), end_date.isoformat()),
    )
    events = cursor.fetchall() or []
    schedules = []
    for ev in events:
        misa_id = ev.get("id")
        cursor.execute(
            """
            SELECT n.id AS role_id, n.role_name, n.required_count,
                   a.id AS assignment_id, a.member_id,
                   COALESCE(ag.nama, CONCAT('ID ', a.member_id)) AS member_name,
                   COALESCE(ag.role, 'user') AS account_role
            FROM misa_besar_names n
            LEFT JOIN misa_besar_assignments a ON a.role_id = n.id
            LEFT JOIN anggota ag ON ag.id = a.member_id
            WHERE n.misa_id = %s
            ORDER BY n.id ASC, a.id ASC
            """,
            (misa_id,),
        )
        grouped: dict[str, dict[str, object]] = {}
        for row in cursor.fetchall() or []:
            role_id = str(row.get("role_id"))
            if role_id not in grouped:
                grouped[role_id] = {
                    "roleId": role_id,
                    "role": normalize_text(row.get("role_name")) or "Role",
                    "requiredCount": max(1, parse_required_int(row.get("required_count"), 1)),
                    "assignments": [],
                }
            if row.get("assignment_id"):
                grouped[role_id]["assignments"].append({
                    "assignmentId": row.get("assignment_id"),
                    "memberId": str(row.get("member_id")),
                    "memberName": normalize_text(row.get("member_name")),
                    "accountRole": normalize_role_value(row.get("account_role") or "user"),
                })
        staff = []
        for role_data in grouped.values():
            assignments = role_data.get("assignments") or []
            required = max(len(assignments), parse_required_int(role_data.get("requiredCount"), 1))
            for idx in range(required):
                assigned = assignments[idx] if idx < len(assignments) else {}
                staff.append({
                    "slotId": f"besar:{misa_id}:{role_data.get('roleId')}:{idx}",
                    "assignmentId": assigned.get("assignmentId"),
                    "role": role_data.get("role"),
                    "roleId": role_data.get("roleId"),
                    "memberId": assigned.get("memberId") or "",
                    "memberName": assigned.get("memberName") or "Belum terisi",
                    "accountRole": assigned.get("accountRole") or "",
                    "attendance": "present" if assigned.get("memberId") else "empty",
                })
        date_text = normalize_text(ev.get("misa_date"))
        time_text = eval_time_to_hhmm(ev.get("misa_time"))
        schedules.append({
            "kind": "misa_besar",
            "kindLabel": "Misa Besar",
            "scheduleKey": f"besar:{misa_id}",
            "id": f"misa_besar|besar:{misa_id}",
            "misaId": misa_id,
            "misaName": normalize_text(ev.get("misa_name")) or "Misa Besar",
            "note": normalize_text(ev.get("misa_note")),
            "allowMemberRequest": bool(ev.get("allow_member_request")),
            "date": date_text,
            "time": time_text,
            "dayName": eval_day_name(date_text),
            "dateText": eval_format_date_id(date_text),
            "displayDateTime": eval_format_period_id(date_text, time_text),
            "staff": staff,
        })
    return schedules


def eval_period_bounds(scale: str, year: int | None = None, month: int | None = None, week: int | None = None) -> tuple[datetime.date, datetime.date]:
    now = datetime.now()
    scale = normalize_text(scale) or "month"
    if scale == "all":
        return datetime(1900, 1, 1).date(), datetime(2999, 12, 31).date()
    if scale == "year":
        y = year or now.year
        return datetime(y, 1, 1).date(), datetime(y, 12, 31).date()
    if scale == "week":
        y = year or now.year
        w = max(1, min(53, parse_required_int(week, int(now.strftime('%V')))))
        start = datetime.strptime(f"{y}-W{w:02d}-1", "%G-W%V-%u").date()
        return start, start + timedelta(days=6)
    y = year or now.year
    m = month or now.month
    m = max(1, min(12, m))
    start = datetime(y, m, 1).date()
    end = datetime(y, m, calendar.monthrange(y, m)[1]).date()
    return start, end


def eval_all_schedules(cursor, start_date: datetime.date, end_date: datetime.date, kind: str = "all") -> list[dict[str, object]]:
    settings = eval_get_settings(cursor)
    min_start = eval_start_date_from_settings(settings)
    if start_date < min_start:
        start_date = min_start
    if end_date < start_date:
        return []
    schedules = []
    if kind in {"all", "misa_biasa", "biasa"}:
        schedules.extend(eval_build_regular_schedules(cursor, start_date, end_date))
    if kind in {"all", "misa_besar", "besar"}:
        schedules.extend(eval_build_misa_besar_schedules(cursor, start_date, end_date))
    schedules.sort(key=lambda item: (item.get("date") or "", item.get("time") or "", item.get("kindLabel") or ""))
    return schedules


def eval_find_schedule(cursor, schedule_id: str) -> dict[str, object] | None:
    raw = normalize_text(schedule_id)
    if "|" not in raw:
        return None
    kind, key = raw.split("|", 1)
    # Build a narrow period where possible.
    if kind == "misa_besar" and key.startswith("besar:"):
        misa_id = key.split(":", 1)[1]
        cursor.execute("SELECT DATE_FORMAT(misa_date, '%Y-%m-%d') AS d FROM misa_besar WHERE id = %s LIMIT 1", (misa_id,))
        row = cursor.fetchone()
        if not row:
            return None
        d = datetime.strptime(row.get("d"), "%Y-%m-%d").date()
        schedules = eval_all_schedules(cursor, d, d, "misa_besar")
    elif kind == "misa_biasa" and key.startswith("biasa:"):
        parts = key.split(":")
        if len(parts) < 3:
            return None
        d = datetime.strptime(parts[1], "%Y-%m-%d").date()
        schedules = eval_all_schedules(cursor, d, d, "misa_biasa")
    else:
        return None
    for schedule in schedules:
        if schedule.get("id") == raw:
            return schedule
    return None


def eval_row_to_dict(row: dict[str, object]) -> dict[str, object]:
    date_text = eval_date_to_iso(row.get("schedule_date"))
    time_text = eval_time_to_hhmm(row.get("schedule_time"))
    return {
        "id": row.get("id"),
        "kind": normalize_text(row.get("schedule_kind")),
        "kindLabel": normalize_text(row.get("misa_type_label")) or ("Misa Besar" if normalize_text(row.get("schedule_kind")) == "misa_besar" else "Misa Biasa"),
        "scheduleKey": normalize_text(row.get("schedule_key")),
        "scheduleId": f"{normalize_text(row.get('schedule_kind'))}|{normalize_text(row.get('schedule_key'))}",
        "date": date_text,
        "time": time_text,
        "dayName": eval_day_name(date_text),
        "displayDateTime": eval_format_period_id(date_text, time_text),
        "misaName": normalize_text(row.get("misa_name")),
        "evaluatorId": row.get("evaluator_id"),
        "evaluatorName": normalize_text(row.get("evaluator_name")),
        "evaluatorRole": normalize_text(row.get("evaluator_role")),
        "staff": eval_safe_json(row.get("staff_json"), []),
        "extraStaff": eval_safe_json(row.get("extra_staff_json"), []),
        "staffEvaluations": eval_safe_json(row.get("staff_evaluations_json"), []),
        "technicalIssue": normalize_text(row.get("technical_issue")),
        "nontechnicalIssue": normalize_text(row.get("nontechnical_issue")),
        "checklist": eval_safe_json(row.get("checklist_json"), {}),
        "finalNote": normalize_text(row.get("final_note")),
        "dynamicAnswers": eval_safe_json(row.get("dynamic_answers_json"), []),
        "generalAssessment": normalize_text(row.get("general_assessment")),
        "submittedAt": row.get("submitted_at").isoformat() if row.get("submitted_at") else None,
    }


def eval_fetch_evaluations(cursor, start_date: str | None = None, end_date: str | None = None, kind: str = "all", search: str = "", sort: str = "date_asc") -> list[dict[str, object]]:
    params = []
    where = []
    if start_date:
        where.append("schedule_date >= %s")
        params.append(start_date)
    if end_date:
        where.append("schedule_date <= %s")
        params.append(end_date)
    if kind in {"misa_biasa", "misa_besar"}:
        where.append("schedule_kind = %s")
        params.append(kind)
    sql = "SELECT * FROM streaming_evaluations"
    if where:
        sql += " WHERE " + " AND ".join(where)
    cursor.execute(sql, tuple(params))
    rows = [eval_row_to_dict(row) for row in cursor.fetchall() or []]
    kw = normalize_text(search).lower()
    if kw:
        rows = [r for r in rows if kw in " ".join([normalize_text(r.get("misaName")), normalize_text(r.get("kindLabel")), normalize_text(r.get("evaluatorName")), normalize_text(r.get("generalAssessment")), normalize_text(r.get("finalNote"))]).lower()]
    if sort == "date_desc":
        rows.sort(key=lambda r: (r.get("date") or "", r.get("time") or ""), reverse=True)
    elif sort == "submitted_desc":
        rows.sort(key=lambda r: r.get("submittedAt") or "", reverse=True)
    elif sort == "submitted_asc":
        rows.sort(key=lambda r: r.get("submittedAt") or "")
    else:
        rows.sort(key=lambda r: (r.get("date") or "", r.get("time") or ""))
    return rows


def eval_upsert_regular_assignment(cursor, schedule, staff_slot):
    date_text = schedule.get("date")
    time_text = schedule.get("time")
    role = normalize_text(staff_slot.get("role"))
    attendance = normalize_text(staff_slot.get("attendance"))
    actual_member_id = normalize_text(staff_slot.get("actualMemberId") or staff_slot.get("memberId"))
    if not role:
        return
    if attendance == "not_attend" or not actual_member_id:
        cursor.execute(
            "DELETE FROM streaming_assignments WHERE schedule_date = %s AND DATE_FORMAT(schedule_time, '%H:%i') = %s AND role_name = %s",
            (date_text, time_text, role),
        )
        return
    cursor.execute(
        """
        INSERT INTO streaming_assignments (schedule_date, schedule_time, role_name, member_id, request_source, created_at)
        VALUES (%s, %s, %s, %s, 'evaluasi', NOW())
        ON DUPLICATE KEY UPDATE member_id = VALUES(member_id), request_source = 'evaluasi', created_at = NOW()
        """,
        (date_text, time_text, role, actual_member_id),
    )


def eval_upsert_misa_besar_assignment(cursor, schedule, staff_slot):
    role_id = normalize_text(staff_slot.get("roleId"))
    assignment_id = normalize_text(staff_slot.get("assignmentId"))
    attendance = normalize_text(staff_slot.get("attendance"))
    actual_member_id = normalize_text(staff_slot.get("actualMemberId") or staff_slot.get("memberId"))
    if not role_id:
        return
    if attendance == "not_attend" or not actual_member_id:
        if assignment_id:
            cursor.execute("DELETE FROM misa_besar_assignments WHERE id = %s", (assignment_id,))
        return
    if assignment_id:
        cursor.execute("UPDATE misa_besar_assignments SET member_id = %s, request_source = 'evaluasi', created_at = NOW() WHERE id = %s", (actual_member_id, assignment_id))
    else:
        cursor.execute("INSERT IGNORE INTO misa_besar_assignments (role_id, member_id, request_source, created_at) VALUES (%s, %s, 'evaluasi', NOW())", (role_id, actual_member_id))


def eval_create_urgent_notifications(cursor, evaluation: dict[str, object]) -> int:
    general = normalize_text(evaluation.get("generalAssessment")).lower()
    if "urgent" not in general and "serius" not in general:
        return 0
    title = f"Evaluasi Urgent: {evaluation.get('misaName') or 'Jadwal Streaming'}"
    body = (
        f"Ada evaluasi streaming berstatus <b>{html.escape(evaluation.get('generalAssessment') or 'Urgent')}</b> "
        f"untuk <b>{html.escape(evaluation.get('misaName') or '-')}</b> pada {html.escape(evaluation.get('displayDateTime') or '-')}. "
        f"Pengisi: <b>{html.escape(evaluation.get('evaluatorName') or '-')}</b>."
    )
    create_notification_once(
        cursor,
        "evaluasi",
        title,
        body,
        "/hasil-evaluasi-streaming.html",
        {"notification_kind": "streaming_evaluation_urgent", "evaluation_id": evaluation.get("id")},
        target_role="admin",
        dedupe_key=f"eval-urgent:admin:{evaluation.get('id')}",
    )
    create_notification_once(
        cursor,
        "evaluasi",
        title,
        body,
        "/hasil-evaluasi-streaming.html",
        {"notification_kind": "streaming_evaluation_urgent", "evaluation_id": evaluation.get("id")},
        target_role="super_admin",
        dedupe_key=f"eval-urgent:super:{evaluation.get('id')}",
    )
    return 2


def eval_validate_submit_payload(payload: dict[str, object]) -> tuple[bool, str]:
    if not normalize_text(payload.get("scheduleId")):
        return False, "Jadwal evaluasi wajib dipilih."
    if not normalize_text(payload.get("evaluatorName")) and not current_user_context().get("logged_in"):
        return False, "Nama pengisi evaluasi wajib dipilih."
    if not normalize_text(payload.get("technicalIssue")):
        return False, "Kendala teknis wajib diisi."
    if not normalize_text(payload.get("nontechnicalIssue")):
        return False, "Kendala non-teknis wajib diisi."
    # Checklist kondisi pelayanan bersifat informatif/opsional.
    # Tidak semua kondisi harus dicentang karena form juga dipakai untuk mencatat sesi yang mengalami kendala.
    if not normalize_text(payload.get("generalAssessment")):
        return False, "Penilaian umum wajib dipilih."
    return True, ""


@app.route("/api/evaluasi-streaming/settings", methods=["GET", "POST"])
def api_streaming_evaluation_settings():
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True, buffered=True)
    try:
        ensure_streaming_evaluation_schema(cursor)
        if request.method == "POST":
            if normalize_role_value(session.get("role") or "") not in {"admin", "super_admin"}:
                return jsonify({"success": False, "error": "Akses ditolak."}), 403
            data = request.get_json(silent=True) or {}
            month = max(1, min(12, parse_required_int(data.get("startMonth"), 5)))
            year = max(2000, parse_required_int(data.get("startYear"), 2026))
            cursor.execute("UPDATE streaming_evaluation_settings SET start_month = %s, start_year = %s WHERE id = 1", (month, year))
            conn.commit()
        return jsonify({"success": True, "settings": eval_get_settings(cursor)})
    finally:
        cursor.close()
        conn.close()


@app.route("/api/evaluasi-streaming/questions", methods=["GET", "POST"])
def api_streaming_evaluation_questions():
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True, buffered=True)
    try:
        ensure_streaming_evaluation_schema(cursor)
        if request.method == "POST":
            if normalize_role_value(session.get("role") or "") not in {"admin", "super_admin"}:
                return jsonify({"success": False, "error": "Akses ditolak."}), 403
            data = request.get_json(silent=True) or {}
            questions = eval_normalize_question_payload(data.get("questions"))
            # Boleh kosong: pertanyaan tambahan evaluasi tidak wajib ada.
            cursor.execute("DELETE FROM streaming_evaluation_questions")
            for idx, q in enumerate(questions, start=1):
                cursor.execute(
                    """
                    INSERT INTO streaming_evaluation_questions
                    (id, label, question_type, required, help_text, options_json, order_index)
                    VALUES (%s, %s, %s, %s, %s, %s, %s)
                    """,
                    (q["id"], q["label"], q["type"], 1 if q.get("required") else 0, q.get("helpText") or "", json.dumps(q.get("options") or [], ensure_ascii=False), idx),
                )
            conn.commit()
        return jsonify({"success": True, "questions": eval_fetch_questions(cursor)})
    finally:
        cursor.close()
        conn.close()


@app.route("/api/evaluasi-streaming/questions/reset", methods=["POST"])
def api_streaming_evaluation_questions_reset():
    if normalize_role_value(session.get("role") or "") not in {"admin", "super_admin"}:
        return jsonify({"success": False, "error": "Akses ditolak."}), 403
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True, buffered=True)
    try:
        ensure_streaming_evaluation_schema(cursor)
        cursor.execute("DELETE FROM streaming_evaluation_questions")
        for idx, q in enumerate(EVAL_DEFAULT_QUESTIONS, start=1):
            cursor.execute(
                """
                INSERT INTO streaming_evaluation_questions
                (id, label, question_type, required, help_text, options_json, order_index)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
                """,
                (q["id"], q["label"], q["type"], 1 if q.get("required") else 0, q.get("helpText") or "", json.dumps(q.get("options") or [], ensure_ascii=False), idx),
            )
        conn.commit()
        return jsonify({"success": True, "questions": eval_fetch_questions(cursor)})
    finally:
        cursor.close()
        conn.close()


@app.route("/api/evaluasi-streaming/members", methods=["GET"])
def api_streaming_evaluation_members():
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True, buffered=True)
    try:
        ensure_auth_schema()
        return jsonify({"success": True, "members": eval_fetch_active_members(cursor), "currentUser": current_user_context()})
    finally:
        cursor.close()
        conn.close()


@app.route("/api/evaluasi-streaming/schedules", methods=["GET"])
def api_streaming_evaluation_schedules():
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True, buffered=True)
    try:
        ensure_streaming_evaluation_schema(cursor)
        now = datetime.now()
        settings = eval_get_settings(cursor)
        start_setting = eval_start_date_from_settings(settings)
        kind = normalize_text(request.args.get("kind") or "all")
        mode = normalize_text(request.args.get("mode") or "form")
        include_evaluated = normalize_text(request.args.get("includeEvaluated") or "0") in {"1", "true", "yes"}
        # Search a useful window: from configured start until now for public form; current year for member/admin if requested.
        start_date = parse_optional_date(request.args.get("startDate"))
        end_date = parse_optional_date(request.args.get("endDate"))
        if start_date:
            start = datetime.strptime(start_date, "%Y-%m-%d").date()
        else:
            start = start_setting
        if end_date:
            end = datetime.strptime(end_date, "%Y-%m-%d").date()
        else:
            end = now.date()
        schedules = eval_all_schedules(cursor, start, end, kind)
        evals = eval_evaluation_map(cursor, start.isoformat(), end.isoformat())
        rows = []
        for schedule in schedules:
            dt = eval_datetime_from_parts(schedule.get("date"), schedule.get("time"))
            is_due = bool(dt and dt <= now)
            has_staff = eval_schedule_has_staff(schedule)
            ev = evals.get((schedule.get("kind"), schedule.get("scheduleKey")))
            schedule["evaluated"] = bool(ev)
            schedule["evaluationId"] = ev.get("id") if ev else None
            schedule["due"] = is_due
            schedule["hasStaff"] = has_staff
            if mode == "form" and (not is_due or not has_staff or (ev and not include_evaluated)):
                continue
            rows.append(schedule)
        return jsonify({"success": True, "schedules": rows, "questions": eval_fetch_questions(cursor), "settings": settings, "currentUser": current_user_context()})
    finally:
        cursor.close()
        conn.close()


@app.route("/api/evaluasi-streaming/schedule-detail", methods=["GET"])
def api_streaming_evaluation_schedule_detail():
    schedule_id = normalize_text(request.args.get("scheduleId"))
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True, buffered=True)
    try:
        ensure_streaming_evaluation_schema(cursor)
        schedule = eval_find_schedule(cursor, schedule_id)
        if not schedule:
            return jsonify({"success": False, "error": "Jadwal tidak ditemukan."}), 404
        ev_map = eval_evaluation_map(cursor, schedule.get("date"), schedule.get("date"))
        ev = ev_map.get((schedule.get("kind"), schedule.get("scheduleKey")))
        return jsonify({"success": True, "schedule": schedule, "evaluation": eval_row_to_dict(ev) if ev else None, "questions": eval_fetch_questions(cursor), "members": eval_fetch_active_members(cursor), "currentUser": current_user_context()})
    finally:
        cursor.close()
        conn.close()


@app.route("/api/evaluasi-streaming/submit", methods=["POST"])
def api_streaming_evaluation_submit():
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True, buffered=True)
    try:
        ensure_streaming_evaluation_schema(cursor)
        payload = request.get_json(silent=True) or {}
        valid, message = eval_validate_submit_payload(payload)
        if not valid:
            return jsonify({"success": False, "error": message}), 400
        schedule = eval_find_schedule(cursor, normalize_text(payload.get("scheduleId")))
        if not schedule:
            return jsonify({"success": False, "error": "Jadwal tidak ditemukan."}), 404
        dt = eval_datetime_from_parts(schedule.get("date"), schedule.get("time"))
        if not dt or dt > datetime.now():
            return jsonify({"success": False, "error": "Evaluasi baru bisa diisi setelah jadwal misa dimulai."}), 400
        if not eval_schedule_has_staff(schedule):
            return jsonify({"success": False, "error": "Jadwal ini belum memiliki petugas."}), 400
        cursor.execute("SELECT id FROM streaming_evaluations WHERE schedule_kind = %s AND schedule_key = %s LIMIT 1", (schedule.get("kind"), schedule.get("scheduleKey")))
        if cursor.fetchone():
            return jsonify({"success": False, "error": "Evaluasi untuk jadwal ini sudah pernah diisi."}), 409

        viewer = current_user_context()
        evaluator_id = viewer.get("user_id") if viewer.get("logged_in") else payload.get("evaluatorId")
        evaluator_name = normalize_text(viewer.get("nama") if viewer.get("logged_in") else payload.get("evaluatorName"))
        evaluator_role = normalize_text(payload.get("evaluatorRole"))
        if viewer.get("logged_in") and not evaluator_role:
            # Find role for this user in selected schedule.
            for staff in schedule.get("staff") or []:
                if normalize_text(staff.get("memberId")) == normalize_text(viewer.get("user_id")):
                    evaluator_role = normalize_text(staff.get("role"))
                    break
        if not evaluator_name:
            return jsonify({"success": False, "error": "Nama pengisi evaluasi wajib diisi."}), 400

        submitted_staff = payload.get("staff") if isinstance(payload.get("staff"), list) else schedule.get("staff") or []
        final_staff = []
        active_members_map = {str(m.get("id")): m for m in eval_fetch_active_members(cursor)}
        for raw_slot in submitted_staff:
            if not isinstance(raw_slot, dict):
                continue
            slot = dict(raw_slot)
            slot.setdefault("role", raw_slot.get("role"))
            slot.setdefault("roleId", raw_slot.get("roleId"))
            slot.setdefault("assignmentId", raw_slot.get("assignmentId"))
            attendance = normalize_text(raw_slot.get("attendance") or "present")
            actual_id = normalize_text(raw_slot.get("actualMemberId") or raw_slot.get("memberId"))
            if attendance == "not_attend":
                slot["actualMemberId"] = ""
                slot["actualMemberName"] = "Tidak datang"
                slot["attendance"] = "not_attend"
            else:
                member_info = active_members_map.get(actual_id)
                slot["actualMemberId"] = actual_id
                slot["actualMemberName"] = normalize_text(member_info.get("name") if member_info else raw_slot.get("actualMemberName") or raw_slot.get("memberName"))
                slot["attendance"] = "present"
            final_staff.append(slot)
            if schedule.get("kind") == "misa_biasa":
                eval_upsert_regular_assignment(cursor, schedule, slot)
            else:
                eval_upsert_misa_besar_assignment(cursor, schedule, slot)

        extra_staff = payload.get("extraStaff") if isinstance(payload.get("extraStaff"), list) else []
        staff_evals = payload.get("staffEvaluations") if isinstance(payload.get("staffEvaluations"), list) else []
        dynamic_answers = payload.get("dynamicAnswers") if isinstance(payload.get("dynamicAnswers"), list) else []
        checklist = payload.get("checklist") if isinstance(payload.get("checklist"), dict) else {}
        general = normalize_text(payload.get("generalAssessment"))

        cursor.execute(
            """
            INSERT INTO streaming_evaluations
            (schedule_kind, schedule_key, schedule_date, schedule_time, misa_name, misa_type_label,
             evaluator_id, evaluator_name, evaluator_role, staff_json, extra_staff_json, staff_evaluations_json,
             technical_issue, nontechnical_issue, checklist_json, final_note, dynamic_answers_json, general_assessment, submitted_at)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, NOW())
            """,
            (
                schedule.get("kind"), schedule.get("scheduleKey"), schedule.get("date"), schedule.get("time"),
                schedule.get("misaName"), schedule.get("kindLabel"), evaluator_id, evaluator_name, evaluator_role,
                json.dumps(final_staff, ensure_ascii=False), json.dumps(extra_staff, ensure_ascii=False), json.dumps(staff_evals, ensure_ascii=False),
                normalize_text(payload.get("technicalIssue")), normalize_text(payload.get("nontechnicalIssue")), json.dumps(checklist, ensure_ascii=False),
                normalize_text(payload.get("finalNote")), json.dumps(dynamic_answers, ensure_ascii=False), general,
            ),
        )
        evaluation_id = cursor.lastrowid
        cursor.execute("SELECT * FROM streaming_evaluations WHERE id = %s", (evaluation_id,))
        evaluation = eval_row_to_dict(cursor.fetchone())
        eval_create_urgent_notifications(cursor, evaluation)
        conn.commit()
        return jsonify({"success": True, "evaluation": evaluation})
    except Exception as exc:
        conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 400
    finally:
        cursor.close()
        conn.close()


@app.route("/api/evaluasi-streaming/member", methods=["GET"])
def api_streaming_evaluation_member():
    if not session.get("logged_in"):
        return jsonify({"success": False, "error": "Unauthorized"}), 401
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True, buffered=True)
    try:
        ensure_streaming_evaluation_schema(cursor)
        viewer = current_user_context()
        member_id = normalize_text(viewer.get("user_id"))
        now = datetime.now()
        month = request.args.get("month")
        year = request.args.get("year")
        selected_month = parse_required_int(month, now.month) if normalize_text(month) not in {"all", ""} else now.month
        selected_year = parse_required_int(year, now.year) if normalize_text(year) not in {"all", ""} else now.year
        start, end = eval_period_bounds("month", selected_year, selected_month)
        kind = normalize_text(request.args.get("kind") or "all")
        status = normalize_text(request.args.get("status") or "all")
        schedules = eval_all_schedules(cursor, start, end, kind)
        evals = eval_evaluation_map(cursor, start.isoformat(), end.isoformat())
        items = []
        total_all = 0
        displayed = 0
        completed = 0
        not_filled = 0
        progress_by_kind: dict[str, dict[str, object]] = {}
        for s in schedules:
            if member_id not in eval_schedule_member_ids(s):
                continue
            total_all += 1
            dt = eval_datetime_from_parts(s.get("date"), s.get("time"))
            is_due = bool(dt and dt <= now)
            ev = evals.get((s.get("kind"), s.get("scheduleKey")))
            if is_due:
                displayed += 1
                if ev:
                    completed += 1
                else:
                    not_filled += 1
            p = progress_by_kind.setdefault(s.get("kind"), {"kind": s.get("kind"), "kindLabel": s.get("kindLabel"), "total": 0, "done": 0})
            p["total"] += 1
            if is_due:
                p["done"] += 1
            s["evaluated"] = bool(ev)
            s["evaluationId"] = ev.get("id") if ev else None
            s["due"] = is_due
            # Form evaluasi anggota hanya boleh tampil setelah jadwal benar-benar dimulai.
            # Total/progress tetap menghitung semua tugas bulan itu, tetapi daftar yang bisa diisi/review
            # tidak menampilkan jadwal masa depan agar user tidak bisa mengisi sebelum hari-H/jam mulai.
            if not is_due:
                continue
            if status == "filled" and not ev:
                continue
            if status == "empty" and ev:
                continue
            items.append(s)
        progress = []
        for row in progress_by_kind.values():
            total = row.get("total") or 0
            done = row.get("done") or 0
            row["percentage"] = round((done / total) * 100) if total else 0
            progress.append(row)
        progress.sort(key=lambda r: r.get("kindLabel") or "")
        items.sort(key=lambda r: (r.get("date") or "", r.get("time") or ""), reverse=normalize_text(request.args.get("sort")) == "date_desc")
        return jsonify({
            "success": True,
            "currentUser": viewer,
            "stats": {"totalSessions": total_all, "displayed": displayed, "filled": completed, "empty": not_filled},
            "progress": progress,
            "items": items,
            "questions": eval_fetch_questions(cursor),
        })
    finally:
        cursor.close()
        conn.close()


@app.route("/api/evaluasi-streaming/admin/results", methods=["GET"])
def api_streaming_evaluation_admin_results():
    if normalize_role_value(session.get("role") or "") not in {"admin", "super_admin"}:
        return jsonify({"success": False, "error": "Akses ditolak."}), 403
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True, buffered=True)
    try:
        ensure_streaming_evaluation_schema(cursor)
        now = datetime.now()
        scale = normalize_text(request.args.get("scale") or "month")
        year = parse_required_int(request.args.get("year"), now.year)
        month = parse_required_int(request.args.get("month"), now.month)
        week = parse_required_int(request.args.get("week"), int(now.strftime("%V")))
        kind = normalize_text(request.args.get("kind") or "all")
        search = normalize_text(request.args.get("search"))
        sort = normalize_text(request.args.get("sort") or "date_asc")
        start, end = eval_period_bounds(scale, year, month, week)
        schedules = eval_all_schedules(cursor, start, end, kind)
        schedule_due = []
        for s in schedules:
            dt = eval_datetime_from_parts(s.get("date"), s.get("time"))
            if dt and dt <= now and eval_schedule_has_staff(s):
                schedule_due.append(s)
        evaluations = eval_fetch_evaluations(cursor, start.isoformat(), end.isoformat(), kind, search, sort)
        eval_keys = {(e.get("kind"), e.get("scheduleKey")) for e in evaluations}
        all_eval_map = eval_evaluation_map(cursor, start.isoformat(), end.isoformat())
        viewer = current_user_context()
        viewer_id = normalize_text(viewer.get("user_id"))
        pending = []
        kw = search.lower()
        for s in schedule_due:
            if (s.get("kind"), s.get("scheduleKey")) in all_eval_map:
                continue
            if kw and kw not in " ".join([normalize_text(s.get("misaName")), normalize_text(s.get("kindLabel")), normalize_text(s.get("displayDateTime"))]).lower():
                continue
            pending_item = dict(s)
            pending_item["canFillByCurrentUser"] = bool(
                viewer_id and any(normalize_text(staff.get("memberId")) == viewer_id for staff in (s.get("staff") or []))
            )
            pending_item["currentUserRole"] = normalize_role_value(viewer.get("role") or "")
            pending.append(pending_item)
        pending.sort(key=lambda r: (r.get("date") or "", r.get("time") or ""))
        urgent_count = sum(1 for e in evaluations if ("urgent" in normalize_text(e.get("generalAssessment")).lower() or "serius" in normalize_text(e.get("generalAssessment")).lower()))
        total_misa = len([s for s in schedules if eval_datetime_from_parts(s.get("date"), s.get("time")) and eval_datetime_from_parts(s.get("date"), s.get("time")) <= now])
        summary = eval_build_admin_summary(evaluations, pending)
        return jsonify({
            "success": True,
            "period": {"start": start.isoformat(), "end": end.isoformat(), "scale": scale, "month": month, "year": year, "week": week},
            "metrics": {"evaluationsIn": len(evaluations), "missed": len(pending), "totalMisa": total_misa, "urgent": urgent_count},
            "summary": summary,
            "evaluations": evaluations,
            "pending": pending,
            "currentUser": viewer,
        })
    finally:
        cursor.close()
        conn.close()


def eval_build_admin_summary(evaluations: list[dict[str, object]], pending: list[dict[str, object]]) -> dict[str, object]:
    count = len(evaluations)
    rating_scores = {"Sangat Baik": 4, "Baik": 3, "Cukup": 2, "Perlu Pendampingan": 1}
    rating_counts: dict[str, int] = {"Sangat Baik": 0, "Baik": 0, "Cukup": 0, "Perlu Pendampingan": 0}
    general_counts: dict[str, int] = {"Aman": 0, "Kendala Ringan": 0, "Kendala Serius (Urgent)": 0}
    checklist_stream = 0
    checklist_coord = 0
    technical_min = 0
    nontechnical_min = 0
    total_staff_score = 0
    total_staff = 0
    dynamic_answer_count = 0
    for e in evaluations:
        general = normalize_text(e.get("generalAssessment")) or "Aman"
        general_counts[general] = general_counts.get(general, 0) + 1
        tech = normalize_text(e.get("technicalIssue")).lower()
        nontech = normalize_text(e.get("nontechnicalIssue")).lower()
        if tech in {"-", "tidak ada", "aman", "none"} or len(tech) < 5:
            technical_min += 1
        if nontech in {"-", "tidak ada", "aman", "none"} or len(nontech) < 5:
            nontechnical_min += 1
        checklist = e.get("checklist") or {}
        if checklist.get("streaming_lancar"):
            checklist_stream += 1
        if checklist.get("koordinasi_baik"):
            checklist_coord += 1
        for staff_eval in e.get("staffEvaluations") or []:
            rating = normalize_text(staff_eval.get("rating") if isinstance(staff_eval, dict) else "")
            if rating in rating_counts:
                rating_counts[rating] += 1
            if rating in rating_scores:
                total_staff_score += rating_scores[rating]
                total_staff += 1
        for ans in e.get("dynamicAnswers") or []:
            if isinstance(ans, dict):
                val = ans.get("answer")
                if (isinstance(val, list) and val) or normalize_text(val):
                    dynamic_answer_count += 1
    avg_staff = round(total_staff_score / total_staff, 2) if total_staff else 0
    checklist_rate = round(((checklist_stream + checklist_coord) / (count * 2)) * 100) if count else 0
    safe_count = general_counts.get("Aman", 0)
    urgent_count = sum(v for k, v in general_counts.items() if "Urgent" in k or "Serius" in k)
    insight = "Belum ada evaluasi pada periode ini."
    if count:
        insight = f"Dalam periode ini terdapat {count} evaluasi masuk. Rata-rata penilaian petugas {avg_staff}/4, kepatuhan checklist {checklist_rate}%, dan {urgent_count} sesi masuk kategori urgent."
    recommendation = "Prioritaskan pengisian evaluasi untuk misa yang sudah lewat dan belum diisi."
    if urgent_count:
        recommendation = "Segera cek evaluasi urgent dan tindak lanjuti kendala teknis/non-teknis yang tercatat."
    elif count and checklist_rate >= 80:
        recommendation = "Pertahankan pola koordinasi dan dokumentasikan praktik yang sudah berjalan baik."
    return {
        "averageStaffScore": avg_staff,
        "checklistRate": checklist_rate,
        "generalSafeRate": round((safe_count / count) * 100) if count else 0,
        "generalUrgentCount": urgent_count,
        "dynamicAnswerCount": dynamic_answer_count,
        "ratingCounts": rating_counts,
        "generalCounts": general_counts,
        "technicalMinRate": round((technical_min / count) * 100) if count else 0,
        "nontechnicalMinRate": round((nontechnical_min / count) * 100) if count else 0,
        "pendingCount": len(pending),
        "insight": insight,
        "recommendation": recommendation,
    }


@app.route("/api/evaluasi-streaming/admin/results/<int:evaluation_id>", methods=["DELETE"])
def api_streaming_evaluation_delete(evaluation_id):
    if normalize_role_value(session.get("role") or "") not in {"admin", "super_admin"}:
        return jsonify({"success": False, "error": "Akses ditolak."}), 403
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True, buffered=True)
    try:
        ensure_streaming_evaluation_schema(cursor)
        cursor.execute("DELETE FROM streaming_evaluations WHERE id = %s", (evaluation_id,))
        conn.commit()
        return jsonify({"success": True})
    finally:
        cursor.close()
        conn.close()



def eval_export_value(value) -> str:
    if value is None:
        return "-"
    text = normalize_text(value)
    return text if text else "-"


def eval_checklist_export_text(checklist: dict[str, object]) -> str:
    checklist = checklist if isinstance(checklist, dict) else {}
    labels = {
        "streaming_lancar": "Streaming berjalan lancar sampai akhir",
        "koordinasi_baik": "Koordinasi tim berjalan baik",
        "rekaman_tersimpan": "Rekaman tersimpan dengan benar",
    }
    parts = []
    for key, label in labels.items():
        if key in checklist:
            parts.append(f"{label}: {'Ya' if checklist.get(key) else 'Tidak'}")
    for key, value in checklist.items():
        if key not in labels:
            parts.append(f"{key}: {'Ya' if value else 'Tidak'}")
    return "; ".join(parts) or "-"


def eval_staff_export_text(staff: list[dict[str, object]]) -> str:
    parts = []
    for slot in staff or []:
        if not isinstance(slot, dict):
            continue
        role = eval_export_value(slot.get("role"))
        scheduled = eval_export_value(slot.get("memberName"))
        actual = eval_export_value(slot.get("actualMemberName") or slot.get("memberName"))
        attendance = normalize_text(slot.get("attendance"))
        attendance_label = "Tidak datang" if attendance == "not_attend" else "Hadir/Digantikan"
        if actual == "Tidak datang":
            attendance_label = "Tidak datang"
        parts.append(f"{role}: jadwal {scheduled}; aktual {actual}; status {attendance_label}")
    return "; ".join(parts) or "-"


def eval_extra_staff_export_text(extra_staff: list[dict[str, object]]) -> str:
    parts = []
    for staff in extra_staff or []:
        if not isinstance(staff, dict):
            continue
        name = eval_export_value(staff.get("name") or staff.get("memberName"))
        role = eval_export_value(staff.get("role") or staff.get("helpRole"))
        parts.append(f"{name} ({role})")
    return "; ".join(parts) or "-"


def eval_staff_evaluation_export_text(staff_evals: list[dict[str, object]]) -> str:
    parts = []
    for item in staff_evals or []:
        if not isinstance(item, dict):
            continue
        name = eval_export_value(item.get("memberName") or item.get("name"))
        role = eval_export_value(item.get("role"))
        rating = eval_export_value(item.get("rating") or item.get("assessment"))
        note = eval_export_value(item.get("note") or item.get("catatan") or item.get("comment"))
        parts.append(f"{name} ({role}) - {rating}; catatan: {note}")
    return "; ".join(parts) or "-"


def eval_dynamic_answers_export_text(dynamic_answers: list[dict[str, object]]) -> str:
    parts = []
    for item in dynamic_answers or []:
        if not isinstance(item, dict):
            continue
        question = eval_export_value(item.get("question") or item.get("label") or item.get("text"))
        answer = item.get("answer")
        if isinstance(answer, list):
            answer_text = ", ".join(normalize_text(v) for v in answer if normalize_text(v))
        else:
            answer_text = normalize_text(answer)
        parts.append(f"{question}: {answer_text or '-'}")
    return "; ".join(parts) or "-"


def eval_export_rows(evaluations: list[dict[str, object]]) -> tuple[list[str], list[list[str]]]:
    headers = [
        "No", "Tanggal", "Jam", "Hari", "Jenis Misa", "Nama Misa", "Pengisi", "Role Pengisi",
        "Penilaian Umum", "Petugas Jadwal & Aktual", "Petugas Tambahan", "Evaluasi Per Petugas",
        "Kendala Teknis", "Kendala Non-Teknis", "Checklist Kondisi Pelayanan", "Pertanyaan Tambahan",
        "Catatan Penutup", "Waktu Submit",
    ]
    rows: list[list[str]] = []
    for idx, e in enumerate(evaluations, start=1):
        rows.append([
            str(idx),
            eval_export_value(e.get("date")),
            eval_export_value(e.get("time")),
            eval_export_value(e.get("dayName")),
            eval_export_value(e.get("kindLabel")),
            eval_export_value(e.get("misaName")),
            eval_export_value(e.get("evaluatorName")),
            eval_export_value(e.get("evaluatorRole")),
            eval_export_value(e.get("generalAssessment")),
            eval_staff_export_text(e.get("staff") or []),
            eval_extra_staff_export_text(e.get("extraStaff") or []),
            eval_staff_evaluation_export_text(e.get("staffEvaluations") or []),
            eval_export_value(e.get("technicalIssue")),
            eval_export_value(e.get("nontechnicalIssue")),
            eval_checklist_export_text(e.get("checklist") or {}),
            eval_dynamic_answers_export_text(e.get("dynamicAnswers") or []),
            eval_export_value(e.get("finalNote")),
            eval_export_value((e.get("submittedAt") or "").replace("T", " ")[:16]),
        ])
    return headers, rows


@app.route("/api/evaluasi-streaming/admin/export.xlsx", methods=["GET"])
def api_streaming_evaluation_export_xlsx():
    if normalize_role_value(session.get("role") or "") not in {"admin", "super_admin"}:
        return jsonify({"success": False, "error": "Akses ditolak."}), 403
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True, buffered=True)
    try:
        ensure_streaming_evaluation_schema(cursor)
        now = datetime.now()
        start, end = eval_period_bounds(
            normalize_text(request.args.get("scale") or "month"),
            parse_required_int(request.args.get("year"), now.year),
            parse_required_int(request.args.get("month"), now.month),
            parse_required_int(request.args.get("week"), int(now.strftime("%V"))),
        )
        evaluations = eval_fetch_evaluations(
            cursor,
            start.isoformat(),
            end.isoformat(),
            normalize_text(request.args.get("kind") or "all"),
            normalize_text(request.args.get("search")),
            normalize_text(request.args.get("sort") or "date_asc"),
        )
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        wb = Workbook()
        ws = wb.active
        ws.title = "Evaluasi Streaming"
        headers, rows = eval_export_rows(evaluations)
        ws.append(headers)
        for row in rows:
            ws.append(row)
        fill = PatternFill("solid", fgColor="7F0000")
        font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 12), 65)
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(buf, as_attachment=True, download_name="hasil-evaluasi-streaming.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    finally:
        cursor.close()
        conn.close()


@app.route("/api/evaluasi-streaming/admin/export.pdf", methods=["GET"])
def api_streaming_evaluation_export_pdf():
    if normalize_role_value(session.get("role") or "") not in {"admin", "super_admin"}:
        return jsonify({"success": False, "error": "Akses ditolak."}), 403
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True, buffered=True)
    try:
        ensure_streaming_evaluation_schema(cursor)
        now = datetime.now()
        scale = normalize_text(request.args.get("scale") or "month")
        year = parse_required_int(request.args.get("year"), now.year)
        month = parse_required_int(request.args.get("month"), now.month)
        week = parse_required_int(request.args.get("week"), int(now.strftime("%V")))
        kind = normalize_text(request.args.get("kind") or "all")
        search = normalize_text(request.args.get("search"))
        sort = normalize_text(request.args.get("sort") or "date_asc")
        start, end = eval_period_bounds(scale, year, month, week)
        evaluations = eval_fetch_evaluations(cursor, start.isoformat(), end.isoformat(), kind, search, sort)

        from reportlab.lib.pagesizes import A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, KeepTogether
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER, TA_LEFT

        def esc_pdf(value) -> str:
            return html.escape(eval_export_value(value)).replace("\n", "<br/>")

        def answer_text(answer) -> str:
            if isinstance(answer, list):
                return ", ".join(normalize_text(v) for v in answer if normalize_text(v)) or "-"
            return normalize_text(answer) or "-"

        def staff_name_from_slot(slot: dict[str, object]) -> str:
            if normalize_text(slot.get("attendance")) == "not_attend":
                return "Tidak datang"
            return normalize_text(slot.get("actualMemberName") or slot.get("actualName") or slot.get("memberName") or slot.get("name")) or "-"

        def compact_period_label() -> str:
            month_names = ["", "Januari", "Februari", "Maret", "April", "Mei", "Juni", "Juli", "Agustus", "September", "Oktober", "November", "Desember"]
            if scale == "all":
                return "Semua Periode"
            if scale == "year":
                return f"Tahun {year}"
            if scale == "week":
                return f"Minggu ke-{week}, {year}"
            return f"{month_names[month] if 1 <= month <= 12 else month} {year}"

        kind_label = {"all": "Misa Biasa & Misa Besar", "misa_biasa": "Misa Biasa", "misa_besar": "Misa Besar"}.get(kind, "Misa Biasa & Misa Besar")

        buf = BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=28, rightMargin=28, topMargin=26, bottomMargin=24)
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle("eval-title", parent=styles["Title"], fontSize=18, leading=22, alignment=TA_CENTER, spaceAfter=8)
        subtitle_style = ParagraphStyle("eval-subtitle", parent=styles["BodyText"], fontSize=9, leading=12, alignment=TA_CENTER, textColor=colors.HexColor("#4b5563"), spaceAfter=14)
        section_style = ParagraphStyle("eval-section", parent=styles["Heading2"], fontSize=11, leading=14, textColor=colors.HexColor("#800000"), spaceBefore=8, spaceAfter=6)
        normal = ParagraphStyle("eval-normal", parent=styles["BodyText"], fontSize=8, leading=10, alignment=TA_LEFT)
        small = ParagraphStyle("eval-small", parent=styles["BodyText"], fontSize=7.2, leading=9, alignment=TA_LEFT)
        tiny = ParagraphStyle("eval-tiny", parent=styles["BodyText"], fontSize=6.6, leading=8.2, alignment=TA_LEFT)

        def P(value, style=small):
            return Paragraph(esc_pdf(value), style)

        def styled_table(data, col_widths=None, header=True):
            table = Table(data, colWidths=col_widths, repeatRows=1 if header else 0, hAlign="LEFT")
            base_style = [
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#d9b8b8")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 5),
                ("RIGHTPADDING", (0, 0), (-1, -1), 5),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
            if header:
                base_style += [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#800000")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ]
            table.setStyle(TableStyle(base_style))
            return table

        elements = [
            Paragraph("Hasil Evaluasi Streaming", title_style),
            Paragraph(f"Periode: {html.escape(compact_period_label())} &nbsp;&nbsp;|&nbsp;&nbsp; Jenis: {html.escape(kind_label)} &nbsp;&nbsp;|&nbsp;&nbsp; Total evaluasi: {len(evaluations)}", subtitle_style),
        ]

        if not evaluations:
            elements.append(Paragraph("Tidak ada data evaluasi sesuai filter yang dipilih.", normal))
        else:
            for idx, e in enumerate(evaluations, start=1):
                heading = f"{idx}. {eval_export_value(e.get('misaName'))} — {eval_export_value(e.get('kindLabel'))}"
                header_tbl = Table(
                    [[Paragraph(html.escape(heading), section_style), Paragraph(html.escape(eval_export_value(e.get('generalAssessment'))), section_style)]],
                    colWidths=[380, 130],
                    hAlign="LEFT",
                )
                header_tbl.setStyle(TableStyle([
                    ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#fff1f2")),
                    ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#800000")),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 7),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 7),
                ]))

                info_rows = [
                    [P("Tanggal"), P(f"{e.get('dayName') or '-'}, {e.get('dateText') or e.get('date') or '-'}")],
                    [P("Jam"), P(f"{e.get('time') or '-'} WIB")],
                    [P("Jenis Misa"), P(e.get("kindLabel"))],
                    [P("Judul / Nama Misa"), P(e.get("misaName"))],
                    [P("Pengisi Evaluasi"), P(f"{e.get('evaluatorName') or '-'} ({e.get('evaluatorRole') or '-'})")],
                    [P("Waktu Submit"), P((e.get("submittedAt") or "").replace("T", " ")[:16])],
                    [P("Penilaian Umum Streaming Keseluruhan"), P(e.get("generalAssessment"))],
                    [P("Kendala Teknis Selama Streaming"), P(e.get("technicalIssue"))],
                    [P("Kendala Misa Non-Teknis"), P(e.get("nontechnicalIssue"))],
                    [P("Checklist Kondisi Pelayanan"), P(eval_checklist_export_text(e.get("checklist") or {}))],
                    [P("Catatan Penutup / Rekomendasi"), P(e.get("finalNote"))],
                ]
                info_tbl = styled_table(info_rows, col_widths=[150, 360], header=False)
                info_tbl.setStyle(TableStyle([("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#fff7f7")), ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold")]))

                staff_header = [[P("Role", tiny), P("Petugas Jadwal", tiny), P("Petugas Aktual", tiny), P("Status Kehadiran", tiny)]]
                staff_rows = []
                for slot in e.get("staff") or []:
                    if not isinstance(slot, dict):
                        continue
                    staff_rows.append([
                        P(slot.get("role"), tiny),
                        P(slot.get("memberName") or slot.get("name"), tiny),
                        P(staff_name_from_slot(slot), tiny),
                        P("Tidak datang" if normalize_text(slot.get("attendance")) == "not_attend" else "Hadir/Digantikan", tiny),
                    ])
                if not staff_rows:
                    staff_rows = [[P("-", tiny), P("-", tiny), P("-", tiny), P("-", tiny)]]
                staff_tbl = styled_table(staff_header + staff_rows, col_widths=[95, 135, 135, 145])

                eval_header = [[P("Nama Petugas", tiny), P("Role", tiny), P("Penilaian Singkat", tiny), P("Catatan Performa", tiny)]]
                eval_rows = []
                for item in e.get("staffEvaluations") or []:
                    if not isinstance(item, dict):
                        continue
                    eval_rows.append([
                        P(item.get("memberName") or item.get("name"), tiny),
                        P(item.get("role"), tiny),
                        P(item.get("rating") or item.get("assessment"), tiny),
                        P(item.get("note") or item.get("catatan") or item.get("comment"), tiny),
                    ])
                if not eval_rows:
                    eval_rows = [[P("-", tiny), P("-", tiny), P("-", tiny), P("-", tiny)]]
                staff_eval_tbl = styled_table(eval_header + eval_rows, col_widths=[125, 95, 115, 175])

                extra_header = [[P("Nama Petugas Tambahan", tiny), P("Role Bantuan", tiny)]]
                extra_rows = []
                for item in e.get("extraStaff") or []:
                    if not isinstance(item, dict):
                        continue
                    extra_rows.append([P(item.get("name") or item.get("memberName"), tiny), P(item.get("role") or item.get("helpRole"), tiny)])
                if not extra_rows:
                    extra_rows = [[P("-", tiny), P("-", tiny)]]
                extra_tbl = styled_table(extra_header + extra_rows, col_widths=[255, 255])

                dynamic_header = [[P("Pertanyaan Tambahan", tiny), P("Jawaban", tiny)]]
                dynamic_rows = []
                for item in e.get("dynamicAnswers") or []:
                    if not isinstance(item, dict):
                        continue
                    dynamic_rows.append([P(item.get("question") or item.get("label") or item.get("text"), tiny), P(answer_text(item.get("answer")), tiny)])
                if not dynamic_rows:
                    dynamic_rows = [[P("-", tiny), P("-", tiny)]]
                dynamic_tbl = styled_table(dynamic_header + dynamic_rows, col_widths=[255, 255])

                block = [
                    header_tbl,
                    Spacer(1, 5),
                    info_tbl,
                    Spacer(1, 6),
                    Paragraph("Petugas Bertugas: Jadwal dan Aktual", section_style),
                    staff_tbl,
                    Spacer(1, 6),
                    Paragraph("Evaluasi Per Petugas", section_style),
                    staff_eval_tbl,
                    Spacer(1, 6),
                    Paragraph("Petugas Tambahan", section_style),
                    extra_tbl,
                    Spacer(1, 6),
                    Paragraph("Jawaban Pertanyaan Tambahan Dinamis", section_style),
                    dynamic_tbl,
                    Spacer(1, 12),
                ]
                elements.extend(block)
                if idx < len(evaluations):
                    elements.append(PageBreak())

        doc.build(elements)
        buf.seek(0)
        return send_file(buf, as_attachment=False, download_name="hasil-evaluasi-streaming.pdf", mimetype="application/pdf")
    finally:
        cursor.close()
        conn.close()


def create_streaming_evaluation_reminder_notifications() -> int:
    """Kirim pengingat evaluasi streaming H+1 setiap hari sampai evaluasi terisi.

    Perbaikan penting:
    - Backfill reminder harian yang terlewat. Misalnya hari ini 14 Mei dan jadwal
      10 Mei belum dievaluasi, sistem akan membuat reminder untuk 11, 12, 13,
      dan 14 Mei jika belum pernah dibuat.
    - Berlaku untuk Misa Biasa dan Misa Besar.
    - Dikirim ke setiap petugas pada sesi tersebut, termasuk user, admin, super_admin.
    - Jika satu evaluasi sudah masuk untuk sesi itu, reminder berhenti untuk semua petugas.
    - Dedupe per tanggal reminder + jadwal + member, sehingga aman dipanggil berkali-kali.
    """
    ensure_streaming_evaluation_schema()
    ensure_notifications_schema()
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True, buffered=True)
    sent = 0
    try:
        today = datetime.now().date()
        settings = eval_get_settings(cursor)
        start = eval_start_date_from_settings(settings)
        # Pengingat mulai H+1, jadi jadwal yang dicek maksimal kemarin.
        end = today - timedelta(days=1)
        if end < start:
            conn.commit()
            return 0

        schedules = eval_all_schedules(cursor, start, end, "all")
        evals = eval_evaluation_map(cursor, start.isoformat(), end.isoformat())

        for schedule in schedules:
            dt = eval_datetime_from_parts(schedule.get("date"), schedule.get("time"))
            if not dt or dt.date() >= today:
                continue
            if not eval_schedule_has_staff(schedule):
                continue
            # Jika salah satu petugas sudah submit evaluasi untuk sesi ini, stop pengingat.
            if evals.get((schedule.get("kind"), schedule.get("scheduleKey"))):
                continue

            schedule_id = normalize_text(schedule.get("id"))
            misa_name = normalize_text(schedule.get("misaName")) or "Jadwal Streaming"
            display_dt = normalize_text(schedule.get("displayDateTime")) or eval_format_period_id(schedule.get("date"), schedule.get("time"))
            kind_label = normalize_text(schedule.get("kindLabel")) or "Jadwal Streaming"

            # Backfill dari H+1 sampai hari ini. Dibatasi 31 hari agar tidak membanjiri
            # database jika ada jadwal sangat lama yang belum dievaluasi.
            first_reminder_day = dt.date() + timedelta(days=1)
            if (today - first_reminder_day).days > 31:
                first_reminder_day = today - timedelta(days=31)

            reminder_days = []
            walk_day = first_reminder_day
            while walk_day <= today:
                reminder_days.append(walk_day)
                walk_day += timedelta(days=1)

            for reminder_day in reminder_days:
                reminder_date_key = reminder_day.isoformat()
                elapsed_days = max((reminder_day - dt.date()).days, 1)
                sent_members_for_schedule: set[str] = set()
                for staff in schedule.get("staff") or []:
                    if not isinstance(staff, dict):
                        continue
                    member_id = normalize_text(staff.get("memberId"))
                    if not member_id or member_id in sent_members_for_schedule:
                        continue
                    sent_members_for_schedule.add(member_id)

                    member_role = normalize_role_value(staff.get("accountRole") or "user")
                    role_name = normalize_text(staff.get("role")) or "Petugas"
                    url = "/form-evaluasi-streaming.html" if member_role in {"admin", "super_admin"} else "/evaluasi-streaming-anggota.html"
                    dedupe_key = f"eval-reminder:{reminder_date_key}:{schedule_id}:{member_id}"

                    cursor.execute(
                        """
                        SELECT `id` FROM `notifications`
                        WHERE `type` = %s
                          AND (`data` LIKE %s OR `data` LIKE %s)
                        LIMIT 1
                        """,
                        (
                            "evaluasi",
                            f'%"dedupe_key": "{dedupe_key}"%',
                            f'%"dedupe_key":"{dedupe_key}"%',
                        ),
                    )
                    already_exists = cursor.fetchone() is not None

                    if elapsed_days == 1:
                        overdue_label = "H+1"
                    else:
                        overdue_label = f"H+{elapsed_days}"

                    title = f"Pengingat Evaluasi Streaming {overdue_label}: {misa_name}"
                    body = (
                        f"Form evaluasi untuk <b>{html.escape(misa_name)}</b> ({html.escape(kind_label)}) pada "
                        f"{html.escape(display_dt)} belum diisi. Anda terdaftar sebagai "
                        f"<b>{html.escape(role_name)}</b>. Silakan isi evaluasi streaming; cukup salah satu petugas "
                        f"pada sesi ini yang mengisi agar pengingat berhenti."
                    )
                    create_notification_once(
                        cursor,
                        "evaluasi",
                        title,
                        body,
                        url,
                        {
                            "target_user_id": member_id,
                            "notification_kind": "streaming_evaluation_reminder",
                            "schedule_id": schedule_id,
                            "schedule_kind": schedule.get("kind"),
                            "schedule_key": schedule.get("scheduleKey"),
                            "misa_type": schedule.get("kind"),
                            "misa_name": misa_name,
                            "misa_date": schedule.get("date"),
                            "misa_time": schedule.get("time"),
                            "role": role_name,
                            "reminder_date": reminder_date_key,
                            "overdue_days": elapsed_days,
                            "overdue_label": overdue_label,
                        },
                        target_role=None,
                        dedupe_key=dedupe_key,
                    )
                    if not already_exists:
                        sent += 1

        conn.commit()
        return sent
    except Exception:
        conn.rollback()
        raise
    finally:
        cursor.close()
        conn.close()


@app.route("/api/evaluasi-streaming/reminders/run", methods=["POST"])
def api_streaming_evaluation_reminders_run():
    """Endpoint manual untuk admin/super_admin menjalankan generator reminder evaluasi."""
    if normalize_role_value(session.get("role") or "") not in {"admin", "super_admin"}:
        return jsonify({"success": False, "error": "Akses ditolak."}), 403
    try:
        created = create_streaming_evaluation_reminder_notifications()
        return jsonify({"success": True, "created": created})
    except Exception as exc:
        return jsonify({"success": False, "error": str(exc)}), 400



# ===== Activity Log Backend =====
ACTIVITY_ROLE_LABELS = {
    "super_admin": "Super Admin",
    "admin": "Admin",
    "user": "Anggota",
}

ACTIVITY_MODULE_BY_PATH = [
    ("/api/password", "Autentikasi"),
    ("/login", "Autentikasi"),
    ("/logout", "Autentikasi"),
    ("/api/anggota", "Keanggotaan"),
    ("/api/admin", "Keanggotaan"),
    ("/api/membership", "Keanggotaan"),
    ("/api/sertifikat", "Sertifikat"),
    ("/api/streaming", "Tugas Streaming"),
    ("/api/misa-besar", "Tugas Streaming"),
    ("/api/request-tugas", "Request Tugas"),
    ("/api/cancel-tugas", "Pembatalan Tugas"),
    ("/api/task-exchanges", "Tukar Jadwal"),
    ("/api/evaluasi-streaming", "Evaluasi Streaming"),
    ("/api/monitoring", "Monitoring Tugas"),
    ("/api/inventory", "Inventaris"),
    ("/api/pengajuan", "Peminjaman"),
    ("/api/loan", "Peminjaman"),
    ("/api/kerusakan", "Kerusakan Barang"),
    ("/api/news", "Pengumuman"),
    ("/api/agenda", "Agenda"),
    ("/api/forms", "Form Pendaftaran"),
    ("/api/form", "Form Pendaftaran"),
    ("/api/tentang", "Konten Website"),
    ("/api/youtube", "Konten Website"),
    ("/api/google", "Konten Website"),
    ("/api/instagram", "Konten Website"),
    ("/api/carousel", "Konten Website"),
]

def ensure_activity_log_schema(cursor=None) -> None:
    own_conn = None
    own_cursor = cursor
    if own_cursor is None:
        own_conn = mysql_connection()
        own_cursor = own_conn.cursor()
    own_cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS `activity_logs` (
          `id` bigint(20) NOT NULL AUTO_INCREMENT,
          `actor_id` int(11) DEFAULT NULL,
          `actor_name` varchar(255) DEFAULT NULL,
          `actor_username` varchar(150) DEFAULT NULL,
          `actor_role` varchar(50) DEFAULT NULL,
          `action` varchar(50) NOT NULL,
          `module` varchar(120) NOT NULL,
          `description` text DEFAULT NULL,
          `route` varchar(255) DEFAULT NULL,
          `method` varchar(12) DEFAULT NULL,
          `target_user_id` int(11) DEFAULT NULL,
          `target_label` varchar(255) DEFAULT NULL,
          `meta` longtext DEFAULT NULL,
          `created_at` datetime NOT NULL DEFAULT CURRENT_TIMESTAMP,
          PRIMARY KEY (`id`),
          KEY `idx_activity_actor` (`actor_id`),
          KEY `idx_activity_role` (`actor_role`),
          KEY `idx_activity_action` (`action`),
          KEY `idx_activity_module` (`module`),
          KEY `idx_activity_created_at` (`created_at`),
          KEY `idx_activity_target_user` (`target_user_id`)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_general_ci
        """
    )
    for col, definition in {
        "actor_id": "`actor_id` int(11) DEFAULT NULL",
        "actor_name": "`actor_name` varchar(255) DEFAULT NULL",
        "actor_username": "`actor_username` varchar(150) DEFAULT NULL",
        "actor_role": "`actor_role` varchar(50) DEFAULT NULL",
        "action": "`action` varchar(50) NOT NULL DEFAULT 'ACTION'",
        "module": "`module` varchar(120) NOT NULL DEFAULT 'Sistem'",
        "description": "`description` text DEFAULT NULL",
        "route": "`route` varchar(255) DEFAULT NULL",
        "method": "`method` varchar(12) DEFAULT NULL",
        "target_user_id": "`target_user_id` int(11) DEFAULT NULL",
        "target_label": "`target_label` varchar(255) DEFAULT NULL",
        "meta": "`meta` longtext DEFAULT NULL",
        "created_at": "`created_at` datetime NOT NULL DEFAULT CURRENT_TIMESTAMP",
    }.items():
        ensure_column(own_cursor, "activity_logs", col, definition)
    if own_conn is not None:
        own_conn.commit()
        own_cursor.close()
        own_conn.close()

def activity_role_label(role_value: str) -> str:
    return ACTIVITY_ROLE_LABELS.get(normalize_role_value(role_value), "Anggota")

def activity_current_role() -> str:
    return normalize_role_value(session.get("role") or "user")

def activity_current_user_id() -> int | None:
    try:
        if session.get("user_id") in (None, ""):
            return None
        return int(session.get("user_id"))
    except (TypeError, ValueError):
        return None

def activity_visible_role_values(current_role: str | None = None) -> list[str]:
    role = normalize_role_value(current_role or session.get("role") or "user")
    if role == "super_admin":
        return ["super_admin", "admin", "user"]
    if role == "admin":
        return ["admin", "user"]
    return ["user"]

def activity_scope_where(alias: str = "l") -> tuple[str, list[object]]:
    role = activity_current_role()
    actor_id = activity_current_user_id()
    prefix = f"{alias}." if alias else ""
    if role == "super_admin":
        return "1=1", []
    if role == "admin":
        return f"({prefix}`actor_role` = 'user' OR {prefix}`actor_id` = %s)", [actor_id or 0]
    return f"{prefix}`actor_id` = %s", [actor_id or 0]

def activity_normalize_sort(sort_value: str) -> str:
    raw = normalize_text(sort_value).lower()
    if raw in {"oldest", "terlama", "date_asc", "asc"}:
        return "oldest"
    return "newest"

def activity_month_year_defaults() -> tuple[int, int]:
    now = datetime.now()
    return now.month, now.year

def activity_period_bounds_from_args(args) -> tuple[datetime, datetime, int | str, int]:
    default_month, default_year = activity_month_year_defaults()
    raw_month = normalize_text(args.get("month"))
    raw_year = normalize_text(args.get("year"))
    year = parse_required_int(raw_year, default_year) or default_year
    if raw_month.lower() in {"all", "semua", "0"}:
        start = datetime(year, 1, 1)
        end = datetime(year + 1, 1, 1)
        return start, end, "all", year
    month = parse_required_int(raw_month, default_month) or default_month
    month = min(max(month, 1), 12)
    start = datetime(year, month, 1)
    if month == 12:
        end = datetime(year + 1, 1, 1)
    else:
        end = datetime(year, month + 1, 1)
    return start, end, month, year

def activity_module_from_path(path_value: str) -> str:
    path_lower = normalize_text(path_value).lower()
    for prefix, module_name in ACTIVITY_MODULE_BY_PATH:
        if path_lower.startswith(prefix):
            return module_name
    return "Sistem"

def activity_action_from_request(method: str, path_value: str, endpoint_name: str | None = None) -> str:
    method = (method or "").upper()
    path_lower = normalize_text(path_value).lower()
    endpoint_lower = normalize_text(endpoint_name).lower()
    if path_lower.startswith("/login") and method == "POST":
        return "LOGIN"
    if path_lower.startswith("/logout"):
        return "LOGOUT"
    if "approve" in path_lower or "setujui" in path_lower or "approve" in endpoint_lower:
        return "APPROVE"
    if "reject" in path_lower or "tolak" in path_lower or "reject" in endpoint_lower:
        return "REJECT"
    if "cancel" in path_lower or "batal" in path_lower or "cancel" in endpoint_lower:
        return "CANCEL"
    if "export" in path_lower or "download" in path_lower:
        return "EXPORT"
    if "reset" in path_lower:
        return "RESET"
    if method == "POST":
        return "CREATE"
    if method in {"PUT", "PATCH"}:
        return "UPDATE"
    if method == "DELETE":
        return "DELETE"
    return "ACCESS"

def activity_description_from_request(action: str, module_name: str, path_value: str) -> str:
    if action == "LOGIN":
        return "Login ke sistem"
    if action == "LOGOUT":
        return "Logout dari sistem"
    if action == "EXPORT":
        return f"Mengekspor data {module_name}"
    if action == "CREATE":
        return f"Menambahkan atau mengirim data pada modul {module_name}"
    if action == "UPDATE":
        return f"Memperbarui data pada modul {module_name}"
    if action == "DELETE":
        return f"Menghapus data pada modul {module_name}"
    if action == "APPROVE":
        return f"Menyetujui data pada modul {module_name}"
    if action == "REJECT":
        return f"Menolak data pada modul {module_name}"
    if action == "CANCEL":
        return f"Membatalkan data pada modul {module_name}"
    if action == "RESET":
        return f"Melakukan reset pada modul {module_name}"
    return f"Aktivitas pada modul {module_name}"

def activity_log_row_to_dict(row: dict[str, object]) -> dict[str, object]:
    created = row.get("created_at")
    created_iso = created.isoformat() if hasattr(created, "isoformat") else normalize_text(created)
    role_value = normalize_role_value(row.get("actor_role") or "user")
    return {
        "id": row.get("id"),
        "createdAt": created_iso,
        "timestamp": created_iso,
        "actorId": row.get("actor_id"),
        "actorName": row.get("actor_name") or "Pengguna",
        "actorUsername": row.get("actor_username") or "",
        "actorRole": role_value,
        "actorRoleLabel": activity_role_label(role_value),
        "action": row.get("action") or "-",
        "module": row.get("module") or "-",
        "description": row.get("description") or "-",
        "route": row.get("route") or "-",
        "method": row.get("method") or "",
        "targetUserId": row.get("target_user_id"),
        "targetLabel": row.get("target_label") or "",
        "meta": safe_json_loads(row.get("meta"), {}),
    }

def insert_activity_log(
    cursor,
    *,
    actor_id=None,
    actor_name: str | None = None,
    actor_username: str | None = None,
    actor_role: str | None = None,
    action: str = "ACTION",
    module: str = "Sistem",
    description: str = "",
    route: str | None = None,
    method: str | None = None,
    target_user_id=None,
    target_label: str | None = None,
    meta: dict | None = None,
):
    ensure_activity_log_schema(cursor)
    cursor.execute(
        """
        INSERT INTO `activity_logs`
          (`actor_id`, `actor_name`, `actor_username`, `actor_role`, `action`, `module`, `description`,
           `route`, `method`, `target_user_id`, `target_label`, `meta`)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """,
        (
            actor_id,
            normalize_text(actor_name) or None,
            normalize_text(actor_username) or None,
            normalize_role_value(actor_role or "user"),
            normalize_text(action).upper() or "ACTION",
            normalize_text(module) or "Sistem",
            normalize_text(description),
            normalize_text(route) or None,
            normalize_text(method).upper() or None,
            target_user_id,
            normalize_text(target_label) or None,
            json.dumps(meta or {}, ensure_ascii=False),
        ),
    )

def record_activity(
    action: str,
    module: str,
    description: str,
    route: str | None = None,
    *,
    method: str | None = None,
    target_user_id=None,
    target_label: str | None = None,
    meta: dict | None = None,
    actor: dict[str, object] | None = None,
):
    if actor is None:
        if not session.get("logged_in"):
            return
        actor = {
            "id": activity_current_user_id(),
            "nama": session.get("nama") or session.get("username") or "Pengguna",
            "username": session.get("username") or "",
            "role": activity_current_role(),
        }
    conn = None
    cursor = None
    try:
        conn = mysql_connection()
        cursor = conn.cursor()
        insert_activity_log(
            cursor,
            actor_id=actor.get("id"),
            actor_name=actor.get("nama") or actor.get("name") or "Pengguna",
            actor_username=actor.get("username") or "",
            actor_role=actor.get("role") or "user",
            action=action,
            module=module,
            description=description,
            route=route,
            method=method or (request.method if request else ""),
            target_user_id=target_user_id,
            target_label=target_label,
            meta=meta,
        )
        conn.commit()
    except Exception as exc:
        if conn:
            try:
                conn.rollback()
            except Exception:
                pass
        print(f"[WARN] Gagal mencatat log aktivitas: {exc}")
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

def record_activity_for_member(member_id, action: str, module: str, description: str, route: str | None = None, *, meta: dict | None = None):
    conn = None
    cursor = None
    try:
        conn = mysql_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT id, nama, username, role FROM anggota WHERE id = %s LIMIT 1", (member_id,))
        member = cursor.fetchone()
        if not member:
            return
        insert_activity_log(
            cursor,
            actor_id=member.get("id"),
            actor_name=member.get("nama") or "Pengguna",
            actor_username=member.get("username") or "",
            actor_role=member.get("role") or "user",
            action=action,
            module=module,
            description=description,
            route=route,
            method=request.method if request else None,
            meta=meta,
        )
        conn.commit()
    except Exception as exc:
        if conn:
            try:
                conn.rollback()
            except Exception:
                pass
        print(f"[WARN] Gagal mencatat log aktivitas member: {exc}")
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

def activity_should_auto_log(response) -> bool:
    try:
        if response.status_code >= 400:
            return False
        path = request.path or ""
        if path.startswith("/static/") or path.startswith("/uploads/") or path == "/favicon.ico":
            return False
        if path.startswith("/api/activity-logs"):
            return False
        if path.startswith("/api/monitoring-kewajiban-tugas/"):
            return False
        if path == "/api/evaluasi-streaming/reminders/run":
            return False
        if not session.get("logged_in"):
            return False
        if request.method in {"POST", "PUT", "PATCH", "DELETE"}:
            return True
        if path in {"/logout"}:
            return True
        return False
    except Exception:
        return False

@app.after_request
def activity_auto_logger(response):
    if not activity_should_auto_log(response):
        return response
    try:
        action = activity_action_from_request(request.method, request.path, request.endpoint)
        module_name = activity_module_from_path(request.path)
        description = activity_description_from_request(action, module_name, request.path)
        record_activity(
            action,
            module_name,
            description,
            route=request.endpoint or request.path,
            method=request.method,
            meta={
                "path": request.path,
                "query": request.query_string.decode("utf-8", errors="ignore"),
                "status": response.status_code,
            },
        )
    except Exception as exc:
        print(f"[WARN] activity_auto_logger skipped: {exc}")
    return response

def require_activity_log_auth():
    if not session.get("logged_in"):
        return jsonify({"ok": False, "message": "Unauthorized"}), 401
    return None

def activity_visible_users(cursor) -> list[dict[str, object]]:
    role = activity_current_role()
    actor_id = activity_current_user_id()
    if role == "super_admin":
        cursor.execute(
            """
            SELECT id, nama, username, role
            FROM anggota
            ORDER BY FIELD(role, 'super_admin', 'admin', 'user'), nama ASC, username ASC
            """
        )
    elif role == "admin":
        cursor.execute(
            """
            SELECT id, nama, username, role
            FROM anggota
            WHERE role = 'user' OR id = %s
            ORDER BY FIELD(role, 'admin', 'user'), nama ASC, username ASC
            """,
            (actor_id or 0,),
        )
    else:
        cursor.execute(
            """
            SELECT id, nama, username, role
            FROM anggota
            WHERE id = %s
            LIMIT 1
            """,
            (actor_id or 0,),
        )
    rows = cursor.fetchall() or []
    return [
        {
            "id": row.get("id"),
            "name": row.get("nama") or row.get("username") or "Pengguna",
            "username": row.get("username") or "",
            "role": normalize_role_value(row.get("role") or "user"),
            "roleLabel": activity_role_label(row.get("role") or "user"),
            "label": f"{row.get('nama') or row.get('username') or 'Pengguna'} ({row.get('username') or '-'}) - {activity_role_label(row.get('role') or 'user')}",
        }
        for row in rows
    ]

def activity_filter_query(args, *, include_pagination: bool = True) -> tuple[str, list[object], dict[str, object]]:
    ensure_auth_schema()
    start, end, month, year = activity_period_bounds_from_args(args)
    where_clauses = []
    params: list[object] = []

    scope_sql, scope_params = activity_scope_where("l")
    where_clauses.append(scope_sql)
    params.extend(scope_params)

    where_clauses.append("l.`created_at` >= %s AND l.`created_at` < %s")
    params.extend([start, end])

    search = normalize_text(args.get("search"))
    if search:
        like = f"%{search}%"
        where_clauses.append(
            """
            (
              l.`actor_name` LIKE %s OR l.`actor_username` LIKE %s OR l.`actor_role` LIKE %s OR
              l.`action` LIKE %s OR l.`module` LIKE %s OR l.`description` LIKE %s OR l.`route` LIKE %s
            )
            """
        )
        params.extend([like, like, like, like, like, like, like])

    user_id_raw = normalize_text(args.get("user_id") or args.get("userId") or args.get("actor_id") or args.get("account"))
    if user_id_raw and user_id_raw.lower() != "all":
        try:
            user_id = int(user_id_raw)
            where_clauses.append("l.`actor_id` = %s")
            params.append(user_id)
        except ValueError:
            pass

    role_raw = normalize_text(args.get("role"))
    if role_raw and role_raw.lower() != "all":
        role_value = normalize_role_value(role_raw)
        where_clauses.append("l.`actor_role` = %s")
        params.append(role_value)

    action_raw = normalize_text(args.get("action"))
    if action_raw and action_raw.lower() != "all":
        where_clauses.append("l.`action` = %s")
        params.append(action_raw.upper())

    module_raw = normalize_text(args.get("module"))
    if module_raw and module_raw.lower() != "all":
        where_clauses.append("l.`module` = %s")
        params.append(module_raw)

    sort_mode = activity_normalize_sort(args.get("sort"))
    order_sql = "l.`created_at` ASC, l.`id` ASC" if sort_mode == "oldest" else "l.`created_at` DESC, l.`id` DESC"

    page = max(1, parse_required_int(args.get("page"), 1))
    per_page = parse_required_int(args.get("limit") or args.get("per_page") or args.get("perPage"), 20)
    if per_page not in {10, 20, 50}:
        per_page = 20

    metadata = {
        "page": page,
        "perPage": per_page,
        "sort": sort_mode,
        "month": month,
        "year": year,
        "start": start,
        "end": end,
    }
    where_sql = " AND ".join(where_clauses) if where_clauses else "1=1"
    sql = f"""
        FROM `activity_logs` l
        WHERE {where_sql}
    """
    return sql, params, metadata | {"orderSql": order_sql}

@app.route("/api/activity-logs/context", methods=["GET"])
def api_activity_logs_context():
    denied = require_activity_log_auth()
    if denied:
        return denied
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        ensure_activity_log_schema(cursor)
        current = current_user_context()
        users = activity_visible_users(cursor)
        scope_label = "Semua pengguna" if activity_current_role() == "super_admin" else ("Admin sendiri + anggota" if activity_current_role() == "admin" else "Diri sendiri")
        role_values = activity_visible_role_values()
        module_scope, module_params = activity_scope_where("l")
        cursor.execute(f"SELECT DISTINCT `module` FROM `activity_logs` l WHERE {module_scope} ORDER BY `module` ASC", tuple(module_params))
        modules = [row.get("module") for row in cursor.fetchall() or [] if row.get("module")]
        cursor.execute(f"SELECT DISTINCT `action` FROM `activity_logs` l WHERE {module_scope} ORDER BY `action` ASC", tuple(module_params))
        actions = [row.get("action") for row in cursor.fetchall() or [] if row.get("action")]
        default_month, default_year = activity_month_year_defaults()
        return jsonify({
            "ok": True,
            "currentUser": current,
            "scopeLabel": scope_label,
            "users": users,
            "roles": [{"value": value, "label": activity_role_label(value)} for value in role_values],
            "modules": modules,
            "actions": actions,
            "defaultMonth": default_month,
            "defaultYear": default_year,
        })
    finally:
        cursor.close()
        conn.close()

@app.route("/api/activity-logs", methods=["GET"])
def api_activity_logs():
    denied = require_activity_log_auth()
    if denied:
        return denied
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        ensure_activity_log_schema(cursor)
        base_sql, params, meta = activity_filter_query(request.args)
        cursor.execute(f"SELECT COUNT(*) AS total {base_sql}", tuple(params))
        total = fetch_scalar_value(cursor.fetchone(), 0) or 0
        offset = (meta["page"] - 1) * meta["perPage"]
        cursor.execute(
            f"""
            SELECT l.*
            {base_sql}
            ORDER BY {meta["orderSql"]}
            LIMIT %s OFFSET %s
            """,
            tuple(params + [meta["perPage"], offset]),
        )
        logs = [activity_log_row_to_dict(row) for row in cursor.fetchall() or []]
        return jsonify({
            "ok": True,
            "logs": logs,
            "total": total,
            "page": meta["page"],
            "perPage": meta["perPage"],
            "totalPages": max(1, (int(total) + meta["perPage"] - 1) // meta["perPage"]),
            "period": {
                "month": meta["month"],
                "year": meta["year"],
                "start": meta["start"].isoformat(),
                "end": meta["end"].isoformat(),
            },
        })
    finally:
        cursor.close()
        conn.close()

def activity_export_rows(logs: list[dict[str, object]]) -> tuple[list[str], list[list[str]]]:
    headers = ["No", "Waktu", "Pengguna", "Username", "Role", "Aksi", "Modul", "Deskripsi", "Route", "Method"]
    rows: list[list[str]] = []
    for idx, row in enumerate(logs, start=1):
        rows.append([
            str(idx),
            normalize_text(row.get("createdAt") or row.get("timestamp")),
            normalize_text(row.get("actorName")),
            normalize_text(row.get("actorUsername")),
            normalize_text(row.get("actorRoleLabel")),
            normalize_text(row.get("action")),
            normalize_text(row.get("module")),
            normalize_text(row.get("description")),
            normalize_text(row.get("route")),
            normalize_text(row.get("method")),
        ])
    return headers, rows

def fetch_activity_logs_for_export(cursor, args) -> list[dict[str, object]]:
    base_sql, params, meta = activity_filter_query(args, include_pagination=False)
    cursor.execute(
        f"""
        SELECT l.*
        {base_sql}
        ORDER BY {meta["orderSql"]}
        """,
        tuple(params),
    )
    return [activity_log_row_to_dict(row) for row in cursor.fetchall() or []]

@app.route("/api/activity-logs/export.xlsx", methods=["GET"])
def api_activity_logs_export_xlsx():
    denied = require_activity_log_auth()
    if denied:
        return denied
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        ensure_activity_log_schema(cursor)
        logs = fetch_activity_logs_for_export(cursor, request.args)
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = "Log Aktivitas"
        headers, rows = activity_export_rows(logs)
        ws.append(headers)
        for row in rows:
            ws.append(row)
        fill = PatternFill("solid", fgColor="800000")
        font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for row_cells in ws.iter_rows(min_row=2):
            for cell in row_cells:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for column_cells in ws.columns:
            max_length = 0
            column_letter = column_cells[0].column_letter
            for cell in column_cells:
                max_length = max(max_length, len(str(cell.value or "")))
            ws.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 45)

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        record_activity("EXPORT", "Log Aktivitas", "Mengekspor log aktivitas ke Excel", "activity_logs.export_xlsx", method="GET")
        return send_file(
            buffer,
            as_attachment=True,
            download_name=f"log-aktivitas-{datetime.now().strftime('%Y%m%d%H%M')}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    finally:
        cursor.close()
        conn.close()

@app.route("/api/activity-logs/export.pdf", methods=["GET"])
def api_activity_logs_export_pdf():
    denied = require_activity_log_auth()
    if denied:
        return denied
    conn = mysql_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        ensure_activity_log_schema(cursor)
        logs = fetch_activity_logs_for_export(cursor, request.args)
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Spacer, Table, TableStyle, Paragraph

        def p(value, style):
            return Paragraph(html.escape(normalize_text(value) or "-").replace("\n", "<br/>"), style)

        headers, rows = activity_export_rows(logs)
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=8 * mm, leftMargin=8 * mm, topMargin=10 * mm, bottomMargin=10 * mm)
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle("activity-title", parent=styles["Title"], fontName="Helvetica-Bold", fontSize=16, textColor=colors.HexColor("#800000"), alignment=1)
        cell_style = ParagraphStyle("activity-cell", parent=styles["BodyText"], fontSize=7, leading=9)
        header_style = ParagraphStyle("activity-header", parent=cell_style, fontName="Helvetica-Bold", textColor=colors.white)

        data = [[p(h, header_style) for h in headers]]
        for row in rows:
            data.append([p(cell, cell_style) for cell in row])

        table = Table(data, repeatRows=1, colWidths=[10*mm, 34*mm, 30*mm, 28*mm, 20*mm, 20*mm, 26*mm, 72*mm, 42*mm, 16*mm])
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#800000")),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#e9d5d5")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#fff7f7")]),
            ("LEFTPADDING", (0, 0), (-1, -1), 3),
            ("RIGHTPADDING", (0, 0), (-1, -1), 3),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]))
        story = [
            Paragraph("Laporan Log Aktivitas", title_style),
            Spacer(1, 6),
            Paragraph(f"Total log: {len(rows)}", styles["Normal"]),
            Spacer(1, 8),
            table,
        ]
        doc.build(story)
        buffer.seek(0)
        record_activity("EXPORT", "Log Aktivitas", "Mengekspor log aktivitas ke PDF", "activity_logs.export_pdf", method="GET")
        return send_file(
            buffer,
            as_attachment=False,
            download_name=f"log-aktivitas-{datetime.now().strftime('%Y%m%d%H%M')}.pdf",
            mimetype="application/pdf",
        )
    finally:
        cursor.close()
        conn.close()


_EVAL_REMINDER_AUTO_RUN_DATE: str | None = None


@app.before_request
def auto_run_streaming_evaluation_reminders_once_per_day():
    """Menjalankan pengingat evaluasi sekali per hari saat aplikasi menerima request.

    Catatan: untuk produksi, endpoint `/api/evaluasi-streaming/reminders/run` tetap bisa dipanggil
    lewat cron harian. Hook ini membantu mode lokal/dev agar notifikasi tetap dibuat otomatis
    tanpa harus membuka dashboard tertentu terlebih dahulu.
    """
    global _EVAL_REMINDER_AUTO_RUN_DATE
    path = request.path or ""
    if path.startswith("/static/") or path.startswith("/uploads/") or path == "/favicon.ico":
        return
    if path == "/api/evaluasi-streaming/reminders/run":
        return
    today_key = datetime.now().date().isoformat()
    if _EVAL_REMINDER_AUTO_RUN_DATE == today_key:
        return
    try:
        create_streaming_evaluation_reminder_notifications()
        _EVAL_REMINDER_AUTO_RUN_DATE = today_key
    except Exception as exc:
        # Jangan sampai halaman utama gagal hanya karena generator notif bermasalah.
        print(f"[WARN] Gagal auto-run pengingat evaluasi streaming: {exc}")


if __name__ == "__main__":
    try:
        ensure_auth_schema()
        ensure_news_schema()
        ensure_agenda_schema()
        ensure_notifications_schema()
        ensure_activity_log_schema()
        ensure_misa_besar_schema()
    except Exception as exc:
        print(f"[WARN] MySQL bootstrap skipped: {exc}")
    app.run(debug=True)